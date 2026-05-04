from odoo import models, fields
import base64
import logging
import io
import requests
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from PIL import Image
import json
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openai import OpenAI
import re
import fitz
 

_logger = logging.getLogger(__name__)

class ProductTemplate(models.Model):
    _inherit = 'product.template'

    vendor_id = fields.Many2one(
        'res.partner',
        string="Vendor"
    )


# ✅ Extend existing model
class ResPartner(models.Model):
    _inherit = 'res.partner'

    #Vendor user role
    is_vendor_user = fields.Boolean(
        string="Vendor User",
        default=False
    )


class VendorImportJob(models.Model):

    _name = "vendor.import.job"
    _description = "Vendor Import Job"

    partner_id = fields.Many2one("res.partner", string="Vendor")  # ✅ LINK instead

    name = fields.Char(default="Vendor Data Import")
    extra_info = fields.Text()

    pdf_file = fields.Binary()
    excel_file = fields.Binary()
    logo_file = fields.Binary()

    extracted_text = fields.Text()
    current_page = fields.Integer(
        string="Current PDF Page",
        default=0
    )
    total_pages = fields.Integer(string="Total Pages", default=0)
    last_ai_page = fields.Integer(string="Last AI Page", default=0)
    ai_response = fields.Text()
    priority = fields.Integer(default=10)
    excel_created_index = fields.Integer(
        string="Excel Created Index",
        default=0
    )

    apify_run_id = fields.Char()
    apify_dataset_id = fields.Char()
   
    last_processed_product_index = fields.Integer(default=0)
    last_created_page = fields.Integer(default=0)
    lock = fields.Boolean(default=False)
    is_excel_parsed = fields.Boolean(default=False)
    excel_ai_index = fields.Integer(default=0)
    upload_signature = fields.Char(string="Upload Signature")
    processed_group_ids = fields.Text(default="[]")

    url_total_batches = fields.Integer(default=0)
    url_batch_index = fields.Integer(default=0)
    data_url = fields.Char()
    url_parse_index = fields.Integer(
        string="URL Parse Index",
        default=0
    )
    url_blocks_json = fields.Text(
        string="URL Blocks JSON"
    )
    

    
    excel_parse_index = fields.Integer(
        default=0
    )

    source_type = fields.Selection([
        ("pdf", "PDF"),
        ("excel", "Excel"),
        ("url", "URL"),
    ])


    state = fields.Selection([
        ('draft', 'Draft'),
        ('processing', 'Processing'),
        ('review', 'Vendor Review'),
        ('done', 'Completed'),
        ('error', 'Error'),
        ('failed', 'Failed'),

         #New
        ('url_scraping', 'URL Scraping'),
        ('url_ai', 'URL AI'),
        ('url_creating', 'URL Creating'),

        ('pdf_extracting', 'PDF Extracting'),
        ('pdf_ai', 'PDF AI'),
        ('pdf_creating', 'PDF Creating'),

        ('excel_parsing', 'Excel Parsing'),
        ('excel_ai', 'Excel AI'),
        ('excel_creating', 'Excel Creating'),

    ], default='draft')


     #============================= MAIN FLOW (process steps) =====================

    def process_import(self):

        _logger.warning(f"PROCESS START → Job {self.id}")

        try:
            # if self.state == 'review':
            #     self.state = 'processing'

            # if self.state == 'draft':
            #     self.state = 'processing'

            self._process_step()

        except Exception as e:
            _logger.error(f"PROCESS FAILED → {str(e)}")
            self.state = "failed"


    #=============Safe commit======================================================
    def _safe_commit_progress(self):

        try:

            self.flush_recordset()

        except Exception as flush_error:

            _logger.warning(
                f"FLUSH FAILED → {flush_error}"
            )

        try:

            self.env.cr.commit()

        except Exception as commit_error:

            _logger.warning(
                f"COMMIT FAILED → {commit_error}"
            )


    #============Processing Jobs===================================================

    def _process_step(self):

        import json
        import re

        self.ensure_one()


        # =================================================
        # SAFETY
        # =================================================

        if self.pdf_file and self.excel_file:

            _logger.error(
                "[PROCESS STEP] BOTH PDF AND EXCEL PROVIDED"
            )

            self.state = "failed"

            self.flush_recordset()
            self.env.cr.commit()

            return


        _logger.warning(
            f"[PROCESS STEP] → state={self.state}"
        )


        # =================================================
        # DONE
        # =================================================

        if self.state == 'done':

            _logger.warning(
                f"JOB {self.id} ALREADY DONE ✅"
            )

            return


        # =================================================
        # REVIEW RECOVERY
        # =================================================

        if self.state == 'review':

            _logger.warning(
                "REVIEW → RESET TO START"
            )

            if self.pdf_file:

                self.state = 'pdf_extracting'

            elif self.excel_file and not self.pdf_file:

                self.state = 'excel_parsing'

            elif self.data_url:

                self.state = 'url_scraping'


            self.flush_recordset()
            self.env.cr.commit()

            return


        # =================================================
        # URL FLOW
        # =================================================
        

        if self.data_url:

            _logger.warning(
                "FLOW → URL"
            )


            # =============================================
            # START
            # =============================================

            if self.state == 'draft':

                _logger.warning(
                    "[URL FLOW] START"
                )

                self.state = 'url_scraping'

                self._safe_commit_progress()

                return


            # =============================================
            # SCRAPE
            # =============================================

            if self.state == 'url_scraping':

                # =========================================
                # RECOVERY
                # =========================================

                if self.url_blocks_json:

                    _logger.warning(

                        "[URL RECOVERY] "

                        "USING SAVED BLOCKS"
                    )

                    self.state = 'url_ai'

                    self._safe_commit_progress()

                    return


                _logger.warning(

                    f"[URL SCRAPE] "

                    f"SENDING TO APIFY "

                    f"| {self.data_url}"
                )


                previous_extract = bool(
                    self.extracted_text
                )


                result = self.parse_url()


                # =========================================
                # APIFY STILL PROCESSING
                # =========================================

                if result is True:

                    _logger.warning(

                        "[APIFY STATUS] "

                        "WAITING FOR RESPONSE"
                    )

                    return


                # =========================================
                # APIFY RESPONSE READY
                # =========================================

                _logger.warning(
                    "[APIFY STATUS] RESPONSE RECEIVED"
                )


                if (

                    self.extracted_text

                    and

                    not previous_extract

                ):

                    _logger.warning(
                        "URL EXTRACTION SUCCESS → url_ai"
                    )

                    self.state = 'url_ai'

                    self._safe_commit_progress()

                    return


                # =========================================
                # FAILED
                # =========================================

                if self.state == 'failed':

                    _logger.warning(
                        "[URL SCRAPE FAILED]"
                    )

                    self._safe_commit_progress()

                    return


                _logger.warning(
                    "[URL SCRAPE] NO DATA EXTRACTED"
                )

                self.state = 'failed'

                self._safe_commit_progress()

                return


            # =============================================
            # AI
            # =============================================

            if self.state == 'url_ai':

                previous_batch = (
                    self.url_batch_index or 0
                )


                _logger.warning(

                    f"[URL AI START] "

                    f"batch={previous_batch}"
                )


                try:

                    self.send_to_openai_url()

                except Exception as e:

                    _logger.exception(

                        f"URL AI FAILED → {str(e)}"
                    )

                    self.state = 'failed'


                    self._safe_commit_progress()

                    return


                new_batch = (
                    self.url_batch_index or 0
                )


                _logger.warning(

                    f"[URL AI CHECK] "

                    f"{previous_batch} -> {new_batch}"
                )


                # =========================================
                # PROGRESS DETECTED
                # =========================================

                if new_batch > previous_batch:

                    _logger.warning(
                        "[URL AI] PROGRESS SAVED"
                    )

                elif self.state != 'url_creating':

                    _logger.warning(
                        "[URL AI] NO PROGRESS DETECTED"
                    )

                    self.state = 'failed'


                self._safe_commit_progress()

                return


            # =============================================
            # CREATE
            # =============================================

            if self.state == 'url_creating':

                if not self.ai_response:

                    self.state = 'failed'

                    _logger.warning(
                        "URL CREATE FAILED → NO AI RESPONSE"
                    )


                    self._safe_commit_progress()

                    return


                previous_index = (
                    self.last_processed_product_index or 0
                )


                _logger.warning(

                    f"[URL CREATE START] "

                    f"{previous_index}"
                )


                try:

                    self.create_products_url()

                except Exception as e:

                    _logger.exception(

                        f"URL CREATE FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self._safe_commit_progress()

                    return


                new_index = (
                    self.last_processed_product_index or 0
                )


                _logger.warning(

                    f"[URL CREATE CHECK] "

                    f"{previous_index} -> {new_index}"
                )


                try:

                    total = len(
                        json.loads(
                            self.ai_response or "[]"
                        )
                    )

                except Exception:

                    total = 0


                _logger.warning(

                    f"[URL TOTAL PRODUCTS] "

                    f"{total}"
                )


                if new_index >= total:

                    self.state = 'done'

                    _logger.warning(
                        "URL COMPLETE ✅"
                    )

                elif new_index > previous_index:

                    self.state = 'url_creating'

                    _logger.warning(
                        "[URL CREATE] CONTINUE"
                    )

                else:

                    _logger.warning(
                        "[URL CREATE] NO PROGRESS"
                    )

                    self.state = 'failed'


                self._safe_commit_progress()

                return


        # =================================================
        # EXCEL FLOW
        # =================================================

        if self.excel_file and not self.pdf_file:

            _logger.warning(
                "FLOW → EXCEL"
            )


            # =============================================
            # START
            # =============================================

            if self.state == 'draft':

                self.state = 'excel_parsing'

                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # PARSE
            # =============================================

            if self.state == 'excel_parsing':

                previous_index = (
                
                    self.excel_parse_index or 0
                )

                _logger.warning(

                    f"[EXCEL PARSE START] "

                    f"previous_index={previous_index}"
                )


                self.parse_excel()


                new_index = (
                   
                    self.excel_parse_index or 0
                )


                _logger.warning(

                    f"[EXCEL PARSE CHECK] "

                    f"{previous_index} -> {new_index}"
                )


                # =========================================
                # NEW ROWS FOUND
                # =========================================

                if new_index > previous_index:

                    _logger.warning(
                        "[EXCEL PARSE] NEW BATCH READY → excel_ai"
                    )

                    # reset AI/create cycle
                    self.excel_ai_index = 0
                    self.excel_created_index = 0
                    self.ai_response = False

                    self.state = 'excel_ai'

                else:

                    _logger.warning(
                        "[EXCEL PARSE] NO NEW ROWS"
                    )


                    # parser itself decides completion
                    if self.is_excel_parsed:

                        _logger.warning(
                            "[EXCEL PARSE] FULLY COMPLETE"
                        )

                        self.state = 'done'

                    else:

                        self.state = 'excel_parsing'


                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # AI
            # =============================================

            if self.state == 'excel_ai':

                try:

                    self.send_to_openai_excel()

                except Exception as e:

                    _logger.exception(

                        f"EXCEL AI FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self.flush_recordset()
                    self.env.cr.commit()

                    return


                total_rows = (
                    self.last_processed_product_index
                    or 0
                )


                _logger.warning(

                    f"[EXCEL AI TOTAL ROWS] "

                    f"{total_rows}"
                )


                _logger.warning(

                    f"[EXCEL AI STATE] "

                    f"{self.state}"
                )


                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # CREATE
            # =============================================

            if self.state == 'excel_creating':

                try:

                    self.create_products_excel()

                except Exception as e:

                    _logger.exception(

                        f"EXCEL CREATE FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self.flush_recordset()
                    self.env.cr.commit()

                    return


                _logger.warning(

                    f"[EXCEL CREATE STATE] "

                    f"{self.state}"
                )


                self.flush_recordset()
                self.env.cr.commit()

                return


        # =================================================
        # PDF FLOW
        # =================================================

        elif self.pdf_file:

            _logger.warning(
                "FLOW → PDF"
            )


            # =============================================
            # START
            # =============================================

            if self.state == 'draft':

                self.state = 'pdf_extracting'

                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # EXTRACT
            # =============================================

            if self.state == 'pdf_extracting':

                try:

                    self.extract_pdf()

                except Exception as e:

                    _logger.exception(

                        f"PDF EXTRACT FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self.flush_recordset()
                    self.env.cr.commit()

                    return


                if (

                    (self.current_page or 0)

                    <

                    (self.total_pages or 0)

                ):

                    _logger.warning(

                        f"PDF EXTRACTION CONTINUES "

                        f"→ PAGE "

                        f"{self.current_page}/"

                        f"{self.total_pages}"
                    )

                    self.state = 'pdf_extracting'

                else:

                    _logger.warning(
                        "PDF EXTRACTION COMPLETE → pdf_ai"
                    )

                    self.state = 'pdf_ai'


                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # PDF AI
            # =============================================

            if self.state == 'pdf_ai':

                try:

                    self.send_to_openai_pdf()

                except Exception as e:

                    _logger.exception(

                        f"PDF AI FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self.flush_recordset()
                    self.env.cr.commit()

                    return


                page_total = self.env[
                    'vendor.import.page'
                ].search_count([

                    ('job_id', '=', self.id)

                ])


                _logger.warning(

                    f"[PDF AI CHECK] "

                    f"{self.last_ai_page}/"

                    f"{page_total}"
                )


                if (

                    (self.last_ai_page or 0)

                    <

                    page_total

                ):

                    _logger.warning(

                        f"PDF AI CONTINUES "

                        f"→ {self.last_ai_page}/"

                        f"{page_total}"
                    )

                    self.state = 'pdf_ai'

                else:

                    _logger.warning(
                        "PDF AI COMPLETE → pdf_creating"
                    )

                    self.state = 'pdf_creating'


                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # PDF CREATE
            # =============================================

            if self.state == 'pdf_creating':

                try:

                    self.create_products_pdf()

                except Exception as e:

                    _logger.exception(

                        f"PDF CREATE FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self.flush_recordset()
                    self.env.cr.commit()

                    return


                try:

                    total_ai_pages = len(
                        json.loads(
                            self.ai_response or "[]"
                        )
                    )

                except Exception:

                    total_ai_pages = 0


                if (

                    (self.last_created_page or 0)

                    <

                    total_ai_pages

                ):

                    _logger.warning(

                        f"PDF CREATE CONTINUES "

                        f"→ {self.last_created_page}/"

                        f"{total_ai_pages}"
                    )

                    self.state = 'pdf_creating'

                else:

                    _logger.warning(
                        "PDF COMPLETE ✅"
                    )

                    self.state = 'done'


                self.flush_recordset()
                self.env.cr.commit()

                return


    #------------parse url------------------------------------

    def parse_url(self):

        import json

        _logger.warning(f"APIFY SCRAPE → {self.data_url}")

        raw_data = self._run_apify_actor(self.data_url)

        # =====================================================
        # APIFY STILL RUNNING
        # =====================================================

        if raw_data is None:

            _logger.warning(
                "APIFY NOT READY → WAIT NEXT CRON"
            )

            self.state = "url_scraping"

            return True

        # =====================================================
        # EMPTY RAW RESPONSE
        # =====================================================

        if not raw_data:

            _logger.error(
                "APIFY FAILED → EMPTY DATASET"
            )

            self.state = "failed"

            return

        # =====================================================
        # SAFE DEBUG LOG
        # =====================================================

        try:

            _logger.warning(
                f"RAW APIFY ITEMS → {len(raw_data)}"
            )

        except Exception:
            pass

        # =====================================================
        # HANDLE STRUCTURED RESPONSES
        # =====================================================

        first = raw_data[0] if raw_data else {}

        response_type = first.get("type")

        # =====================================================
        # BLOCKED
        # =====================================================

        if response_type == "BLOCKED":

            reason = first.get(
                "reason",
                "Unknown block detected"
            )

            status_code = first.get(
                "status_code"
            )

            _logger.error(
                f"URL BLOCKED → {reason}"
            )

            if status_code:
                _logger.error(
                    f"BLOCK STATUS CODE → {status_code}"
                )

            self.state = "failed"

            return

        # =====================================================
        # EMPTY
        # =====================================================

        if response_type == "EMPTY":

            reason = first.get(
                "reason",
                "No products extracted"
            )

            debug = first.get("debug", {})

            _logger.error(
                f"URL EXTRACTION EMPTY → {reason}"
            )

            if debug:

                _logger.error(
                    f"PAGE TITLE → {debug.get('title')}"
                )

                _logger.error(
                    f"IMAGES FOUND → {debug.get('images_found')}"
                )

                _logger.error(
                    f"LINKS FOUND → {debug.get('links_found')}"
                )

                _logger.error(
                    f"POSSIBLE PRODUCT BLOCKS → "
                    f"{debug.get('possible_product_blocks')}"
                )

                _logger.error(
                    f"COOKIE DETECTED → "
                    f"{debug.get('cookie_detected')}"
                )

                preview = debug.get(
                    'body_preview',
                    ''
                )

                _logger.error(
                    f"BODY PREVIEW → {preview[:300]}"
                )

            self.state = "failed"

            return

        # ===================================================
        # PRODUCTS
        # ===================================================

        structured_data = []

        for block in raw_data:

            # ==============================================
            # FORMAT 1 → ORIGINAL EB FORMAT
            # ==============================================

            if block.get("text"):

                structured_data.append({
                    "text": block.get("text"),
                    "image": block.get("image")
                })

                continue

            # ==============================================
            # FORMAT 2 → STRUCTURED FORMAT
            # ==============================================

            if block.get("type") == "PRODUCTS":

                items = block.get("items", [])

                if not items:
                    continue

                structured_data.extend(items)

        # =====================================================
        # NO PRODUCTS AFTER PARSE
        # =====================================================

        if not structured_data:

            _logger.error(
                "NO VALID PRODUCTS FOUND AFTER PARSING"
            )

            self.state = "failed"

            return

        # =====================================================
        # LIMIT SIZE (VERY IMPORTANT)
        # =====================================================

        # structured_data = structured_data[:40]

        # ============================================
        # URL BATCHING
        # ============================================

        BATCH_SIZE = 40

        start = (
            self.url_parse_index or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(structured_data)
        )


        structured_data = structured_data[
            start:end
        ]


        _logger.warning(

            f"[URL PARSE BATCH] "

            f"{start} -> {end} "

            f"| total={len(normalized if 'normalized' in locals() else structured_data)}"
        )


        # =====================================================
        # NORMALIZE
        # =====================================================

        normalized = self._normalize_url_data(
            structured_data
        )

        # ============================================
        # SAVE URL PARSE PROGRESS
        # ============================================

        self.url_parse_index = end


        _logger.warning(

            f"[URL PARSE SAVE] "

            f"{self.url_parse_index}"
        )


        if not normalized:

            _logger.error(
                "NORMALIZATION FAILED → EMPTY DATA"
            )

            self.state = "failed"

            return

        # =====================================================
        # STORE SAFELY
        # =====================================================

        #payload = json.dumps(normalized)

        # 🔥 LIMIT STORAGE SIZE
        #self.extracted_text = payload[:50000]

        payload = json.dumps(normalized)

        # ============================================
        # PERSIST URL BLOCKS
        # ============================================

        self.url_blocks_json = payload

        # compatibility
        self.extracted_text = payload[:50000]


        _logger.warning(

            f"[URL STORE] "

            f"saved_blocks={len(normalized)}"
        )


        _logger.warning(
            f"APIFY DONE → {len(normalized)} ITEMS"
        )

        # =====================================================
        # MOVE TO NEXT STEP
        # =====================================================

        # self.state = "url_ai"
        if self.url_parse_index >= len(raw_data):

            _logger.warning(
                "[URL PARSE] FINAL BATCH READY"
            )

        else:

            _logger.warning(
                "[URL PARSE] MORE BATCHES REMAIN"
            )


        self.state = "url_ai"

    #------excel parsing method---------------
    
    def parse_excel(self):

        _logger.warning(
            "EXCEL → START PARSING (BATCH MODE)"
        )

        excel_bytes = base64.b64decode(
            self.excel_file
        )

        wb = load_workbook(
            filename=BytesIO(excel_bytes)
        )

        headers = {
            "User-Agent": "Mozilla/5.0"
        }

        pages = []

        # =====================================
        # SAFE BATCH CONTROL
        # =====================================

        BATCH_SIZE = 8
        MAX_BUFFER = 150


        start_index = (
            self.excel_parse_index or 0
        )


        current_count = 0

        global_index = 0

        _logger.warning(
            f"EXCEL RESUME FROM INDEX "
            f"→ {start_index}"
        )


        # =====================================
        # TOTAL ROWS
        # =====================================

        total_rows = 0

        for sheet in wb.worksheets:

            for idx, row in enumerate(

                sheet.iter_rows()

            ):

                if idx == 0:
                    continue

                row_text_parts = [

                    str(cell.value or "").strip()

                    for cell in row

                    if str(
                        cell.value or ""
                    ).strip()
                ]

                if not row_text_parts:
                    continue

                total_rows += 1


        _logger.warning(
            f"[DEBUG] REAL TOTAL ROWS "
            f"→ {total_rows}"
        )


        # =====================================
        # MAIN LOOP
        # =====================================

        for sheet in wb.worksheets:

            _logger.warning(
                f"PROCESSING SHEET → "
                f"{sheet.title}"
            )

            image_loader = SheetImageLoader(
                sheet
            )


            for idx, row in enumerate(

                sheet.iter_rows()

            ):

                if idx == 0:
                    continue


                row_text_parts = [

                    str(cell.value or "").strip()

                    for cell in row

                    if str(
                        cell.value or ""
                    ).strip()
                ]


                if not row_text_parts:
                    continue


                # =================================
                # GLOBAL INDEX TRACKING
                # =================================

                global_index += 1


                # =================================
                # SKIP OLD ROWS
                # =================================

                if global_index <= start_index:
                    continue


                # =================================
                # BATCH LIMIT
                # =================================

                if current_count >= BATCH_SIZE:

                    _logger.warning(
                        "BATCH LIMIT REACHED "
                        "→ NEXT CRON"
                    )

                    break


                # =================================
                # PRICE/STOCK DETECTION
                # =================================

                price = ""
                stock = ""

                numeric_candidates = []


                for col_idx, cell in enumerate(row):

                    raw_val = str(
                        cell.value or ""
                    ).strip()

                    if not raw_val:
                        continue


                    # skip ranges
                    if (
                        "-" in raw_val
                        and not raw_val.startswith("-")
                    ):
                        continue


                    try:

                        clean = raw_val.replace(
                            ",",
                            "."
                        )

                        clean = re.sub(

                            r"[^\d.]",

                            "",

                            clean
                        )


                        if not clean:
                            continue


                        num = float(clean)

                        is_real_decimal = False


                        if "." in clean:

                            decimal_part = (
                                clean.split(".")[-1]
                            )

                            if decimal_part not in [

                                "0",

                                "00"
                            ]:

                                is_real_decimal = True


                        numeric_candidates.append({

                            "col": col_idx,

                            "num": num,

                            "raw": raw_val,

                            "is_decimal":
                                is_real_decimal
                        })

                    except:
                        continue


                # =================================
                # PRICE
                # =================================

                price_candidates = [

                    x for x in numeric_candidates

                    if (

                        x["is_decimal"]

                        and

                        0 < x["num"] < 1000
                    )
                ]


                best_price = None


                if price_candidates:

                    best_price = sorted(

                        price_candidates,

                        key=lambda x: x["col"]

                    )[-1]

                    price = str(
                        best_price["num"]
                    )


                # =================================
                # STOCK
                # =================================

                if best_price:

                    price_col = (
                        best_price["col"]
                    )

                    stock_candidates = []


                    for item in numeric_candidates:

                        if item["is_decimal"]:
                            continue


                        val = item["num"]


                        if val > 9999:
                            continue


                        if item["col"] >= price_col:
                            continue


                        stock_candidates.append(
                            item
                        )


                    if stock_candidates:

                        best_stock = max(

                            stock_candidates,

                            key=lambda x: x["num"]
                        )

                        stock = str(

                            int(
                                best_stock["num"]
                            )
                        )


                _logger.warning(
                    f'''
                    EXCEL RAW ROW →

                    TEXT=
                    {" | ".join(row_text_parts)}

                    PRICE={price}

                    STOCK={stock}
                    '''
                )


                # =================================
                # ROW TEXT
                # =================================

                row_text = f"""
                ROW_DATA:
                {" | ".join(row_text_parts)}

                RULE:
                - THIS IS EXACTLY ONE PRODUCT
                - DO NOT SPLIT THIS ROW
                - THIS ROW MAY BE A VARIANT
                - USE SIMILAR ID/SKU
                """


                row_images = []


                # =================================
                # EMBEDDED IMAGE
                # =================================

                for cell in row:

                    try:

                        if image_loader.image_in(
                            cell.coordinate
                        ):

                            pil_img = (
                                image_loader.get(
                                    cell.coordinate
                                )
                            )

                            buffer = BytesIO()

                            pil_img.save(
                                buffer,
                                format="JPEG"
                            )

                            img_base64 = (
                                base64.b64encode(

                                    buffer.getvalue()

                                ).decode("utf-8")
                            )

                            row_images.append(
                                img_base64
                            )

                            break

                    except:
                        continue


                # =================================
                # URL IMAGE
                # =================================

                if not row_images:

                    for cell in row:

                        val = str(
                            cell.value or ""
                        ).strip()

                        if val.startswith("http"):

                            try:

                                response = requests.get(

                                    val,

                                    headers=headers,

                                    timeout=5
                                )

                                if (

                                    response.status_code
                                    == 200

                                    and

                                    "image"

                                    in response.headers.get(
                                        "Content-Type",
                                        ""
                                    )
                                ):

                                    img_base64 = (
                                        base64.b64encode(

                                            response.content

                                        ).decode("utf-8")
                                    )

                                    row_images.append(
                                        img_base64
                                    )

                                    break

                            except:
                                continue


                # =================================
                # STORE
                # =================================

                pages.append({

                    "page": global_index,

                    "text": row_text,

                    "images": row_images,

                    "row_index": global_index,

                    "price": price,

                    "stock": stock,
                })


                current_count += 1


                # =================================
                # MEMORY SAFETY
                # =================================

                if len(pages) >= MAX_BUFFER:

                    _logger.warning(
                        f"EXCEL SAFETY BREAK "
                        f"→ {len(pages)} rows"
                    )

                    break


            if (
                current_count >= BATCH_SIZE
                or
                len(pages) >= MAX_BUFFER
            ):
                break


        # =====================================
        # STORE
        # =====================================

        existing = []

        if self.extracted_text:

            try:

                existing = json.loads(
                    self.extracted_text
                )

            except:
                existing = []


        # combined = existing + pages
        
        existing_map = {

            item.get("row_index"): item

            for item in existing
        }


        for item in pages:

            existing_map[
                item.get("row_index")
            ] = item


        combined = sorted(

            existing_map.values(),

            key=lambda x: x.get(
                "row_index",
                0
            )
        )
    

        self.extracted_text = json.dumps(
            combined
        )


        # =====================================
        # SAVE PROGRESS
        # =====================================

        new_index = (
            start_index
            +
            current_count
        )

        self.excel_parse_index = (
            new_index
        )

        _logger.warning(

            f"[EXCEL PARSE INDEX SAVE] "

            f"{self.excel_parse_index}"
        )


        # =====================================
        # DEBUG
        # =====================================

        remaining = max(
            total_rows - new_index,
            0
        )

        progress = round(

            (new_index / total_rows) * 100,

            2

        ) if total_rows else 0


        _logger.warning(
            f"[DEBUG] CURRENT INDEX "
            f"→ {new_index}"
        )

        _logger.warning(
            f"[DEBUG] REMAINING ROWS "
            f"→ {remaining}"
        )

        _logger.warning(
            f"[DEBUG] PROGRESS "
            f"→ {progress}%"
        )

        _logger.warning(
            f"EXCEL NEW INDEX "
            f"→ {new_index}"
        )

        _logger.warning(
            f"EXCEL BATCH STORED "
            f"→ {len(pages)} rows"
        )


        # =====================================
        # COMPLETION
        # =====================================

        if new_index >= total_rows:

            _logger.warning(
                "EXCEL → PARSING COMPLETED ✅"
            )

            self.is_excel_parsed = True

            self.state = "excel_ai"

        else:

            _logger.warning(
                "EXCEL → MORE DATA REMAIN "
                "→ NEXT CRON"
            )

            self.state = "excel_parsing"


        wb.close()


    # ---------------- Extract PDF ----------------
 
    def extract_pdf(self):

        import gc
        import json
        import io
        import re
        import fitz
        import base64
        import requests

        _logger.warning(
            f"[PDF EXTRACT] START "
            f"| job={self.id}"
        )

        MAX_RETRIES = 3

        # balanced batch size
        BATCH_SIZE = 3

        doc = None

        try:

            pdf_bytes = base64.b64decode(
                self.pdf_file
            )

        except Exception as e:

            _logger.exception(
                f"[PDF EXTRACT ERROR] "
                f"PDF DECODE FAILED "
                f"| {str(e)}"
            )

            self.state = "failed"

            return


        # =========================================
        # OPEN PDF
        # =========================================

        try:

            doc = fitz.open(
                stream=pdf_bytes,
                filetype="pdf"
            )

        except Exception as e:

            _logger.exception(
                f"[PDF EXTRACT ERROR] "
                f"PDF OPEN FAILED "
                f"| {str(e)}"
            )

            self.state = "failed"

            return


        try:

            total_pages = len(doc)

            self.total_pages = total_pages


            _logger.warning(
                f"[PDF EXTRACT] "
                f"TOTAL PAGES={total_pages}"
            )


            # =====================================
            # CRASH SAFE RECOVERY
            # =====================================

            existing_pages = self.env[
                'vendor.import.page'
            ].search([

                ('job_id', '=', self.id)

            ], order='page_number desc', limit=1)


            if existing_pages:

                # move to NEXT page
                start_page = (
                    existing_pages.page_number
                )

                _logger.warning(
                    f"[PDF RECOVERY] "
                    f"LAST SAVED PAGE="
                    f"{existing_pages.page_number}"
                )

            else:

                start_page = (
                    self.current_page or 0
                )

                _logger.warning(
                    f"[PDF RECOVERY] "
                    f"NO SAVED PAGES"
                )


            # =====================================
            # SAFETY CLAMP
            # =====================================

            if start_page >= total_pages:

                start_page = total_pages


            end_page = min(
                start_page + BATCH_SIZE,
                total_pages
            )


            _logger.warning(
                f"[PDF BATCH] "
                f"START={start_page + 1} "
                f"| END={end_page}"
            )


            processed_count = 0


            # =====================================
            # PROCESS PAGES
            # =====================================

            for i in range(start_page, end_page):

                _logger.warning(
                    f"[PDF PAGE] "
                    f"START PAGE={i + 1}"
                )

                page_success = False


                # =================================
                # SKIP IF ALREADY EXISTS
                # =================================

                existing = self.env[
                    'vendor.import.page'
                ].search([

                    ('job_id', '=', self.id),

                    ('page_number', '=', i + 1)

                ], limit=1)


                if existing:

                    _logger.warning(
                        f"[PDF PAGE] "
                        f"SKIP EXISTING "
                        f"| page={i + 1}"
                    )

                    self.current_page = i + 1

                    continue


                for attempt in range(MAX_RETRIES):

                    single_pdf = None
                    pdf_bytes_io = None

                    try:

                        _logger.warning(
                            f"[PDF API] "
                            f"PAGE={i + 1} "
                            f"| ATTEMPT={attempt + 1}"
                        )


                        # =========================
                        # SINGLE PAGE PDF
                        # =========================

                        single_pdf = fitz.open()

                        single_pdf.insert_pdf(

                            doc,

                            from_page=i,

                            to_page=i
                        )


                        pdf_bytes_io = io.BytesIO()

                        single_pdf.save(
                            pdf_bytes_io
                        )

                        pdf_bytes_io.seek(0)


                        # =========================
                        # API CALL
                        # =========================

                        response = requests.post(

                            "https://pdf-extractor-staging.onrender.com/extract",

                            files={

                                "file": (

                                    "page.pdf",

                                    pdf_bytes_io,

                                    "application/pdf"
                                )
                            },

                            timeout=45
                        )


                        _logger.warning(
                            f"[PDF API] "
                            f"STATUS="
                            f"{response.status_code} "
                            f"| page={i + 1}"
                        )


                        if response.status_code != 200:

                            continue


                        page_data = response.json()


                        # =========================
                        # RESPONSE FORMAT
                        # =========================

                        if isinstance(page_data, dict):

                            pages = page_data.get(
                                "pages",
                                []
                            )

                        elif isinstance(
                            page_data,
                            list
                        ):

                            pages = page_data

                        else:

                            pages = []


                        if not pages:

                            _logger.warning(
                                f"[PDF PAGE] "
                                f"EMPTY RESPONSE "
                                f"| page={i + 1}"
                            )

                            continue


                        normalized_blocks = []


                        # =========================
                        # NORMALIZE
                        # =========================

                        for p in pages:

                            text = p.get(
                                "text",
                                ""
                            )

                            images = p.get(
                                "images",
                                []
                            )


                            if (
                                not text
                                and
                                not images
                            ):
                                continue


                            price = ""

                            stock = ""


                            price_match = re.search(

                                r'(\$|€|£)\s?\d+[.,]?\d*',

                                text
                            )


                            if price_match:

                                price = (
                                    price_match.group(0)
                                )


                            stock_match = re.search(

                                r'(stock|available)'
                                r'\s*:?\s*'
                                r'(\d+)'
                                r'\s*(pcs|pieces)?',

                                text,

                                re.I
                            )


                            if stock_match:

                                stock = (
                                    stock_match.group(2)
                                )


                            normalized_blocks.append({

                                "page": i + 1,

                                "text": text,

                                "price": price,

                                "stock": stock,

                                "images": images
                            })


                        if not normalized_blocks:

                            _logger.warning(
                                f"[PDF PAGE] "
                                f"NO VALID BLOCKS "
                                f"| page={i + 1}"
                            )

                            continue


                        # =========================
                        # SAVE PAGE
                        # =========================

                        self.env[
                            'vendor.import.page'
                        ].create({

                            'job_id': self.id,

                            'page_number': i + 1,

                            'extracted_json': json.dumps(
                                normalized_blocks
                            )
                        })


                        _logger.warning(
                            f"[PDF PAGE] "
                            f"SAVED "
                            f"| page={i + 1}"
                        )


                        self.current_page = i + 1

                        processed_count += 1

                        page_success = True

                        break


                    except Exception as e:

                        _logger.exception(
                            f"[PDF PAGE ERROR] "
                            f"page={i + 1} "
                            f"| {str(e)}"
                        )


                    finally:

                        try:

                            if pdf_bytes_io:
                                pdf_bytes_io.close()

                        except Exception:
                            pass


                        try:

                            if single_pdf:
                                single_pdf.close()

                        except Exception:
                            pass


                if not page_success:

                    _logger.error(
                        f"[PDF PAGE FAILED] "
                        f"page={i + 1}"
                    )


            # =====================================
            # SAVE BATCH ONCE
            # =====================================

            _logger.warning(
                f"[PDF BATCH] "
                f"PROCESSED="
                f"{processed_count}"
            )


            if self.current_page < total_pages:

                self.state = "pdf_extracting"

            else:

                self.state = "pdf_ai"


            try:

                self.flush_recordset()

                self.env.cr.commit()

                _logger.warning(
                    f"[PDF SAVE] "
                    f"SUCCESS "
                    f"| state={self.state} "
                    f"| current={self.current_page}"
                )

            except Exception as e:

                _logger.exception(
                    f"[PDF SAVE ERROR] "
                    f"{str(e)}"
                )


        finally:

            try:

                if doc:
                    doc.close()

            except Exception:
                pass


            gc.collect()

            _logger.warning(
                "[PDF GC] COMPLETE"
            )


    # ---------------- Send to OPENAI URL ----------------
    def send_to_openai_url(self):

        import re
        import json
        import math

        api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')

        if not api_key:
            raise Exception("OpenAI API key not configured")

        client = OpenAI(api_key=api_key)

        # ================= LOAD PAGES =================
        try:
            # pages = json.loads(self.extracted_text or "[]")

            pages = json.loads(

                self.url_blocks_json

                or

                self.extracted_text

                or

                "[]"
            )

        except Exception:
            _logger.error("INVALID extracted_text JSON")
            return

        if not pages:
            _logger.error("NO PAGES TO PROCESS")
            return

        # ================= LOAD EXISTING =================
        existing_products = []
        if self.ai_response:
            try:
                data = json.loads(self.ai_response)
                if isinstance(data, list):
                    existing_products = data
            except Exception as e:
                _logger.warning(f"AI RESPONSE LOAD FAILED → {str(e)}")
                existing_products = []

        current_batch = self.url_batch_index or 0

        # ================= FLATTEN =================
        all_blocks = [b for p in pages for b in p.get("blocks", [])]

        _logger.warning(f"RAW BLOCKS → {len(all_blocks)}")

        # ================= CLEAN =================
        cleaned_blocks = self._clean_scraped_blocks(all_blocks)

        _logger.warning(f"CLEAN BLOCKS → {len(cleaned_blocks)}")
        _logger.warning(f"REMOVED BLOCKS → {len(all_blocks) - len(cleaned_blocks)}")

        cleaned_blocks = sorted(cleaned_blocks, key=lambda x: (x.get("text") or "")[:50])

        # ================= BATCH =================
        BLOCK_BATCH_SIZE = 8

        batched_blocks = [
            cleaned_blocks[i:i + BLOCK_BATCH_SIZE]
            for i in range(0, len(cleaned_blocks), BLOCK_BATCH_SIZE)
        ]

        total_batches = len(batched_blocks)
        self.url_total_batches = total_batches

        _logger.warning(f"TOTAL BLOCK BATCHES → {total_batches}")
        _logger.warning(f"CURRENT BATCH → {current_batch}")

        # ================= STOP IF DONE =================
        if current_batch >= total_batches:
            _logger.warning("ALL URL BATCHES PROCESSED ✅")
            self.state = "url_creating"
            return

        # ================= PROCESS ONE BATCH =================
        block_batch = batched_blocks[current_batch]

        _logger.warning(f"PROCESSING BLOCK COUNT → {len(block_batch)}")
        _logger.warning(f"AI → PROCESSING BLOCK BATCH {current_batch + 1}")

      
        combined_text = "\n\n---\n\n".join([
            f"""
            TEXT:
            {b.get('text','')}

            PRICE:
            {b.get('price','')}

            STOCK:
            {b.get('stock','')}

            IMAGE_URL:
            {b.get('image','')}
            """

            for b in block_batch
        ])


        if not combined_text.strip():
            _logger.warning("EMPTY COMBINED TEXT → SKIP")
            self.url_batch_index += 1
            return

        if len(combined_text) > 15000:
            combined_text = combined_text[:15000]
            _logger.warning("TEXT TRIMMED → PREVENT TOKEN OVERFLOW")

        # ================= PROMPT =================
        prompt = f"""
            You are a highly precise e-commerce product extraction engine.

            You are processing raw scraped website content.

            =====================================
            YOUR GOAL
            =====================================

            Extract ALL individual products from the text.

            IMPORTANT:
            - A single block may contain MULTIPLE products → you MUST split them
            - Each product MUST be returned separately
            - NEVER merge multiple products into one

            =====================================
            HOW TO IDENTIFY A PRODUCT
            =====================================

            A product usually contains:
            - product name
            - price (e.g. $, £, €, numbers)
            - specs or description
            - sometimes reviews

            Strong indicators:
            - currency symbols ($, £, €)
            - model names (Dell, Lenovo, etc.)
            - numbers like GB, inch, SSD, RAM, etc.

            =====================================
            STRICT RULES
            =====================================

            1. RETURN ONLY VALID JSON ARRAY
            2. NO explanation
            3. NO markdown
            4. NO text outside JSON

            5. EACH product must be unique
            6. REMOVE duplicates
            7. DO NOT skip products
            8. SPLIT combined text into multiple products

            =====================================
            REMOVE THIS NOISE
            =====================================

            Ignore anything related to:
            - navigation (Home, Login, Pricing)
            - cookies/privacy
            - footer text
            - categories only (no product info)
            - repeated headers

            =====================================
            OUTPUT FORMAT
            =====================================

            [
            {{
                "name": "Clean product name",
                "description": "Short product description (max 30 words)",
                "category": "Best guess category",
                "price": "",
                "stock": "",
                "image": "image_url_or_null"
            }}
            ]

            =====================================
            EXTRA RULES
            =====================================

            - Keep names SHORT and CLEAN
            - Description must be concise
            - Infer category intelligently
            - If no image exists → return null
            - If price exists → extract it
            - If stock exists → extract it
            - NEVER invent stock or price
            - If unsure → still extract

            =====================================
            TEXT TO PROCESS
            =====================================

        {combined_text}
        """

        # ================= OPENAI =================
        try:
            response = client.responses.create(
                model="gpt-4.1-mini",
                input=prompt,
                temperature=0,
                timeout=60
            )

            result = response.output_text.strip()
            result = re.sub(r"^```(?:json)?|```$", "", result).strip()

            parsed = json.loads(result)

            if isinstance(parsed, list):

                cleaned = [p for p in parsed if p.get("name")]

                _logger.warning(f"AI RETURNED → {len(cleaned)} PRODUCTS")

                # 🔥 DEDUPE BY NAME
                existing_map = {p.get("name"): p for p in existing_products}

                for p in cleaned:
                    existing_map[p.get("name")] = p

                existing_products = list(existing_map.values())

                _logger.warning(f"TOTAL ACCUMULATED → {len(existing_products)}")

            else:
                _logger.warning("AI RESPONSE NOT LIST")

        except Exception as e:
            _logger.warning(f"AI ERROR → {str(e)}")
            return

        # ================= SAVE =================
        self.ai_response = json.dumps(existing_products)
        self.url_batch_index = current_batch + 1

        _logger.warning(f"URL AI PROGRESS → {self.url_batch_index}/{self.url_total_batches}")

        # ================= STATE =================
        if self.url_batch_index < self.url_total_batches:
            self.state = "url_ai"
        else:
            _logger.warning("URL AI FINISHED ALL BATCHES")
            self.state = "url_creating"

        # 🔥 IMPORTANT: COMMIT FOR CRON CONTINUITY
        # self.env.cr.commit()

        try:

            self.env.cr.commit()

        except Exception as commit_error:

            _logger.warning(
                f"COMMIT SKIPPED → {commit_error}"
            )

        return


    # =========== PDF OPENAI =========================

    def send_to_openai_pdf(self):

        import json

        api_key = self.env[
            'ir.config_parameter'
        ].sudo().get_param(
            'openai.api.key'
        )

        if not api_key:

            raise Exception(
                "OpenAI API key not configured"
            )


        client = OpenAI(
            api_key=api_key
        )


        _logger.warning(
            "[PDF AI] START"
        )


        # =====================================================
        # LOAD PAGE RECORDS
        # =====================================================

        page_records = self.env[
            'vendor.import.page'
        ].search([

            ('job_id', '=', self.id)

        ], order='page_number asc')


        total_available_pages = len(
            page_records
        )


        _logger.warning(

            f"[PDF AI] "

            f"TOTAL PAGE RECORDS="

            f"{total_available_pages}"
        )


        if total_available_pages <= 0:

            _logger.warning(
                "[PDF AI] "
                "NO PAGE RECORDS FOUND"
            )

            return


        # =====================================================
        # LOAD EXISTING AI RESPONSE
        # =====================================================

        existing_pages = []


        if self.ai_response:

            try:

                loaded = json.loads(
                    self.ai_response
                )

                if isinstance(
                    loaded,
                    list
                ):

                    existing_pages = loaded


            except Exception as e:

                _logger.warning(

                    f"[PDF AI] "

                    f"LOAD EXISTING FAILED "

                    f"| {str(e)}"
                )

                existing_pages = []


        # =====================================================
        # FIND ALREADY PROCESSED PAGES
        # =====================================================

        processed_pages = set()


        for p in existing_pages:

            page_num = p.get("page")

            if page_num:

                processed_pages.add(
                    page_num
                )


        _logger.warning(

            f"[PDF AI] "

            f"PROCESSED PAGES="

            f"{sorted(list(processed_pages))}"
        )


        # =====================================================
        # FIND NEXT UNPROCESSED PAGE
        # =====================================================

        next_record = None


        for record in page_records:

            if (

                record.page_number

                not in processed_pages

            ):

                next_record = record

                break


        # =====================================================
        # ALL COMPLETE
        # =====================================================

        if not next_record:

            _logger.warning(
                "[PDF AI] COMPLETE ✅"
            )

            self.last_ai_page = (
                total_available_pages
            )

            self.state = "pdf_creating"

            self.flush_recordset()

            self.env.cr.commit()

            return


        _logger.warning(

            f"[PDF AI] "

            f"PROCESSING PAGE "

            f"{next_record.page_number}"
        )


        # =====================================================
        # LOAD PAGE DATA
        # =====================================================

        try:

            page_blocks = json.loads(

                next_record.extracted_json
                or
                "[]"
            )

        except Exception as e:

            _logger.warning(

                f"[PDF AI] "

                f"PAGE LOAD FAILED "

                f"| PAGE "

                f"{next_record.page_number} "

                f"| {str(e)}"
            )

            return


        if not page_blocks:

            _logger.warning(

                f"[PDF AI] "

                f"EMPTY PAGE BLOCKS "

                f"| PAGE "

                f"{next_record.page_number}"
            )

            return


        # =====================================================
        # BUILD PAGE DATA
        # =====================================================

        page_text = "\n".join([

            p.get("text", "")

            for p in page_blocks

        ])


        page_images = []


        for p in page_blocks:

            page_images.extend(
                p.get("images", [])
            )


        page_price = ""

        page_stock = ""


        for p in page_blocks:

            if (

                not page_price

                and

                p.get("price")

            ):

                page_price = (
                    p.get("price")
                )


            if (

                not page_stock

                and

                p.get("stock")

            ):

                page_stock = (
                    p.get("stock")
                )


        # =====================================================
        # PROMPT
        # =====================================================

        prompt = f"""
        You are an advanced product extraction and interpretation engine for catalog PDFs.

        =====================
        CORE RULES (STRICT)
        =====================

        1. RETURN ONLY VALID JSON
        2. NO explanation
        3. NO markdown
        4. NO text outside JSON
        5. DO NOT duplicate products WITHIN THE SAME PAGE
        6. DO NOT skip any product
        7. EACH product must appear exactly once PER PAGE

        IMPORTANT GLOBAL RULE:

        - This input represents ONLY ONE PAGE of a catalog
        - You MUST extract ONLY products visible on THIS PAGE
        - DO NOT repeat products from previous pages
        - DO NOT assume products continue across pages

        =====================
        PRODUCT DETECTION LOGIC
        =====================

        A page may contain:

        (A) ONE large product (hero layout)
        (B) MULTIPLE products (grid/catalog layout)
        (C) MIX of large + small supporting products

        You MUST:

        - If SINGLE main product:
        → return ONE product

        - If MULTIPLE distinct products:
        → extract EACH product separately

        IMPORTANT:

        - If multiple images exist → assume multiple products
        - DO NOT collapse multiple items into one product
        - If unsure → split into multiple products instead of merging

        =====================
        VARIANT DETECTION LOGIC
        =====================

        - Similar products with same structure/design
        should be grouped as variants.

        - Different products MUST stay separate.

        =====================
        OUTPUT FORMAT
        =====================

        [
            {{
                "name": "",
                "description": "",
                "category": "",
                "price": "",
                "stock": "",
                "variant_group": "",
                "variants": [
                    {{
                        "attributes": {{
                            "Color": ""
                        }},
                        "image_index": 0,
                        "stock": null
                    }}
                ]
            }}
        ]

        PAGE TEXT:
        {page_text}

        DETECTED PRICE:
        {page_price}

        DETECTED STOCK:
        {page_stock}
        """


        # =====================================================
        # AI CALL
        # =====================================================

        try:

            image_inputs = [
                {
                    "type": "input_image",
                    "image_url": f"data:image/jpeg;base64,{img}"
                }
                for img in page_images[:10]
            ]


            response = client.responses.create(

                model="gpt-4.1-mini",

                input=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": prompt
                        }
                    ] + image_inputs
                }],

                timeout=60
            )


            result = (
                response.output_text or ""
            ).strip()


            result = result.replace(
                "```json",
                ""
            )

            result = result.replace(
                "```",
                ""
            ).strip()


            if not result:

                raise Exception(
                    "EMPTY AI RESPONSE"
                )


            parsed = json.loads(
                result
            )


            if not isinstance(
                parsed,
                list
            ):

                parsed = []


        except Exception as e:

            _logger.warning(

                f"[PDF AI] "

                f"FAILED "

                f"| PAGE "

                f"{next_record.page_number} "

                f"| {str(e)}"
            )

            return


        # =====================================================
        # ATTACH IMAGES
        # =====================================================

        for p_index, prod in enumerate(
            parsed
        ):

            if (

                page_images

                and

                p_index < len(page_images)

            ):

                prod["image"] = (
                    page_images[p_index]
                )


        # =====================================================
        # MERGE RESULTS
        # =====================================================

        existing_map = {}


        for p in existing_pages:

            existing_map[
                p.get("page")
            ] = p


        existing_map[
            next_record.page_number
        ] = {

            "page": next_record.page_number,

            "products": parsed
        }


        combined_pages = sorted(

            list(existing_map.values()),

            key=lambda x: x.get(
                "page",
                0
            )
        )


        # =====================================================
        # SAVE
        # =====================================================

        self.ai_response = json.dumps(
            combined_pages
        )


        self.last_ai_page = len(
            combined_pages
        )


        _logger.warning(

            f"[PDF AI] "

            f"PAGE SAVED "

            f"| PAGE "

            f"{next_record.page_number}"
        )


        # =====================================================
        # NEXT STATE
        # =====================================================

        if (

            self.last_ai_page

            <

            total_available_pages

        ):

            self.state = "pdf_ai"

            _logger.warning(

                f"[PDF AI] CONTINUE "

                f"{self.last_ai_page}/"

                f"{total_available_pages}"
            )

        else:

            _logger.warning(
                "[PDF AI] COMPLETE ✅"
            )

            self.state = "pdf_creating"


        self.flush_recordset()

        self.env.cr.commit()

        return


    #===========Excel Open AI================================
    def send_to_openai_excel(self):

        import json

        api_key = self.env[
            'ir.config_parameter'
        ].sudo().get_param(
            'openai.api.key'
        )

        if not api_key:

            raise Exception(
                "OpenAI API key not configured"
            )


        client = OpenAI(
            api_key=api_key
        )


        _logger.warning(
            "[EXCEL AI] START"
        )


        # =====================================================
        # LOAD EXTRACTED DATA
        # =====================================================

        try:

            pages = json.loads(
                self.extracted_text or "[]"
            )

        except Exception as e:

            _logger.error(

                f"[EXCEL AI] "

                f"INVALID extracted_text JSON "

                f"| {str(e)}"
            )

            return


        if not pages:

            _logger.error(
                "[EXCEL AI] NO ROWS TO PROCESS"
            )

            return


        _logger.warning(

            f"[EXCEL AI] "

            f"TOTAL ROWS={len(pages)}"
        )


        # =====================================================
        # BATCH
        # =====================================================

        BATCH_SIZE = 5

        start = (
            self.excel_ai_index or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(pages)
        )


        batch = pages[start:end]


        _logger.warning(

            f"[EXCEL AI BATCH] "

            f"{start} → {end}"
        )


        # =====================================================
        # LOAD EXISTING PRODUCTS
        # =====================================================

        existing_products = []


        if self.ai_response:

            try:

                existing_ai = json.loads(
                    self.ai_response
                )

                if (
                    isinstance(existing_ai, list)
                    and existing_ai
                ):

                    existing_products = (
                        existing_ai[0].get(
                            "products",
                            []
                        )
                    )


                _logger.warning(

                    f"[EXCEL AI] "

                    f"EXISTING PRODUCTS="

                    f"{len(existing_products)}"
                )

            except Exception as e:

                _logger.warning(

                    f"[EXCEL AI] "

                    f"FAILED LOAD EXISTING "

                    f"| {str(e)}"
                )

                existing_products = []


        new_products = []


        # =====================================================
        # PROCESS ROWS
        # =====================================================

        for idx, row in enumerate(

            batch,

            start=start

        ):

            try:

                row_text = row.get(
                    "text",
                    ""
                )

                row_price = row.get(
                    "price",
                    ""
                )

                row_stock = row.get(
                    "stock",
                    ""
                )

                images = row.get(
                    "images",
                    []
                )


                _logger.warning(

                    f"[EXCEL AI ROW] "

                    f"idx={idx} "

                    f"| images={len(images)}"
                )


                prompt = f"""
                You are a structured Excel product parser.

                Each input represents EXACTLY ONE ROW = ONE PRODUCT.

                =====================================
                COLUMN UNDERSTANDING (CRITICAL)
                =====================================

                The row contains mixed values like:

                - ID (e.g. 94601, 12345)
                - Range (e.g. 2-66, 11-00)
                - Stock numbers
                - Prices
                - Links (http...)
                - Image references

                YOU MUST:

                1. IDENTIFY PRODUCT ID
                - Usually numeric (e.g. 94601)
                - Column name may vary (KOD, SKU, ID, CODE)

                2. IDENTIFY PRODUCT NAME
                - MUST NOT be:
                    - pure numbers
                    - ranges
                    - links
                    - dates
                    - headers

                - If unclear:
                    → generate:
                    Product <ID>

                =====================================
                VARIANT GROUPING
                =====================================

                - SAME ID = SAME variant_group
                - DIFFERENT ID = DIFFERENT PRODUCT
                - NEVER leave variant_group empty

                =====================================
                VARIANT DETECTION
                =====================================

                If rows share same ID:

                → they are variants

                Put differences into:

                "attributes"

                =====================================
                OUTPUT FORMAT
                =====================================

                [
                    {{
                        "name": "",
                        "description": "",
                        "category": "",
                        "price": "",
                        "stock": "",
                        "variant_group": "",
                        "variants": [
                            {{
                                "attributes": {{
                                    "Variant": ""
                                }},
                                "image_index": 0,
                                "stock": null
                            }}
                        ]
                    }}
                ]

                ROW TEXT:
                {row_text}

                DETECTED PRICE:
                {row_price}

                DETECTED STOCK:
                {row_stock}
                """


                response = client.responses.create(

                    model="gpt-4.1-mini",

                    input=prompt,

                    timeout=60
                )


                result = (
                    response.output_text or ""
                ).strip()


                result = result.replace(
                    "```json",
                    ""
                )

                result = result.replace(
                    "```",
                    ""
                ).strip()


                if not result:

                    raise Exception(
                        "EMPTY AI RESPONSE"
                    )


                parsed = json.loads(
                    result
                )


                if (

                    isinstance(parsed, list)

                    and parsed

                ):

                    parsed = parsed[0]


                if not isinstance(
                    parsed,
                    dict
                ):

                    _logger.warning(

                        f"[EXCEL AI] "

                        f"INVALID STRUCTURE "

                        f"| idx={idx}"
                    )

                    continue


                # =================================================
                # IMAGE
                # =================================================

                if images:

                    parsed["image"] = (
                        images[0]
                    )


                # =================================================
                # DEBUG
                # =================================================

                _logger.warning(

                    f"[EXCEL AI PRODUCT] "

                    f"name={parsed.get('name')} "

                    f"| group={parsed.get('variant_group')}"
                )


                new_products.append(
                    parsed
                )


            except Exception as e:

                _logger.exception(

                    f"[EXCEL AI ERROR] "

                    f"idx={idx} "

                    f"| {str(e)}"
                )


        # =====================================================
        # MERGE PRODUCTS SAFELY
        # =====================================================

        combined_products = (
            existing_products
            +
            new_products
        )


        _logger.warning(

            f"[EXCEL AI MERGE] "

            f"existing={len(existing_products)} "

            f"| new={len(new_products)} "

            f"| total={len(combined_products)}"
        )


        # =====================================================
        # SAVE
        # =====================================================

        self.ai_response = json.dumps([{

            "page": 1,

            "products": combined_products

        }])


        self.excel_ai_index = end


        _logger.warning(

            f"[EXCEL AI SAVE] "

            f"{self.excel_ai_index}/"

            f"{len(pages)}"
        )


        # =====================================================
        # NEXT STATE
        # =====================================================

        if end < len(pages):

            self.state = "excel_ai"

        else:

            _logger.warning(
                "[EXCEL AI COMPLETE]"
            )

            self.state = (
                "excel_creating"
            )


        self.flush_recordset()

        self.env.cr.commit()

        return
    

    #-----------scoring image before picking best/quality image (inage logic)-------------
    def pick_best_image(self, images):

                best_img = None
                best_score = 0
                seen_hashes = set()

                for img in images:

                    try:
                        img_bytes = base64.b64decode(img)
                        size = len(img_bytes)

                        # ❌ Skip tiny images (logos/icons)
                        if size < 10000:
                            continue

                        # ❌ Skip extremely large (full lifestyle pages)
                        if size > 800000:
                            continue

                        # ================= IMAGE ANALYSIS =================
                        pil_img = Image.open(BytesIO(img_bytes))
                        width, height = pil_img.size

                        # ❌ Skip extremely small resolution
                        if width < 150 or height < 150:
                            continue

                        # ❌ Skip extreme aspect ratios (banners, strips)
                        aspect_ratio = width / height if height else 1

                        if aspect_ratio > 3 or aspect_ratio < 0.3:
                            continue

                        # ================= DUPLICATE CHECK =================
                        img_hash = hash(img_bytes[:100])  # fast partial hash

                        if img_hash in seen_hashes:
                            continue

                        seen_hashes.add(img_hash)

                        # ================= SCORING =================
                        score = 0

                        # ✅ Prefer medium-good sizes
                        if 20000 < size < 300000:
                            score += 200

                        # ✅ Prefer square-ish product images
                        if 0.7 < aspect_ratio < 1.5:
                            score += 150

                        # ✅ Prefer decent resolution
                        if width > 400 and height > 400:
                            score += 150

                        # ❌ Penalize too wide/tall
                        if aspect_ratio > 2 or aspect_ratio < 0.5:
                            score -= 100

                        # ❌ Penalize very large images (likely lifestyle)
                        if size > 500000:
                            score -= 150

                        # Base size contribution
                        score += size / 2000

                        # ================= SELECT BEST =================
                        if score > best_score:
                            best_score = score
                            best_img = img

                    except Exception:
                        continue

                return best_img


    #----marchin AI-----------------------------------
    def match_image_with_ai(self, product_name, images):

        api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')
        client = OpenAI(api_key=api_key)

        if not images:
            return None

        # limit images for performance
        images = images[:5]
        image_inputs = [
        {
                "type": "input_image",
                "image_url": f"data:image/jpeg;base64,{img}"
            }
            for img in images
        ]

        prompt = f"""
        You are an expert product image matcher.

        Select the image that BEST represents this product:

        PRODUCT:
        {product_name}

        RULES:
        - Return ONLY the index (0-based integer)
        - No explanation
        - No text

        PRIORITY:
        1. Clean product image (plain background)
        2. Product centered and clearly visible
        3. No human interaction preferred
        4. If only lifestyle images exist, choose the clearest one

        DO NOT PICK:
        - logos
        - icons
        - background-only images
        - cropped fragments
        """

        try:
            response = client.responses.create(
                model="gpt-4.1-mini",
                input=[{
                    "role": "user",
                    "content": [{"type": "input_text", "text": prompt}] + image_inputs
                }],
                timeout=30
            )

            result = response.output_text.strip()

            index = int(result)

            if 0 <= index < len(images):
                return images[index]

        except Exception as e:
            _logger.warning(f"AI IMAGE MATCH FAILED → {str(e)}")

        return None


    # ================= PRODUCT CREATION URL ================

    def create_products_url(self):

        import requests
        import base64
        import json

        if not self.ai_response:
            _logger.warning("NO AI RESPONSE")
            return

        product_obj = self.env['product.template']
        category_obj = self.env['product.category']

        try:
            products = json.loads(self.ai_response)

            if isinstance(products, dict):
                products = [products]

        except Exception as e:
            _logger.error(f"INVALID AI JSON → {str(e)}")
            return

        TOTAL_PRODUCTS = len(products)
        start_index = self.last_processed_product_index or 0

        _logger.warning(f"TOTAL AI PRODUCTS → {TOTAL_PRODUCTS}")
        _logger.warning(f"START INDEX → {start_index}")

        created_count = 0
        skipped_count = 0

        MAX_PRODUCTS_PER_RUN = 5

        CATEGORY_MAPPING = {
            "t-shirt": "Apparel",
            "shirt": "Apparel",
            "polo": "Apparel",
            "bag": "Bags",
            "backpack": "Bags",
            "cap": "Headwear",
            "hat": "Headwear",
            "bottle": "Drinkware",
            "cup": "Drinkware",
            "drinkware": "Drinkware",
            "pen": "Stationery",
            "notebook": "Stationery",
            "powerbank": "Electronics",
            "charger": "Electronics",
            "laptop": "Electronics",
            "football": "Football Fever",
            "Wristband": "Football Fever",
            "sports t-shirt": "Football Fever",
            "sports towel": "Football Fever",
            "Sports Bottles": "Football Fever"
        }

        parent_category = category_obj.search([('name', '=', "All Products")], limit=1)
        vendor_id = self.partner_id.id if self.partner_id else False

        if not parent_category:
            parent_category = category_obj.create({'name': "All Products"})

        end_index = min(start_index + MAX_PRODUCTS_PER_RUN, TOTAL_PRODUCTS)

        _logger.warning(f"PROCESSING RANGE → {start_index} to {end_index}")

        for idx in range(start_index, end_index):

            product = products[idx]

            name = product.get("name")
            if not name:
                skipped_count += 1
                continue

            description = product.get("description", "")
            raw_category = (product.get("category") or "").lower()

            # ================= CATEGORY =================
            mapped_category = "General"
            for key, val in CATEGORY_MAPPING.items():
                if key in raw_category:
                    mapped_category = val
                    break

            category = category_obj.search([
                ('name', '=', mapped_category),
                ('parent_id', '=', parent_category.id)
            ], limit=1)

            if not category:
                category = category_obj.create({
                    'name': mapped_category,
                    'parent_id': parent_category.id
                })

            # ================= DUPLICATE CHECK =================
            existing = product_obj.search([
                ('name', 'ilike', name.strip()),
                ('vendor_id', '=', vendor_id)
            ], limit=1)

            if existing:
                _logger.warning(f"SKIPPED DUPLICATE → {name}")
                skipped_count += 1
                continue

            vals = {
                'name': name.strip(),
                'description_sale': description,
                'categ_id': category.id,
                'sale_ok': True,
                'website_published': False,
                'vendor_id': vendor_id,
            }

            # ================= IMAGE =================
            image_url = product.get("image")

            if image_url and isinstance(image_url, str) and image_url.startswith("http"):

                try:
                    _logger.warning(f"FETCHING IMAGE → {image_url}")

                    res = requests.get(image_url, timeout=5, stream=True)

                    if res.status_code != 200:
                        _logger.warning(f"IMAGE HTTP ERROR → {res.status_code}")
                        continue

                    content_type = res.headers.get("Content-Type", "")
                    if "image" not in content_type:
                        _logger.warning(f"NOT AN IMAGE → {content_type}")
                        continue

                    content = res.raw.read(500000, decode_content=True)

                    if not content:
                        _logger.warning("EMPTY IMAGE CONTENT")
                        continue

                    vals['image_1920'] = base64.b64encode(content).decode("utf-8")

                    _logger.warning("IMAGE STORED SUCCESSFULLY")

                except Exception as e:
                    _logger.warning(f"IMAGE FAILED → {str(e)}")

            else:
                _logger.warning(f"NO VALID IMAGE URL → {image_url}")

            # ================= CREATE =================
            try:
                _logger.warning(
                    f"[URL CREATE PRODUCT] → "
                    f"NAME={name} | "
                    f"VENDOR_ID={vendor_id}"
                )

                # product_obj.create(vals)
                product = product_obj.with_context(
                    mail_create_nolog=True,
                    mail_notify_force_send=False,
                    tracking_disable=True
                ).create(vals)
                created_count += 1

            except Exception as e:
                _logger.error(f"CREATE FAILED → {name} | {str(e)}")
                skipped_count += 1
                continue

            if created_count % 10 == 0:
                self.env.cr.commit()

        # ================= SAVE PROGRESS =================
        self.last_processed_product_index = end_index

        _logger.warning(f"CREATED THIS RUN → {created_count}")
        _logger.warning(f"SKIPPED THIS RUN → {skipped_count}")
        _logger.warning(f"NEXT START INDEX → {self.last_processed_product_index}")

        if self.last_processed_product_index >= TOTAL_PRODUCTS:
            _logger.warning("ALL PRODUCTS CREATED ✅")
            self.state = "done"
        else:
            _logger.warning("MORE PRODUCTS REMAIN → CONTINUE CREATION")
            self.state = "url_creating"

        self.env.cr.commit()


    #==========create pdf product==========================
  
    def create_products_pdf(self):

        import json
        import re

        _logger.warning(
            "[PDF CREATE] START"
        )


        # =====================================================
        # VALIDATION
        # =====================================================

        if not self.ai_response:

            _logger.warning(
                "[PDF CREATE] NO AI RESPONSE"
            )

            return


        try:

            ai_pages = json.loads(
                self.ai_response
            )

        except Exception as e:

            _logger.error(

                f"[PDF CREATE] INVALID AI JSON "

                f"| {str(e)}"
            )

            return


        if not isinstance(ai_pages, list):

            _logger.warning(
                "[PDF CREATE] INVALID AI FORMAT"
            )

            return


        # =====================================================
        # MODELS
        # =====================================================

        product_obj = self.env[
            'product.template'
        ]

        category_obj = self.env[
            'product.category'
        ]


        # =====================================================
        # CATEGORY MAP
        # =====================================================

        CATEGORY_MAPPING = {

            "t-shirt": "Apparel",
            "shirt": "Apparel",
            "polo": "Apparel",

            "bag": "Bags",
            "backpack": "Bags",

            "cap": "Headwear",
            "hat": "Headwear",

            "bottle": "Drinkware",
            "drinkware": "Drinkware",

            "pen": "Stationery",
            "notebook": "Stationery",

            "powerbank": "Electronics",
            "charger": "Electronics",
            "laptop": "Electronics",
        }


        # =====================================================
        # ROOT CATEGORY
        # =====================================================

        parent_category = category_obj.search([

            ('name', '=', "All Products")

        ], limit=1)


        if not parent_category:

            parent_category = category_obj.create({

                'name': "All Products"

            })


        # =====================================================
        # VENDOR
        # =====================================================

        vendor_id = (

            self.partner_id.id

            if self.partner_id

            else False
        )


        # =====================================================
        # BATCHING
        # =====================================================

        BATCH_SIZE = 3

        start = (
            self.last_created_page or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(ai_pages)
        )


        _logger.warning(

            f"[PDF CREATE RANGE] "

            f"{start} -> {end} "

            f"| total={len(ai_pages)}"
        )


        created_count = 0

        skipped_count = 0


        # =====================================================
        # MAIN LOOP
        # =====================================================

        for page_index in range(start, end):

            try:

                page_data = ai_pages[
                    page_index
                ]

            except Exception as e:

                _logger.warning(

                    f"[PDF PAGE LOAD ERROR] "

                    f"{str(e)}"
                )

                continue


            page_number = page_data.get(
                "page"
            )


            products = page_data.get(
                "products",
                []
            )


            _logger.warning(

                f"[PDF PAGE] "

                f"page={page_number} "

                f"| products={len(products)}"
            )


            # =================================================
            # PRODUCTS
            # =================================================

            for product_data in products:

                try:

                    name = (

                        product_data.get(
                            "name"
                        )

                        or ""

                    ).strip()


                    if not name:

                        _logger.warning(
                            "[PDF SKIP] EMPTY NAME"
                        )

                        continue


                    description = (
                        product_data.get(
                            "description"
                        ) or ""
                    )


                    raw_category = (

                        product_data.get(
                            "category"
                        ) or ""

                    ).lower()


                    variants = product_data.get(
                        "variants",
                        []
                    )


                    variant_group = (

                        product_data.get(
                            "variant_group"
                        )

                        or

                        name
                    )


                    variant_group = str(
                        variant_group
                    ).strip().upper()


                    # =========================================
                    # CATEGORY MAP
                    # =========================================

                    mapped_category = (
                        "General"
                    )


                    for key, val in CATEGORY_MAPPING.items():

                        if key in raw_category:

                            mapped_category = val

                            break


                    category = category_obj.search([

                        ('name', '=', mapped_category),

                        (
                            'parent_id',
                            '=',
                            parent_category.id
                        )

                    ], limit=1)


                    if not category:

                        category = category_obj.create({

                            'name': mapped_category,

                            'parent_id':
                                parent_category.id
                        })


                    # =========================================
                    # DUPLICATE CHECK
                    # =========================================

                    product = product_obj.search([

                        (
                            'default_code',
                            '=',
                            variant_group
                        ),

                        (
                            'vendor_id',
                            '=',
                            vendor_id
                        )

                    ], limit=1)


                    # =========================================
                    # CREATE PRODUCT
                    # =========================================

                    if not product:

                        vals = {

                            'name': name,

                            'default_code':
                                variant_group,

                            'description_sale':
                                description,

                            'categ_id':
                                category.id,

                            'sale_ok': True,

                            'website_published':
                                False,

                            'vendor_id':
                                vendor_id,
                        }


                        image = product_data.get(
                            "image"
                        )


                        if image:

                            vals[
                                'image_1920'
                            ] = image


                        product = product_obj.with_context(

                            mail_create_nolog=True,

                            mail_notify_force_send=False,

                            tracking_disable=True

                        ).create(vals)


                        created_count += 1


                        _logger.warning(

                            f"[PDF CREATED] "

                            f"{variant_group}"
                        )

                    else:

                        skipped_count += 1


                        _logger.warning(

                            f"[PDF EXISTS] "

                            f"{variant_group} "

                            f"| vendor={vendor_id}"
                        )


                    # =========================================
                    # FALLBACK VARIANT
                    # =========================================

                    if not variants:

                        variants = [{

                            "attributes": {

                                "Variant": name
                            }

                        }]


                    # =========================================
                    # VARIANTS
                    # =========================================

                    for variant in variants:

                        attributes = variant.get(
                            "attributes",
                            {}
                        )


                        for attr_name, attr_value in attributes.items():

                            if not attr_value:

                                continue


                            attribute = self.env[
                                'product.attribute'
                            ].search([

                                (
                                    'name',
                                    '=',
                                    attr_name
                                )

                            ], limit=1)


                            if not attribute:

                                attribute = self.env[
                                    'product.attribute'
                                ].create({

                                    'name': attr_name

                                })


                            value = self.env[
                                'product.attribute.value'
                            ].search([

                                (
                                    'name',
                                    '=',
                                    attr_value
                                ),

                                (
                                    'attribute_id',
                                    '=',
                                    attribute.id
                                )

                            ], limit=1)


                            if not value:

                                value = self.env[
                                    'product.attribute.value'
                                ].create({

                                    'name': attr_value,

                                    'attribute_id':
                                        attribute.id
                                })


                            line = self.env[
                                'product.template.attribute.line'
                            ].search([

                                (
                                    'product_tmpl_id',
                                    '=',
                                    product.id
                                ),

                                (
                                    'attribute_id',
                                    '=',
                                    attribute.id
                                )

                            ], limit=1)


                            if not line:

                                self.env[
                                    'product.template.attribute.line'
                                ].create({

                                    'product_tmpl_id':
                                        product.id,

                                    'attribute_id':
                                        attribute.id,

                                    'value_ids': [(6, 0, [

                                        value.id

                                    ])]
                                })

                            else:

                                if (

                                    value.id

                                    not in

                                    line.value_ids.ids

                                ):

                                    line.value_ids = [

                                        (4, value.id)

                                    ]


                        # =====================================
                        # VARIANT IMAGE
                        # =====================================

                        image_index = variant.get(
                            "image_index",
                            0
                        )


                        variant_record = self.env[
                            'product.product'
                        ].search([

                            (
                                'product_tmpl_id',
                                '=',
                                product.id
                            )

                        ], limit=1)


                        if (

                            variant_record

                            and

                            product_data.get(
                                "image"
                            )

                        ):

                            variant_record.image_1920 = (

                                product_data.get(
                                    "image"
                                )
                            )


                except Exception as e:

                    _logger.exception(

                        f"[PDF PRODUCT ERROR] "

                        f"{str(e)}"
                    )

                    continue


            # =================================================
            # SAVE PAGE PROGRESS
            # =================================================

            self.last_created_page = (
                page_index + 1
            )


            _logger.warning(

                f"[PDF TRACK] "

                f"last_created_page="

                f"{self.last_created_page}"
            )


            self.flush_recordset()

            self.env.cr.commit()


        # =====================================================
        # FINAL LOG
        # =====================================================

        _logger.warning(

            f"[PDF CREATE COMPLETE] "

            f"created={created_count} "

            f"| skipped={skipped_count}"
        )


        # =====================================================
        # NEXT STATE
        # =====================================================

        if self.last_created_page >= len(ai_pages):

            _logger.warning(
                "[PDF FLOW] DONE ✅"
            )

            self.state = 'done'

        else:

            _logger.warning(
                "[PDF FLOW] CONTINUE"
            )

            self.state = 'pdf_creating'


        self.flush_recordset()

        self.env.cr.commit()

  
    # =====================================================
    # TRANSLATION HELPER
    # =====================================================

    def _translate_product_fields(
        self,
        name,
        description=""
    ):

        from openai import OpenAI

        api_key = self.env[
            'ir.config_parameter'
        ].sudo().get_param(
            'openai.api.key'
        )

        if not api_key:

            return {
                "en_US": {
                    "name": name,
                    "description": description,
                }
            }

        try:

            client = OpenAI(
                api_key=api_key
            )

            prompt = f"""
            Translate this product content.

            RULES:
            - Keep branding unchanged
            - Keep SKU unchanged
            - Keep technical values unchanged
            - Return JSON ONLY

            OUTPUT:

            {{
                "en_US": {{
                    "name": "",
                    "description": ""
                }},
                "ru_RU": {{
                    "name": "",
                    "description": ""
                }},
                "az_AZ": {{
                    "name": "",
                    "description": ""
                }}
            }}

            PRODUCT NAME:
            {name}

            DESCRIPTION:
            {description}
            """

            response = client.responses.create(

                model="gpt-4.1-mini",

                input=prompt,

                timeout=40
            )

            result = (
                response.output_text or ""
            ).strip()

            result = result.replace(
                "```json",
                ""
            )

            result = result.replace(
                "```",
                ""
            ).strip()

            translated = json.loads(result)

            return translated

        except Exception as e:

            _logger.exception(
                f"[TRANSLATION ERROR] {str(e)}"
            )

            return {
                "en_US": {
                    "name": name,
                    "description": description,
                }
            }

    #==========create excel product==========================
    def create_products_excel(self):

        import json
        import re

        _logger.warning(
            "[EXCEL CREATE] START"
        )


        # =====================================================
        # VALIDATION
        # =====================================================

        if not self.ai_response:

            _logger.warning(
                "[EXCEL CREATE] NO AI RESPONSE"
            )

            return


        try:

            ai_pages = json.loads(
                self.ai_response or "[]"
            )

        except Exception as e:

            _logger.exception(

                f"[EXCEL CREATE] "

                f"INVALID AI JSON "

                f"| {str(e)}"
            )

            return


        if not ai_pages:

            _logger.warning(
                "[EXCEL CREATE] EMPTY AI"
            )

            return


        ai_page = ai_pages[0]

        products = ai_page.get(
            "products",
            []
        )


        _logger.warning(

            f"[EXCEL CREATE] "

            f"RAW PRODUCTS={len(products)}"
        )


        if not products:

            return


        # =====================================================
        # MODELS
        # =====================================================

        product_obj = self.env[
            'product.template'
        ]

        category_obj = self.env[
            'product.category'
        ]

        attribute_obj = self.env[
            'product.attribute'
        ]

        attribute_value_obj = self.env[
            'product.attribute.value'
        ]

        line_obj = self.env[
            'product.template.attribute.line'
        ]


        # =====================================================
        # ROOT CATEGORY
        # =====================================================

        parent_category = category_obj.search([

            ('name', '=', "All Products")

        ], limit=1)


        if not parent_category:

            parent_category = category_obj.create({

                'name': "All Products"

            })


        # =====================================================
        # CATEGORY MAP
        # =====================================================

        CATEGORY_MAPPING = {

            "t-shirt": "Apparel",
            "shirt": "Apparel",
            "polo": "Apparel",
            "bag": "Bags",
            "backpack": "Bags",
            "cap": "Headwear",
            "hat": "Headwear",
            "bottle": "Drinkware",
            "drinkware": "Drinkware",
            "pen": "Stationery",
            "notebook": "Stationery",
            "powerbank": "Electronics",
            "charger": "Electronics",
            "laptop": "Electronics",
        }


        # =====================================================
        # GROUP PRODUCTS
        # =====================================================

        grouped_products = {}


        for p in products:

            raw_name = (
                p.get("name") or ""
            ).strip()


            variant_group = (
                p.get("variant_group")
            )


            if variant_group:

                group_id = str(
                    variant_group
                ).strip().upper()

            else:

                match = re.search(

                    r'(?:Product\s*)?([A-Z]*\d+)',

                    raw_name,

                    re.I
                )


                if match:

                    group_id = (
                        match.group(1)
                        .upper()
                    )

                else:

                    group_id = (
                        raw_name.upper()
                    )


            grouped_products.setdefault(

                group_id,

                []

            ).append(p)


        grouped_keys = list(
            grouped_products.keys()
        )


        _logger.warning(

            f"[EXCEL GROUPS] "

            f"TOTAL={len(grouped_keys)}"
        )


        # =====================================================
        # BATCH GROUPS
        # =====================================================

        BATCH_SIZE = 10

        start = (
            self.excel_created_index or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(grouped_keys)
        )


        _logger.warning(

            f"[EXCEL BATCH] "

            f"{start} → {end}"
        )


        created_count = 0
        merged_count = 0


        # =====================================================
        # PROCESS GROUPS
        # =====================================================

        for group_idx in range(start, end):

            try:

                group_id = grouped_keys[
                    group_idx
                ]

                group_items = grouped_products[
                    group_id
                ]


                _logger.warning(

                    f"[EXCEL GROUP] "

                    f"{group_id} "

                    f"| items={len(group_items)}"
                )


                main_product = (
                    group_items[0]
                )


                name = (
                    main_product.get(
                        "name"
                    ) or ""
                ).strip()


                description = (
                    main_product.get(
                        "description"
                    ) or ""
                )


                # =================================================
                # TRANSLATIONS
                # =================================================

                translations = self._translate_product_fields(
                    name=name,
                    description=description
                )


                raw_category = (

                    main_product.get(
                        "category"
                    ) or ""

                ).lower()


                mapped_category = (
                    "General"
                )


                for key, val in CATEGORY_MAPPING.items():

                    if key in raw_category:

                        mapped_category = val

                        break


                category = category_obj.search([

                    (
                        'name',
                        '=',
                        mapped_category
                    ),

                    (
                        'parent_id',
                        '=',
                        parent_category.id
                    )

                ], limit=1)


                if not category:

                    category = category_obj.create({

                        'name':
                            mapped_category,

                        'parent_id':
                            parent_category.id
                    })


                # =================================================
                # FIND PARENT PRODUCT
                # =================================================

                vendor_id = (
                    self.partner_id.id
                    if self.partner_id
                    else False
                )


                # ================================================
                # FIND BY PRODUCT CODE FIRST
                # ================================================

                existing_products = product_obj.search([

                    (
                        'default_code',
                        '=',
                        group_id
                    )

                ])


                product = False


                for existing in existing_products:

                    existing_vendor = (

                        existing.vendor_id.id

                        if existing.vendor_id

                        else False
                    )


                    _logger.warning(

                        f"[EXCEL CHECK] "

                        f"group={group_id} "

                        f"| existing_product={existing.id} "

                        f"| existing_vendor={existing_vendor} "

                        f"| current_vendor={vendor_id}"
                    )


                    # ============================================
                    # SAME PRODUCT + SAME VENDOR
                    # ============================================

                    if existing_vendor == vendor_id:

                        product = existing

                        _logger.warning(

                            f"[EXCEL DUPLICATE FOUND] "

                            f"{group_id} "

                            f"| vendor={vendor_id} "

                            f"| product_id={existing.id}"
                        )

                        break


                # =================================================
                # CREATE PARENT
                # =================================================

                if not product:

                    vals = {

                        'name': translations.get(
                            'en_US',
                            {}
                        ).get(
                            'name',
                            name
                        ),

                        'default_code':
                            group_id,

                        'description_sale':
                            description,

                        'categ_id':
                            category.id,

                        'sale_ok': True,

                        'website_published':
                            False,

                        # =====================================
                        # SAVE VENDOR LINK
                        # =====================================

                        'vendor_id':
                            vendor_id,

                        'list_price':
                            self._safe_float(
                                main_product.get("price")
                            ),
                    }


                    image = main_product.get(
                        "image"
                    )


                    if image:

                        vals[
                            'image_1920'
                        ] = image


                    product = product_obj.create(
                        vals
                    )


                    # =================================================
                    # SAVE TRANSLATIONS
                    # =================================================

                    try:

                        # RUSSIAN

                        ru = translations.get(
                            'ru_RU',
                            {}
                        )

                        if ru:

                            product.with_context(
                                lang='ru_RU'
                            ).write({

                                'name': ru.get(
                                    'name',
                                    name
                                ),

                                'description_sale': ru.get(
                                    'description',
                                    description
                                )
                            })


                        # AZERBAIJANI

                        az = translations.get(
                            'az_AZ',
                            {}
                        )

                        if az:

                            product.with_context(
                                lang='az_AZ'
                            ).write({

                                'name': az.get(
                                    'name',
                                    name
                                ),

                                'description_sale': az.get(
                                    'description',
                                    description
                                )
                            })


                        _logger.warning(

                            f"[TRANSLATIONS SAVED] "

                            f"{product.id}"
                        )


                    except Exception as e:

                        _logger.exception(

                            f"[TRANSLATION SAVE ERROR] "

                            f"{str(e)}"
                        )


                    created_count += 1


                    _logger.warning(

                        f"[EXCEL CREATED] "

                        f"{group_id} "

                        f"| vendor={vendor_id}"
                    )


                else:

                    merged_count += 1


                    _logger.warning(

                        f"[EXCEL EXISTING PRODUCT] "

                        f"{group_id} "

                        f"| vendor={vendor_id} "

                        f"| product_id={product.id}"
                    )              

                # =================================================
                # VARIANTS
                # =================================================

                for idx, item in enumerate(
                    group_items
                ):

                    attr_value = (
                        f"Variant {idx+1}"
                    )


                    attribute = attribute_obj.search([

                        (
                            'name',
                            '=',
                            "Variant"
                        )

                    ], limit=1)


                    if not attribute:

                        attribute = (
                            attribute_obj.create({

                                'name':
                                    "Variant"
                            })
                        )


                    value = (
                        attribute_value_obj.search([

                            (
                                'name',
                                '=',
                                attr_value
                            ),

                            (
                                'attribute_id',
                                '=',
                                attribute.id
                            )

                        ], limit=1)
                    )


                    if not value:

                        value = (
                            attribute_value_obj.create({

                                'name':
                                    attr_value,

                                'attribute_id':
                                    attribute.id
                            })
                        )


                    line = line_obj.search([

                        (
                            'product_tmpl_id',
                            '=',
                            product.id
                        ),

                        (
                            'attribute_id',
                            '=',
                            attribute.id
                        )

                    ], limit=1)


                    if not line:

                        line_obj.create({

                            'product_tmpl_id':
                                product.id,

                            'attribute_id':
                                attribute.id,

                            'value_ids': [

                                (
                                    6,
                                    0,
                                    [value.id]
                                )
                            ]
                        })


                        _logger.warning(

                            f"[VARIANT LINE CREATED] "

                            f"{group_id}"
                        )

                    else:

                        if value.id not in line.value_ids.ids:

                            line.value_ids = [

                                (
                                    4,
                                    value.id
                                )
                            ]


                            _logger.warning(

                                f"[VARIANT ADDED] "

                                f"{group_id} "

                                f"| {attr_value}"
                            )


                    # =============================================
                    # VARIANT IMAGE
                    # =============================================

                    variant_record = self.env[
                        'product.product'
                    ].search([

                        (
                            'product_tmpl_id',
                            '=',
                            product.id
                        ),

                        (
                            'product_template_attribute_value_ids.product_attribute_value_id',
                            '=',
                            value.id
                        )

                    ], limit=1)


                    if variant_record:

                        variant_image = item.get(
                            "image"
                        )


                        if variant_image:

                            variant_record.image_1920 = (
                                variant_image
                            )


                            _logger.warning(

                                f"[VARIANT IMAGE] "

                                f"{group_id} "

                                f"| {attr_value}"
                            )


                # =================================================
                # SAVE PROGRESS
                # =================================================

                self.excel_created_index = (
                    group_idx + 1
                )


                self.flush_recordset()

                self.env.cr.commit()


                _logger.warning(

                    f"[EXCEL SAVE] "

                    f"index="

                    f"{self.excel_created_index}"
                )


            except Exception as e:

                _logger.exception(

                    f"[EXCEL GROUP ERROR] "

                    f"group_idx={group_idx} "

                    f"| {str(e)}"
                )

                self.env.cr.rollback()


        # =====================================================
        # FINAL LOG
        # =====================================================

        _logger.warning(

            f"[EXCEL COMPLETE] "

            f"created={created_count} "

            f"| merged={merged_count}"
        )


        # ======================================================
        # NEXT STATE
        # ======================================================

        if self.excel_created_index >= len(grouped_keys):

            _logger.warning(

                "[EXCEL FLOW] "

                "GROUP BATCH COMPLETE "
                "→ RETURN TO excel_parsing"
            )

            # reset AI/create cycle
            self.excel_created_index = 0
            self.excel_ai_index = 0

            # IMPORTANT
            # clear AI batch only
            self.ai_response = False

            # continue parser batching
            self.state = 'excel_parsing'

            _logger.warning(

                "[EXCEL FLOW] "

                f"NEXT PARSE INDEX="

                f"{self.excel_parse_index}"
            )


        else:

            self.state = 'excel_creating'


        self.flush_recordset()

        self.env.cr.commit()
    

    #-----URL API FLOW-------------------------------------------

    def scrape_with_playwright(self):

        from playwright.sync_api import sync_playwright
        import subprocess
        import os

        _logger.warning(f"PLAYWRIGHT SCRAPE → {self.data_url}")

        #✅ CHECK IF BROWSER EXISTS FIRST
        browser_path = os.path.expanduser("~/.cache/ms-playwright")

        if not os.path.exists(browser_path):
            _logger.warning("PLAYWRIGHT → INSTALLING BROWSER (FIRST RUN)")

            try:
                subprocess.run(
                    ["python", "-m", "playwright", "install", "chromium"],
                    check=True
                )
                _logger.warning("PLAYWRIGHT BROWSER INSTALLED")
            except Exception as e:
                _logger.error(f"PLAYWRIGHT INSTALL FAILED → {str(e)}")
                return

        products = []

        with sync_playwright() as p:

            browser = p.chromium.launch(headless=True)
            page = browser.new_page()

            page.goto(self.data_url, timeout=60000)
            page.wait_for_timeout(5000)

            # ✅ Cookie handling
            try:
                page.locator("button:has-text('Accept')").click(timeout=3000)
            except:
                pass

            items = page.query_selector_all("a, div")

            _logger.warning(f"PLAYWRIGHT ELEMENTS FOUND → {len(items)}")

            for item in items[:200]:

                try:
                    text = item.inner_text().strip()

                    if not text or len(text) < 5:
                        continue

                    img_el = item.query_selector("img")
                    img_url = img_el.get_attribute("src") if img_el else None

                    if img_url and img_url.startswith("//"):
                        img_url = "https:" + img_url

                    img_base64 = None

                    if img_url:
                        try:
                            res = requests.get(img_url, timeout=10)
                            if res.status_code == 200:
                                img_base64 = base64.b64encode(res.content).decode("utf-8")
                        except:
                            pass

                    products.append({
                        "name": text[:120],
                        "image": img_base64
                    })

                except:
                    continue

            browser.close()

        _logger.warning(f"PLAYWRIGHT PRODUCTS → {len(products)}")

        if not products:
            _logger.error("PLAYWRIGHT FAILED → NO PRODUCTS")
            return

        pages = [{
            "page": 1,
            "text": "\n".join([p["name"] for p in products]),
            "images": [p["image"] for p in products if p["image"]]
        }]

        self.extracted_text = json.dumps(pages)

        _logger.warning(f"PLAYWRIGHT DONE → {len(products)} PRODUCTS")

    # =====================================================
    # CRON PROCESSOR
    # =====================================================
    def run_pending_jobs(self):

        from odoo import fields

        _logger.warning(
            "🔥 CRON HEARTBEAT → RUNNING"
        )


        active_states = [

            'draft',

            'excel_parsing',
            'excel_ai',
            'excel_creating',

            'pdf_extracting',
            'pdf_ai',
            'pdf_creating',

            'url_scraping',
            'url_ai',
            'url_creating',
        ]


        # =================================================
        # LOAD JOBS
        # =================================================

        jobs = self.search(

            [('state', 'in', active_states)],

            order="id desc"
        )


        _logger.warning(

            f"[CRON] ACTIVE JOBS "

            f"→ {len(jobs)}"
        )


        # =================================================
        # REMOVE DUPLICATES
        # =================================================

        seen = {}

        duplicates = self.env[
            'vendor.import.job'
        ]


        for j in jobs:

            sig = j.upload_signature

            if not sig:

                continue


            if sig not in seen:

                seen[sig] = j

            else:

                if j.id > seen[sig].id:

                    duplicates |= seen[sig]

                    seen[sig] = j

                else:

                    duplicates |= j


        if duplicates:

            _logger.warning(

                f"[CRON] REMOVING DUPLICATES "

                f"→ {len(duplicates)}"
            )

            try:

                duplicates.unlink()

                self.env.cr.commit()

                _logger.warning(
                    "[CRON] DUPLICATES REMOVED"
                )

            except Exception as e:

                _logger.exception(

                    f"[CRON ERROR] "

                    f"DUPLICATE DELETE FAILED "

                    f"→ {str(e)}"
                )


        # =================================================
        # RECOVER STALE LOCKS
        # =================================================

        stale_jobs = self.search([

            ('state', 'in', active_states),

            ('lock', '=', True)

        ])


        for stale in stale_jobs:

            try:

                delta = (

                    fields.Datetime.now()

                    - stale.write_date

                ).total_seconds()

            except Exception:

                delta = 0


            _logger.warning(

                f"[LOCK CHECK] "

                f"job={stale.id} "

                f"| seconds={delta}"
            )


            if delta > 60:

                _logger.warning(

                    f"[STALE LOCK RESET] "

                    f"job={stale.id}"
                )

                try:

                    stale.lock = False

                    self.env.cr.commit()

                except Exception as e:

                    _logger.exception(

                        f"[STALE LOCK ERROR] "

                        f"{str(e)}"
                    )


        # =================================================
        # GET NEXT JOB
        # =================================================

        job = self.search(

            [

                ('state', 'in', active_states),

                ('lock', '=', False)

            ],

            order="create_date asc, id asc",

            limit=1
        )


        if not job:

            _logger.warning(

                "[CRON] NO AVAILABLE JOBS "
                "(all locked or done)"
            )

            return


        # =================================================
        # PROCESS
        # =================================================

        try:

            # =============================================
            # LOCK
            # =============================================

            job.lock = True

            self.env.cr.commit()


            _logger.warning(

                f"[CRON] JOB LOCKED "

                f"| job={job.id}"
            )


            # =============================================
            # SAFER CHAIN
            # =============================================

            MAX_CHAIN = 1


            for step in range(MAX_CHAIN):

                # =========================================
                # REFRESH
                # =========================================

                try:

                    job.invalidate_cache()

                except Exception:

                    pass


                job = self.env[
                    'vendor.import.job'
                ].browse(job.id)


                _logger.warning(

                    f"[CHAIN] STEP "

                    f"{step + 1} "

                    f"| state={job.state}"
                )


                # =========================================
                # STOP STATES
                # =========================================

                if job.state in [

                    'done',
                    'failed'

                ]:

                    _logger.warning(

                        f"[CHAIN STOP] "

                        f"terminal state "

                        f"→ {job.state}"
                    )

                    break


                # =========================================
                # TRACK BEFORE
                # =========================================

                previous_state = (
                    job.state
                )

                previous_page = (
                    job.current_page or 0
                )

                previous_ai_page = (
                    job.last_ai_page or 0
                )

                previous_created = (
                    job.last_created_page or 0
                )

                previous_excel_ai = (
                    job.excel_ai_index or 0
                )

                previous_excel_created = (
                    job.excel_created_index or 0
                )

            
                previous_url_batch = (
                    job.url_batch_index or 0
                )

                previous_url_created = (
                    job.last_processed_product_index or 0
                )
             

                _logger.warning(

                    f"[CHAIN BEFORE] "

                    f"state={previous_state} "

                    f"| extract={previous_page} "

                    f"| ai={previous_ai_page} "

                    f"| create={previous_created} "

                    f"| excel_ai={previous_excel_ai} "

                    f"| excel_create={previous_excel_created}"
                )


                # =========================================
                # PROCESS
                # =========================================

                try:

                    job._process_step()

                except Exception as e:

                    _logger.exception(

                        f"[PROCESS ERROR] "

                        f"job={job.id} "

                        f"| {str(e)}"
                    )

                    try:

                        job.state = 'failed'

                        self.env.cr.commit()

                    except Exception:

                        _logger.warning(

                            "[PROCESS ERROR] "

                            "FAILED SAVE FAILED"
                        )

                    break


                # =========================================
                # REFRESH AFTER
                # =========================================

                try:

                    job.invalidate_cache()

                except Exception:

                    pass


                job = self.env[
                    'vendor.import.job'
                ].browse(job.id)


                _logger.warning(

                    f"[CHAIN AFTER] "

                    f"state={job.state} "

                    f"| extract={job.current_page} "

                    f"| ai={job.last_ai_page} "

                    f"| create={job.last_created_page} "

                    f"| excel_ai={job.excel_ai_index} "

                    f"| excel_create={job.excel_created_index}"
                )


                # =========================================
                # PROGRESS DETECTION
                # =========================================

                progress_detected = False


                # PDF extract progress

                if (

                    (job.current_page or 0)

                    >

                    previous_page

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] PDF "

                        f"{previous_page}"

                        f" → "

                        f"{job.current_page}"
                    )


                # PDF AI progress

                if (

                    (job.last_ai_page or 0)

                    >

                    previous_ai_page

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] PDF AI "

                        f"{previous_ai_page}"

                        f" → "

                        f"{job.last_ai_page}"
                    )


                # PDF create progress

                if (

                    (job.last_created_page or 0)

                    !=

                    previous_created

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] PDF CREATE "

                        f"{previous_created}"

                        f" → "

                        f"{job.last_created_page}"
                    )


                # PDF create state continuation

                elif job.state == 'pdf_creating':

                    progress_detected = True

                    _logger.warning(

                        "[PROGRESS] PDF CREATE LOOP ACTIVE"
                    )
            

                # Excel AI progress

                if (

                    (job.excel_ai_index or 0)

                    >

                    previous_excel_ai

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] EXCEL AI "

                        f"{previous_excel_ai}"

                        f" → "

                        f"{job.excel_ai_index}"
                    )


                # Excel create progress

                if (

                    (job.excel_created_index or 0)

                    >

                    previous_excel_created

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] EXCEL CREATE "

                        f"{previous_excel_created}"

                        f" → "

                        f"{job.excel_created_index}"
                    )

             
                # =========================================
                # URL AI progress
                # =========================================

                if (

                    (job.url_batch_index or 0)

                    >

                    previous_url_batch

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] URL AI "

                        f"{previous_url_batch}"

                        f" → "

                        f"{job.url_batch_index}"
                    )


                # =========================================
                # URL create progress
                # =========================================

                if (

                    (job.last_processed_product_index or 0)

                    >

                    previous_url_created

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] URL CREATE "

                        f"{previous_url_created}"

                        f" → "

                        f"{job.last_processed_product_index}"
                    )


                # =========================================
                # APIFY WAIT STATE
                # =========================================

                if job.state == 'url_scraping':

                    progress_detected = True

                    _logger.warning(

                        "[PROGRESS] APIFY WAITING"
                    )
              


                # state transition

                if previous_state != job.state:

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] STATE "

                        f"{previous_state}"

                        f" → "

                        f"{job.state}"
                    )


                # =========================================
                # STOP IF NO PROGRESS
                # =========================================

                if not progress_detected:

                    _logger.warning(

                        "[CHAIN STOP] "

                        "NO PROGRESS DETECTED"
                    )

                    break


                _logger.warning(

                    "[CHAIN CONTINUE] "

                    "PROGRESS DETECTED"
                )


                # =========================================
                # COMMIT
                # =========================================

                try:

                    self.env.cr.commit()

                    _logger.warning(
                        "[CHAIN] COMMIT OK"
                    )

                except Exception as e:

                    _logger.exception(

                        f"[CHAIN COMMIT ERROR] "

                        f"{str(e)}"
                    )

                    break


            _logger.warning(
                "[CRON] PROCESS LOOP COMPLETE"
            )


        except Exception as e:

            _logger.exception(

                f"[CRON FATAL ERROR] "

                f"{str(e)}"
            )


            try:

                self.env.cr.rollback()

                _logger.warning(
                    "[CRON] ROLLBACK OK"
                )

            except Exception:

                _logger.warning(
                    "[CRON] ROLLBACK FAILED"
                )


        finally:

            # =============================================
            # UNLOCK
            # =============================================

            try:

                if (

                    job

                    and

                    job.exists()

                ):

                    job.lock = False

                    self.env.cr.commit()


                    _logger.warning(

                        f"[CRON] JOB UNLOCKED "

                        f"| job={job.id}"
                    )

            except Exception:

                _logger.warning(

                    "[CRON] UNLOCK FAILED"
                )


        _logger.warning(
            "[CRON] RUN COMPLETE"
        )

        return


   #=============flask setup/installation=================== 
    def ping_flask_server(self):
      
        try:
            requests.get("https://pdf-extractor-staging.onrender.com", timeout=10)
            _logger.info("FLASK PING SUCCESS")
        except Exception:
            _logger.warning("FLASK PING FAILED")


    # #---------------clean_scraped_blocks-------------------------------
    def _clean_scraped_blocks(self, raw_blocks):
        """
        Clean Apify output before sending to AI
        """

        cleaned = []
        seen = set()

        for item in raw_blocks:

            text = (item.get("text") or "").strip()
            image = item.get("image")

            # ❌ REMOVE NOISE
            if not text:
                continue

            if len(text) < 15:
                continue

            if any(x in text.lower() for x in [
                "privacy", "cookie", "terms", "login",
                "menu", "navigation", "home"
            ]):
                continue

            # ❌ REMOVE DUPLICATES
            key = text[:120]  # allow more variation

            if key in seen:
                continue

            seen.add(key)

            # cleaned.append({
            #     "text": text,
            #     "image": image
            # })

            cleaned.append({
                "text": text,
                "image": image,
                "price": item.get("price", ""),
                "stock": item.get("stock", "")
            })

        return cleaned

    #---------------normalizer-------------------------------

    def _normalize_url_data(self, items):

        blocks = []

        for item in items:

            # =====================================================
            # FORMAT 1 → ORIGINAL WORKING FORMAT
            # =====================================================
            # {
            #   "text": "...",
            #   "image": "..."
            # }
            # =====================================================

            if item.get("text"):

                text = (item.get("text") or "").strip()
                image = item.get("image")

                # 🔥 STRICT VALIDATION
                if not text or len(text) < 5:
                    continue

                if (
                    image and
                    isinstance(image, str) and
                    not image.startswith("http")
                ):
                    image = None

                blocks.append({
                    "text": text,
                    "image": image,
                    "price": item.get("price", ""),
                    "stock": item.get("stock", "")
                })

                continue

            # =====================================================
            # FORMAT 2 → STRUCTURED FORMAT
            # =====================================================

            if item.get("type") == "PRODUCTS":

                for sub in item.get("items", []):

                    text = (
                        sub.get("text") or ""
                    ).strip()

                    image = sub.get("image")

                    if not text or len(text) < 5:
                        continue

                    if (
                        image and
                        isinstance(image, str) and
                        not image.startswith("http")
                    ):
                        image = None

                    # blocks.append({
                    #     "text": text,
                    #     "image": image
                    # })

                    blocks.append({
                        "text": text,
                        "image": image,
                        "price": sub.get("price", ""),
                        "stock": sub.get("stock", "")
                    })

            # =====================================================
            # DEBUG TYPES
            # =====================================================

            elif item.get("type") in [
                "EMPTY",
                "BLOCKED"
            ]:

                _logger.error(
                    f"URL DEBUG → "
                    f"{item.get('reason')}"
                )

        _logger.warning(f"NORMALIZED BLOCKS → {len(blocks)}")

        # =====================================================
        # 🔥 SPLIT INTO MULTIPLE PAGES (CRITICAL FIX)
        # =====================================================

        PAGE_SIZE = 20  # 🔥 prevents AI overload

        pages = []

        for i in range(0, len(blocks), PAGE_SIZE):

            chunk = blocks[i:i + PAGE_SIZE]

            pages.append({
                "page": len(pages) + 1,
                "blocks": chunk
            })

        _logger.warning(f"NORMALIZED PAGES → {len(pages)}")

        return pages
    

    #======apify url fetch/scrapp products===============
    
    def _run_apify_actor(self, url):

        token = self.env['ir.config_parameter'].sudo().get_param('apify.api_token')

        if not token:
            raise Exception("Apify API token not configured")

        ACTOR_ID = "selectad~my-actor"

        # =====================================================
        # 🔥 STEP 1: START ACTOR (ONLY IF NOT STARTED)
        # =====================================================

        if not getattr(self, "apify_run_id", False):

            run_url = f"https://api.apify.com/v2/acts/{ACTOR_ID}/runs?token={token}"

            payload = {
                "startUrls": [{"url": url}]
            }

            headers = {
                "Content-Type": "application/json"
            }

            response = requests.post(run_url, json=payload, headers=headers, timeout=30)

            if response.status_code != 201:
                raise Exception(f"Apify run failed: {response.text}")

            run_data = response.json()

            # ✅ SAVE FOR NEXT CRON
            self.apify_run_id = run_data["data"]["id"]
            self.apify_dataset_id = run_data["data"]["defaultDatasetId"]

            _logger.warning(f"APIFY STARTED → RUN ID {self.apify_run_id}")

            # 🔥 IMPORTANT: STOP HERE (NON-BLOCKING)
            return None

        # =====================================================
        # 🔥 STEP 2: CHECK STATUS
        # =====================================================

        status_url = f"https://api.apify.com/v2/actor-runs/{self.apify_run_id}?token={token}"

        status_res = requests.get(status_url, timeout=20).json()
        status = status_res["data"]["status"]

        _logger.warning(f"APIFY STATUS → {status}")

        if status in ["RUNNING", "READY"]:
            _logger.warning("APIFY STILL RUNNING → WAIT NEXT CRON")
            return None

        if status in ["FAILED", "ABORTED", "TIMED-OUT"]:
            raise Exception(f"Apify run failed with status: {status}")

        # =====================================================
        # 🔥 STEP 3: FETCH DATA (ONLY WHEN DONE)
        # =====================================================

        dataset_url = f"https://api.apify.com/v2/datasets/{self.apify_dataset_id}/items"

        params = {
            "token": token,
            "limit": 1000,
            "clean": "true"
        }

        dataset_res = requests.get(dataset_url, params=params, timeout=30)

        if dataset_res.status_code != 200:
            raise Exception(f"Failed to fetch dataset: {dataset_res.text}")

        data = dataset_res.json()

        _logger.warning(f"APIFY ITEMS FETCHED → {len(data)}")

        if not data:
            _logger.warning("APIFY RETURNED EMPTY → MARK JOB AS DONE")

            self.state = 'done'   # 🔥 STOP LOOP COMPLETELY
            self.env.cr.commit()
            return

        # 🔥 CLEAN UP (IMPORTANT)
        self.apify_run_id = False
        self.apify_dataset_id = False

        return data

    #=======validation===================
    def validate_ai_output(products):
        for p in products:
            if "variants" in p:
                if not isinstance(p["variants"], list):
                    p["variants"] = []

                for v in p["variants"]:
                    if "attributes" not in v:
                        v["attributes"] = {"Variant": "Default"}

        return products
    
    #=======keep cron alive================
    def keep_alive(self):
        _logger.warning("KEEP ALIVE PING")

   
   #=========gloat numbers=============
    def _safe_float(self, value):

        try:

            if value is None:
                return 0.0

            value = str(value)

            value = value.replace('$', '')
            value = value.replace('€', '')
            value = value.replace('£', '')
            value = value.replace(',', '')

            return float(value.strip())

        except:
            return 0.0