#old working backup copy

from odoo import models, fields
import base64
import logging
import io
import requests
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from PIL import Image
import time
import json
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openai import OpenAI
import re
import fitz

 

_logger = logging.getLogger(__name__)

# ✅ Extend existing model
class ResPartner(models.Model):
    _inherit = 'res.partner'

    #Vendor user role
    is_vendor_user = fields.Boolean(
        string="Vendor User",
        default=False
    )


class VendorImportJob(models.Model):

    _name = "vendor.import.job"
    _description = "Vendor Import Job"

    partner_id = fields.Many2one("res.partner", string="Vendor")  # ✅ LINK instead

    name = fields.Char(default="Vendor Data Import")

    data_url = fields.Char()
    #data_url = fields.Char(string="Vendor URL")
    extra_info = fields.Text()

    pdf_file = fields.Binary()
    excel_file = fields.Binary()
    logo_file = fields.Binary()

    extracted_text = fields.Text()
    current_page = fields.Integer(
        string="Current PDF Page",
        default=0
    )
    total_pages = fields.Integer(string="Total Pages", default=0)
    last_ai_page = fields.Integer(string="Last AI Page", default=0)
    ai_response = fields.Text()
    priority = fields.Integer(default=10)

    apify_run_id = fields.Char()
    apify_dataset_id = fields.Char()
    url_batch_index = fields.Integer(default=0)
    last_processed_product_index = fields.Integer(default=0)
    last_created_page = fields.Integer(default=0)
    lock = fields.Boolean(default=False)
    is_excel_parsed = fields.Boolean(default=False)
    excel_ai_index = fields.Integer(default=0)
    upload_signature = fields.Char(string="Upload Signature")

    state = fields.Selection([
        ('draft', 'Draft'),
        ('processing', 'Processing'),
        ('ai_processing', 'AI Processing'),
        ('review', 'Vendor Review'),
        ('done', 'Completed'),
        ('error', 'Error'),
         ('failed', 'Failed'),
    ], default='draft')

    
    #------------parse url----------------------------
    def parse_url(self):

        _logger.warning(f"APIFY SCRAPE → {self.data_url}")

        raw_data = self._run_apify_actor(self.data_url)

        _logger.warning(f"RAW APIFY DATA SAMPLE → {str(raw_data)[:200]}")

        # 🔥 VERY IMPORTANT FIX
        if raw_data is None:
            _logger.warning("APIFY NOT READY → WAIT NEXT CRON")
            return

        if not raw_data:
            _logger.error("APIFY FAILED → EMPTY DATASET")
            return

        structured_data = self._normalize_url_data(raw_data)

        if not structured_data:
            _logger.error("NORMALIZATION FAILED → EMPTY DATA")
            return

        # ✅ Convert to same format used by Excel/PDF
        self.extracted_text = json.dumps(structured_data)

        _logger.warning(f"APIFY DONE → {len(structured_data)} ITEMS")
      


    #------excel processing method---------------
    
    def parse_excel(self):

        _logger.warning("EXCEL → START PARSING")

        excel_bytes = base64.b64decode(self.excel_file)

        wb = load_workbook(filename=BytesIO(excel_bytes))

        headers = {"User-Agent": "Mozilla/5.0"}

        pages = []
        page_number = 1

        # 🔥 PROCESS ALL SHEETS
        for sheet in wb.worksheets:

            _logger.warning(f"PROCESSING SHEET → {sheet.title}")

            image_loader = SheetImageLoader(sheet)

            for idx, row in enumerate(sheet.iter_rows()):

                row_text_parts = []
                row_images = []

                # ================= TEXT =================
                for cell in row:
                    val = str(cell.value or "").strip()
                    if val:
                        row_text_parts.append(val)

                # skip empty rows
                if not row_text_parts:
                    continue

                # skip header
                if idx == 0:
                    _logger.warning(f"SKIP HEADER ROW → {sheet.title}")
                    continue

                # ================= STRUCTURED TEXT =================
                row_text = f"""
                ROW_DATA:
                {" | ".join(row_text_parts)}

                RULE:
                - THIS IS EXACTLY ONE PRODUCT
                - DO NOT SPLIT THIS ROW
                - THIS ROW MAY BE A VARIANT OF ANOTHER ROW
                - USE SIMILAR ID/SKU TO GROUP VARIANTS
                """

                # ================= IMAGE (EMBEDDED FIRST) =================
                for cell in row:
                    try:
                        if image_loader.image_in(cell.coordinate):

                            pil_img = image_loader.get(cell.coordinate)

                            buffer = BytesIO()
                            pil_img.save(buffer, format="JPEG")

                            img_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

                            row_images.append(img_base64)

                            _logger.warning(f"ROW {idx} → EMBED IMAGE FOUND")
                            break

                    except Exception:
                        continue

                # ================= IMAGE (URL FALLBACK) =================
                if not row_images:

                    for cell in row:
                        val = str(cell.value or "").strip()

                        if val.startswith("http"):
                            try:
                                response = requests.get(val, headers=headers, timeout=10)

                                if response.status_code != 200:
                                    continue

                                if "image" in response.headers.get("Content-Type", ""):
                                    img_base64 = base64.b64encode(response.content).decode("utf-8")
                                    row_images.append(img_base64)

                                    _logger.warning(f"ROW {idx} → IMAGE FROM URL")
                                    break

                            except Exception:
                                _logger.warning(f"ROW {idx} → IMAGE URL FAILED")

                # ================= DEBUG =================
                _logger.warning(f"SHEET → {sheet.title} | ROW → {idx}")
                _logger.warning(f"TEXT PARTS → {len(row_text_parts)}")
                _logger.warning(f"IMAGES FOUND → {len(row_images)}")

                # ================= STORE =================
                pages.append({
                    "page": page_number,
                    "text": row_text,
                    "images": row_images,
                    "row_index": idx,
                    "sheet": sheet.title
                })

                page_number += 1

        # ================= FINAL =================
        self.extracted_text = json.dumps(pages)

        _logger.warning(f"EXCEL DONE → TOTAL ROWS: {len(pages)}")


    #---------------- MAIN FLOW ----------------
   
    def process_import(self):

        _logger.warning(f"PROCESS START → Job {self.id}")

        try:

            #================= URL FLOW =================
            if self.data_url:
                _logger.warning("FLOW → URL")

                self.parse_url()

                if not self.extracted_text:
                    _logger.warning("URL NOT READY → WAIT NEXT CRON")
                    return

                _logger.warning("STEP → SEND TO AI (URL)")
                self.send_to_openai_url()

                if not self.ai_response:
                    _logger.error("URL AI FAILED → STOP")
                    return

                _logger.warning("STEP → CREATE PRODUCTS (URL)")
                self.create_products_url()

            # ================= EXCEL FLOW =================
            elif self.excel_file:
                _logger.warning("FLOW → EXCEL")

                self.parse_excel()

                if not self.extracted_text:
                    _logger.error("EXCEL PARSE FAILED → STOP")
                    return

                _logger.warning("STEP → SEND TO AI (EXCEL)")
                self.send_to_openai_pdf_excel()

                # 🔥 CRITICAL FIX
                if not self.ai_response:
                    _logger.error("EXCEL AI FAILED → STOP")
                    return

                _logger.warning("STEP → CREATE PRODUCTS (EXCEL)")
                self.create_products_pdf_excel()

            # ================= PDF FLOW =================
            elif self.pdf_file:
                _logger.warning("FLOW → PDF")

                self.extract_pdf()

                if not self.extracted_text:
                    _logger.error("PDF EXTRACTION FAILED → STOP")
                    return

                _logger.warning("STEP → SEND TO AI (PDF)")
                self.send_to_openai_pdf_excel()

                if not self.ai_response:
                    _logger.error("PDF AI FAILED → STOP")
                    return

                _logger.warning("STEP → CREATE PRODUCTS (PDF)")
                self.create_products_pdf_excel()

            else:
                raise Exception("No input found")

            # ================= FINAL STATE CONTROL =================

            # ✅ URL FLOW
            if self.data_url:
                total_batches = getattr(self, "url_total_batches", 0)
                current_batch = getattr(self, "url_batch_index", 0)

                if total_batches and current_batch >= total_batches:
                    _logger.warning("URL → ALL BATCHES COMPLETED ✅")
                    self.state = 'done'
                else:
                    _logger.warning("URL → WAITING FOR NEXT BATCH")
                    self.state = 'processing'

            # ✅ PDF / EXCEL FLOW
            else:
                if self.current_page >= self.total_pages:
                    _logger.warning("PROCESS IMPORT → ALL PAGES COMPLETED")
                    self.state = 'done'
                else:
                    _logger.warning("PROCESS IMPORT → WAITING FOR NEXT BATCH")
                    self.state = 'processing'

        except Exception as e:
            _logger.error(f"PROCESS FAILED → {str(e)}")
            self.state = "failed"
   

    # ---------------- PDF ----------------

    def extract_pdf(self):

        _logger.warning("PDF → START EXTRACTION (BATCH MODE)")
        pdf_bytes = base64.b64decode(self.pdf_file)

        MAX_RETRIES = 3

        # 🔥 BATCH CONFIG
        BATCH_SIZE = 3
        PAGE_DELAY = 2
        BATCH_DELAY = 5

        all_pages = []

        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        except Exception as e:
            _logger.exception(f"PDF OPEN FAILED → {str(e)}")
            self.state = "failed"
            return

        total_pages = len(doc)
        self.total_pages = total_pages

        # 🔥 RESUME LOG (NEW)
        _logger.warning(f"RESUMING FROM PAGE → {self.current_page or 0}")

        start_page = self.current_page or 0
        end_page = min(start_page + BATCH_SIZE, total_pages)

        _logger.warning(f"PDF TOTAL PAGES → {total_pages}")
        _logger.warning(f"BATCH → Processing pages {start_page+1} to {end_page}")

        for i in range(start_page, end_page):

            page = doc[i]
            _logger.warning(f"PROCESSING PAGE {i + 1}")

            # ================= CREATE SINGLE PAGE PDF =================
            try:
                single_pdf = fitz.open()
                single_pdf.insert_pdf(doc, from_page=i, to_page=i)

                pdf_bytes_io = io.BytesIO()
                single_pdf.save(pdf_bytes_io)
                pdf_bytes_io.seek(0)

            except Exception as e:
                _logger.exception(f"FAILED TO SPLIT PAGE {i+1} → {str(e)}")
                continue

            # ================= CALL FLASK =================
            page_success = False

            for attempt in range(MAX_RETRIES):

                try:
                    _logger.warning(f"FLASK CALL PAGE {i+1} → ATTEMPT {attempt + 1}")

                    response = requests.post(
                        "https://pdf-extractor-staging.onrender.com/extract",
                        files={"file": ("page.pdf", pdf_bytes_io, "application/pdf")},
                        timeout=120
                    )

                    if response.status_code != 200:
                        _logger.warning(f"FLASK ERROR PAGE {i+1}: {response.status_code}")
                        continue

                    page_data = response.json()

                    # ================= FORMAT SUPPORT =================
                    if isinstance(page_data, dict):
                        pages = page_data.get("pages", [])
                    elif isinstance(page_data, list):
                        pages = page_data
                    else:
                        pages = []

                    if not pages:
                        _logger.warning(f"EMPTY PAGE DATA PAGE {i+1}")
                        continue

                    _logger.warning(f"PAGE {i+1} → RECEIVED {len(pages)} BLOCKS")

                    # ================= NORMALIZATION =================
                    for p in pages:

                        text = p.get("text", "")
                        images = p.get("images", [])

                        if not text and not images:
                            continue

                        all_pages.append({
                            "page": i + 1,
                            "text": text,
                            "images": images
                        })

                    page_success = True
                    break

                except Exception as e:
                    _logger.exception(f"FLASK CALL FAILED PAGE {i+1} → {str(e)}")

                time.sleep(5)

            if not page_success:
                _logger.error(f"PAGE {i+1} FAILED AFTER RETRIES")

            time.sleep(PAGE_DELAY)

        # ================= UPDATE PROGRESS =================
        self.current_page = end_page

        # ================= STORE =================
        try:
            existing = []

            if self.extracted_text:
                try:
                    existing = json.loads(self.extracted_text)
                except Exception:
                    existing = []

            combined = existing + all_pages

            self.extracted_text = json.dumps(combined)

            _logger.warning(f"BATCH STORED → {len(all_pages)} pages")
            _logger.warning(f"TOTAL STORED → {len(combined)} pages")
            _logger.warning(f"EXTRACTED DATA SIZE → {len(self.extracted_text)}")

        except Exception as e:
            _logger.exception(f"FAILED TO STORE DATA → {str(e)}")
            self.state = "failed"
            return

        # ================= 🔥 CRITICAL FIX =================
        if self.current_page < total_pages:

            _logger.warning(f"JOB NOT FINISHED → NEXT START PAGE {self.current_page + 1}")

            # 🔥 FORCE JOB TO CONTINUE
            self.state = "processing"

            time.sleep(BATCH_DELAY)

        else:

            _logger.warning("ALL PAGES PROCESSED ✅")
            self.state = "done"

        _logger.warning("PDF EXTRACTION BATCH COMPLETED")


    # ---------------- OPENAI ----------------
    def send_to_openai_url(self):

        import time
        import re
        import json

        self.state = "ai_processing"

        api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')

        if not api_key:
            raise Exception("OpenAI API key not configured")

        client = OpenAI(api_key=api_key)

        try:
            pages = json.loads(self.extracted_text or "[]")
        except Exception:
            _logger.error("INVALID extracted_text JSON")
            return

        if not pages:
            _logger.error("NO PAGES TO PROCESS")
            return

        # ================= LOAD EXISTING =================
        existing_products = []
        if self.ai_response:
            try:
                existing_products = json.loads(self.ai_response)
            except:
                existing_products = []

        current_batch = getattr(self, "url_batch_index", 0)

        # ================= FLATTEN =================
        all_blocks = [
            b for p in pages for b in p.get("blocks", [])
        ]

        _logger.warning(f"RAW BLOCKS → {len(all_blocks)}")

        # ================= CLEAN =================
        cleaned_blocks = self._clean_scraped_blocks(all_blocks)

        _logger.warning(f"CLEAN BLOCKS → {len(cleaned_blocks)}")
        _logger.warning(f"REMOVED BLOCKS → {len(all_blocks) - len(cleaned_blocks)}")

        # Sort for consistency
        cleaned_blocks = sorted(cleaned_blocks, key=lambda x: (x.get("text") or "")[:50])

        # ================= BATCH =================
        BLOCK_BATCH_SIZE = 20  # 🔥 SAFE for memory

        batched_blocks = [
            cleaned_blocks[i:i + BLOCK_BATCH_SIZE]
            for i in range(0, len(cleaned_blocks), BLOCK_BATCH_SIZE)
        ]

        total_batches = len(batched_blocks)

        _logger.warning(f"TOTAL BLOCK BATCHES → {total_batches}")
        _logger.warning(f"CURRENT BATCH → {current_batch}")

        # ================= STOP IF DONE =================
        if current_batch >= total_batches:
            _logger.warning("ALL URL BATCHES PROCESSED ✅")
            return

        # ================= PROCESS ONE BATCH =================
        block_batch = batched_blocks[current_batch]

        _logger.warning(f"PROCESSING BLOCK COUNT → {len(block_batch)}")
        _logger.warning(f"AI → PROCESSING BLOCK BATCH {current_batch + 1}")

        # 🔥 DO NOT LIMIT AGAIN (FIXED)
        combined_text = "\n\n---\n\n".join([
            f"{b.get('text','')}\nIMAGE_URL: {b.get('image','')}"
            for b in block_batch
        ])

        if not combined_text.strip():
            _logger.warning("EMPTY COMBINED TEXT → SKIP")
            return

        # ================= SAFETY LIMIT =================
        if len(combined_text) > 15000:
            combined_text = combined_text[:15000]
            _logger.warning("TEXT TRIMMED → PREVENT TOKEN OVERFLOW")

        # ================= PROMPT =================
        prompt = f"""
            You are a highly precise e-commerce product extraction engine.

            You are processing raw scraped website content.

            =====================================
            YOUR GOAL
            =====================================

            Extract ALL individual products from the text.

            IMPORTANT:
            - A single block may contain MULTIPLE products → you MUST split them
            - Each product MUST be returned separately
            - NEVER merge multiple products into one

            =====================================
            HOW TO IDENTIFY A PRODUCT
            =====================================

            A product usually contains:
            - product name
            - price (e.g. $, £, €, numbers)
            - specs or description
            - sometimes reviews

            Strong indicators:
            - currency symbols ($, £, €)
            - model names (Dell, Lenovo, etc.)
            - numbers like GB, inch, SSD, RAM, etc.

            =====================================
            STRICT RULES
            =====================================

            1. RETURN ONLY VALID JSON ARRAY
            2. NO explanation
            3. NO markdown
            4. NO text outside JSON

            5. EACH product must be unique
            6. REMOVE duplicates
            7. DO NOT skip products
            8. SPLIT combined text into multiple products

            =====================================
            REMOVE THIS NOISE
            =====================================

            Ignore anything related to:
            - navigation (Home, Login, Pricing)
            - cookies/privacy
            - footer text
            - categories only (no product info)
            - repeated headers

            =====================================
            OUTPUT FORMAT
            =====================================

            [
            {{
                "name": "Clean product name",
                "description": "Short product description (max 30 words)",
                "category": "Best guess category",
                "image": "image_url_or_null"
            }}
            ]

            =====================================
            EXTRA RULES
            =====================================

            - Keep names SHORT and CLEAN
            - Description must be concise
            - Infer category intelligently
            - If no image exists → return null
            - If unsure → still extract

            =====================================
            TEXT TO PROCESS
            =====================================

        {combined_text}
        """

        # ================= OPENAI =================
        try:
            response = client.responses.create(
                model="gpt-4.1-mini",
                input=prompt,
                temperature=0,
                timeout=60
            )

            result = response.output_text.strip()
            result = re.sub(r"^```(?:json)?|```$", "", result).strip()

            parsed = json.loads(result)

            if isinstance(parsed, list):

                cleaned = [p for p in parsed if p.get("name")]

                _logger.warning(f"AI RETURNED → {len(cleaned)} PRODUCTS")

                if len(cleaned) < 5:
                    _logger.warning("⚠️ LOW EXTRACTION → CHECK BLOCK QUALITY")

                existing_products.extend(cleaned)

                _logger.warning(f"TOTAL ACCUMULATED → {len(existing_products)}")

            else:
                _logger.warning("AI RESPONSE NOT LIST")

        except Exception as e:
            _logger.warning(f"AI ERROR → {str(e)}")
            return

        # ================= SAVE PROGRESS =================
        self.ai_response = json.dumps(existing_products)
        self.url_batch_index = current_batch + 1

        _logger.warning(f"NEXT BATCH INDEX → {self.url_batch_index}")

        # ================= SAFE EXIT =================
        _logger.warning("CRON EXIT → CONTINUE NEXT RUN")
        return

    
    #===========pdf and excel open ai OPENAI=========================

    def send_to_openai_pdf_excel(self):

        import json
        import time

        self.state = "ai_processing"

        api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')

        if not api_key:
            raise Exception("OpenAI API key not configured")

        client = OpenAI(api_key=api_key)

        try:
            pages = json.loads(self.extracted_text or "[]")
        except Exception:
            _logger.error("INVALID extracted_text JSON")
            return

        if not pages:
            _logger.error("NO PAGES TO PROCESS")
            return

        # 🔥 detect excel
        is_excel = any("row_index" in p for p in pages)

        _logger.warning(f"MODE DETECTED → {'EXCEL' if is_excel else 'PDF'}")

        # ================= EXCEL MODE =================
        if is_excel:

            _logger.warning(f"EXCEL MODE → TOTAL ROWS: {len(pages)}")

            products = []

            for idx, row in enumerate(pages):

                row_text = row.get("text", "")
                images = row.get("images", [])

                _logger.warning(f"ROW {idx} → PROCESSING")
                _logger.warning(f"ROW {idx} → IMAGES: {len(images)}")

                
                prompt = f"""
                You are a structured Excel product parser.

                Each input represents EXACTLY ONE ROW = ONE PRODUCT.

                =====================================
                COLUMN UNDERSTANDING (CRITICAL)
                =====================================

                The row contains mixed values like:

                - ID (e.g. 94601, 12345)
                - Range (e.g. 2-66, 11-00)
                - Stock numbers
                - Prices
                - Links (http...)
                - Image references

                YOU MUST:

                1. IDENTIFY PRODUCT ID
                - Usually numeric (e.g. 94601)
                - Column name may vary (KOD, SKU, ID, CODE)

                2. IDENTIFY PRODUCT NAME
                - MUST NOT be:
                    - pure numbers
                    - ranges (e.g. 2-66)
                    - links
                    - dates
                    - column headers like FOTO

                - If no clear name:
                    → GENERATE NAME like:
                    "Product <ID>"

                3. DESCRIPTION:
                - Short summary from row

                4. CATEGORY:
                - Guess intelligently (e.g. bottle → Drinkware)

                =====================================
                VARIANT DETECTION (VERY IMPORTANT)
                =====================================

                - If multiple rows share SAME ID
                → they are VARIANTS of same product
                - ALSO extract variant attributes from the row:
    
                Examples:
                - Colors → Black, Blue, Red
                - Sizes → S, M, L
                - Range values (e.g. 2-66) → treat as Size or Option

                - If row contains variation info:
                    → put inside "variants"

                - If no clear attribute:
                    → create:
                       "attributes": {{
                            "Variant": "<value from row>"
                        }}

               =====================================
                VARIANT GROUPING (MANDATORY - STRICT)
                =====================================

                - Every product MUST have "variant_group"

                - Extract product ID from the row:
                    (examples: 94601, 92070, ANT021)

                - That ID MUST be used as variant_group

                - RULES:
                    - SAME ID → SAME variant_group
                    - DIFFERENT ID → DIFFERENT product
                    - NEVER leave variant_group empty
                    - NEVER return null

                - If ID exists → use it
                - If ID is a mixture of numerical data and string → use it, but never use date or range
                - If ID unclear → use first numeric value in row

                =====================================
                OUTPUT FORMAT (STRICT)
                =====================================

                 [
                    {{
                        "name": "",
                        "description": "",
                        "category": "",
                         "price": "",
                        "stock": "",
                        "variants": [
                            {{
                                "attributes": {{
                                    "Color": ""
                                }},
                                "image_index": 0,
                                "stock": null
                            }}
                        ]
                    }}
                ]

                  =====================================
                ROW DATA
                =====================================

                ROW TEXT:
                {row_text}

                DETECTED PRICE:
                {row_price}

                DETECTED STOCK:
                {row_stock}
                """


                try:
                    response = client.responses.create(
                        model="gpt-4.1-mini",
                        input=prompt,
                        timeout=60
                    )

                    result = response.output_text.strip()

                    if "```" in result:
                        result = result.split("```")[1]

                    if result.lower().startswith("json"):
                        result = result[4:]

                    parsed = json.loads(result)

                    # 🔥 enforce ONE product per row
                    if isinstance(parsed, list) and parsed:
                        parsed = parsed[0]

                    if not isinstance(parsed, dict):
                        continue

                    # 🔥 attach image (CRITICAL FIX)
                    if images:
                        parsed["image"] = images[0]

                    products.append(parsed)

                    _logger.warning(f"ROW {idx} → PRODUCT PARSED WITH IMAGE")

                except Exception as e:
                    _logger.warning(f"ROW {idx} FAILED → {str(e)}")
                    continue

            # final structure
            self.ai_response = json.dumps([{
                "page": 1,
                "products": products
            }])

            _logger.warning(f"EXCEL PRODUCTS TOTAL → {len(products)}")

            return

        # ================= PDF MODE (UNCHANGED) =================

        page_products = []

        start_index = self.last_ai_page or 0
        _logger.warning(f"AI RESUME FROM PAGE INDEX → {start_index}")

        for i, page in enumerate(pages[start_index:], start=start_index):

            page_no = page.get("page")
            page_text = page.get("text", "")
            images = page.get("images", [])

            if not page_text.strip() and not images:
                continue

            _logger.warning(f"AI → PROCESSING PAGE {page_no}")

            
            prompt = f""" You are an advanced product extraction and interpretation engine for catalog PDFs.

            =====================
            CORE RULES (STRICT)
            =====================

            1. RETURN ONLY VALID JSON
            2. NO explanation
            3. NO markdown
            4. NO text outside JSON
            5. DO NOT duplicate products WITHIN THE SAME PAGE
            6. DO NOT skip any product
            7. EACH product must appear exactly once PER PAGE

            IMPORTANT GLOBAL RULE:

            - This input represents ONLY ONE PAGE of a catalog
            - You MUST extract ONLY products visible on THIS PAGE
            - DO NOT repeat products from previous pages
            - DO NOT assume products continue across pages

            =====================
            PRODUCT DETECTION LOGIC
            =====================

            A page may contain:

            (A) ONE large product (hero layout)
            (B) MULTIPLE products (grid/catalog layout)
            (C) MIX of large + small supporting products

            You MUST:

            - If SINGLE main product:
            → return ONE product

            - If MULTIPLE distinct products:
            → extract EACH product separately

            IMPORTANT:

            - If multiple images exist → assume multiple products
            - DO NOT collapse multiple items into one product
            - If unsure → split into multiple products instead of merging

            CRITICAL:

            - Products are typically aligned with images
            - Each image or grouped images usually represent a product
            - DO NOT treat the whole page as one product

            =====================
            VARIANT DETECTION LOGIC (CRITICAL)
            =====================

            Products in catalogs may appear as multiple similar images WITHOUT explicit labels like "color" or "size".

            You MUST detect variants using visual, structural, and contextual clues.

            A product HAS VARIANTS ONLY IF:

            - The items are clearly the SAME product design
            - Differences are ONLY color, size, or minor variation
            - The items would share the same product name in a store

            DO NOT group items as variants if:
            - They are different product types
            - They have different shapes or purposes
            - They would be listed separately in an e-commerce store

            =====================
            🔥 CRITICAL VARIANT COUNT RULE (NEW)
            =====================

            - If multiple similar items are displayed in a row or grid:
            → EACH visible item MUST be treated as a variant

            - You MUST COUNT the number of visible items
            - DO NOT estimate
            - DO NOT reduce the count

            EXAMPLES:

            - If 10 shirts are visible → return 10 variants
            - If 6 bottles are shown → return 6 variants

            GRID RULE:

            - Each grid item = 1 variant

            VISUAL PRIORITY:

            - Images override text
            - If images show more items than text → trust images

            FAIL CONDITION:

            - Returning fewer variants than visible items is WRONG

            =====================
            VARIANT RULES
            =====================

            - Each product appears ONLY ONCE per page
            - Variants must be grouped under "variants"
            - If no variants exist → DO NOT include "variants"

            ATTRIBUTE INFERENCE:

            - If difference looks like color → use "Color"
            - If difference looks like size → use "Size"
            - If unclear → use "Variant"

            =====================
            VARIANT IMAGE MAPPING (VERY IMPORTANT)
            =====================

            - Each variant MUST map to an image

            - Provide:
            "image_index": index of image (starting from 0)

            STRICT RULE:

            - Number of variants MUST NOT exceed number of images
            - If multiple variants exist → distribute across images

            =====================
            WHEN TO SPLIT PRODUCTS
            =====================

            Treat items as SEPARATE products ONLY IF:
            - Names are clearly different
            - Designs are significantly different
            - They are unrelated items

            If unsure:
            → Prefer splitting into separate products

            =====================
            ANTI-REPETITION RULE
            =====================

            - If same product appears multiple times on THIS PAGE → return it ONLY ONCE

            =====================
            MINIMUM EXTRACTION RULE
            =====================

            - You MUST extract at least ONE product
            - NEVER return empty list

            =====================
            OUTPUT FORMAT
            =====================

            [
                {{
                    "name": "",
                    "description": "",
                    "category": "",
                    "price": "",
                    "stock": "",
                    "variant_group": "",
                    "variants": [
                        {{
                            "attributes": {{
                                "Color": ""
                            }},
                            "image_index": 0,
                            "stock": null
                        }}
                    ]
                }}
            ]

            PAGE CONTEXT:
            - This page contains product images

            TEXT TO ANALYZE:
            PAGE TEXT:
            {page_text}

            DETECTED PRICE:
            {page_price}

            DETECTED STOCK:
            {page_stock}
            """



            try:
                image_inputs = [
                    {
                        "type": "input_image",
                        "image_url": f"data:image/jpeg;base64,{img}"
                    }
                    for img in images[:10]
                ]

                response = client.responses.create(
                    model="gpt-4.1-mini",
                    input=[{
                        "role": "user",
                        "content": [{"type": "input_text", "text": prompt}] + image_inputs
                    }],
                    timeout=60
                )

                result = response.output_text.strip()

            except Exception as e:
                _logger.warning(f"PAGE {page_no} FAILED → {str(e)}")
                continue

            if "```" in result:
                result = result.split("```")[1]

            if result.lower().startswith("json"):
                result = result[4:]

            try:
                parsed = json.loads(result)
            except Exception:
                parsed = []

            if not isinstance(parsed, list):
                parsed = []

            page_products.append({
                "page": page_no,
                "products": parsed
            })

            self.last_ai_page = i + 1

            _logger.warning(f"PAGE {page_no} → STORED: {len(parsed)}")

            time.sleep(1)

        self.ai_response = json.dumps(page_products)

        _logger.warning(f"AI TOTAL PAGES STORED: {len(page_products)}")

    
    #-----------scoring image before picking best/quality image (inage logic)-------------
    def pick_best_image(self, images):

                best_img = None
                best_score = 0
                seen_hashes = set()

                for img in images:

                    try:
                        img_bytes = base64.b64decode(img)
                        size = len(img_bytes)

                        # ❌ Skip tiny images (logos/icons)
                        if size < 10000:
                            continue

                        # ❌ Skip extremely large (full lifestyle pages)
                        if size > 800000:
                            continue

                        # ================= IMAGE ANALYSIS =================
                        pil_img = Image.open(BytesIO(img_bytes))
                        width, height = pil_img.size

                        # ❌ Skip extremely small resolution
                        if width < 150 or height < 150:
                            continue

                        # ❌ Skip extreme aspect ratios (banners, strips)
                        aspect_ratio = width / height if height else 1

                        if aspect_ratio > 3 or aspect_ratio < 0.3:
                            continue

                        # ================= DUPLICATE CHECK =================
                        img_hash = hash(img_bytes[:100])  # fast partial hash

                        if img_hash in seen_hashes:
                            continue

                        seen_hashes.add(img_hash)

                        # ================= SCORING =================
                        score = 0

                        # ✅ Prefer medium-good sizes
                        if 20000 < size < 300000:
                            score += 200

                        # ✅ Prefer square-ish product images
                        if 0.7 < aspect_ratio < 1.5:
                            score += 150

                        # ✅ Prefer decent resolution
                        if width > 400 and height > 400:
                            score += 150

                        # ❌ Penalize too wide/tall
                        if aspect_ratio > 2 or aspect_ratio < 0.5:
                            score -= 100

                        # ❌ Penalize very large images (likely lifestyle)
                        if size > 500000:
                            score -= 150

                        # Base size contribution
                        score += size / 2000

                        # ================= SELECT BEST =================
                        if score > best_score:
                            best_score = score
                            best_img = img

                    except Exception:
                        continue

                return best_img


    #----marchin AI-----------------------------------
    def match_image_with_ai(self, product_name, images):

        api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')
        client = OpenAI(api_key=api_key)

        if not images:
            return None

        # limit images for performance
        images = images[:5]
        image_inputs = [
        {
                "type": "input_image",
                "image_url": f"data:image/jpeg;base64,{img}"
            }
            for img in images
        ]

        prompt = f"""
        You are an expert product image matcher.

        Select the image that BEST represents this product:

        PRODUCT:
        {product_name}

        RULES:
        - Return ONLY the index (0-based integer)
        - No explanation
        - No text

        PRIORITY:
        1. Clean product image (plain background)
        2. Product centered and clearly visible
        3. No human interaction preferred
        4. If only lifestyle images exist, choose the clearest one

        DO NOT PICK:
        - logos
        - icons
        - background-only images
        - cropped fragments
        """

        try:
            response = client.responses.create(
                model="gpt-4.1-mini",
                input=[{
                    "role": "user",
                    "content": [{"type": "input_text", "text": prompt}] + image_inputs
                }],
                timeout=30
            )

            result = response.output_text.strip()

            index = int(result)

            if 0 <= index < len(images):
                return images[index]

        except Exception as e:
            _logger.warning(f"AI IMAGE MATCH FAILED → {str(e)}")

        return None


    # ---------------- PRODUCT CREATION URL----------------

    def create_products_url(self):

        import requests
        import base64
        import json

        if not self.ai_response:
            _logger.warning("NO AI RESPONSE")
            return

        product_obj = self.env['product.template']
        category_obj = self.env['product.category']

        try:
            products = json.loads(self.ai_response)

            if isinstance(products, dict):
                products = [products]

        except Exception as e:
            _logger.error(f"INVALID AI JSON → {str(e)}")
            return

        TOTAL_PRODUCTS = len(products)
        start_index = self.last_processed_product_index or 0

        _logger.warning(f"TOTAL AI PRODUCTS → {TOTAL_PRODUCTS}")
        _logger.warning(f"START INDEX → {start_index}")

        created_count = 0
        skipped_count = 0

        MAX_PRODUCTS_PER_RUN = 30  # ✅ safe batching

        CATEGORY_MAPPING = {
            "t-shirt": "Apparel",
            "shirt": "Apparel",
            "polo": "Apparel",
            "bag": "Bags",
            "backpack": "Bags",
            "cap": "Headwear",
            "hat": "Headwear",
            "bottle": "Drinkware",
            "cup": "Drinkware",
            "drinkware": "Drinkware",
            "pen": "Stationery",
            "notebook": "Stationery",
            "powerbank": "Electronics",
            "charger": "Electronics",
            "laptop": "Electronics",
        }

        parent_category = category_obj.search([('name', '=', "All Products")], limit=1)
        if not parent_category:
            parent_category = category_obj.create({'name': "All Products"})

        end_index = min(start_index + MAX_PRODUCTS_PER_RUN, TOTAL_PRODUCTS)

        _logger.warning(f"PROCESSING RANGE → {start_index} to {end_index}")

        for idx in range(start_index, end_index):

            product = products[idx]

            name = product.get("name")
            if not name:
                skipped_count += 1
                continue

            description = product.get("description", "")
            raw_category = (product.get("category") or "").lower()

            # ================= CATEGORY =================
            mapped_category = "General"
            for key, val in CATEGORY_MAPPING.items():
                if key in raw_category:
                    mapped_category = val
                    break

            category = category_obj.search([
                ('name', '=', mapped_category),
                ('parent_id', '=', parent_category.id)
            ], limit=1)

            if not category:
                category = category_obj.create({
                    'name': mapped_category,
                    'parent_id': parent_category.id
                })

            # ================= DUPLICATE CHECK =================
            existing = product_obj.search([
                ('name', 'ilike', name.strip())
            ], limit=1)

            if existing:
                _logger.warning(f"SKIPPED DUPLICATE → {name}")
                skipped_count += 1
                continue

            vals = {
                'name': name.strip(),
                'description_sale': description,
                'categ_id': category.id,
                'sale_ok': True,
                'website_published': False,
                'vendor_import_job_id': self.id,
            }

            # ================= IMAGE =======================
            image_url = product.get("image")

            if image_url and isinstance(image_url, str) and image_url.startswith("http"):

                try:
                    _logger.warning(f"FETCHING IMAGE → {image_url}")

                    res = requests.get(image_url, timeout=5, stream=True)

                    # ✅ STATUS CHECK
                    if res.status_code != 200:
                        _logger.warning(f"IMAGE HTTP ERROR → {res.status_code}")
                        return

                    # ✅ CONTENT TYPE CHECK
                    content_type = res.headers.get("Content-Type", "")
                    if "image" not in content_type:
                        _logger.warning(f"NOT AN IMAGE → {content_type}")
                        return

                    # ✅ MEMORY SAFE READ (LIMIT SIZE)
                    content = res.raw.read(500000, decode_content=True)

                    if not content:
                        _logger.warning("EMPTY IMAGE CONTENT")
                        return

                    vals['image_1920'] = base64.b64encode(content).decode("utf-8")

                    _logger.warning("IMAGE STORED SUCCESSFULLY")

                except Exception as e:
                    _logger.warning(f"IMAGE FAILED → {str(e)}")

            else:
                _logger.warning(f"NO VALID IMAGE URL → {image_url}")

            # ================= CREATE =================
            try:
                product_obj.create(vals)
                created_count += 1

            except Exception as e:
                _logger.error(f"CREATE FAILED → {name} | {str(e)}")
                skipped_count += 1
                continue

            # ================= SAFE COMMIT =================
            if created_count % 10 == 0:
                self.env.cr.commit()

        # ================= SAVE PROGRESS =================
        self.last_processed_product_index = end_index

        _logger.warning(f"CREATED THIS RUN → {created_count}")
        _logger.warning(f"SKIPPED THIS RUN → {skipped_count}")
        _logger.warning(f"NEXT START INDEX → {self.last_processed_product_index}")

        # ================= FINAL =================
        if self.last_processed_product_index >= TOTAL_PRODUCTS:
            _logger.warning("ALL PRODUCTS CREATED ✅")
        else:
            _logger.warning("MORE PRODUCTS REMAIN → NEXT CRON WILL CONTINUE")

        self.env.cr.commit()


    #==========old create pdf and excel product======================
    
    def create_products_pdf_excel(self):

        import json
        import re

        if not self.ai_response or not self.extracted_text:
            _logger.warning("NO AI OR EXTRACTED DATA → STOP")
            return

        product_obj = self.env['product.template']
        category_obj = self.env['product.category']

        try:
            pages = json.loads(self.extracted_text)
            ai_pages = json.loads(self.ai_response)
        except Exception:
            _logger.error("INVALID JSON → STOP")
            return

        _logger.warning("CREATING PRODUCTS (PDF + EXCEL FINAL MODE)")

        created_count = 0

        CATEGORY_MAPPING = {
            "t-shirt": "Apparel",
            "shirt": "Apparel",
            "polo": "Apparel",
            "bag": "Bags",
            "backpack": "Bags",
            "cap": "Headwear",
            "hat": "Headwear",
            "bottle": "Drinkware",
            "drinkware": "Drinkware",
            "pen": "Stationery",
            "notebook": "Stationery",
            "powerbank": "Electronics",
            "charger": "Electronics",
            "laptop": "Electronics",
        }

        parent_category = category_obj.search([('name', '=', "All Products")], limit=1)
        if not parent_category:
            parent_category = category_obj.create({'name': "All Products"})

        for page_data in pages:

            page_no = page_data.get("page")

            ai_page = next((p for p in ai_pages if p.get("page") == page_no), None)
            if not ai_page:
                continue

            products = ai_page.get("products", [])
            if not products:
                continue

            #================= EXCEL FLOW ===========================
            if self.excel_file:

                grouped_products = {}

                for p in products:
                    raw_name = (p.get("name") or "").strip()

                    match = re.search(r'(?:Product\s*)?([A-Z]*\d+)', raw_name, re.I)

                    if match:
                        group_id = match.group(1).upper()
                    else:
                        group_id = raw_name.upper()

                    grouped_products.setdefault(group_id, []).append(p)

                _logger.warning(f"[EXCEL] GROUPS → {len(grouped_products)}")

                for group_id, group_items in grouped_products.items():

                    main_product = group_items[0]

                    name = (main_product.get("name") or "").strip()
                    description = main_product.get("description", "")
                    raw_category = (main_product.get("category") or "").lower()

                    mapped_category = "General"
                    for key, val in CATEGORY_MAPPING.items():
                        if key in raw_category:
                            mapped_category = val
                            break

                    category = category_obj.search([
                        ('name', '=', mapped_category),
                        ('parent_id', '=', parent_category.id)
                    ], limit=1)

                    if not category:
                        category = category_obj.create({
                            'name': mapped_category,
                            'parent_id': parent_category.id
                        })

                    product = product_obj.search([
                        ('default_code', '=', group_id)
                    ], limit=1)

                    if not product:
                        vals = {
                            'name': name,
                            'default_code': group_id,
                            'description_sale': description,
                            'categ_id': category.id,
                            'sale_ok': True,
                            'website_published': False,
                            'vendor_import_job_id': self.id,
                        }

                        image = main_product.get("image")
                        if image:
                            vals['image_1920'] = image

                        product = product_obj.create(vals)
                        created_count += 1

                    for idx, item in enumerate(group_items):

                        attr_value = f"Variant {idx+1}"

                        attribute = self.env['product.attribute'].search([
                            ('name', '=', "Variant")
                        ], limit=1)

                        if not attribute:
                            attribute = self.env['product.attribute'].create({
                                'name': "Variant"
                            })

                        value = self.env['product.attribute.value'].search([
                            ('name', '=', attr_value),
                            ('attribute_id', '=', attribute.id)
                        ], limit=1)

                        if not value:
                            value = self.env['product.attribute.value'].create({
                                'name': attr_value,
                                'attribute_id': attribute.id
                            })

                        line = self.env['product.template.attribute.line'].search([
                            ('product_tmpl_id', '=', product.id),
                            ('attribute_id', '=', attribute.id)
                        ], limit=1)

                        if not line:
                            self.env['product.template.attribute.line'].create({
                                'product_tmpl_id': product.id,
                                'attribute_id': attribute.id,
                                'value_ids': [(6, 0, [value.id])]
                            })
                        else:
                            if value.id not in line.value_ids.ids:
                                line.value_ids = [(4, value.id)]

                        # ✅ EXCEL VARIANT IMAGE FIX

                            variant_record = self.env['product.product'].search([
                                ('product_tmpl_id', '=', product.id),
                                ('product_template_attribute_value_ids.product_attribute_value_id', '=', value.id)
                            ], limit=1)

                            if variant_record:
                                variant_image = item.get("image")

                                if variant_image:
                                    variant_record.image_1920 = variant_image
                                    _logger.warning(f"[EXCEL] VARIANT IMAGE SET → {group_id} | {value.name}")
                                else:
                                    _logger.warning(f"[EXCEL] NO IMAGE FOR VARIANT → {group_id} | {value.name}")

                continue  # 🔥 protect PDF

            # ================= PDF FLOW =================
            images = page_data.get("images", [])
            _logger.warning(f"[PDF] IMAGES FOUND → {len(images)}")

            for product_data in products:

                try:
                    name = (product_data.get("name") or "").strip()
                    description = product_data.get("description", "")
                    raw_category = (product_data.get("category") or "").lower()
                    variants = product_data.get("variants", [])

                    variant_group = product_data.get("variant_group") or name
                    variant_group = str(variant_group).strip().upper()

                    mapped_category = "General"
                    for key, val in CATEGORY_MAPPING.items():
                        if key in raw_category:
                            mapped_category = val
                            break

                    category = category_obj.search([
                        ('name', '=', mapped_category),
                        ('parent_id', '=', parent_category.id)
                    ], limit=1)

                    if not category:
                        category = category_obj.create({
                            'name': mapped_category,
                            'parent_id': parent_category.id
                        })

                    product = product_obj.search([
                        ('default_code', '=', variant_group)
                    ], limit=1)

                    if not product:
                        vals = {
                            'name': name,
                            'default_code': variant_group,
                            'description_sale': description,
                            'categ_id': category.id,
                            'sale_ok': True,
                            'website_published': False,
                        }

                        if images:
                            vals['image_1920'] = images[0]
                            _logger.warning("PRODUCT IMAGE SET")

                        product = product_obj.create(vals)
                        created_count += 1

                    if not variants:
                        variants = [{"attributes": {"Variant": name}}]

                    for idx, variant in enumerate(variants):

                        attributes = variant.get("attributes", {})

                        for attr_name, attr_value in attributes.items():

                            if not attr_value:
                                continue

                            attribute = self.env['product.attribute'].search([
                                ('name', '=', attr_name)
                            ], limit=1)

                            if not attribute:
                                attribute = self.env['product.attribute'].create({
                                    'name': attr_name
                                })

                            value = self.env['product.attribute.value'].search([
                                ('name', '=', attr_value),
                                ('attribute_id', '=', attribute.id)
                            ], limit=1)

                            if not value:
                                value = self.env['product.attribute.value'].create({
                                    'name': attr_value,
                                    'attribute_id': attribute.id
                                })

                            line = self.env['product.template.attribute.line'].search([
                                ('product_tmpl_id', '=', product.id),
                                ('attribute_id', '=', attribute.id)
                            ], limit=1)

                            if not line:
                                self.env['product.template.attribute.line'].create({
                                    'product_tmpl_id': product.id,
                                    'attribute_id': attribute.id,
                                    'value_ids': [(6, 0, [value.id])]
                                })
                            else:
                                if value.id not in line.value_ids.ids:
                                    line.value_ids = [(4, value.id)]

                        if images and idx < len(images):
                            variant_record = self.env['product.product'].search([
                                ('product_tmpl_id', '=', product.id)
                            ], limit=1)

                            if variant_record:
                                variant_record.image_1920 = images[idx]
                                _logger.warning(f"VARIANT IMAGE SET → {idx}")

                    if created_count % 10 == 0:
                        self.env.cr.commit()

                except Exception as e:
                    _logger.error(f"PRODUCT FAILED → {str(e)}")
                    self.env.cr.rollback()
                    continue

            self.env.cr.commit()

        _logger.warning(f"TOTAL PRODUCTS CREATED: {created_count}")


    #-----URL API FLOW-------------------------------------------

    def scrape_with_playwright(self):

        from playwright.sync_api import sync_playwright
        import subprocess
        import os

        _logger.warning(f"PLAYWRIGHT SCRAPE → {self.data_url}")

        #✅ CHECK IF BROWSER EXISTS FIRST
        browser_path = os.path.expanduser("~/.cache/ms-playwright")

        if not os.path.exists(browser_path):
            _logger.warning("PLAYWRIGHT → INSTALLING BROWSER (FIRST RUN)")

            try:
                subprocess.run(
                    ["python", "-m", "playwright", "install", "chromium"],
                    check=True
                )
                _logger.warning("PLAYWRIGHT BROWSER INSTALLED")
            except Exception as e:
                _logger.error(f"PLAYWRIGHT INSTALL FAILED → {str(e)}")
                return

        products = []

        with sync_playwright() as p:

            browser = p.chromium.launch(headless=True)
            page = browser.new_page()

            page.goto(self.data_url, timeout=60000)
            page.wait_for_timeout(5000)

            # ✅ Cookie handling
            try:
                page.locator("button:has-text('Accept')").click(timeout=3000)
            except:
                pass

            items = page.query_selector_all("a, div")

            _logger.warning(f"PLAYWRIGHT ELEMENTS FOUND → {len(items)}")

            for item in items[:200]:

                try:
                    text = item.inner_text().strip()

                    if not text or len(text) < 5:
                        continue

                    img_el = item.query_selector("img")
                    img_url = img_el.get_attribute("src") if img_el else None

                    if img_url and img_url.startswith("//"):
                        img_url = "https:" + img_url

                    img_base64 = None

                    if img_url:
                        try:
                            res = requests.get(img_url, timeout=10)
                            if res.status_code == 200:
                                img_base64 = base64.b64encode(res.content).decode("utf-8")
                        except:
                            pass

                    products.append({
                        "name": text[:120],
                        "image": img_base64
                    })

                except:
                    continue

            browser.close()

        _logger.warning(f"PLAYWRIGHT PRODUCTS → {len(products)}")

        if not products:
            _logger.error("PLAYWRIGHT FAILED → NO PRODUCTS")
            return

        pages = [{
            "page": 1,
            "text": "\n".join([p["name"] for p in products]),
            "images": [p["image"] for p in products if p["image"]]
        }]

        self.extracted_text = json.dumps(pages)

        _logger.warning(f"PLAYWRIGHT DONE → {len(products)} PRODUCTS")

    #---------------- CRON ---------------

    def run_pending_jobs(self):

        jobs = self.search(
            [('state', 'in', ['draft', 'processing'])],
            order="priority asc, id asc",
            limit=1
        )

        _logger.warning(f"CRON → Found {len(jobs)} jobs")

        for job in jobs:
            try:
                _logger.warning(f"CRON → START JOB {job.id}")
                _logger.warning(f"CRON → JOB {job.id} CURRENT STATE: {job.state}")

                job.state = 'processing'

                job.process_import()

                _logger.warning(f"CRON → JOB {job.id} FINAL STATE: {job.state}")

                # ❌ DO NOT TOUCH STATE HERE

                _logger.warning(f"CRON → JOB {job.id} DONE")

            except Exception as e:
                _logger.exception(f"PROCESS FAILED → {str(e)}")
                job.state = 'failed'


   #=============flask setup/installation=================== 
    def ping_flask_server(self):
      
        try:
            requests.get("https://pdf-extractor-staging.onrender.com", timeout=10)
            _logger.info("FLASK PING SUCCESS")
        except Exception:
            _logger.warning("FLASK PING FAILED")

    #---------------normalizer-------------------------------
   
    def _normalize_url_data(self, items):

        blocks = []

        for item in items:

            text = (item.get("text") or "").strip()
            image = item.get("image")

            # 🔥 STRICT VALIDATION
            if not text or len(text) < 5:
                continue

            if image and isinstance(image, str) and not image.startswith("http"):
                image = None

            blocks.append({
                "text": text,
                "image": image
            })

        _logger.warning(f"NORMALIZED BLOCKS → {len(blocks)}")

        # =====================================================
        # 🔥 SPLIT INTO MULTIPLE PAGES (CRITICAL FIX)
        # =====================================================

        PAGE_SIZE = 20  # 🔥 prevents AI overload

        pages = []

        for i in range(0, len(blocks), PAGE_SIZE):

            chunk = blocks[i:i + PAGE_SIZE]

            pages.append({
                "page": len(pages) + 1,
                "blocks": chunk
            })

        _logger.warning(f"NORMALIZED PAGES → {len(pages)}")

        return pages

    #---------------clean_scraped_blocks-------------------------------
    def _clean_scraped_blocks(self, raw_blocks):
        """
        Clean Apify output before sending to AI
        """

        cleaned = []
        seen = set()

        for item in raw_blocks:

            text = (item.get("text") or "").strip()
            image = item.get("image")

            # ❌ REMOVE NOISE
            if not text:
                continue

            if len(text) < 15:
                continue

            if any(x in text.lower() for x in [
                "privacy", "cookie", "terms", "login",
                "menu", "navigation", "home"
            ]):
                continue

            # ❌ REMOVE DUPLICATES
            key = text[:120]  # allow more variation

            if key in seen:
                continue

            seen.add(key)

            cleaned.append({
                "text": text,
                "image": image
            })

        return cleaned
    

    #======apify url fetch/scrapp products===============
    
    def _run_apify_actor(self, url):

        token = self.env['ir.config_parameter'].sudo().get_param('apify.api_token')

        if not token:
            raise Exception("Apify API token not configured")

        ACTOR_ID = "selectad~my-actor"

        # =====================================================
        # 🔥 STEP 1: START ACTOR (ONLY IF NOT STARTED)
        # =====================================================

        if not getattr(self, "apify_run_id", False):

            run_url = f"https://api.apify.com/v2/acts/{ACTOR_ID}/runs?token={token}"

            payload = {
                "startUrls": [{"url": url}]
            }

            headers = {
                "Content-Type": "application/json"
            }

            response = requests.post(run_url, json=payload, headers=headers, timeout=30)

            if response.status_code != 201:
                raise Exception(f"Apify run failed: {response.text}")

            run_data = response.json()

            # ✅ SAVE FOR NEXT CRON
            self.apify_run_id = run_data["data"]["id"]
            self.apify_dataset_id = run_data["data"]["defaultDatasetId"]

            _logger.warning(f"APIFY STARTED → RUN ID {self.apify_run_id}")

            # 🔥 IMPORTANT: STOP HERE (NON-BLOCKING)
            return None

        # =====================================================
        # 🔥 STEP 2: CHECK STATUS
        # =====================================================

        status_url = f"https://api.apify.com/v2/actor-runs/{self.apify_run_id}?token={token}"

        status_res = requests.get(status_url, timeout=20).json()
        status = status_res["data"]["status"]

        _logger.warning(f"APIFY STATUS → {status}")

        if status in ["RUNNING", "READY"]:
            _logger.warning("APIFY STILL RUNNING → WAIT NEXT CRON")
            return None

        if status in ["FAILED", "ABORTED", "TIMED-OUT"]:
            raise Exception(f"Apify run failed with status: {status}")

        # =====================================================
        # 🔥 STEP 3: FETCH DATA (ONLY WHEN DONE)
        # =====================================================

        dataset_url = f"https://api.apify.com/v2/datasets/{self.apify_dataset_id}/items"

        params = {
            "token": token,
            "limit": 1000,
            "clean": "true"
        }

        dataset_res = requests.get(dataset_url, params=params, timeout=30)

        if dataset_res.status_code != 200:
            raise Exception(f"Failed to fetch dataset: {dataset_res.text}")

        data = dataset_res.json()

        _logger.warning(f"APIFY ITEMS FETCHED → {len(data)}")

        if not data:
            raise Exception("Apify returned empty dataset")

        # 🔥 CLEAN UP (IMPORTANT)
        self.apify_run_id = False
        self.apify_dataset_id = False

        return data

    #=======validation===================
    def validate_ai_output(products):
        for p in products:
            if "variants" in p:
                if not isinstance(p["variants"], list):
                    p["variants"] = []

                for v in p["variants"]:
                    if "attributes" not in v:
                        v["attributes"] = {"Variant": "Default"}

        return products
    
    #=======keep cron alive================
    def keep_alive(self):
        _logger.warning("KEEP ALIVE PING")










#===========================================================================
#Old previous models
#=============================================================================

from odoo import models, fields
import base64
import logging
import io
import requests
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import json
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openai import OpenAI
import re
import fitz
import hashlib
import psycopg2

from PIL import (
    Image,
    ImageOps,
    ImageChops
)

import cv2
import numpy as np

 

_logger = logging.getLogger(__name__)

class ProductTemplate(models.Model):

    _inherit = 'product.template'

    vendor_id = fields.Many2one(
        'res.partner',
        string="Vendor"
    )

    vendor_fingerprint = fields.Char(
        index=True,
        copy=False
    )

    vendor_import_job_id = fields.Many2one(
        'vendor.import.job',
        string='Vendor Import Job',
        index=True,
        ondelete='set null'
    )

    vendor_stock_qty = fields.Integer()

# ✅ Extend existing model
class ResPartner(models.Model):
    _inherit = 'res.partner'

    #Vendor user role
    is_vendor_user = fields.Boolean(
        string="Vendor User",
        default=False
    )


class VendorImportJob(models.Model):

    _name = "vendor.import.job"
    _description = "Vendor Import Job"

    partner_id = fields.Many2one("res.partner", string="Vendor")  # ✅ LINK instead

    name = fields.Char(default="Vendor Data Import")
    extra_info = fields.Text()

    pdf_file = fields.Binary()
    excel_file = fields.Binary()
    logo_file = fields.Binary()

    extracted_text = fields.Text()
    current_page = fields.Integer(
        string="Current PDF Page",
        default=0
    )
    total_pages = fields.Integer(string="Total Pages", default=0)
    last_ai_page = fields.Integer(string="Last AI Page", default=0)
    ai_response = fields.Text()
    priority = fields.Integer(default=10)
    excel_created_index = fields.Integer(
        string="Excel Created Index",
        default=0
    )

    apify_run_id = fields.Char()
    apify_dataset_id = fields.Char()
   
    last_processed_product_index = fields.Integer(default=0)
    last_created_page = fields.Integer(default=0)
    lock = fields.Boolean(default=False)
    completion_email_sent = fields.Boolean(
        default=False
    )
    is_excel_parsed = fields.Boolean(default=False)
    excel_ai_index = fields.Integer(default=0)
    upload_signature = fields.Char(string="Upload Signature")
    processed_group_ids = fields.Text(default="[]")

    url_total_batches = fields.Integer(default=0)
    url_batch_index = fields.Integer(default=0)
    data_url = fields.Char()
    url_parse_index = fields.Integer(
        string="URL Parse Index",
        default=0
    )
    url_blocks_json = fields.Text(
        string="URL Blocks JSON"
    )
        
    excel_parse_index = fields.Integer(
        default=0
    )

    source_type = fields.Selection([
        ("pdf", "PDF"),
        ("excel", "Excel"),
        ("url", "URL"),
    ])


    excel_url_queue = fields.Text()

    excel_url_index = fields.Integer(
        default=0
    )

    excel_url_processing = fields.Boolean(
        default=False
    )

    completion_email_sent = fields.Boolean(
        default=False
    )

   
    state = fields.Selection([
        ('draft', 'Draft'),
        ('processing', 'Processing'),
        ('review', 'Vendor Review'),
        ('done', 'Completed'),
        ('error', 'Error'),
        ('failed', 'Failed'),

         #New
        ('url_scraping', 'URL Scraping'),
        ('url_ai', 'URL AI'),
        ('url_creating', 'URL Creating'),

        ('pdf_extracting', 'PDF Extracting'),
        ('pdf_ai', 'PDF AI'),
        ('pdf_creating', 'PDF Creating'),

        ('excel_parsing', 'Excel Parsing'),
        ('excel_ai', 'Excel AI'),
        ('excel_creating', 'Excel Creating'),

    ], default='draft')


     #============================= MAIN FLOW (process steps) =====================

    def process_import(self):

        _logger.warning(f"PROCESS START → Job {self.id}")

        try:

            self._process_step()

        except Exception as e:
            _logger.error(f"PROCESS FAILED → {str(e)}")
            self.state = "failed"


    #=============Safe commit======================================================
    def _safe_commit_progress(self):

        try:

            self.flush_recordset()

        except Exception as flush_error:

            _logger.warning(
                f"FLUSH FAILED → {flush_error}"
            )

        try:

            self.env.cr.commit()

        except Exception as commit_error:

            _logger.warning(
                f"COMMIT FAILED → {commit_error}"
            )


    #========vendor email notification==========
    def send_completion_email(self):

        if not self.partner_id.email:
            return

        subject = f"Import Completed - {self.name}"

        body = f"""
        Hello {self.partner_id.name},

        Your import job has completed successfully.

        File: {self.name}
        Source: {self.source_type}
        Upload Date: {self.create_date}

        Status: Completed

        Regards
        """

        mail = self.env['mail.mail'].create({
            'subject': subject,
            'body_html': body,
            'email_to': self.partner_id.email,
        })

        mail.send()

        self.completion_email_sent = True

    #============Processing Jobs===================================================
    def _process_step(self):

        import json
        import re

        self.ensure_one()


        # =================================================
        # SAFETY
        # =================================================

        if self.pdf_file and self.excel_file:

            _logger.error(
                "[PROCESS STEP] BOTH PDF AND EXCEL PROVIDED"
            )

            self.state = "failed"

            self.flush_recordset()
            self.env.cr.commit()

            return


        _logger.warning(
            f"[PROCESS STEP] → state={self.state}"
        )


        # =================================================
        # DONE
        # =================================================

        if self.state == 'done':

            _logger.warning(
                f"JOB {self.id} ALREADY DONE ✅"
            )

            return


        # =================================================
        # REVIEW RECOVERY
        # =================================================

        if self.state == 'review':

            _logger.warning(
                "REVIEW → RESET TO START"
            )

            if self.pdf_file:

                self.state = 'pdf_extracting'

            elif self.excel_file and not self.pdf_file:

                self.state = 'excel_parsing'

            elif self.data_url:

                self.state = 'url_scraping'


            self.flush_recordset()
            self.env.cr.commit()

            return


        # =================================================
        # URL FLOW
        # =================================================
        

        if self.data_url:

            _logger.warning(
                "FLOW → URL"
            )


            # =============================================
            # START
            # =============================================

            if self.state == 'draft':

                _logger.warning(
                    "[URL FLOW] START"
                )

                self.state = 'url_scraping'

                self._safe_commit_progress()

                return


            # =============================================
            # SCRAPE
            # =============================================

            if self.state == 'url_scraping':

                # =========================================
                # RECOVERY
                # =========================================

                if self.url_blocks_json:

                    _logger.warning(

                        "[URL RECOVERY] "

                        "USING SAVED BLOCKS"
                    )

                    self.state = 'url_ai'

                    self._safe_commit_progress()

                    return


                _logger.warning(

                    f"[URL SCRAPE] "

                    f"SENDING TO APIFY "

                    f"| {self.data_url}"
                )


                previous_extract = bool(
                    self.extracted_text
                )


                result = self.parse_url()


                # =========================================
                # APIFY STILL PROCESSING
                # =========================================

                if result is True:

                    _logger.warning(

                        "[APIFY STATUS] "

                        "WAITING FOR RESPONSE"
                    )

                    return


                # =========================================
                # APIFY RESPONSE READY
                # =========================================

                _logger.warning(
                    "[APIFY STATUS] RESPONSE RECEIVED"
                )


                if (

                    self.extracted_text

                    and

                    not previous_extract

                ):

                    _logger.warning(
                        "URL EXTRACTION SUCCESS → url_ai"
                    )

                    self.state = 'url_ai'

                    self._safe_commit_progress()

                    return


                # =========================================
                # FAILED
                # =========================================

                if self.state == 'failed':

                    _logger.warning(
                        "[URL SCRAPE FAILED]"
                    )

                    self._safe_commit_progress()

                    return


                _logger.warning(
                    "[URL SCRAPE] NO DATA EXTRACTED"
                )

                self.state = 'failed'

                self._safe_commit_progress()

                return


            # =============================================
            # AI
            # =============================================

            if self.state == 'url_ai':

                previous_batch = (
                    self.url_batch_index or 0
                )


                _logger.warning(

                    f"[URL AI START] "

                    f"batch={previous_batch}"
                )


                try:

                    self.send_to_openai_url()

                except Exception as e:

                    _logger.exception(

                        f"URL AI FAILED → {str(e)}"
                    )

                    self.state = 'failed'


                    self._safe_commit_progress()

                    return


                new_batch = (
                    self.url_batch_index or 0
                )


                _logger.warning(

                    f"[URL AI CHECK] "

                    f"{previous_batch} -> {new_batch}"
                )


                # =========================================
                # PROGRESS DETECTED
                # =========================================

                if new_batch > previous_batch:

                    _logger.warning(
                        "[URL AI] PROGRESS SAVED"
                    )

                elif self.state != 'url_creating':

                    _logger.warning(
                        "[URL AI] NO PROGRESS DETECTED"
                    )

                    self.state = 'failed'


                self._safe_commit_progress()

                return


            # =============================================
            # CREATE
            # =============================================

            if self.state == 'url_creating':

                if not self.ai_response:

                    self.state = 'failed'

                    _logger.warning(
                        "URL CREATE FAILED → NO AI RESPONSE"
                    )


                    self._safe_commit_progress()

                    return


                previous_index = (
                    self.last_processed_product_index or 0
                )


                _logger.warning(

                    f"[URL CREATE START] "

                    f"{previous_index}"
                )


                try:

                    self.create_products_url()

                except Exception as e:

                    _logger.exception(

                        f"URL CREATE FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self._safe_commit_progress()

                    return


                new_index = (
                    self.last_processed_product_index or 0
                )


                _logger.warning(

                    f"[URL CREATE CHECK] "

                    f"{previous_index} -> {new_index}"
                )


                try:

                    total = len(
                        json.loads(
                            self.ai_response or "[]"
                        )
                    )

                except Exception:

                    total = 0


                _logger.warning(

                    f"[URL TOTAL PRODUCTS] "

                    f"{total}"
                )


                if new_index >= total:

                    self.state = 'done'

                    _logger.warning(
                        "URL COMPLETE ✅"
                    )

                    if not self.completion_email_sent:

                        self.send_completion_email()

                elif new_index > previous_index:

                    self.state = 'url_creating'

                    _logger.warning(
                        "[URL CREATE] CONTINUE"
                    )

                else:

                    _logger.warning(
                        "[URL CREATE] NO PROGRESS"
                    )

                    self.state = 'failed'


                self._safe_commit_progress()

                return


        # =================================================
        # EXCEL FLOW
        # =================================================

        if self.excel_file and not self.pdf_file:

            _logger.warning(
                "FLOW → EXCEL"
            )


            # =============================================
            # START
            # =============================================

            if self.state == 'draft':

                self.state = 'excel_parsing'

                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # PARSE
            # =============================================

            if self.state == 'excel_parsing':

                previous_index = (
                
                    self.excel_parse_index or 0
                )

                _logger.warning(

                    f"[EXCEL PARSE START] "

                    f"previous_index={previous_index}"
                )


                self.parse_excel()


                new_index = (
                   
                    self.excel_parse_index or 0
                )


                _logger.warning(

                    f"[EXCEL PARSE CHECK] "

                    f"{previous_index} -> {new_index}"
                )


                # =========================================
                # NEW ROWS FOUND
                # =========================================

                if new_index > previous_index:

                    _logger.warning(
                        "[EXCEL PARSE] NEW BATCH READY → excel_ai"
                    )

                    self.state = 'excel_ai'

                else:

                    _logger.warning(
                        "[EXCEL PARSE] NO NEW ROWS"
                    )


                    if self.is_excel_parsed:

                        _logger.warning(
                            "[EXCEL IMPORT COMPLETE] ✅"
                        )

                        self.state = 'done'

                        if not self.completion_email_sent:

                            self.send_completion_email()

                        # cleanup URL queue
                        self.excel_url_processing = False

                        self.excel_url_queue = False

                        self.excel_url_index = 0

                    else:

                        self.state = 'excel_parsing'


                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # AI
            # =============================================

            if self.state == 'excel_ai':

                try:

                    self.send_to_openai_excel()

                except Exception as e:

                    _logger.exception(

                        f"EXCEL AI FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self.flush_recordset()
                    self.env.cr.commit()

                    return


                try:

                    extracted_rows = json.loads(
                        self.extracted_text or "[]"
                    )

                    total_rows = len(extracted_rows)

                except Exception:

                    total_rows = 0


                _logger.warning(

                    f"[EXCEL AI TOTAL ROWS] "

                    f"{total_rows}"
                )


                _logger.warning(

                    f"[EXCEL AI STATE] "

                    f"{self.state}"
                )


                self.flush_recordset()
                self.env.cr.commit()

                return

            # =============================================
            # CREATE
            # =============================================

           
            #-----------EXCEL URL QUEUE PROCESSOR-------------
          
            if self.excel_url_processing:

                try:

                    _logger.warning(
                        "[URL QUEUE PROCESSING]"
                    )

                    self.process_excel_url_queue()

                except Exception as e:

                    _logger.exception(
                        f"[URL QUEUE ERROR] {str(e)}"
                    )

            #-----------EXCEL ROW PROCESSOR-------------

            if self.state == 'excel_creating':

                try:

                    self.create_products_excel()

                except Exception as e:

                    _logger.exception(
                        f"EXCEL CREATE FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self.flush_recordset()
                    self.env.cr.commit()

                    return

                _logger.warning(

                    f"[EXCEL CREATE STATE] "

                    f"{self.state}"
                )

                self.flush_recordset()
                self.env.cr.commit()


            # =============================================
            # RETURN
            # =============================================

            if self.state == 'excel_creating' \
                    or self.excel_url_processing:

                return
        
        # =================================================
        # PDF FLOW
        # =================================================

        elif self.pdf_file:

            _logger.warning(
                "FLOW → PDF"
            )


            # =============================================
            # START
            # =============================================

            if self.state == 'draft':

                self.state = 'pdf_extracting'

                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # EXTRACT
            # =============================================

            if self.state == 'pdf_extracting':

                try:

                    self.extract_pdf()

                except Exception as e:

                    _logger.exception(

                        f"PDF EXTRACT FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self.flush_recordset()
                    self.env.cr.commit()

                    return


                if (

                    (self.current_page or 0)

                    <

                    (self.total_pages or 0)

                ):

                    _logger.warning(

                        f"PDF EXTRACTION CONTINUES "

                        f"→ PAGE "

                        f"{self.current_page}/"

                        f"{self.total_pages}"
                    )

                    self.state = 'pdf_extracting'

                else:

                    _logger.warning(
                        "PDF EXTRACTION COMPLETE → pdf_ai"
                    )

                    self.state = 'pdf_ai'


                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # PDF AI
            # =============================================

            if self.state == 'pdf_ai':

                try:

                    self.send_to_openai_pdf()

                except Exception as e:

                    _logger.exception(

                        f"PDF AI FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self.flush_recordset()
                    self.env.cr.commit()

                    return


                page_total = self.env[
                    'vendor.import.page'
                ].search_count([

                    ('job_id', '=', self.id)

                ])


                _logger.warning(

                    f"[PDF AI CHECK] "

                    f"{self.last_ai_page}/"

                    f"{page_total}"
                )


                if (

                    (self.last_ai_page or 0)

                    <

                    page_total

                ):

                    _logger.warning(

                        f"PDF AI CONTINUES "

                        f"→ {self.last_ai_page}/"

                        f"{page_total}"
                    )

                    self.state = 'pdf_ai'

                else:

                    _logger.warning(
                        "PDF AI COMPLETE → pdf_creating"
                    )

                    self.state = 'pdf_creating'


                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # PDF CREATE
            # =============================================

            if self.state == 'pdf_creating':

                try:

                    self.create_products_pdf()

                except Exception as e:

                    _logger.exception(

                        f"PDF CREATE FAILED → {str(e)}"
                    )

                    self.state = 'failed'

                    self.flush_recordset()
                    self.env.cr.commit()

                    return


                try:

                    total_ai_pages = len(
                        json.loads(
                            self.ai_response or "[]"
                        )
                    )

                except Exception:

                    total_ai_pages = 0


                if (

                    (self.last_created_page or 0)

                    <

                    total_ai_pages

                ):

                    _logger.warning(

                        f"PDF CREATE CONTINUES "

                        f"→ {self.last_created_page}/"

                        f"{total_ai_pages}"
                    )

                    self.state = 'pdf_creating'

                else:

                    _logger.warning(
                        "PDF COMPLETE ✅"
                    )

                    self.state = 'done'

                    if not self.completion_email_sent:

                        self.send_completion_email()


                self.flush_recordset()
                self.env.cr.commit()

                return


    #------------parse url------------------------------------

    def parse_url(self):

        import json

        _logger.warning(f"APIFY SCRAPE → {self.data_url}")

        raw_data = self._run_apify_actor(self.data_url)

        # =====================================================
        # APIFY STILL RUNNING
        # =====================================================

        if raw_data is None:

            _logger.warning(
                "APIFY NOT READY → WAIT NEXT CRON"
            )

            self.state = "url_scraping"

            return True

        # =====================================================
        # EMPTY RAW RESPONSE
        # =====================================================

        if not raw_data:

            _logger.error(
                "APIFY FAILED → EMPTY DATASET"
            )

            self.state = "failed"

            return

        # =====================================================
        # SAFE DEBUG LOG
        # =====================================================

        try:

            _logger.warning(
                f"RAW APIFY ITEMS → {len(raw_data)}"
            )

        except Exception:
            pass

        # =====================================================
        # HANDLE STRUCTURED RESPONSES
        # =====================================================

        first = raw_data[0] if raw_data else {}

        response_type = first.get("type")

        # =====================================================
        # BLOCKED
        # =====================================================

        if response_type == "BLOCKED":

            reason = first.get(
                "reason",
                "Unknown block detected"
            )

            status_code = first.get(
                "status_code"
            )

            _logger.error(
                f"URL BLOCKED → {reason}"
            )

            if status_code:
                _logger.error(
                    f"BLOCK STATUS CODE → {status_code}"
                )

            self.state = "failed"

            return

        # =====================================================
        # EMPTY
        # =====================================================

        if response_type == "EMPTY":

            reason = first.get(
                "reason",
                "No products extracted"
            )

            debug = first.get("debug", {})

            _logger.error(
                f"URL EXTRACTION EMPTY → {reason}"
            )

            if debug:

                _logger.error(
                    f"PAGE TITLE → {debug.get('title')}"
                )

                _logger.error(
                    f"IMAGES FOUND → {debug.get('images_found')}"
                )

                _logger.error(
                    f"LINKS FOUND → {debug.get('links_found')}"
                )

                _logger.error(
                    f"POSSIBLE PRODUCT BLOCKS → "
                    f"{debug.get('possible_product_blocks')}"
                )

                _logger.error(
                    f"COOKIE DETECTED → "
                    f"{debug.get('cookie_detected')}"
                )

                preview = debug.get(
                    'body_preview',
                    ''
                )

                _logger.error(
                    f"BODY PREVIEW → {preview[:300]}"
                )

            self.state = "failed"

            return

        # ===================================================
        # PRODUCTS
        # ===================================================

        structured_data = []

        for block in raw_data:

            # ==============================================
            # FORMAT 1 → ORIGINAL EB FORMAT
            # ==============================================

            if block.get("text"):

                structured_data.append({
                    "text": block.get("text"),
                    "image": block.get("image")
                })

                continue

            # ==============================================
            # FORMAT 2 → STRUCTURED FORMAT
            # ==============================================

            if block.get("type") == "PRODUCTS":

                items = block.get("items", [])

                if not items:
                    continue

                structured_data.extend(items)

        # =====================================================
        # NO PRODUCTS AFTER PARSE
        # =====================================================

        if not structured_data:

            _logger.error(
                "NO VALID PRODUCTS FOUND AFTER PARSING"
            )

            self.state = "failed"

            return

        # =====================================================
        # LIMIT SIZE (VERY IMPORTANT)
        # =====================================================

        # structured_data = structured_data[:40]

        # ============================================
        # URL BATCHING
        # ============================================

        BATCH_SIZE = 40

        start = (
            self.url_parse_index or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(structured_data)
        )


        structured_data = structured_data[
            start:end
        ]


        _logger.warning(

            f"[URL PARSE BATCH] "

            f"{start} -> {end} "

            f"| total={len(normalized if 'normalized' in locals() else structured_data)}"
        )


        # =====================================================
        # NORMALIZE
        # =====================================================

        normalized = self._normalize_url_data(
            structured_data
        )

        # ============================================
        # SAVE URL PARSE PROGRESS
        # ============================================

        self.url_parse_index = end


        _logger.warning(

            f"[URL PARSE SAVE] "

            f"{self.url_parse_index}"
        )


        if not normalized:

            _logger.error(
                "NORMALIZATION FAILED → EMPTY DATA"
            )

            self.state = "failed"

            return

        # =====================================================
        # STORE SAFELY
        # =====================================================

        #payload = json.dumps(normalized)

        # 🔥 LIMIT STORAGE SIZE
        #self.extracted_text = payload[:50000]

        payload = json.dumps(normalized)

        # ============================================
        # PERSIST URL BLOCKS
        # ============================================

        self.url_blocks_json = payload

        # compatibility
        self.extracted_text = payload[:50000]


        _logger.warning(

            f"[URL STORE] "

            f"saved_blocks={len(normalized)}"
        )


        _logger.warning(
            f"APIFY DONE → {len(normalized)} ITEMS"
        )

        # =====================================================
        # MOVE TO NEXT STEP
        # =====================================================

        if self.url_parse_index >= len(structured_data):

            _logger.warning(
                "[URL PARSE] FINAL BATCH READY"
            )

        else:

            _logger.warning(
                "[URL PARSE] MORE BATCHES REMAIN"
            )


        self.state = "url_ai"

    #------excel parsing method---------------
    
    def parse_excel(self):

        _logger.warning(
            "EXCEL → START PARSING (BATCH MODE)"
        )

        excel_bytes = base64.b64decode(
            self.excel_file
        )

        wb = load_workbook(
            filename=BytesIO(excel_bytes)
        )

        headers = {
            "User-Agent": "Mozilla/5.0"
        }

        pages = []

        # =====================================
        # SAFE BATCH CONTROL
        # =====================================

        BATCH_SIZE = 8
        MAX_BUFFER = 150


        start_index = (
            self.excel_parse_index or 0
        )


        current_count = 0

        global_index = 0

        _logger.warning(
            f"EXCEL RESUME FROM INDEX "
            f"→ {start_index}"
        )


        # =====================================
        # TOTAL ROWS
        # =====================================

        total_rows = 0

        for sheet in wb.worksheets:

            for idx, row in enumerate(

                sheet.iter_rows()

            ):

                if idx == 0:
                    continue

                row_text_parts = [

                    str(cell.value or "").strip()

                    for cell in row

                    if str(
                        cell.value or ""
                    ).strip()
                ]

                if not row_text_parts:
                    continue

                total_rows += 1


        _logger.warning(
            f"[DEBUG] REAL TOTAL ROWS "
            f"→ {total_rows}"
        )


        # =====================================
        # MAIN LOOP
        # =====================================

        for sheet in wb.worksheets:

            _logger.warning(
                f"PROCESSING SHEET → "
                f"{sheet.title}"
            )

            image_loader = SheetImageLoader(
                sheet
            )


            for idx, row in enumerate(

                sheet.iter_rows()

            ):

                if idx == 0:
                    continue


                row_text_parts = [

                    str(cell.value or "").strip()

                    for cell in row

                    if str(
                        cell.value or ""
                    ).strip()
                ]


                if not row_text_parts:
                    continue

                # =================================
                # EARLY URL INTERCEPTION
                # =================================

                detected_url = None

                for part in row_text_parts:

                    part = str(part or "")

                    urls = re.findall(

                        r'https?://[^\s|]+',

                        part
                    )

                    if urls:

                        detected_url = urls[0].strip()

                        break


                if detected_url:

                    _logger.warning(

                        f"[URL ROW DETECTED] "

                        f"{detected_url}"
                    )


                    existing_queue = []

                    if self.excel_url_queue:

                        try:

                            existing_queue = json.loads(
                                self.excel_url_queue
                            )

                        except Exception:

                            existing_queue = []


                    existing_queue.append({

                        "detected_url": detected_url,

                        "row_index": global_index + 1,

                        "vendor_id": (
                            self.partner_id.id
                            if self.partner_id
                            else False
                        ),
                    })


                    self.excel_url_queue = json.dumps(
                        existing_queue
                    )

                    self.excel_url_processing = True


                    _logger.warning(

                        f"[URL QUEUED] "

                        f"total={len(existing_queue)}"
                    )


                    global_index += 1

                    current_count += 1

                    continue

                # =================================
                # GLOBAL INDEX TRACKING
                # =================================

                global_index += 1


                # =================================
                # SKIP OLD ROWS
                # =================================

                if global_index <= start_index:
                    continue


                # =================================
                # BATCH LIMIT
                # =================================

                if current_count >= BATCH_SIZE:

                    _logger.warning(
                        "BATCH LIMIT REACHED "
                        "→ NEXT CRON"
                    )

                    break


                # =================================
                # PRICE/STOCK DETECTION
                # =================================

                price = ""
                stock = ""

                numeric_candidates = []


                for col_idx, cell in enumerate(row):

                    raw_val = str(
                        cell.value or ""
                    ).strip()

                    if not raw_val:
                        continue


                    # skip ranges
                    if (
                        "-" in raw_val
                        and not raw_val.startswith("-")
                    ):
                        continue


                    try:

                        clean = raw_val.replace(
                            ",",
                            "."
                        )

                        clean = re.sub(

                            r"[^\d.]",

                            "",

                            clean
                        )


                        if not clean:
                            continue


                        num = float(clean)

                        is_real_decimal = False


                        if "." in clean:

                            decimal_part = (
                                clean.split(".")[-1]
                            )

                            if decimal_part not in [

                                "0",

                                "00"
                            ]:

                                is_real_decimal = True


                        numeric_candidates.append({

                            "col": col_idx,

                            "num": num,

                            "raw": raw_val,

                            "is_decimal":
                                is_real_decimal
                        })

                    except:
                        continue


                # =================================
                # PRICE
                # =================================

                price_candidates = [

                    x for x in numeric_candidates

                    if (

                        x["is_decimal"]

                        and

                        0 < x["num"] < 1000
                    )
                ]


                best_price = None


                if price_candidates:

                    best_price = sorted(

                        price_candidates,

                        key=lambda x: x["col"]

                    )[-1]

                    price = str(
                        best_price["num"]
                    )


                # =================================
                # STOCK
                # =================================

                if best_price:

                    price_col = (
                        best_price["col"]
                    )

                    stock_candidates = []


                    for item in numeric_candidates:

                        if item["is_decimal"]:
                            continue


                        val = item["num"]


                        if val > 9999:
                            continue


                        if item["col"] >= price_col:
                            continue


                        stock_candidates.append(
                            item
                        )


                    if stock_candidates:

                        best_stock = max(

                            stock_candidates,

                            key=lambda x: x["num"]
                        )

                        stock = str(

                            int(
                                best_stock["num"]
                            )
                        )


                _logger.warning(
                    f'''
                    EXCEL RAW ROW →

                    TEXT=
                    {" | ".join(row_text_parts)}

                    PRICE={price}

                    STOCK={stock}
                    '''
                )


                # =================================
                # ROW TEXT
                # =================================

                row_text = f"""
                ROW_DATA:
                {" | ".join(row_text_parts)}

                RULE:
                - THIS IS EXACTLY ONE PRODUCT
                - DO NOT SPLIT THIS ROW
                - THIS ROW MAY BE A VARIANT
                - USE SIMILAR ID/SKU
                """

                row_images = []


                # ==================================
                # EMBEDDED IMAGE
                # ==================================

                for cell in row:

                    try:

                        if image_loader.image_in(
                            cell.coordinate
                        ):

                            pil_img = (
                                image_loader.get(
                                    cell.coordinate
                                )
                            )

                            buffer = BytesIO()

                            pil_img.save(
                                buffer,
                                format="JPEG"
                            )

                            img_base64 = (
                                base64.b64encode(

                                    buffer.getvalue()

                                ).decode("utf-8")
                            )

                            row_images.append(
                                img_base64
                            )

                            break

                    except:
                        continue


                # =================================
                # URL IMAGE
                # =================================

                if not row_images:

                    for cell in row:

                        val = str(
                            cell.value or ""
                        ).strip()

                        if val.startswith("http"):

                            try:

                                response = requests.get(

                                    val,

                                    headers=headers,

                                    timeout=5
                                )

                                if (

                                    response.status_code
                                    == 200

                                    and

                                    "image"

                                    in response.headers.get(
                                        "Content-Type",
                                        ""
                                    )
                                ):

                                    img_base64 = (
                                        base64.b64encode(

                                            response.content

                                        ).decode("utf-8")
                                    )

                                    row_images.append(
                                        img_base64
                                    )

                                    break

                            except:
                                continue


                # =================================
                # STORE
                # =================================

                pages.append({

                    "page": global_index,

                    "text": row_text,

                    "images": row_images,

                    "row_index": global_index,

                    "price": price,

                    "stock": stock,
                })


                current_count += 1


                # =================================
                # MEMORY SAFETY
                # =================================

                if len(pages) >= MAX_BUFFER:

                    _logger.warning(
                        f"EXCEL SAFETY BREAK "
                        f"→ {len(pages)} rows"
                    )

                    break


            if (
                current_count >= BATCH_SIZE
                or
                len(pages) >= MAX_BUFFER
            ):
                break


        # =====================================
        # STORE
        # =====================================

        existing = []

        if self.extracted_text:

            try:

                existing = json.loads(
                    self.extracted_text
                )

            except:
                existing = []


        # combined = existing + pages
        
        existing_map = {

            item.get("row_index"): item

            for item in existing
        }


        for item in pages:

            existing_map[
                item.get("row_index")
            ] = item


        combined = sorted(

            existing_map.values(),

            key=lambda x: x.get(
                "row_index",
                0
            )
        )
    

        self.extracted_text = json.dumps(
            combined
        )


        # =====================================
        # SAVE PROGRESS
        # =====================================

        new_index = (
            start_index
            +
            current_count
        )

        self.excel_parse_index = (
            new_index
        )

        _logger.warning(

            f"[EXCEL PARSE INDEX SAVE] "

            f"{self.excel_parse_index}"
        )


        # =====================================
        # DEBUG
        # =====================================

        remaining = max(
            total_rows - new_index,
            0
        )

        progress = round(

            (new_index / total_rows) * 100,

            2

        ) if total_rows else 0


        _logger.warning(
            f"[DEBUG] CURRENT INDEX "
            f"→ {new_index}"
        )

        _logger.warning(
            f"[DEBUG] REMAINING ROWS "
            f"→ {remaining}"
        )

        _logger.warning(
            f"[DEBUG] PROGRESS "
            f"→ {progress}%"
        )

        _logger.warning(
            f"EXCEL NEW INDEX "
            f"→ {new_index}"
        )

        _logger.warning(
            f"EXCEL BATCH STORED "
            f"→ {len(pages)} rows"
        )


        # =====================================
        # COMPLETION
        # =====================================

        if new_index >= total_rows:

            _logger.warning(
                "EXCEL → PARSING COMPLETED ✅"
            )

            self.is_excel_parsed = True

            self.state = "excel_ai"

        else:

            _logger.warning(
                "EXCEL → MORE DATA REMAIN "
                "→ NEXT CRON"
            )

            self.state = "excel_parsing"


        wb.close()

    # =====================================================
    # REMOVE TEXT AREAS
    # =====================================================

    def _trim_catalog_whitespace(self, pil_image):

        try:

            bg = Image.new(
                pil_image.mode,
                pil_image.size,
                pil_image.getpixel((0, 0))
            )

            diff = ImageChops.difference(
                pil_image,
                bg
            )

            bbox = diff.getbbox()

            if bbox:
                pil_image = pil_image.crop(bbox)

            return pil_image

        except Exception:

            return pil_image

    # =====================================================
    # SEGMENT CATALOG PAGE INTO CLEAN PRODUCT ASSETS
    # =====================================================

    def _segment_catalog_images(self, images):

        segmented_images = []

        if not images:
            return segmented_images

        for img_b64 in images:

            try:

                img_data = base64.b64decode(img_b64)

                pil_image = Image.open(
                    BytesIO(img_data)
                ).convert("RGB")

                original_width, original_height = pil_image.size

                # =========================================
                # CONVERT TO OPENCV
                # =========================================

                cv_image = cv2.cvtColor(
                    np.array(pil_image),
                    cv2.COLOR_RGB2BGR
                )

                gray = cv2.cvtColor(
                    cv_image,
                    cv2.COLOR_BGR2GRAY
                )

                # =========================================
                # THRESHOLD
                # =========================================

                _, thresh = cv2.threshold(
                    gray,
                    245,
                    255,
                    cv2.THRESH_BINARY_INV
                )

                # =========================================
                # DILATION
                # =========================================

                kernel = cv2.getStructuringElement(
                    cv2.MORPH_RECT,
                    (9, 9)
                )

                dilated = cv2.dilate(
                    thresh,
                    kernel,
                    iterations=2
                )

                # =========================================
                # FIND CONTOURS
                # =========================================

                contours, _ = cv2.findContours(
                    dilated,
                    cv2.RETR_EXTERNAL,
                    cv2.CHAIN_APPROX_SIMPLE
                )

                filtered_contours = []

                for contour in contours:

                    area = cv2.contourArea(contour)

                    if area < 8000:
                        continue

                    x, y, w, h = cv2.boundingRect(contour)

                    # reject ultra-thin text columns
                    if w < 120 or h < 120:
                        continue

                    ratio = w / float(h)

                    # reject long text strips
                    if ratio > 4.5 or ratio < 0.22:
                        continue

                    filtered_contours.append(contour)

                contours = filtered_contours[:24]

                candidate_crops = []

                for contour in contours:

                    x, y, w, h = cv2.boundingRect(contour)

                    # ======================================
                    # SIZE FILTERS
                    # ======================================

                    if w < 120 or h < 120:
                        continue

                    # reject huge full page
                    if (
                        w > original_width * 0.95
                        and
                        h > original_height * 0.95
                    ):
                        continue

                    area = w * h

                    # reject tiny fragments
                    if area < 25000:
                        continue

                    # =====================================
                    # CROP
                    # =====================================

                    pad = 12

                    x1 = max(x - pad, 0)
                    y1 = max(y - pad, 0)
                    x2 = min(x + w + pad, original_width)
                    y2 = min(y + h + pad, original_height)

                    crop = pil_image.crop(
                        (x1, y1, x2, y2)
                    )

                    crop = self._trim_catalog_whitespace(
                        crop
                    )

                    # =====================================
                    # VALIDATE
                    # =====================================

                    if not self._is_valid_product_crop(crop):
                        continue

                    # =====================================
                    # OCR-LIKE TEXT REJECTION
                    # =====================================

                    crop_gray = crop.convert("L")

                    crop_arr = np.array(crop_gray)

                    dark_pixels = np.mean(
                        crop_arr < 90
                    )

                    if dark_pixels < 0.002:
                            continue

                    # =====================================
                    # IMAGE ANALYSIS
                    # =====================================

                    crop_width, crop_height = crop.size

                    crop_area = crop_width * crop_height

                    page_area = (
                        original_width * original_height
                    )

                    coverage_ratio = (
                        crop_area / float(page_area)
                    )

                    # =====================================
                    # COLLAGE DETECTION
                    # =====================================

                    is_collage = False

                    if len(filtered_contours) >= 6:

                        is_collage = True

                    # =====================================
                    # CENTER DETECTION
                    # =====================================

                    centered_object = False

                    crop_center_x = x + (w / 2.0)
                    crop_center_y = y + (h / 2.0)

                    page_center_x = (
                        original_width / 2.0
                    )

                    page_center_y = (
                        original_height / 2.0
                    )

                    distance_x = abs(
                        crop_center_x - page_center_x
                    )

                    distance_y = abs(
                        crop_center_y - page_center_y
                    )

                    if (

                        distance_x < original_width * 0.18

                        and

                        distance_y < original_height * 0.18
                    ):

                        centered_object = True

                    # =====================================
                    # SCORE
                    # =====================================
                    human_penalty = 0
                    score = 0

                    # big clean product bonus
                    score += int(
                        coverage_ratio * 140
                    )

                    # centered ecommerce product
                    if centered_object:
                        score += 55

                    # strong collage penalty
                    if is_collage:
                        score -= 70

                    # portrait product bonus
                    if crop_height > crop_width:
                        score += 18

                    # isolated product bonus
                    edge_density = cv2.Canny(
                        crop_arr,
                        80,
                        160
                    ).mean()

                    # =====================================
                    # CLEAN CENTER HERO DETECTION
                    # =====================================

                    background_ratio = np.mean(
                        crop_arr > 235
                    )

                    # strong ecommerce isolated render
                    if (
                        centered_object
                        and
                        background_ratio > 0.45
                        and
                        not is_collage
                    ):
                        score += 120

                    # medium clean product
                    elif (
                        background_ratio > 0.30
                        and
                        not is_collage
                    ):
                        score += 60

                    # dark/lifestyle penalty
                    if background_ratio < 0.12:
                        score -= 55

                    # excessive visual noise
                    if edge_density > 55:
                        score -= 35


                    # =====================================
                    # HUMAN / LIFESTYLE APPROXIMATION
                    # =====================================

                    human_penalty = 0

                    rgb_arr = np.array(crop)

                    r = rgb_arr[:, :, 0]
                    g = rgb_arr[:, :, 1]
                    b = rgb_arr[:, :, 2]

                    skin_mask = (

                        (r > 95)

                        &

                        (g > 40)

                        &

                        (b > 20)

                        &

                        (r > g)

                        &

                        (r > b)

                        &

                        (np.abs(r - g) > 15)
                    )

                    skin_ratio = np.mean(skin_mask)

                    if skin_ratio > 0.28:

                        human_penalty = 40

                    score -= human_penalty

                    if 8 < edge_density < 35:
                        score += 25

                    # =====================================
                    # SAVE
                    # =====================================

                    buffer = BytesIO()

                    crop.save(
                        buffer,
                        format="JPEG",
                        quality=92
                    )

                    encoded = base64.b64encode(
                        buffer.getvalue()
                    ).decode("utf-8")

                    _logger.warning(

                        f"[IMAGE SCORE DEBUG] "

                        f"size={crop_width}x{crop_height} "

                        f"| coverage={coverage_ratio:.3f} "

                        f"| bg={background_ratio:.3f} "

                        f"| edge={edge_density:.2f} "

                        f"| centered={centered_object} "

                        f"| collage={is_collage} "

                        f"| skin={skin_ratio:.3f} "

                        f"| human_penalty={human_penalty} "

                        f"| final={score}"
                    )

                    candidate_crops.append({

                        "image": encoded,

                        "score": score,

                        "is_collage": is_collage
                    })

                    candidate_crops.append({

                        "image": encoded,

                        "score": score,

                        "is_collage": False,

                        "fallback_fullpage": True
                    })

                    _logger.warning(

                        f"[CROP DETECTED] "

                        f"{w}x{h} "

                        f"| score={score} "

                        f"| collage={is_collage}"
                    )

                # =========================================
                # FALLBACK
                # =========================================

                if not candidate_crops:

                    buffer = BytesIO()

                    pil_image.save(
                        buffer,
                        format="JPEG"
                    )

                    encoded = base64.b64encode(
                        buffer.getvalue()
                    ).decode("utf-8")

                    candidate_crops.append({

                        "image": encoded,

                        "score": 45,

                        "is_collage": False
                    })

                segmented_images.extend(
                    candidate_crops
                )

            except Exception as e:

                _logger.warning(
                    f"[SEGMENTATION FAILED] {str(e)}"
                )

        # =============================================
        # DEDUPE
        # =============================================

        deduped = []
        hashes = {}

        for asset in segmented_images:

            try:

                img = asset.get("image")

                image_hash = hashlib.md5(
                    img.encode("utf-8")
                ).hexdigest()


                existing_score = hashes.get(
                    image_hash
                )

                if existing_score == asset.get(
                    "score"
                ):
                    continue

                hashes[image_hash] = asset.get(
                    "score"
                )

                deduped.append(asset)

            except Exception:
                continue

        return deduped


    # =====================================================
    # VARIANTS IMAGES CONTROLLER/DETECTOR
    # =====================================================

    def _split_grid_products(self, image):

        try:

            import cv2
            import numpy as np
            import base64

            gray = cv2.cvtColor(
                image,
                cv2.COLOR_BGR2GRAY
            )

            thresh = cv2.adaptiveThreshold(
                gray,
                255,
                cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY_INV,
                21,
                5
            )

            contours, _ = cv2.findContours(
                thresh,
                cv2.RETR_EXTERNAL,
                cv2.CHAIN_APPROX_SIMPLE
            )

            results = []

            for contour in contours:

                area = cv2.contourArea(contour)

                if area < 3500:
                    continue

                x, y, w, h = cv2.boundingRect(contour)

                if w < 80 or h < 80:
                    continue

                ratio = w / float(h)

                # reject text strips
                if ratio > 4.0 or ratio < 0.25:
                    continue

                sub = image[
                    y:y+h,
                    x:x+w
                ]

                success, buffer = cv2.imencode(
                    '.jpg',
                    sub
                )

                if not success:
                    continue

                results.append(
                    base64.b64encode(
                        buffer
                    ).decode()
                )

            return results[:12]

        except Exception as e:

            _logger.warning(
                f"[GRID SPLIT FAILED] {str(e)}"
            )

            return []

    # =====================================================
    # VALIDATE CROPPED IMAGE
    # =====================================================

    def _is_valid_product_crop(self, pil_image):

        try:

            width, height = pil_image.size

            # too tiny
            if width < 120 or height < 120:
                return False

            # aspect safety
            ratio = width / float(height)

            if ratio > 5 or ratio < 0.15:
                return False

            gray = pil_image.convert("L")

            arr = np.array(gray)

            # blank image rejection
            if arr.std() < 14:
                return False

            # excessive dark block rejection
            dark_ratio = np.mean(arr < 25)

            if dark_ratio > 0.92:
                return False

            return True

        except Exception:

            return False

    #=========VALIDATE AI IMAGE====================================
    def _is_valid_ai_image(self, image_data):

        try:

            if not image_data:
                return False

            import base64
            import io

            from PIL import Image

            # remove data url prefix
            if ',' in image_data:
                image_data = image_data.split(',')[1]

            decoded = base64.b64decode(image_data)

            img = Image.open(
                io.BytesIO(decoded)
            )

            img.verify()

            return True

        except Exception as e:

            _logger.warning(

                f"[INVALID AI IMAGE] {str(e)}"
            )

            return False
        
    # ---------------- Extract PDF ----------------
 
    def extract_pdf(self):

        import gc
        import json
        import io
        import re
        import fitz
        import base64
        import requests

        _logger.warning(
            f"[PDF EXTRACT] START "
            f"| job={self.id}"
        )

        MAX_RETRIES = 3

        # balanced batch size
        BATCH_SIZE = 3

        doc = None

        try:

            pdf_bytes = base64.b64decode(
                self.pdf_file
            )

        except Exception as e:

            _logger.exception(
                f"[PDF EXTRACT ERROR] "
                f"PDF DECODE FAILED "
                f"| {str(e)}"
            )

            self.state = "failed"

            return


        # =========================================
        # OPEN PDF
        # =========================================

        try:

            doc = fitz.open(
                stream=pdf_bytes,
                filetype="pdf"
            )

        except Exception as e:

            _logger.exception(
                f"[PDF EXTRACT ERROR] "
                f"PDF OPEN FAILED "
                f"| {str(e)}"
            )

            self.state = "failed"

            return


        try:

            total_pages = len(doc)

            self.total_pages = total_pages


            _logger.warning(
                f"[PDF EXTRACT] "
                f"TOTAL PAGES={total_pages}"
            )


            # =====================================
            # CRASH SAFE RECOVERY
            # =====================================

            existing_pages = self.env[
                'vendor.import.page'
            ].search([

                ('job_id', '=', self.id)

            ], order='page_number desc', limit=1)


            if existing_pages:

                # move to NEXT page
                start_page = (
                    existing_pages.page_number
                )

                _logger.warning(
                    f"[PDF RECOVERY] "
                    f"LAST SAVED PAGE="
                    f"{existing_pages.page_number}"
                )

            else:

                start_page = (
                    self.current_page or 0
                )

                _logger.warning(
                    f"[PDF RECOVERY] "
                    f"NO SAVED PAGES"
                )


            # =====================================
            # SAFETY CLAMP
            # =====================================

            if start_page >= total_pages:

                start_page = total_pages


            end_page = min(
                start_page + BATCH_SIZE,
                total_pages
            )


            _logger.warning(
                f"[PDF BATCH] "
                f"START={start_page + 1} "
                f"| END={end_page}"
            )


            processed_count = 0


            # =====================================
            # PROCESS PAGES
            # =====================================

            for i in range(start_page, end_page):

                _logger.warning(
                    f"[PDF PAGE] "
                    f"START PAGE={i + 1}"
                )

                page_success = False


                # =================================
                # SKIP IF ALREADY EXISTS
                # =================================

                existing = self.env[
                    'vendor.import.page'
                ].search([

                    ('job_id', '=', self.id),

                    ('page_number', '=', i + 1)

                ], limit=1)


                if existing:

                    _logger.warning(
                        f"[PDF PAGE] "
                        f"SKIP EXISTING "
                        f"| page={i + 1}"
                    )

                    self.current_page = i + 1

                    continue


                for attempt in range(MAX_RETRIES):

                    single_pdf = None
                    pdf_bytes_io = None

                    try:

                        _logger.warning(
                            f"[PDF API] "
                            f"PAGE={i + 1} "
                            f"| ATTEMPT={attempt + 1}"
                        )


                        # =========================
                        # SINGLE PAGE PDF
                        # =========================

                        single_pdf = fitz.open()

                        single_pdf.insert_pdf(

                            doc,

                            from_page=i,

                            to_page=i
                        )


                        pdf_bytes_io = io.BytesIO()

                        single_pdf.save(
                            pdf_bytes_io
                        )

                        pdf_bytes_io.seek(0)


                        # =========================
                        # API CALL
                        # =========================

                        response = requests.post(

                            "https://pdf-extractor-staging.onrender.com/extract",

                            files={

                                "file": (

                                    "page.pdf",

                                    pdf_bytes_io,

                                    "application/pdf"
                                )
                            },

                            timeout=45
                        )


                        _logger.warning(
                            f"[PDF API] "
                            f"STATUS="
                            f"{response.status_code} "
                            f"| page={i + 1}"
                        )


                        if response.status_code != 200:

                            continue


                        page_data = response.json()


                        # =========================
                        # RESPONSE FORMAT
                        # =========================

                        if isinstance(page_data, dict):

                            pages = page_data.get(
                                "pages",
                                []
                            )

                        elif isinstance(
                            page_data,
                            list
                        ):

                            pages = page_data

                        else:

                            pages = []


                        if not pages:

                            _logger.warning(
                                f"[PDF PAGE] "
                                f"EMPTY RESPONSE "
                                f"| page={i + 1}"
                            )

                            continue


                        normalized_blocks = []


                        # =========================
                        # NORMALIZE
                        # =========================

                        for p in pages:

                            text = p.get(
                                "text",
                                ""
                            )

                            images = p.get(
                                "images",
                                []
                            )

                            # ===========================
                            # CLEAN CATALOG SEGMENTATION
                            # ===========================

                            images = self._segment_catalog_images(
                                images
                            )


                            if (
                                not text
                                and
                                not images
                            ):
                                continue


                            price = ""

                            stock = ""


                            price_match = re.search(

                                r'(\$|€|£)\s?\d+[.,]?\d*',

                                text
                            )


                            if price_match:

                                price = (
                                    price_match.group(0)
                                )


                            stock_match = re.search(

                                r'(stock|available)'
                                r'\s*:?\s*'
                                r'(\d+)'
                                r'\s*(pcs|pieces)?',

                                text,

                                re.I
                            )


                            if stock_match:

                                stock = (
                                    stock_match.group(2)
                                )


                            normalized_blocks.append({

                                "page": i + 1,

                                "text": text,

                                "price": price,

                                "stock": stock,

                                "images": images
                            })


                        if not normalized_blocks:

                            _logger.warning(
                                f"[PDF PAGE] "
                                f"NO VALID BLOCKS "
                                f"| page={i + 1}"
                            )

                            continue


                        # ===========================
                        # SAVE PAGE
                        # ===========================


                        all_page_images = []

                        for block in normalized_blocks:

                            all_page_images.extend(
                                block.get("images", [])
                            )

                            self._safe_commit_progress()

                        self.env[
                            'vendor.import.page'
                        ].create({

                            'job_id': self.id,

                            'page_number': i + 1,

                            'extracted_json': json.dumps(
                                normalized_blocks
                            ),

                            'page_images_json': json.dumps(
                                all_page_images
                            )
                        })


                        _logger.warning(
                            f"[PDF PAGE] "
                            f"SAVED "
                            f"| page={i + 1}"
                        )


                        self.current_page = i + 1

                        processed_count += 1

                        page_success = True

                        break


                    except Exception as e:

                        _logger.exception(
                            f"[PDF PAGE ERROR] "
                            f"page={i + 1} "
                            f"| {str(e)}"
                        )


                    finally:

                        try:

                            if pdf_bytes_io:
                                pdf_bytes_io.close()

                        except Exception:
                            pass


                        try:

                            if single_pdf:
                                single_pdf.close()

                        except Exception:
                            pass


                if not page_success:

                    _logger.error(
                        f"[PDF PAGE FAILED] "
                        f"page={i + 1}"
                    )


            # =====================================
            # SAVE BATCH ONCE
            # =====================================

            _logger.warning(
                f"[PDF BATCH] "
                f"PROCESSED="
                f"{processed_count}"
            )


            if self.current_page < total_pages:

                self.state = "pdf_extracting"

            else:

                self.state = "pdf_ai"


            try:

                self._safe_commit_progress()

                _logger.warning(
                    f"[PDF SAVE] "
                    f"SUCCESS "
                    f"| state={self.state} "
                    f"| current={self.current_page}"
                )

            except Exception as e:

                _logger.exception(
                    f"[PDF SAVE ERROR] "
                    f"{str(e)}"
                )


        finally:

            try:

                if doc:
                    doc.close()

            except Exception:
                pass


            gc.collect()

            _logger.warning(
                "[PDF GC] COMPLETE"
            )


    # ---------------- Send to OPENAI URL ----------------
    def send_to_openai_url(self):

        import re
        import json
        import math

        api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')

        if not api_key:
            raise Exception("OpenAI API key not configured")

        client = OpenAI(api_key=api_key)

        # ================= LOAD PAGES =================
        try:
            # pages = json.loads(self.extracted_text or "[]")

            pages = json.loads(

                self.url_blocks_json

                or

                self.extracted_text

                or

                "[]"
            )

        except Exception:
            _logger.error("INVALID extracted_text JSON")
            return

        if not pages:
            _logger.error("NO PAGES TO PROCESS")
            return

        # ================= LOAD EXISTING =================
        existing_products = []
        if self.ai_response:
            try:
                data = json.loads(self.ai_response)
                if isinstance(data, list):
                    existing_products = data
            except Exception as e:
                _logger.warning(f"AI RESPONSE LOAD FAILED → {str(e)}")
                existing_products = []

        current_batch = self.url_batch_index or 0

        # ================= FLATTEN =================
        all_blocks = [b for p in pages for b in p.get("blocks", [])]

        _logger.warning(f"RAW BLOCKS → {len(all_blocks)}")

        # ================= CLEAN =================
        cleaned_blocks = self._clean_scraped_blocks(all_blocks)

        _logger.warning(f"CLEAN BLOCKS → {len(cleaned_blocks)}")
        _logger.warning(f"REMOVED BLOCKS → {len(all_blocks) - len(cleaned_blocks)}")

        cleaned_blocks = sorted(cleaned_blocks, key=lambda x: (x.get("text") or "")[:50])

        # ================= BATCH =================
        BLOCK_BATCH_SIZE = 8

        batched_blocks = [
            cleaned_blocks[i:i + BLOCK_BATCH_SIZE]
            for i in range(0, len(cleaned_blocks), BLOCK_BATCH_SIZE)
        ]

        total_batches = len(batched_blocks)
        self.url_total_batches = total_batches

        _logger.warning(f"TOTAL BLOCK BATCHES → {total_batches}")
        _logger.warning(f"CURRENT BATCH → {current_batch}")

        # ================= STOP IF DONE =================
        if current_batch >= total_batches:
            _logger.warning("ALL URL BATCHES PROCESSED ✅")
            self.state = "url_creating"
            return

        # ================= PROCESS ONE BATCH =================
        block_batch = batched_blocks[current_batch]

        _logger.warning(f"PROCESSING BLOCK COUNT → {len(block_batch)}")
        _logger.warning(f"AI → PROCESSING BLOCK BATCH {current_batch + 1}")

      
        combined_text = "\n\n---\n\n".join([
            f"""
            TEXT:
            {b.get('text','')}

            PRICE:
            {b.get('price','')}

            STOCK:
            {b.get('stock','')}

            IMAGE_URL:
            {b.get('image','')}
            """

            for b in block_batch
        ])


        if not combined_text.strip():
            _logger.warning("EMPTY COMBINED TEXT → SKIP")
            self.url_batch_index += 1
            return

        if len(combined_text) > 15000:
            combined_text = combined_text[:15000]
            _logger.warning("TEXT TRIMMED → PREVENT TOKEN OVERFLOW")

        # ================= PROMPT =================
        prompt = f"""
            You are a highly precise e-commerce product extraction engine.

            You are processing raw scraped website content.

            =====================================
            YOUR GOAL
            =====================================

            Extract ALL individual products from the text.

            IMPORTANT:
            - A single block may contain MULTIPLE products → you MUST split them
            - Each product MUST be returned separately
            - NEVER merge multiple products into one

            =====================================
            HOW TO IDENTIFY A PRODUCT
            =====================================

            A product usually contains:
            - product name
            - price (e.g. $, £, €, numbers)
            - specs or description
            - sometimes reviews

            Strong indicators:
            - currency symbols ($, £, €)
            - model names (Dell, Lenovo, etc.)
            - numbers like GB, inch, SSD, RAM, etc.

            =====================================
            STRICT RULES
            =====================================

            1. RETURN ONLY VALID JSON ARRAY
            2. NO explanation
            3. NO markdown
            4. NO text outside JSON

            5. EACH product must be unique
            6. REMOVE duplicates
            7. DO NOT skip products
            8. SPLIT combined text into multiple products

            =====================================
            REMOVE THIS NOISE
            =====================================

            Ignore anything related to:
            - navigation (Home, Login, Pricing)
            - cookies/privacy
            - footer text
            - categories only (no product info)
            - repeated headers

           =====================================
            VARIANT DETECTION
            =====================================

            IMPORTANT:

            Use PRODUCT IMAGES as the PRIMARY
            source for detecting variants.

            Also use:
            - product title
            - description
            - SKU
            - repeated patterns
            - packaging labels
            - text printed on product
            - visible size/capacity markings

            Detect REAL differences such as:
            - color
            - material
            - finish
            - texture
            - pattern
            - lid type
            - bottle type
            - packaging
            - shape
            - capacity
            - dimensions
            - style
            - design
            - print variation

            IMPORTANT RULES:

            1. NEVER generate:
            - Variant 1
            - Variant 2
            - Default
            - Standard
            - Option A
            - Option B

            2. ALWAYS return meaningful
            attribute names and values.

            GOOD EXAMPLES:

            {{
            "Color": "Black"
            }}

            {{
            "Material": "Bamboo"
            }}

            {{
            "Capacity": "750ml"
            }}

            {{
            "Design": "Football Print"
            }}

            {{
            "Finish": "Matte Silver"
            }}

            3. If a SINGLE IMAGE contains
            MULTIPLE product colors/designs:

            Create SEPARATE variants for EACH
            visible product variation.

            Example:
            - black bottle
            - blue bottle
            - red bottle

            MUST become:

            [
            {{
                "attributes": {{
                "Color": "Black"
                }}
            }},
            {{
                "attributes": {{
                "Color": "Blue"
                }}
            }},
            {{
                "attributes": {{
                "Color": "Red"
                }}
            }}
            ]

            4. If products differ by:
            - artwork
            - printed graphics
            - pattern
            - branding
            - sports design
            - texture

            Use:
            {{
            "Design": "..."
            }}

            5. If products differ mainly by:
            - size
            - dimensions
            - capacity

            Use:
            {{
            "Size": "..."
            }}

            OR

            {{
            "Capacity": "..."
            }}

            6. NEVER invent attributes that
            cannot be visually or textually
            supported.

            7. If uncertainty exists:
            Prefer:
            - Color
            - Design
            - Material
            - Capacity

            based on strongest visible evidence.

            8. If NO meaningful difference exists:
            Return ONE variant only.

            9. IMPORTANT:
            When multiple products appear in
            one image, treat each visible
            variation as a separate variant,
            even if no explicit text exists.

            10. Preserve consistency across
            all variants for the same product.

            BAD EXAMPLE:
            [
            {{
                "attributes": {{
                "Variant": "Variant 1"
                }}
            }}
            ]

            GOOD EXAMPLE:
            [
            {{
                "attributes": {{
                "Color": "White"
                }}
            }},
            {{
                "attributes": {{
                "Color": "Black"
                }}
            }}
            ]


            =====================================
            OUTPUT FORMAT
            =====================================

            [
                {{
                    "name": "Clean product name",
                    "description": "Short product description (max 30 words)",
                    "category": "Best guess category",
                    "price": "",
                    "stock": "",
                    "image": "image_url_or_null",
                    "variants": [
                                {{
                                    "attributes": {{
                                        "Variant": ""
                                    }},
                                    "image_index": 0,
                                    "stock": null
                                }}
                            ]
                }}
            ]

            =====================================
            EXTRA RULES
            =====================================

            - Keep names SHORT and CLEAN
            - Description must be concise
            - Infer category intelligently
            - If no image exists → return null
            - If price exists → extract it
            - If stock exists → extract it
            - NEVER invent stock or price
            - If unsure → still extract

            =====================================
            TEXT TO PROCESS
            =====================================

        {combined_text}
        """

        # ================= OPENAI =================
        try:
            response = client.responses.create(
                model="gpt-4.1-mini",
                input=prompt,
                temperature=0,
                timeout=60
            )

            result = response.output_text.strip()
            result = re.sub(r"^```(?:json)?|```$", "", result).strip()

            parsed = json.loads(result)

            if isinstance(parsed, list):

                cleaned = [p for p in parsed if p.get("name")]

                _logger.warning(f"AI RETURNED → {len(cleaned)} PRODUCTS")

                # 🔥 DEDUPE BY NAME
                existing_map = {p.get("name"): p for p in existing_products}

                for p in cleaned:
                    existing_map[p.get("name")] = p

                existing_products = list(existing_map.values())

                _logger.warning(f"TOTAL ACCUMULATED → {len(existing_products)}")

            else:
                _logger.warning("AI RESPONSE NOT LIST")

        except Exception as e:
            _logger.warning(f"AI ERROR → {str(e)}")
            return

        # ================= SAVE =================
        self.ai_response = json.dumps(existing_products)
        self.url_batch_index = current_batch + 1

        _logger.warning(f"URL AI PROGRESS → {self.url_batch_index}/{self.url_total_batches}")

        # ================= STATE =================
        if self.url_batch_index < self.url_total_batches:
            self.state = "url_ai"
        else:
            _logger.warning("URL AI FINISHED ALL BATCHES")
            self.state = "url_creating"

        # 🔥 IMPORTANT: COMMIT FOR CRON CONTINUITY

        try:

            self.env.cr.commit()

        except Exception as commit_error:

            _logger.warning(
                f"COMMIT SKIPPED → {commit_error}"
            )

        return


    # =========== PDF OPENAI =========================

    def send_to_openai_pdf(self):

        import json

        api_key = self.env[
            'ir.config_parameter'
        ].sudo().get_param(
            'openai.api.key'
        )

        if not api_key:

            raise Exception(
                "OpenAI API key not configured"
            )


        client = OpenAI(
            api_key=api_key
        )


        _logger.warning(
            "[PDF AI] START"
        )


        # =====================================================
        # LOAD PAGE RECORDS
        # =====================================================

        page_records = self.env[
            'vendor.import.page'
        ].search([

            ('job_id', '=', self.id)

        ], order='page_number asc')


        total_available_pages = len(
            page_records
        )


        _logger.warning(

            f"[PDF AI] "

            f"TOTAL PAGE RECORDS="

            f"{total_available_pages}"
        )


        if total_available_pages <= 0:

            _logger.warning(
                "[PDF AI] "
                "NO PAGE RECORDS FOUND"
            )

            return


        # =====================================================
        # LOAD EXISTING AI RESPONSE
        # =====================================================

        existing_pages = []


        if self.ai_response:

            try:

                loaded = json.loads(
                    self.ai_response
                )

                if isinstance(
                    loaded,
                    list
                ):

                    existing_pages = loaded


            except Exception as e:

                _logger.warning(

                    f"[PDF AI] "

                    f"LOAD EXISTING FAILED "

                    f"| {str(e)}"
                )

                existing_pages = []


        # =====================================================
        # FIND ALREADY PROCESSED PAGES
        # =====================================================

        processed_pages = set()


        for p in existing_pages:

            page_num = p.get("page")

            if page_num:

                processed_pages.add(
                    page_num
                )


        _logger.warning(

            f"[PDF AI] "

            f"PROCESSED PAGES="

            f"{sorted(list(processed_pages))}"
        )


        # =====================================================
        # FIND NEXT UNPROCESSED PAGE
        # =====================================================

        next_record = None


        for record in page_records:

            if (

                record.page_number

                not in processed_pages

            ):

                next_record = record

                break


        # =====================================================
        # ALL COMPLETE
        # =====================================================

        if not next_record:

            _logger.warning(
                "[PDF AI] COMPLETE ✅"
            )

            self.last_ai_page = (
                total_available_pages
            )

            self.state = "pdf_creating"

            self.flush_recordset()

            self.env.cr.commit()

            return


        _logger.warning(

            f"[PDF AI] "

            f"PROCESSING PAGE "

            f"{next_record.page_number}"
        )


        # =====================================================
        # LOAD PAGE DATA
        # =====================================================

        try:

            page_blocks = json.loads(

                next_record.extracted_json
                or
                "[]"
            )

        except Exception as e:

            _logger.warning(

                f"[PDF AI] "

                f"PAGE LOAD FAILED "

                f"| PAGE "

                f"{next_record.page_number} "

                f"| {str(e)}"
            )

            return


        if not page_blocks:

            _logger.warning(

                f"[PDF AI] "

                f"EMPTY PAGE BLOCKS "

                f"| PAGE "

                f"{next_record.page_number}"
            )

            return


        # =====================================================
        # BUILD PAGE DATA
        # =====================================================

        page_text = "\n".join([

            p.get("text", "")

            for p in page_blocks

        ])


        page_images = []

        for p in page_blocks:

            raw_images = p.get("images", [])

            # =====================================
            # NORMALIZE STRUCTURED ASSETS
            # =====================================

            for img in raw_images:

                if isinstance(img, dict):

                    if img.get("image"):

                        page_images.append(img)

                elif isinstance(img, str):

                    page_images.append({

                        "image": img,

                        "score": 0,

                        "is_collage": False
                    })

        # =====================================================
        # VALIDATE PAGE IMAGES
        # =====================================================

        valid_page_images = []

        for asset in page_images:

            try:

                # segmented assets are now dicts
                if isinstance(asset, dict):

                    image_data = asset.get(
                        "image"
                    )

                else:

                    image_data = asset

                if not image_data:
                    continue

                if not self._is_valid_ai_image(
                    image_data
                ):

                    _logger.warning(

                        f"[PDF AI] INVALID IMAGE "

                        f"| PAGE "

                        f"{next_record.page_number}"
                    )

                    continue

                valid_page_images.append(
                    asset
                )

            except Exception as e:

                _logger.warning(

                    f"[PDF AI IMAGE ERROR] "

                    f"{str(e)}"
                )

        page_images = valid_page_images
        # =========================================
        # REBUILD CLEAN IMAGE INDEX MAP
        # =========================================

        normalized_page_images = []

        for idx, asset in enumerate(page_images):

            if isinstance(asset, dict):

                asset["clean_index"] = idx

                normalized_page_images.append(asset)


        page_images = normalized_page_images

        _logger.warning(

            f"[PDF AI IMAGES] "

            f"PAGE={next_record.page_number} "

            f"| valid={len(page_images)}"
        )

        page_price = ""

        page_stock = ""


        for p in page_blocks:

            if (

                not page_price

                and

                p.get("price")

            ):

                page_price = (
                    p.get("price")
                )


            if (

                not page_stock

                and

                p.get("stock")

            ):

                page_stock = (
                    p.get("stock")
                )


        # =====================================================
        # PROMPT
        # =====================================================
       
        prompt = f"""
        You are an advanced AI product extraction engine for catalog PDF pages.

        You analyze BOTH:
        - page text
        - catalog product images

        Your job:
        extract ALL visible products accurately.

        ==================================================
        STRICT OUTPUT RULES
        ==================================================

        1. RETURN ONLY VALID JSON ARRAY
        2. NO markdown
        3. NO explanation
        4. NO text outside JSON
        5. NEVER invent products not visible
        6. NEVER skip visible products
        7. NEVER duplicate products
        8. EACH product must appear ONLY ONCE
        9. ALWAYS preserve product grouping correctly

        ==================================================
        CATALOG UNDERSTANDING RULES
        ==================================================

        This input represents ONLY ONE catalog page.

        DO NOT:
        - continue products from previous pages
        - assume future pages
        - merge unrelated products

        A page may contain:
        - one hero product
        - multiple products
        - one product with variants
        - one product with gallery/supporting images

        ==================================================
        PRODUCT DETECTION RULES
        ==================================================

        If a page contains:
        - visually separated products
        - different product names
        - different product codes
        - different structures/shapes

        Then:
        extract them as SEPARATE products.

        IMPORTANT:

        If unsure:
        it is BETTER to slightly over-detect
        than to miss products.

        NEVER silently ignore visible products.

        ==================================================
        STOCK EXTRACTION RULES:
        ==================================================

        Extract stock quantity ONLY when
        actual available inventory is explicitly stated.

        Examples:
        - "Stock: 11 pcs"
        - "Available: 25"
        - "In stock: 8"

        DO NOT extract:
        - delivery times
        - MOQ
        - carton quantity
        - package quantity
        - shipping quantity
        - lead times
        - dimensions
        - capacity values

        If no real stock quantity exists:
        set:

        "stock_qty": 0

        ==================================================
        VARIANT DETECTION RULES
        ==================================================

        VARIANT GROUPING RULES:

        Products MUST be grouped as variants when:

        - same product shape
        - same structure
        - same branding
        - same dimensions
        - same material
        - only color changes
        - only size changes
        - only minor style changes

        EXAMPLES:
        - same cap in multiple colors
        - same polo shirt in different colors
        - same bottle with color variations

        DO NOT create separate products for:
        - color-only changes
        - size-only changes

        Instead:
        create ONE parent product with variants.

        Each variant should contain:

        {{
            "attributes": {{
                "Color": "",
                "Size": ""
            }},

            "image_index": null
        }}

        ==================================================
        ECOMMERCE IMAGE UNDERSTANDING RULES
        ==================================================

        You are NOT selecting the most artistic image.

        You are selecting the BEST PROFESSIONAL
        ECOMMERCE PRODUCT IMAGE.

        Your goal:
        produce Amazon/Alibaba/Shopify-style
        product merchandising quality.

        --------------------------------------------------
        PRIORITY ORDER (VERY IMPORTANT)
        --------------------------------------------------

        ALWAYS prioritize:

        1. isolated standalone product
        2. clean white/plain background
        3. centered product
        4. full product visibility
        5. variant color visibility
        6. clean catalog render
        7. multiple isolated color options

        NEVER prioritize:
        - humans/models
        - lifestyle scenes
        - promotional layouts
        - infographic compositions
        - text-heavy blocks
        - banners
        - decorative graphics

        ==================================================
        HERO IMAGE RULES
        ==================================================

        hero_image_index MUST point to:

        - ONE isolated product
        - clean/plain background
        - centered product
        - professional ecommerce shot
        - no text overlays
        - no large text areas
        - no promotional layout
        - no infographic composition

        DO NOT use:
        - humans wearing products
        - lifestyle photography
        - catalog cover layouts
        - multi-product collages
        - pages with large text blocks
        - specification layouts
        - promotional graphics

        VERY IMPORTANT:

        If isolated product variants exist anywhere
        on the page,
        ALWAYS prefer them over:
        - human/model photos
        - lifestyle shots
        - promotional scenes

        Example:
        If a cap page contains:
        - woman wearing cap
        - isolated cap colors

        hero_image_index MUST use:
        isolated cap color image

        NOT the woman/model image.

        ==================================================
        GALLERY IMAGE RULES
        ==================================================

        gallery_image_indexes should contain ONLY:

        - isolated alternate angles
        - isolated closeups
        - isolated detail shots
        - isolated side/back views

        DO NOT include:
        - banners
        - specification layouts
        - infographic graphics
        - text-heavy images
        - decorative layouts
        - icons
        - logos
        - promotional compositions

        ==================================================
        VARIANT IMAGE RULES
        ==================================================

        Variants MUST be created when:

        - same product
        - same shape
        - same structure
        - same dimensions
        - only color/material/style changes

        IMPORTANT:

        If multiple isolated product colors exist,
        they MUST become variants.

        Example:
        - black cap
        - blue cap
        - red cap

        MUST become:
        ONE product
        with multiple color variants.

        DO NOT create separate products.

        Each variant should contain:
        - correct Color/Material attribute
        - correct image_index

        ==================================================
        COLLAGE UNDERSTANDING RULES
        ==================================================

        Supplier catalog pages often contain:
        - one large lifestyle image
        - multiple smaller isolated products

        IMPORTANT:

        The smaller isolated products are usually
        the CORRECT ecommerce assets.

        DO NOT automatically prefer the largest image.

        Prefer:
        isolated product renders
        over:
        visually dominant lifestyle graphics.

        ==================================================
        PRICE/STOCK RULES
        ==================================================

        Extract:
        - visible product price
        - visible stock quantity
        - visible product code

        If stock/price belongs to a specific variant:
        assign it to that variant.

        DO NOT invent prices or stock.

        ==================================================
        OUTPUT FORMAT
        ==================================================

        Return JSON ARRAY:

        [
            {{
                "name": "",
                "description": "",
                "stock_qty": 0,
                "price": "",
                "product_code": "",
                "hero_image_index": null,
                "gallery_image_indexes": [],
                "variants": [
                    {{
                        "attributes": {{
                            "Color": ""
                        }},

                        "image_index": null,
                        "stock_qty": 0,
                        "price": ""
                    }}
                ]
            }}
        ]

        ==================================================
        PAGE TEXT
        ==================================================

        {page_text}

        ==================================================
        DETECTED PRICE
        ==================================================

        {page_price}

        ==================================================
        DETECTED STOCK
        ==================================================

        {page_stock}
        """

        # =====================================================
        # AI CALL
        # =====================================================

        try:
            
            MAX_IMAGES = 15

            image_inputs = []

            sorted_page_images = sorted(

                page_images,

                key=lambda x: x.get(
                    "score",
                    0
                ),

                reverse=True
            )

            for asset in sorted_page_images[:MAX_IMAGES]:

                try:

                    # =====================================
                    # SUPPORT DICT ASSETS
                    # =====================================

                    if isinstance(asset, dict):

                        image_data = asset.get(
                            "image"
                        )

                    else:

                        image_data = asset

                    if not image_data:
                        continue

                    image_inputs.append({

                        "type": "input_image",

                        "image_url":

                            f"data:image/jpeg;base64,{image_data}"
                    })

                except Exception as e:

                    _logger.warning(

                        f"[IMAGE INPUT BUILD FAILED] "

                        f"{str(e)}"
                    )


            response = client.responses.create(

                model="gpt-4.1",

                input=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": prompt
                        }
                    ] + image_inputs
                }],

                timeout=60
            )


            result = (
                response.output_text or ""
            ).strip()


            result = result.replace(
                "```json",
                ""
            )

            result = result.replace(
                "```",
                ""
            ).strip()


            if not result:

                raise Exception(
                    "EMPTY AI RESPONSE"
                )


            try:

                parsed = json.loads(result)

            except Exception as e:

                _logger.warning(

                    f"[PDF AI JSON FAILED] "

                    f"PAGE={next_record.page_number} "

                    f"| {str(e)}"
                )

                _logger.warning(

                    f"[PDF AI RAW OUTPUT] "

                    f"{result[:1200]}"
                )

                next_record.write({

                    'state': 'failed'
                })

                self._safe_commit_progress()

                return

            if not parsed:

                _logger.warning(

                    f"[PDF AI EMPTY RESPONSE] "

                    f"PAGE={next_record.page_number}"
                )

                next_record.write({

                    'state': 'failed'
                })

                self._safe_commit_progress()

                return


            if not isinstance(
                parsed,
                list
            ):

                parsed = []

            _logger.warning(
                f"[PDF AI PARSED COUNT] "
                f"{len(parsed)} products"
            )

            for product in parsed:

                _logger.warning(

                    f"[PDF AI PRODUCT] "

                    f"name={product.get('name')} "

                    f"| hero={product.get('hero_image_index')} "

                    f"| gallery={len(product.get('gallery_image_indexes', []))} "

                    f"| variants={len(product.get('variants', []))}"
                )


        except Exception as e:

            _logger.warning(

                f"[PDF AI] "

                f"FAILED "

                f"| PAGE "

                f"{next_record.page_number} "

                f"| {str(e)}"
            )

            return


        # =====================================================
        # SMART IMAGE MATCHING
        # =====================================================

        for prod in parsed:

            try:

                product_name = (
                    prod.get("name")
                    or ""
                )


                best_index = prod.get(
                    "hero_image_index"
                )

                # =====================================
                # VALIDATE CLEAN INDEX
                # =====================================

                valid_indexes = [

                    a.get("clean_index")

                    for a in page_images

                    if isinstance(a, dict)
                ]

                if (

                    best_index is None

                    or

                    not isinstance(best_index, int)

                    or

                    best_index not in valid_indexes
                ):

                    best_index = (
                        self.match_image_index_with_ai(
                            product_name,
                            page_images
                        )
                    )

                    if isinstance(best_index, int):

                        try:

                            matched_asset = page_images[
                                best_index
                            ]

                            if isinstance(matched_asset, dict):

                                best_index = matched_asset.get(
                                    "clean_index"
                                )

                        except Exception:
                            pass

                if best_index is not None:

                    prod["hero_image_index"] = (
                        best_index
                    )

                    _logger.warning(

                        f"[PDF HERO INDEX] "

                        f"{product_name} "

                        f"-> {best_index}"
                    )

            except Exception as e:

                _logger.warning(

                    f"[PDF IMAGE MATCH FAILED] "

                    f"{str(e)}"

                )

        # =====================================================
        # MERGE RESULTS
        # =====================================================

        existing_map = {}


        for p in existing_pages:

            existing_map[
                p.get("page")
            ] = p


        existing_map[
            next_record.page_number
        ] = {

            "page": next_record.page_number,

            "products": parsed,

            "images": page_images
        }


        combined_pages = sorted(

            list(existing_map.values()),

            key=lambda x: x.get(
                "page",
                0
            )
        )


        # =====================================================
        # SAVE
        # =====================================================

        self.ai_response = json.dumps(
            combined_pages
        )


        self.last_ai_page = len(
            combined_pages
        )


        _logger.warning(

            f"[PDF AI] "

            f"PAGE SAVED "

            f"| PAGE "

            f"{next_record.page_number}"
        )


        # =====================================================
        # NEXT STATE
        # =====================================================

        if (

            self.last_ai_page

            <

            total_available_pages

        ):

            self.state = "pdf_ai"

            _logger.warning(

                f"[PDF AI] CONTINUE "

                f"{self.last_ai_page}/"

                f"{total_available_pages}"
            )

        else:

            _logger.warning(
                "[PDF AI] COMPLETE ✅"
            )

            self.state = "pdf_creating"


        self.flush_recordset()
        self.env.cr.commit()

        return
    

    #===========Excel Open AI================================
    def send_to_openai_excel(self):

        import json

        api_key = self.env[
            'ir.config_parameter'
        ].sudo().get_param(
            'openai.api.key'
        )

        if not api_key:

            raise Exception(
                "OpenAI API key not configured"
            )


        client = OpenAI(
            api_key=api_key
        )


        _logger.warning(
            "[EXCEL AI] START"
        )


        # =====================================================
        # LOAD EXTRACTED DATA
        # =====================================================

        try:

            pages = json.loads(
                self.extracted_text or "[]"
            )

        except Exception as e:

            _logger.error(

                f"[EXCEL AI] "

                f"INVALID extracted_text JSON "

                f"| {str(e)}"
            )

            return


        if not pages:

            _logger.error(
                "[EXCEL AI] NO ROWS TO PROCESS"
            )

            return


        _logger.warning(

            f"[EXCEL AI] "

            f"TOTAL ROWS={len(pages)}"
        )


        # =====================================================
        # BATCH
        # =====================================================

        BATCH_SIZE = 5

        start = (
            self.excel_ai_index or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(pages)
        )


        batch = pages[start:end]


        _logger.warning(

            f"[EXCEL AI BATCH] "

            f"{start} → {end}"
        )


        # =====================================================
        # LOAD EXISTING PRODUCTS
        # =====================================================

        existing_products = []


        if self.ai_response:

            try:

                existing_ai = json.loads(
                    self.ai_response
                )

                if (
                    isinstance(existing_ai, list)
                    and existing_ai
                ):

                    existing_products = (
                        existing_ai[0].get(
                            "products",
                            []
                        )
                    )


                _logger.warning(

                    f"[EXCEL AI] "

                    f"EXISTING PRODUCTS="

                    f"{len(existing_products)}"
                )

            except Exception as e:

                _logger.warning(

                    f"[EXCEL AI] "

                    f"FAILED LOAD EXISTING "

                    f"| {str(e)}"
                )

                existing_products = []


        new_products = []


        # =====================================================
        # PROCESS ROWS
        # =====================================================

        for idx, row in enumerate(

            batch,

            start=start

        ):

            try:

                row_text = row.get(
                    "text",
                    ""
                )

                row_price = row.get(
                    "price",
                    ""
                )

                row_stock = row.get(
                    "stock",
                    ""
                )

                images = row.get(
                    "images",
                    []
                )


                _logger.warning(

                    f"[EXCEL AI ROW] "

                    f"idx={idx} "

                    f"| images={len(images)}"
                )

                prompt = f"""
                You are a structured Excel product parser.

                Each input represents EXACTLY ONE ROW = ONE PRODUCT.

                =====================================
                COLUMN UNDERSTANDING (CRITICAL)
                =====================================

                The row could contain mixed values like:

                - ID (e.g. 94601, 12345)
                - Range (e.g. 2-66, 11-00)
                - Stock numbers
                - Prices
                - Links (http...)
                - Image references

                YOU MUST:

                1. IDENTIFY PRODUCT ID
                - Usually numeric (e.g. 94601)
                - Column name may vary:
                    - KOD
                    - SKU
                    - ID
                    - CODE

                2. IDENTIFY PRODUCT NAME
                - MUST NOT be:
                    - pure numbers
                    - ranges
                    - links
                    - dates
                    - headers

                - Product names should describe the ACTUAL product type.

               GOOD:
                - Sports Bottle
                - Metal Pen
                - Travel Mug
                - Drawstring Bag

                If the Excel already contains a valid product name:
                - preserve and use it

                If the Excel does NOT contain a real product name:
                - intelligently generate one using:
                    - Product <ID>
                    - category clues
                    - image appearance
                    - surrounding row data

                Fallback naming is allowed when necessary.

                GOOD fallback examples:
                - Product 94601
                - Bottle 94646
                - Pen 92070

                However:

                If rows belong to the SAME variant_group,
                you MUST still detect and extract the REAL variant difference.

                Example:

                Parent:
                Product 94646

                Variants:
                - White
                - Orange
                - Black

                DO NOT return:
                - Variant 1
                - Variant 2

                when a real difference can be visually or textually identified.
                
                =====================================
                VARIANT GROUPING (VERY IMPORTANT)
                =====================================

                - SAME ID = SAME variant_group
                - DIFFERENT ID = DIFFERENT PRODUCT
                - NEVER leave variant_group empty

                Rows sharing the same:
                - ID
                - grouped code
                - SKU group

                should be treated as variants of ONE parent product.

                 =====================================
                VARIANT DETECTION
                =====================================

                If rows share same PRODUCT ID:

                → they belong to the SAME product family.

                IMPORTANT:

                Use PRODUCT IMAGES as the PRIMARY
                source for identifying variants.

                Look for visual differences such as:

                - color
                - material
                - finish
                - lid type
                - texture
                - shape
                - capacity
                - packaging

                THEN use nearby codes/numbers
                as supporting evidence.

                Example:

                Rows may contain:

                106
                103
                128

                These MAY represent:
                - color codes
                - material codes
                - size codes

                DO NOT assume globally.

                Infer meaning from:
                - image differences
                - repeated patterns
                - product appearance

                If uncertain:

                Use safe fallback:

                "attributes": {{
                    "Vendor Code": "106"
                }}

                NEVER return:
                - Variant 1
                - Variant 2
                - Variant 3

                ALWAYS return meaningful attributes.

                =====================================
                VISUAL DIFFERENCE DETECTION
                =====================================

                If product images exist:

                You MUST visually inspect the images
                to identify the distinguishing feature.

                Example:

                If grouped products show:
                - white bottle
                - orange bottle
                - black bottle

                Return:

                {{
                    "name": "Sports Bottle",
                    "color": "White"
                }}

                {{
                    "name": "Sports Bottle",
                    "color": "Orange"
                }}

                DO NOT return:
                - Variant 1
                - Variant 2
                - Product 94601

                =====================================
                PARENT PRODUCT CONSISTENCY
                =====================================

                When multiple rows belong to the same
                variant_group:

                - The parent product name MUST remain consistent.
                - ONLY the variant fields should change.

                GOOD:

                Sports Bottle
                → White
                → Orange
                → Black

                BAD:

                White Bottle
                Orange Bottle
                Black Bottle

                =====================================
                ATTRIBUTE EXTRACTION
                =====================================

                Put distinguishing values into:

                - color
                - material
                - size
                - capacity
                - style

                Only use generic "Variant"
                if absolutely no real difference can be detected.

                =====================================
                PRICE & STOCK
                =====================================

                - Extract numeric price carefully
                - Extract stock carefully
                - Ignore ranges like:
                    - 2-66
                    - 11-00

                =====================================
                LINKS
                =====================================

                If a row contains a product URL:
                - preserve it
                - never use URL as product name

                =====================================
                OUTPUT FORMAT
                =====================================

                [
                    {{
                        "name": "",
                        "description": "",
                        "category": "",
                        "price": "",
                        "stock": "",
                        "variant_group": "",
                        "color": "",
                        "material": "",
                        "size": "",
                        "capacity": "",
                        "style": "",
                        "url": "",
                        "variants": [
                            {{
                                "attributes": {{
                                    "Variant": ""
                                }},
                                "image_index": 0,
                                "stock": null
                            }}
                        ]
                    }}
                ]

                =====================================
                IMPORTANT RULES
                =====================================

                - Return ONLY valid JSON
                - No markdown
                - No explanations
                - No comments
                - No trailing commas

                ROW TEXT:
                {row_text}

                DETECTED PRICE:
                {row_price}

                DETECTED STOCK:
                {row_stock}
                """

                response = client.responses.create(

                    model="gpt-4.1-mini",

                    input=prompt,

                    timeout=60
                )


                result = (
                    response.output_text or ""
                ).strip()


                result = result.replace(
                    "```json",
                    ""
                )

                result = result.replace(
                    "```",
                    ""
                ).strip()


                if not result:

                    raise Exception(
                        "EMPTY AI RESPONSE"
                    )


                parsed = json.loads(
                    result
                )


                if (

                    isinstance(parsed, list)

                    and parsed

                ):

                    parsed = parsed[0]


                if not isinstance(
                    parsed,
                    dict
                ):

                    _logger.warning(

                        f"[EXCEL AI] "

                        f"INVALID STRUCTURE "

                        f"| idx={idx}"
                    )

                    continue


                # =================================================
                # IMAGE
                # =================================================

                if images:

                    parsed["image"] = (
                        images[0]
                    )


                # =================================================
                # DEBUG
                # =================================================

                _logger.warning(

                    f"[EXCEL AI PRODUCT] "

                    f"name={parsed.get('name')} "

                    f"| group={parsed.get('variant_group')}"
                )


                new_products.append(
                    parsed
                )


            except Exception as e:

                _logger.exception(

                    f"[EXCEL AI ERROR] "

                    f"idx={idx} "

                    f"| {str(e)}"
                )


        # =====================================================
        # MERGE PRODUCTS SAFELY
        # =====================================================

        combined_products = (
            existing_products
            +
            new_products
        )


        _logger.warning(

            f"[EXCEL AI MERGE] "

            f"existing={len(existing_products)} "

            f"| new={len(new_products)} "

            f"| total={len(combined_products)}"
        )


        # =====================================================
        # SAVE
        # =====================================================

        self.ai_response = json.dumps([{

            "page": 1,

            "products": combined_products

        }])


        self.excel_ai_index = end


        _logger.warning(

            f"[EXCEL AI SAVE] "

            f"{self.excel_ai_index}/"

            f"{len(pages)}"
        )


        # =====================================================
        # NEXT STATE
        # =====================================================

        if end < len(pages):

            self.state = "excel_ai"

        else:

            _logger.warning(
                "[EXCEL AI COMPLETE]"
            )

            self.state = (
                "excel_creating"
            )


        self.flush_recordset()

        self.env.cr.commit()

        return
    

    #-----------scoring image before picking best/quality image (inage logic)-------------
    def pick_best_image(self, images):

                best_img = None
                best_score = 0
                seen_hashes = set()

                for img in images:

                    try:
                        img_bytes = base64.b64decode(img)
                        size = len(img_bytes)

                        # ❌ Skip tiny images (logos/icons)
                        if size < 10000:
                            continue

                        # ❌ Skip extremely large (full lifestyle pages)
                        if size > 800000:
                            continue

                        # ================= IMAGE ANALYSIS =================
                        pil_img = Image.open(BytesIO(img_bytes))
                        width, height = pil_img.size

                        # ❌ Skip extremely small resolution
                        if width < 150 or height < 150:
                            continue

                        # ❌ Skip extreme aspect ratios (banners, strips)
                        aspect_ratio = width / height if height else 1

                        if aspect_ratio > 3 or aspect_ratio < 0.3:
                            continue

                        # ================= DUPLICATE CHECK =================
                        img_hash = hash(img_bytes[:100])  # fast partial hash

                        if img_hash in seen_hashes:
                            continue

                        seen_hashes.add(img_hash)

                        # ================= SCORING =================
                        score = 0

                        # ✅ Prefer medium-good sizes
                        if 20000 < size < 300000:
                            score += 200

                        # ✅ Prefer square-ish product images
                        if 0.7 < aspect_ratio < 1.5:
                            score += 150

                        # ✅ Prefer decent resolution
                        if width > 400 and height > 400:
                            score += 150

                        # ❌ Penalize too wide/tall
                        if aspect_ratio > 2 or aspect_ratio < 0.5:
                            score -= 100

                        # ❌ Penalize very large images (likely lifestyle)
                        if size > 500000:
                            score -= 150

                        # Base size contribution
                        score += size / 2000

                        # ================= SELECT BEST =================
                        if score > best_score:
                            best_score = score
                            best_img = img

                    except Exception:
                        continue

                return best_img

    #=================Centralized Rusable Image=======================
    def _prepare_asset_pool(self, images):

        prepared = []

        seen = {}

        for asset in (images or []):

            try:

                if not asset:
                    continue

                # =====================================
                # SUPPORT OLD + NEW FORMAT
                # =====================================

                if isinstance(asset, dict):

                    img = asset.get("image")

                    score = asset.get(
                        "score",
                        0
                    )

                    is_collage = asset.get(
                        "is_collage",
                        False
                    )

                else:

                    img = asset

                    score = 0

                    is_collage = False

                if not img:

                    _logger.warning(
                        "[ASSET SKIPPED] EMPTY IMAGE"
                    )

                    continue

                image_hash = hashlib.md5(

                    img.encode('utf-8')

                ).hexdigest()

                # =====================================
                # SAFE COLOR DETECTION
                # =====================================

                dominant_color = ""

                try:

                    dominant_color = (

                        self._detect_dominant_color(
                            img
                        ) or ""
                    )

                except Exception as color_error:

                    _logger.warning(

                        f"[COLOR DETECT FAILED] "

                        f"{str(color_error)}"
                    )

                # =====================================
                # ONLY REMOVE TRUE DUPLICATES
                # =====================================

                existing_asset = seen.get(
                    image_hash
                )

                # SAME HASH + SAME SCORE + SAME COLLAGE
                # = real duplicate only

                if existing_asset:

                    if (

                        existing_asset.get("score") == score

                        and

                        existing_asset.get("is_collage") == is_collage

                    ):

                        _logger.warning(

                            f"[ASSET SKIPPED] TRUE DUPLICATE"
                        )

                        continue

                _logger.warning(

                    f"[ASSET DEBUG] "

                    f"type={type(asset)} "

                    f"score={score} "

                    f"collage={is_collage} "

                    f"color={dominant_color}"
                )

                prepared.append({

                    "image": img,

                    "score": score,

                    "is_collage": is_collage,

                    "dominant_color":
                        dominant_color
                })

                seen[image_hash] = {

                    "score": score,

                    "is_collage": is_collage,

                    "dominant_color": dominant_color
                }

                _logger.warning(

                    f"[ASSET ADDED] "

                    f"score={score} "

                    f"collage={is_collage} "

                    f"color={dominant_color}"
                )

            except Exception as e:

                _logger.warning(

                    f"[ASSET POOL ERROR] "

                    f"{str(e)}"
                )

        # =====================================
        # SORT BEST FIRST
        # =====================================

        prepared = sorted(

            prepared,

            key=lambda x: (

                x.get("score", 0),

                not x.get(
                    "is_collage",
                    False
                )
            ),

            reverse=True
        )

        # =====================================
        # REBUILD INDEXES AFTER SORT
        # =====================================

        for idx, asset in enumerate(prepared):

            asset["index"] = idx

        _logger.warning(

            f"[ASSET POOL READY] "

            f"{len(prepared)} assets"
        )

        return prepared

    # =====================================
    # ADVANCED DOMINANT COLOR DETECTION
    # =====================================

    def _detect_dominant_color(

        self,

        image_base64
    ):

        try:

            import base64
            import colorsys
            import numpy as np

            from io import BytesIO
            from PIL import Image

            image_data = base64.b64decode(
                image_base64
            )

            image = Image.open(

                BytesIO(image_data)

            ).convert("RGB")

            image = image.resize((120, 120))

            pixels = np.array(image)

            # =====================================
            # REMOVE VERY BRIGHT BACKGROUND
            # =====================================

            pixels = pixels.reshape(-1, 3)

            filtered_pixels = []

            for r, g, b in pixels:

                # remove white bg
                if r > 235 and g > 235 and b > 235:
                    continue

                filtered_pixels.append([r, g, b])

            if not filtered_pixels:
                return "white"

            pixels = np.array(filtered_pixels)

            avg = pixels.mean(axis=0)

            r, g, b = avg

            # =====================================
            # RGB → HSV
            # =====================================

            h, s, v = colorsys.rgb_to_hsv(

                r / 255.0,
                g / 255.0,
                b / 255.0
            )

            h = h * 360
            s = s * 100
            v = v * 100

            # =====================================
            # BLACK
            # =====================================

            if v < 18:
                return "black"

            # =====================================
            # WHITE
            # =====================================

            if v > 92 and s < 10:
                return "white"

            # =====================================
            # GRAY / GREY
            # =====================================

            if s < 15:

                if v < 55:
                    return "gray"

                return "grey"

            # =====================================
            # RED
            # =====================================

            if h < 15 or h >= 345:
                return "red"

            # =====================================
            # ORANGE
            # =====================================

            if 15 <= h < 40:
                return "orange"

            # =====================================
            # YELLOW
            # =====================================

            if 40 <= h < 70:
                return "yellow"

            # =====================================
            # GREEN
            # =====================================

            if 70 <= h < 170:
                return "green"

            # =====================================
            # BLUE
            # =====================================

            if 170 <= h < 260:

                if v < 45:
                    return "navy"

                if s < 35:
                    return "light blue"

                return "blue"

            # =====================================
            # PURPLE
            # =====================================

            if 260 <= h < 320:
                return "purple"

            # =====================================
            # PINK
            # =====================================

            if 320 <= h < 345:
                return "pink"

            return "unknown"

        except Exception as e:

            _logger.warning(

                f"[DOMINANT COLOR FAILED] "

                f"{str(e)}"
            )

            return "unknown"

    # =====================================
    # PROFESSIONAL VARIANT IMAGE MATCHER
    # =====================================
  
    def _match_variant_image(
         self,
        variant,
        asset_pool,
        used_asset_indexes=None
    ):

        try:
            if used_asset_indexes is None:

                used_asset_indexes = set()

            if not asset_pool:
                return False

            best_asset = None

            best_score = -999

            if used_asset_indexes is None:

                used_asset_indexes = set()
            variant_text = ""

            attributes = variant.get(
                "attributes",
                {}
            )

            if isinstance(attributes, dict):

                variant_text = " ".join([

                    str(v)

                    for v in attributes.values()

                ]).lower()

                real_assets = [

                    a for a in asset_pool

                    if not a.get(
                        "is_lifestyle",
                        False
                    )
                ]

                _logger.warning(

                    f"[VARIANT POOLS] "

                    f"real={len(real_assets)} "

                    f"lifestyle={len(lifestyle_assets)}"
                )

                lifestyle_assets = [

                    a for a in asset_pool

                    if a.get(
                        "is_lifestyle",
                        False
                    )
                ]

            # =====================================
            # SCORE ASSETS
            # =====================================

            for asset in asset_pool:
                if asset.get("index") in used_asset_indexes:
                    continue

                asset_index = asset.get(
                    "index"
                )

                asset_score = asset.get(
                    "score",
                    0
                )

                dominant_color = asset.get(
                    "dominant_color",
                    ""
                )

                if dominant_color == "unknown":

                    asset_score -= 45

                # ---------------------------------
                # COLLAGE PENALTY
                # ---------------------------------

                if asset.get("is_collage"):

                    asset_score -= 80

                # ---------------------------------
                # COLOR MATCHING
                # ---------------------------------

                color_map = [

                    "red",
                    "blue",
                    "green",
                    "lime",
                    "yellow",
                    "orange",
                    "white",
                    "black",
                    "gray",
                    "grey",
                    "purple",
                    "pink",
                    "brown"
                ]


                for color in color_map:

                    if color not in variant_text:
                        continue

                    # exact match
                    if color == dominant_color:

                        asset_score += 180

                    # gray/grey normalization
                    elif (

                        color in [

                            "blue",
                            "royal blue",
                            "navy",
                            "light blue"
                        ]

                        and

                        dominant_color in [

                            "blue",
                            "navy",
                            "light blue"
                        ]

                    ):

                        asset_score += 180

                    # strong mismatch penalty

                    elif (

                        color in [

                            "blue",
                            "royal blue",
                            "light blue"
                        ]

                        and

                        dominant_color in [

                            "green",
                            "lime",
                            "yellow",
                            "orange"
                        ]

                    ):

                        asset_score -= 220

                    # white/silver/light handling
                    elif (

                        color == "white"

                        and

                        dominant_color in [
                            "white",
                            "gray"
                        ]
                    ):

                        asset_score += 90

                    # dark product approximation
                    elif (

                        color == "black"

                        and

                        dominant_color in [
                            "black",
                            "gray"
                        ]
                    ):

                        asset_score += 90

                # ---------------------------------
                # HERO BONUS
                # ---------------------------------

                if asset.get("score", 0) >= 70:

                    asset_score += 10

                # ---------------------------------
                # BEST MATCH
                # ---------------------------------

                if asset_score > best_score:

                    best_score = asset_score

                    best_asset = asset

            # =====================================
            # SAFE FALLBACK
            # =====================================

            if not best_asset:

                best_asset = sorted(

                    asset_pool,

                    key=lambda x: x.get(
                        "score",
                        0
                    ),

                    reverse=True

                )[0]

            if best_asset:

                used_asset_indexes.add(

                    best_asset.get("index")
                )

            return best_asset

        except Exception as e:

            _logger.warning(

                f"[VARIANT MATCH FAILED] "

                f"{str(e)}"
            )

            return False


    #======score_segmented_image ==========================
    def _score_segmented_image(

        self,

        image_base64
    ):

        try:

            import base64
            import io

            import numpy as np

            from PIL import Image

            image_bytes = base64.b64decode(
                image_base64
            )

            img = Image.open(

                io.BytesIO(image_bytes)

            ).convert("RGB")

            width, height = img.size

            # ==========================================
            # REJECT VERY SMALL CROPS
            # ==========================================

            if width < 120 or height < 120:

                return -999

            np_img = np.array(img)

            score = 0

            # ==========================================
            # LARGE IMAGE BONUS
            # ==========================================

            score += (
                width * height
            ) / 10000

            # ==========================================
            # TEXT HEAVY PENALTY
            # ==========================================

            dark_ratio = np.mean(
                np_img < 70
            )

            score -= dark_ratio * 200

            # ==========================================
            # GOOD PRODUCT ASPECT BONUS
            # ==========================================

            aspect = width / float(height)

            if 0.7 <= aspect <= 1.5:

                score += 50

            # ==========================================
            # CLEAN BACKGROUND BONUS
            # ==========================================

            white_ratio = np.mean(
                np_img > 230
            )

            score += white_ratio * 80

            return score

        except Exception as e:

            _logger.warning(

                f"[IMAGE SCORE ERROR] "

                f"{str(e)}"
            )

            return 0    
        
    #=============variant color enhancement 1=================
    def _get_dominant_color_name(

        self,

        image_base64
    ):

        try:

            import base64
            import io
            import numpy as np

            from PIL import Image

            image_bytes = base64.b64decode(
                image_base64
            )

            img = Image.open(

                io.BytesIO(image_bytes)

            ).convert("RGB")

            img = img.resize((80, 80))

            np_img = np.array(img)

            pixels = np_img.reshape(
                (-1, 3)
            )

            avg = pixels.mean(axis=0)

            r, g, b = avg

            # =====================================
            # COLOR CLASSIFICATION
            # =====================================

            if r > 200 and g > 200 and b > 200:
                return "white"

            if r < 60 and g < 60 and b < 60:
                return "black"

            if r > 160 and g < 120 and b < 120:
                return "red"

            if r > 180 and g > 180 and b < 120:
                return "yellow"

            if b > r and b > g:
                return "blue"

            if g > r and g > b:
                return "green"

            if r > 120 and b > 120:
                return "purple"

            if r > 150 and g > 120 and b < 100:
                return "orange"

            if (
                abs(r - g) < 20
                and
                abs(g - b) < 20
            ):
                return "grey"

            return "unknown"

        except Exception as e:

            _logger.warning(

                f"[COLOR DETECTION ERROR] "

                f"{str(e)}"
            )

            return "unknown"

    #=================Centralized Rusable Image resolver==============

    def _resolve_asset_image(
        self,
        asset_pool,
        index
    ):

        try:

            if index is None:
                return False

            for asset in asset_pool:

                if isinstance(asset, dict):

                    if asset.get("index") == index:

                        return asset.get("image")

                elif isinstance(asset, str):

                    return asset

            return False

        except Exception as e:

            _logger.warning(

                f"[RESOLVE ASSET ERROR] "

                f"{str(e)}"
            )

            return False

    #============marchin AI===================================================
    # LEGACY IMAGE PAYLOAD MATCHER
    # Deprecated after migration to
    # index-based asset orchestration.
    # Keep temporarily for rollback safety.
    # ========================================================================
    def match_image_with_ai(self, product_name, images):

        api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')
        client = OpenAI(api_key=api_key)

        if not images:
            return None

        # limit images for performance
        images = images[:5]
        image_inputs = [
        {
                "type": "input_image",
                "image_url": f"data:image/jpeg;base64,{img}"
            }
            for img in images
        ]

        prompt = f"""
        You are an expert product image matcher.

        Select the image that BEST represents this product:

        PRODUCT:
        {product_name}

        RULES:
        - Return ONLY the index (0-based integer)
        - No explanation
        - No text

            PRIORITY:
        1. Prefer isolated product on plain/white background
        2. Prefer centered single-product image
        3. Prefer image showing full product clearly
        4. Prefer clean studio product photos
        5. Prefer catalog hero product image
        6. Avoid lifestyle scenes if isolated image exists
        7. Avoid collages whenever possible
        8. Avoid infographic layouts
        9. Avoid text-heavy graphics
        10. Avoid multi-product overview images
        11. Avoid images containing large text blocks

        DO NOT PICK:
        - logos
        - icons
        - banners
        - cropped fragments
        - specification charts
        - text-heavy graphics
        - tiny thumbnails
        """

        try:
            response = client.responses.create(
                model="gpt-4.1",
                input=[{
                    "role": "user",
                    "content": [{"type": "input_text", "text": prompt}] + image_inputs
                }],
                timeout=30
            )

            result = response.output_text.strip()

            index = int(result)

            if 0 <= index < len(images):
                return images[index]

        except Exception as e:
            _logger.warning(f"AI IMAGE MATCH FAILED → {str(e)}")

        return None
    
    #======== returning images indexes========================================
    def match_image_index_with_ai( self, product_name, images):

        api_key = self.env[
            'ir.config_parameter'
        ].sudo().get_param(
            'openai.api.key'
        )

        client = OpenAI(api_key=api_key)

        if not images:
            return None


        filtered_images = []

        for img in images:

            try:

                if not img:
                    continue

                img_lower = img.lower()

                bad_keywords = [

                    "banner",

                    "lifestyle",

                    "infographic",

                    "specification",

                    "sizechart",

                    "dimensions"
                ]

                if any(
                    k in img_lower
                    for k in bad_keywords
                ):
                    continue

                filtered_images.append(img)

            except Exception:
                continue

        if filtered_images:

            images = filtered_images

        images = images[:8]

        image_inputs = []

        for idx, img in enumerate(images):

            image_inputs.append({
                "type": "input_text",
                "text": f"IMAGE INDEX: {idx}"
            })

            image_inputs.append({
                "type": "input_image",
                "image_url":
                    f"data:image/jpeg;base64,{img}"
            })

        prompt = f"""
        You are an ecommerce
        product image selector.

        PRODUCT:
        {product_name}

        Return ONLY the BEST
        image index.

        PRIORITY:
        - isolated product
        - plain background
        - centered object
        - clean catalog render

        AVOID:
        - people
        - lifestyle scenes
        - infographics
        - collages
        - banners
        - text-heavy graphics

        Return ONLY integer index.
        """

        try:

            response = client.responses.create(

                model="gpt-4.1",

                input=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": prompt
                        }
                    ] + image_inputs
                }],

                timeout=30
            )

            result = (
                response.output_text or ""
            ).strip()

            index = int(result)

            if 0 <= index < len(images):

                return index

        except Exception as e:

            _logger.warning(

                f"[IMAGE INDEX MATCH FAILED] "

                f"{str(e)}"
            )

        return None

    #============enforce translation=========================================
    def _force_translate(self, text, target_lang):

        from openai import OpenAI

        if not text:
            return text

        try:

            api_key = self.env[
                'ir.config_parameter'
            ].sudo().get_param(
                'openai.api.key'
            )

            if not api_key:

                _logger.warning(
                    "[OPENAI TRANSLATE] MISSING API KEY"
                )

                return text


            client = OpenAI(
                api_key=api_key
            )


            prompt = f"""
            Translate the following text into {target_lang}.

            Rules:
            - Return ONLY the translated text
            - Preserve formatting
            - Preserve product terminology
            - Do not explain anything

            TEXT:
            {text}
            """


            response = client.responses.create(

                model="gpt-4.1-mini",

                input=prompt
            )


            translated = (
                response.output_text or ''
            ).strip()


            if not translated:

                _logger.warning(
                    "[OPENAI TRANSLATE EMPTY]"
                )

                return text


            _logger.warning(

                f"[OPENAI TRANSLATION SUCCESS] "

                f"lang={target_lang}"
            )

            return translated


        except Exception as e:

            _logger.warning(

                f"[OPENAI TRANSLATE FAILED] "

                f"{str(e)}"
            )

            return text

    #=========Translation new logic==========================================

    def _apply_product_translation(self, product):

        if not product:
            return

        name = product.name or ''
        desc = product.description_sale or ''

        # ----------------------------
        # DEBUG
        # ----------------------------
        _logger.warning(
            f"[TRANSLATION INPUT] product={product.id} | name={name} | desc_len={len(desc)}"
        )

        # ----------------------------
        # ALWAYS TRANSLATE NAME (cheap)
        # ----------------------------
        ru_name = self._force_translate(name, "ru")
        az_name = self._force_translate(name, "az")

        # ----------------------------
        # ONLY TRANSLATE DESCRIPTION IF EXISTS
        # ----------------------------
        if desc and len(desc.strip()) > 10:

            ru_desc = self._smart_translate(desc, "ru")
            az_desc = self._smart_translate(desc, "az")

        else:
            ru_desc = ''
            az_desc = ''

            _logger.warning(
                f"[TRANSLATION SKIPPED DESC] product={product.id}"
            )

        # ----------------------------
        # SAVE
        # ----------------------------
        product.with_context(lang='ru_RU').write({
            'name': ru_name,
            'description_sale': ru_desc
        })

        product.with_context(lang='az_AZ').write({
            'name': az_name,
            'description_sale': az_desc
        })


        # =========================================
        # DETECT REAL TRANSLATION
        # =========================================

        translation_changed = False

        try:

            # =====================================
            # SAFE LANGUAGE
            # =====================================

            lang_code = 'ru_RU'

            translated_product = product.with_context(
                lang=lang_code
            )

            translated_name = translated_product.name or ''

            original_name = product.name or ''

            translated_desc = (
                translated_product.description_sale or ''
            )

            original_desc = (
                product.description_sale or ''
            )

            # =====================================
            # NAME CHANGED
            # =====================================

            if translated_name != original_name:

                translation_changed = True

            # =====================================
            # DESCRIPTION CHANGED
            # =====================================

            if translated_desc != original_desc:

                translation_changed = True

        except Exception as e:

            _logger.warning(

                f"[TRANSLATION CHECK FAILED] "

                f"{str(e)}"
            )


        if translation_changed:


            # =========================================
            # SHOW REAL TRANSLATED VALUES
            # =========================================

            try:

                ru_name = product.with_context(
                    lang='ru_RU'
                ).name or ''

                az_name = product.with_context(
                    lang='az_AZ'
                ).name or ''

            except Exception:

                ru_name = ''
                az_name = ''


            _logger.warning(

                f"[TRANSLATION SUCCESS] "

                f"product={product.id} | "

                f"RU={ru_name[:120]} | "

                f"AZ={az_name[:120]}"
            )

        else:

            _logger.warning(

                f"[TRANSLATION NO-CHANGE] "

                f"product={product.id}"
            )


    #============product translation extended================
    def _smart_translate(self, text, lang):

        # fallback if needed
        if len(text) < 5:
            return self._force_translate(text, lang)

        try:
            api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')
            client = OpenAI(api_key=api_key)

            prompt = f"Translate to {lang} and improve clarity:\n{text}"

            response = client.responses.create(
                model="gpt-4.1-mini",
                input=prompt
            )

            return response.output_text.strip()

        except:
            return self._force_translate(text, lang)


    # ================= PRODUCT CREATION URL ================

    def create_products_url(self):

        import requests
        import base64
        import json

        if not self.ai_response:
            _logger.warning("NO AI RESPONSE")
            return

        product_obj = self.env['product.template']
        category_obj = self.env['product.category']

        try:
            products = json.loads(self.ai_response)

            if isinstance(products, dict):
                products = [products]

        except Exception as e:
            _logger.error(f"INVALID AI JSON → {str(e)}")
            return

        TOTAL_PRODUCTS = len(products)
        start_index = self.last_processed_product_index or 0

        _logger.warning(f"TOTAL AI PRODUCTS → {TOTAL_PRODUCTS}")
        _logger.warning(f"START INDEX → {start_index}")

        created_count = 0
        skipped_count = 0

        MAX_PRODUCTS_PER_RUN = 5

        CATEGORY_MAPPING = {
            "t-shirt": "Apparel",
            "shirt": "Apparel",
            "polo": "Apparel",
            "bag": "Bags",
            "backpack": "Bags",
            "cap": "Headwear",
            "hat": "Headwear",
            "bottle": "Drinkware",
            "cup": "Drinkware",
            "drinkware": "Drinkware",
            "pen": "Stationery",
            "notebook": "Stationery",
            "powerbank": "Electronics",
            "charger": "Electronics",
            "laptop": "Electronics",
            "football": "Football Fever",
            "Wristband": "Football Fever",
            "sports t-shirt": "Football Fever",
            "sports towel": "Football Fever",
            "Sports Bottles": "Football Fever"
        }

        parent_category = category_obj.search([('name', '=', "All Products")], limit=1)
        vendor_id = self.partner_id.id if self.partner_id else False

        if not parent_category:
            parent_category = category_obj.create({'name': "All Products"})

        end_index = min(start_index + MAX_PRODUCTS_PER_RUN, TOTAL_PRODUCTS)

        _logger.warning(f"PROCESSING RANGE → {start_index} to {end_index}")

        for idx in range(start_index, end_index):

            product_data = products[idx]

            name = product_data.get("name")
            if not name:
                skipped_count += 1
                continue

            description = product_data.get("description", "")
            raw_category = (product_data.get("category") or "").lower()

            # ================= CATEGORY =================
            mapped_category = "General"
            for key, val in CATEGORY_MAPPING.items():
                if key in raw_category:
                    mapped_category = val
                    break

            category = category_obj.search([
                ('name', '=ilike', mapped_category),
                ('parent_id', '=', parent_category.id)
            ], limit=1)

            if not category:
                category = category_obj.create({
                    'name': mapped_category,
                    'parent_id': parent_category.id
                })

            # ================= FINGERPRINT=================

            variant_group = (

                product_data.get("variant_group")

                or

                name
            )

            variant_group = str(
                variant_group
            ).strip().upper()


            vendor_fingerprint = (
                f"{vendor_id}_{variant_group}"
            )


            # ================= DUPLICATE CHECK =================

            existing = product_obj.search([

                (
                    'vendor_fingerprint',
                    '=',
                    vendor_fingerprint
                )

            ], limit=1)


            if existing:

                _logger.warning(

                    f"[URL DUPLICATE SKIP] "

                    f"{vendor_fingerprint}"
                )

                skipped_count += 1

                continue

            # =========================================
            # FORMAT DESCRIPTION
            # =========================================

            formatted_description = description or ""

            formatted_description = formatted_description.replace(
                "STOCK NEXT ARRIVAL IN PRODUCTION",
                "</p><h4>Stock Information</h4><p>"
            )

            formatted_description = formatted_description.replace(
                "PRODUCT CUSTOMISATION",
                "</p><h4>Customisation Options</h4><p>"
            )

            formatted_description = formatted_description.replace(
                "POSSIBILITIES FOR PRODUCT CUSTOMISATION",
                "</p><h4>Customisation Options</h4><p>"
            )

            formatted_description = formatted_description.replace(
                "Dimensions:",
                "</p><h4>Dimensions</h4><p>Dimensions:"
            )

            formatted_description = (
                f"<p>{formatted_description}</p>"
            )

            vals = {
                'name': name.strip(),
                'description_sale': formatted_description,
                'type': 'consu',
                'categ_id': category.id,
                'sale_ok': True,
                'website_published': False,
                'vendor_id': vendor_id,
                'vendor_fingerprint': vendor_fingerprint,
                'vendor_import_job_id': self.id,
            }

            # ================= IMAGE =================
            image_url = product_data.get("image")

            if image_url and isinstance(image_url, str) and image_url.startswith("http"):

                try:
                    _logger.warning(f"FETCHING IMAGE → {image_url}")

                    res = requests.get(image_url, timeout=5, stream=True)

                    if res.status_code != 200:
                        _logger.warning(f"IMAGE HTTP ERROR → {res.status_code}")
                        continue

                    content_type = res.headers.get("Content-Type", "")
                    if "image" not in content_type:
                        _logger.warning(f"NOT AN IMAGE → {content_type}")
                        continue

                    content = res.raw.read(500000, decode_content=True)

                    if not content:
                        _logger.warning("EMPTY IMAGE CONTENT")
                        continue

                    vals['image_1920'] = base64.b64encode(content).decode("utf-8")

                    _logger.warning("IMAGE STORED SUCCESSFULLY")

                except Exception as e:
                    _logger.warning(f"IMAGE FAILED → {str(e)}")

            else:
                _logger.warning(f"NO VALID IMAGE URL → {image_url}")

            # ================= CREATE =================
            try:
                _logger.warning(
                    f"[URL CREATE PRODUCT] → "
                    f"NAME={name} | "
                    f"VENDOR_ID={vendor_id}"
                )

                # product_obj.create(vals)
                product = product_obj.with_context(
                    mail_create_nolog=True,
                    mail_notify_force_send=False,
                    tracking_disable=True
                ).create(vals)

                self._apply_product_translation(product)

                # =========================================
                # URL VARIANTS
                # =========================================

                variants = product_data.get(
                    "variants",
                    []
                )


                # =========================================
                # FALLBACK VARIANT
                # =========================================

                if not variants:

                    variants = [{

                        "attributes": {

                            "Variant": name
                        }

                    }]


                # =========================================
                # PROCESS VARIANTS
                # =========================================

                for variant in variants:

                    attributes = variant.get(
                        "attributes",
                        {}
                    )


                    for attr_name, attr_value in attributes.items():


                        if not attr_value:
                            continue


                        attr_value = str(attr_value).strip()


                        # =====================================
                        # NORMALIZE BAD VARIANTS
                        # =====================================

                        bad_variants = [

                            'variant 1',

                            'variant 2',

                            'variant 3',

                            'default',

                            'option a',

                            'option b'
                        ]


                        if attr_value.lower() in bad_variants:

                            attr_name = "Design"

                            attr_value = name


                        # =====================================
                        # ATTRIBUTE
                        # =====================================

                        attribute = self.env[
                            'product.attribute'
                        ].search([

                            ('name', '=', attr_name)

                        ], limit=1)

                        # =====================================
                        # TRANSLATE ATTRIBUTE NAME
                        # =====================================

                        try:

                            for lang_code in [

                                'ru_RU',

                                'az_AZ'
                            ]:

                                translated_attr = self._force_translate(

                                    str(attr_name),

                                    lang_code
                                )


                                if translated_attr:

                                    attribute.with_context(
                                        lang=lang_code
                                    ).write({

                                        'name': translated_attr
                                    })


                                    _logger.warning(

                                        f"[URL ATTRIBUTE TRANSLATED] "

                                        f"{attr_name} "

                                        f"-> "

                                        f"{translated_attr} "

                                        f"({lang_code})"
                                    )

                        except Exception as e:

                            _logger.warning(

                                f"[URL ATTRIBUTE TRANSLATION ERROR] "

                                f"{str(e)}"
                            )


                        if not attribute:

                            attribute = self.env[
                                'product.attribute'
                            ].create({

                                'name': attr_name
                            })


                        # =====================================
                        # ATTRIBUTE VALUE
                        # =====================================

                        value = self.env[
                            'product.attribute.value'
                        ].search([

                            ('name', '=', attr_value),

                            (
                                'attribute_id',
                                '=',
                                attribute.id
                            )

                        ], limit=1)


                        if not value:

                            value = self.env[
                                'product.attribute.value'
                            ].create({

                                'name': attr_value,

                                'attribute_id':
                                    attribute.id
                            })


                            # =================================
                            # TRANSLATE VARIANT VALUE
                            # =================================

                            try:

                                for lang_code in [

                                    'ru_RU',

                                    'az_AZ'
                                ]:

                                    translated_variant = (

                                        self._force_translate(

                                            str(attr_value),

                                            lang_code
                                        )
                                    )


                                    if translated_variant:

                                        value.with_context(
                                            lang=lang_code
                                        ).write({

                                            'name':
                                                translated_variant
                                        })


                                        _logger.warning(

                                            f"[URL VARIANT TRANSLATED] "

                                            f"{attr_value} "

                                            f"-> "

                                            f"{translated_variant} "

                                            f"({lang_code})"
                                        )

                            except Exception as e:

                                _logger.warning(

                                    f"[URL VARIANT TRANSLATION ERROR] "

                                    f"{str(e)}"
                                )


                        # =====================================
                        # ATTRIBUTE LINE
                        # =====================================

                        line = self.env[
                            'product.template.attribute.line'
                        ].search([

                            (
                                'product_tmpl_id',
                                '=',
                                product.id
                            ),

                            (
                                'attribute_id',
                                '=',
                                attribute.id
                            )

                        ], limit=1)


                        if not line:

                            self.env[
                                'product.template.attribute.line'
                            ].create({

                                'product_tmpl_id':
                                    product.id,

                                'attribute_id':
                                    attribute.id,

                                'value_ids': [(6, 0, [

                                    value.id

                                ])]
                            })

                        else:

                            if (

                                value.id

                                not in

                                line.value_ids.ids

                            ):

                                line.value_ids = [

                                    (4, value.id)

                                ]
               
                created_count += 1

            except Exception as e:
                _logger.error(f"CREATE FAILED → {name} | {str(e)}")
                skipped_count += 1
                continue

            if created_count % 10 == 0:
                self._safe_commit_progress()

        # ================= SAVE PROGRESS =================
        self.last_processed_product_index = end_index

        _logger.warning(f"CREATED THIS RUN → {created_count}")
        _logger.warning(f"SKIPPED THIS RUN → {skipped_count}")
        _logger.warning(f"NEXT START INDEX → {self.last_processed_product_index}")

        if self.last_processed_product_index >= TOTAL_PRODUCTS:
            _logger.warning("ALL PRODUCTS CREATED ✅")
            self.state = "done"
        else:
            _logger.warning("MORE PRODUCTS REMAIN → CONTINUE CREATION")
            self.state = "url_creating"

        self._safe_commit_progress()


    #==========create pdf product====================================
  
    def create_products_pdf(self):

        import json

        _logger.warning(
            "[PDF CREATE] START"
        )

        if not self.ai_response:

            _logger.warning(
                "[PDF CREATE] NO AI RESPONSE"
            )

            return

        try:

            ai_pages = json.loads(
                self.ai_response
            )

        except Exception as e:

            _logger.error(

                f"[PDF CREATE] INVALID AI JSON "

                f"| {str(e)}"
            )

            return

        if not isinstance(ai_pages, list):

            _logger.warning(
                "[PDF CREATE] INVALID AI FORMAT"
            )

            return

        product_obj = self.env[
            'product.template'
        ]

        category_obj = self.env[
            'product.category'
        ]

        stock_quant_obj = self.env[
            'stock.quant'
        ]

        stock_location = self.env[
            'stock.location'
        ].search([

            ('usage', '=', 'internal')

        ], limit=1)

        CATEGORY_MAPPING = {

            "t-shirt": "Apparel",
            "shirt": "Apparel",
            "polo": "Apparel",

            "bag": "Bags",
            "backpack": "Bags",

            "cap": "Headwear",
            "hat": "Headwear",

            "bottle": "Drinkware",
            "drinkware": "Drinkware",

            "pen": "Stationery",
            "notebook": "Stationery",

            "powerbank": "Electronics",
            "charger": "Electronics",
            "laptop": "Electronics",
        }

        parent_category = category_obj.search([

            ('name', '=', "All Products")

        ], limit=1)

        if not parent_category:

            parent_category = category_obj.create({

                'name': "All Products"

            })

        vendor_id = (

            self.partner_id.id

            if self.partner_id

            else False
        )

        BATCH_SIZE = 3

        start = (
            self.last_created_page or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(ai_pages)
        )

        created_count = 0
        skipped_count = 0

        for page_index in range(start, end):

            try:

                page_data = ai_pages[
                    page_index
                ]

            except Exception as e:

                _logger.warning(

                    f"[PDF PAGE LOAD ERROR] "

                    f"{str(e)}"
                )

                continue

            page_number = page_data.get(
                "page"
            )

            page_record = self.env[
                'vendor.import.page'
            ].search([

                ('job_id', '=', self.id),

                ('page_number', '=', page_number)

            ], limit=1)


            # =====================================
            # LOAD AI-PERSISTED IMAGES
            # =====================================

            page_images = page_data.get(
                "images",
                []
            )

            if not page_images:

                _logger.warning(

                    f"[PDF CREATE] "

                    f"NO IMAGES FOUND "

                    f"| PAGE {page_number}"
                )


            products = page_data.get(
                "products",
                []
            )

            for product_data in products:

                # =====================================
                # PRODUCT IMAGE PREP
                # =====================================

                product_images = product_data.get(
                    "images",
                    []
                )

                _logger.warning(

                    f"[PRODUCT IMAGE COUNT] "

                    f"{product_data.get('name')} "

                    f"| images={len(product_images)}"
                )

                # =====================================
                # FALLBACK TO PAGE IMAGES
                # =====================================

                if not product_images:

                    product_images = page_images

                    _logger.warning(

                        f"[PAGE IMAGE FALLBACK] "

                        f"{product_data.get('name')}"
                    )

                segmented_assets = []

                for img in product_images:

                    # ---------------------------------
                    # ALREADY STRUCTURED
                    # ---------------------------------

                    if isinstance(img, dict):

                        if img.get("image"):

                            segmented_assets.append(img)

                    # ---------------------------------
                    # RAW BASE64 FALLBACK
                    # ---------------------------------

                    elif isinstance(img, str):

                        segmented_assets.append({

                            "image": img,

                            "score": 0,

                            "is_collage": False
                        })

                # =====================================
                # BUILD ASSET POOL
                # =====================================

                asset_pool = self._prepare_asset_pool(
                    segmented_assets
                )

                _logger.warning(

                    f"[PDF ASSET POOL] "

                    f"product={product_data.get('name')} "

                    f"| assets={len(asset_pool)}"
                )

                try:

                    name = (

                        product_data.get(
                            "name"
                        )

                        or ""

                    ).strip()

                    if not name:

                        continue

                    raw_category = (

                        product_data.get(
                            "category"
                        ) or ""

                    ).lower()

                    variants = product_data.get(
                        "variants",
                        []
                    )

                    variant_group = (

                        product_data.get(
                            "variant_group"
                        )

                        or

                        name
                    )

                    variant_group = str(
                        variant_group
                    ).strip().upper()

                    category = (
                        self._get_or_create_pdf_category(

                            raw_category,

                            category_obj,

                            parent_category,

                            CATEGORY_MAPPING
                        )
                    )

                    vendor_fingerprint = (
                        f"{vendor_id}_{variant_group}"
                    )

                    product, created = (

                        self._get_or_create_pdf_product(

                            product_data,

                            variant_group,

                            vendor_id,

                            vendor_fingerprint,

                            category,

                            asset_pool,

                            product_obj
                        )
                    )


                    if created:

                        self._apply_product_translation(
                            product
                        )

                        self._create_pdf_gallery(

                            product,

                            product_data,

                            asset_pool
                        )

                        # =====================================
                        # APPLY REAL INVENTORY STOCK
                        # =====================================

                        try:

                            stock_qty = int(

                                product_data.get(
                                    "stock_qty",
                                    0
                                ) or 0
                            )

                            if stock_qty > 0:

                                quant = stock_quant_obj.search([

                                    (
                                        'product_id',
                                        '=',
                                        product.product_variant_id.id
                                    ),

                                    (
                                        'location_id',
                                        '=',
                                        stock_location.id
                                    )

                                ], limit=1)

                                if quant:

                                    quant.inventory_quantity = (
                                        stock_qty
                                    )

                                    quant.action_apply_inventory()

                                else:

                                    quant = stock_quant_obj.create({

                                        'product_id':
                                            product.product_variant_id.id,

                                        'location_id':
                                            stock_location.id,

                                        'inventory_quantity':
                                            stock_qty
                                    })

                                    quant.action_apply_inventory()

                                _logger.warning(

                                    f"[STOCK APPLIED] "

                                    f"{product.name} "

                                    f"| qty={stock_qty}"
                                )

                        except Exception as e:

                            _logger.warning(

                                f"[STOCK APPLY FAILED] "

                                f"{str(e)}"
                            )

                        created_count += 1

                    else:

                        skipped_count += 1

                    if not variants:

                        variants = [{

                            "attributes": {

                                "Variant": name
                            }

                        }]

                    # =======================================
                    # PASS 1:
                    # BUILD ALL ATTRIBUTE LINES FIRST
                    # =======================================

                    for variant in variants:

                        attributes = variant.get(
                            "attributes",
                            {}
                        )

                        for attr_name, attr_value in attributes.items():

                            if not attr_value:
                                continue

                            attribute = self.env[
                                'product.attribute'
                            ].search([

                                (
                                    'name',
                                    '=',
                                    attr_name
                                )

                            ], limit=1)

                            if not attribute:

                                attribute = self.env[
                                    'product.attribute'
                                ].create({

                                    'name': attr_name

                                })

                            value = self.env[
                                'product.attribute.value'
                            ].search([

                                (
                                    'name',
                                    '=',
                                    attr_value
                                ),

                                (
                                    'attribute_id',
                                    '=',
                                    attribute.id
                                )

                            ], limit=1)

                            if not value:

                                value = self.env[
                                    'product.attribute.value'
                                ].create({

                                    'name': attr_value,

                                    'attribute_id':
                                        attribute.id
                                })

                            line = self.env[
                                'product.template.attribute.line'
                            ].search([

                                (
                                    'product_tmpl_id',
                                    '=',
                                    product.id
                                ),

                                (
                                    'attribute_id',
                                    '=',
                                    attribute.id
                                )

                            ], limit=1)

                            if not line:

                                self.env[
                                    'product.template.attribute.line'
                                ].create({

                                    'product_tmpl_id':
                                        product.id,

                                    'attribute_id':
                                        attribute.id,

                                    'value_ids': [(6, 0, [

                                        value.id

                                    ])]
                                })

                            else:

                                if (

                                    value.id

                                    not in

                                    line.value_ids.ids

                                ):

                                    line.value_ids = [

                                        (4, value.id)

                                    ]

                    # =======================================
                    # PASS 2:
                    # GENERATE ALL VARIANTS ONCE
                    # =======================================

                    product._create_variant_ids()

                    used_asset_indexes = set()

                    # =======================================
                    # PASS 3:
                    # MATCH REAL VARIANTS TO IMAGES
                    # =======================================

                    for variant in variants:

                        # =====================================
                        # MATCH REAL GENERATED VARIANT
                        # =====================================

                        variant_record = False

                        product_variants = (
                            product.product_variant_ids
                        )

                        variant_name = ""

                        attributes = variant.get(
                            "attributes",
                            {}
                        )

                        if isinstance(attributes, dict):

                            variant_name = " ".join([

                                str(v)

                                for v in attributes.values()

                            ]).lower()

                        for pv in product_variants:

                            combo = " ".join([

                                v.name.lower()

                                for v in (
                                    pv.product_template_variant_value_ids
                                )

                            ])

                            if combo:

                                combo_words = combo.split()

                                variant_words = (
                                    variant_name.split()
                                )

                                match_count = 0

                                for word in variant_words:

                                    if word in combo_words:

                                        match_count += 1

                                if (

                                    variant_words

                                    and

                                    match_count >= 1
                                ):

                                    variant_record = pv
                                    break

                        # ---------------------------------
                        # SAFE FALLBACK
                        # ---------------------------------

                        if (

                            not variant_record

                            and

                            product_variants
                        ):

                            variant_record = (
                                product_variants[0]
                            )
                      
                        # =====================================
                        # PROFESSIONAL VARIANT IMAGE MATCHING
                        # =====================================

                        if variant_record:

                            try:
                                
                                matched_asset = self._match_variant_image(

                                    variant,

                                    asset_pool,

                                    used_asset_indexes
                                )

                                # =====================================
                                # APPLY
                                # =====================================

                                if matched_asset:

                                    variant_record.image_1920 = (
                                        matched_asset.get(
                                            "image"
                                        )
                                    )

                                    used_asset_indexes.add(
                                        matched_asset.get("index")
                                    )

                                    _logger.warning(

                                        f"[VARIANT IMAGE APPLIED] "

                                        f"{variant_name} "

                                        f"| asset={matched_asset.get('index')}"
                                    )

                            except Exception as e:

                                _logger.warning(

                                    f"[VARIANT IMAGE FAILED] "

                                    f"{str(e)}"
                                )

                except Exception as e:

                    _logger.exception(

                        f"[PDF PRODUCT ERROR] "

                        f"{str(e)}"
                    )

                    continue

            try:

                self.last_created_page = (
                    page_index + 1
                )

                self._safe_commit_progress()

            except Exception as e:

                _logger.exception(

                    f"[PAGE COMMIT FAILED] "

                    f"{str(e)}"
                )

        _logger.warning(

            f"[PDF CREATE COMPLETE] "

            f"created={created_count} "

            f"| skipped={skipped_count}"
        )

        if self.last_created_page >= len(ai_pages):

            self.state = 'done'

        else:

            self.state = 'pdf_creating'

        self._safe_commit_progress()


    #==========create pdf CATEGORY RESOLVER====================================
    
    def _get_or_create_pdf_category(

        self,

        raw_category,

        category_obj,

        parent_category,

        category_mapping
    ):

        mapped_category = "General"

        raw_category = (
            raw_category or ""
        ).lower()

        for key, val in category_mapping.items():

            if key in raw_category:

                mapped_category = val

                break

        category = category_obj.search([

            ('name', '=ilike', mapped_category),

            (
                'parent_id',
                '=',
                parent_category.id
            )

        ], limit=1)

        if not category:

            category = category_obj.create({

                'name': mapped_category,

                'parent_id': parent_category.id
            })

        return category


    #==========pdf product PRODUCT CREATE/GET====================================
    def _get_or_create_pdf_product(

        self,

        product_data,

        variant_group,

        vendor_id,

        vendor_fingerprint,

        category,

        asset_pool,

        product_obj
    ):

        product = product_obj.search([

            (
                'vendor_fingerprint',
                '=',
                vendor_fingerprint
            )

        ], limit=1)

        if product:

            return product, False

        vals = {

            'name': (
                product_data.get("name")
                or ""
            ).strip(),

            'default_code': variant_group,

            'description_sale': (
                product_data.get(
                    "description"
                ) or ""
            ),

            'type': 'consu',

            'categ_id': category.id,

            'sale_ok': True,

            'website_published': False,

            'vendor_id': vendor_id,

            'vendor_fingerprint':
                vendor_fingerprint,

            'vendor_import_job_id':
                self.id,

            'vendor_stock_qty': int(

                product_data.get(
                    "stock_qty",
                    0
                ) or 0
            ),
        }



        hero_index = product_data.get(
            "hero_image_index"
        )
        


        # =====================================
        # PROFESSIONAL HERO IMAGE SELECTION
        # =====================================

        hero_asset = None

        # =====================================
        # AI SELECTED HERO
        # =====================================

        if hero_index is not None:

            for asset in asset_pool:

                if asset.get("clean_index") == hero_index:

                    # reject collages as hero
                    if asset.get("is_collage"):

                        continue

                    hero_asset = asset

                    break

        # =====================================
        # FALLBACK TO BEST CLEAN IMAGE
        # =====================================

        if not hero_asset:

            sorted_assets = sorted(

                asset_pool,

                key=lambda x: x.get(
                    "score",
                    0
                ),

                reverse=True
            )

            for asset in sorted_assets:

                # reject collage sheets
                if asset.get("is_collage"):
                    continue

                # require strong quality
                if asset.get("score", 0) >= 45:

                    hero_asset = asset

                    break

        # =====================================
        # FINAL SAFE FALLBACK
        # =====================================

        if not hero_asset and asset_pool:

            for asset in asset_pool:

                if not asset.get("is_collage"):

                    hero_asset = asset

                    break

            if not hero_asset:

                hero_asset = asset_pool[0]

        # =====================================
        # APPLY HERO IMAGE
        # =====================================

        if hero_asset:

            vals['image_1920'] = hero_asset.get(
                "image"
            )

            _logger.warning(

                f"[PDF HERO APPLIED] "

                f"score={hero_asset.get('score')} "

                f"color={hero_asset.get('dominant_color')}"
            )

        product = product_obj.with_context(

            mail_create_nolog=True,

            mail_notify_force_send=False,

            tracking_disable=True

        ).create(vals)

        return product, True


    #=========pdf product GALLERY CREATOR=======================
    def _create_pdf_gallery(

        self,

        product,

        product_data,

        asset_pool
    ):

        gallery_indexes = product_data.get(
            "gallery_image_indexes",
            []
        )

        # =====================================
        # FALLBACK GALLERY EXPANSION
        # =====================================

        if (

            len(gallery_indexes) < 3

            and

            asset_pool
        ):

            extra_indexes = [

                a.get("index")

                for a in asset_pool

                if a.get("score", 0) >= 45
            ]

            gallery_indexes.extend(
                extra_indexes
            )

            gallery_indexes = list(

                dict.fromkeys(
                    gallery_indexes
                )
            )[:6]

        used_images = set()
        used_hashes = set()

        if product.image_1920:

            used_images.add(
                product.image_1920
            )

        for index in gallery_indexes:

            try:

                gallery_image = (
                    self._resolve_asset_image(

                        asset_pool,

                        index
                    )
                )

                if not gallery_image:
                    continue

                image_hash = hashlib.md5(

                    gallery_image.encode('utf-8')

                ).hexdigest()

                if image_hash in used_hashes:
                    continue

                self.env[
                    'product.image'
                ].create({

                    'name':
                        f"{product.name} Gallery",

                    'product_tmpl_id':
                        product.id,

                    'image_1920':
                        gallery_image
                })

                used_hashes.add(
                    image_hash
                )

            except Exception as e:

                _logger.warning(

                    f"[GALLERY IMAGE FAILED] "

                    f"{product.name} "

                    f"| {str(e)}"
                )
    
    #=========pdf product STOCK APPLY=======================
    def _apply_pdf_stock(

        self,

        variant_record,

        stock_qty,

        stock_quant_obj,

        stock_location
    ):

        if (

            not stock_qty

            or

            not variant_record

            or

            not stock_location
        ):

            return

        try:

            quant = stock_quant_obj.search([

                (
                    'product_id',
                    '=',
                    variant_record.id
                ),

                (
                    'location_id',
                    '=',
                    stock_location.id
                )

            ], limit=1)

            if quant:

                quant.inventory_quantity = (
                    stock_qty
                )

                quant.action_apply_inventory()

            else:

                quant = stock_quant_obj.create({

                    'product_id':
                        variant_record.id,

                    'location_id':
                        stock_location.id,

                    'inventory_quantity':
                            stock_qty
                })

                quant.action_apply_inventory()

            _logger.warning(

                f"[PDF STOCK SET] "

                f"{variant_record.display_name} "

                f"-> {stock_qty}"
            )

        except Exception as e:

            _logger.warning(

                f"[PDF STOCK FAILED] "

                f"{str(e)}"
            )

    #===============fingerprint================================
    def _build_vendor_fingerprint(self, product_data):

        import re
        import hashlib

        name = (
            product_data.get("name") or ""
        ).strip().lower()

        sku = (
            product_data.get("sku")
            or product_data.get("code")
            or product_data.get("product_code")
            or ""
        ).strip().lower()

        url = (
            product_data.get("url")
            or product_data.get("link")
            or ""
        ).strip().lower()

        # normalize
        def clean(v):
            return re.sub(r'[^a-z0-9]', '', v or '')

        base = "|".join([
            clean(name),
            clean(sku),
            clean(url),
        ])

        return hashlib.md5(
            base.encode("utf-8")
        ).hexdigest()

    #==========Excel url detect workflo=======================
    def _extract_product_url(self, row):

        possible_keys = [

            "url",
            "link",
            "product_url",
            "product link",
            "website",
            "href",

        ]

        for key in possible_keys:

            value = row.get(key)

            if not value:
                continue

            value = str(value).strip()

            if value.startswith(
                ("http://", "https://")
            ):
                return value

        return False
    
    #======Excel url detection router=======================
    def _route_excel_rows(self, products):

        normal_products = []
        url_products = []

        for row in products:

            url = self._extract_product_url(row)

            if url:

                row["detected_url"] = url

                url_products.append(row)

            else:

                normal_products.append(row)

        return {
            "normal": normal_products,
            "url": url_products,
        }

    
    #==========Excel URl queue logic========================
    def _queue_excel_urls(self, url_products):

        import json

        if not url_products:
            return

        seen = set()
        cleaned = []

        for row in url_products:

            url = row.get("detected_url")

            if not url:
                continue

            if url in seen:
                continue

            seen.add(url)

            cleaned.append(row)

        url_products = cleaned

        self.excel_url_queue = json.dumps(
            url_products
        )

        self.excel_url_processing = True

        self.excel_url_index = 0

    
    #============Excel URL processor==========================
    def process_excel_url_queue(self):

        import json

        if not self.excel_url_queue:

            _logger.warning(
                "[URL QUEUE] EMPTY"
            )

            return


        rows = json.loads(
            self.excel_url_queue
        )


        start = self.excel_url_index or 0

        BATCH_SIZE = 5

        end = min(
            start + BATCH_SIZE,
            len(rows)
        )


        _logger.warning(

            f"[URL QUEUE START] "

            f"{start} -> {end} | "

            f"total={len(rows)}"
        )

        vendor_id = (
            self.partner_id.id
            if self.partner_id
            else False
        )

        for idx in range(start, end):

            try:

                row = rows[idx]

                product_url = row.get(
                    "detected_url"
                )


                if not product_url:

                    _logger.warning(
                        f"[URL QUEUE SKIP] "
                        f"NO URL AT INDEX {idx}"
                    )

                    continue

                existing_job = self.env[
                    'vendor.import.job'
                ].search([

                    ('data_url', '=', product_url),

                    ('state', '!=', 'failed')

                ], limit=1)


                if existing_job:

                    _logger.warning(

                        f"[URL JOB EXISTS] "

                        f"{product_url}"
                    )

                    # ====================================
                    # ADVANCE QUEUE INDEX
                    # ====================================

                    self.excel_url_index = idx + 1

                    self._safe_commit_progress()

                    continue

                # ====================================
                # CREATE ISOLATED URL JOB
                # ====================================

                new_job = self.env[
                    'vendor.import.job'
                ].create({

                    'name':
                        f"URL Import - {idx}",

                    'partner_id':
                        vendor_id,

                    'source_type':
                        'url',

                    'data_url':
                        product_url,

                    'state':
                        'url_scraping',
                })


                _logger.warning(

                    f"[URL JOB CREATED] "

                    f"job={new_job.id} | "

                    f"url={product_url}"
                )


                # ====================================
                # SAVE PROGRESS
                # ====================================

                self.excel_url_index = idx + 1

                self._safe_commit_progress()


            except Exception as e:

                _logger.exception(
                    f"[EXCEL URL ERROR] {str(e)}"
                )

                self.env.cr.rollback()

 
        # =========================================
        # COMPLETE
        # =========================================

        if self.excel_url_index >= len(rows):

            _logger.warning(
                "[URL QUEUE COMPLETE]"
            )

            self.excel_url_queue = False

            self.excel_url_processing = False

            self.excel_url_index = 0

            self._safe_commit_progress()

            self.env.invalidate_all()


    #==========create excel product===========================
    def create_products_excel(self):

        import json
        import re

        _logger.warning(
            "[EXCEL CREATE] START"
        )


        # =====================================================
        # VALIDATION
        # =====================================================

        if not self.ai_response:

            _logger.warning(
                "[EXCEL CREATE] NO AI RESPONSE"
            )

            return


        try:

            ai_pages = json.loads(
                self.ai_response or "[]"
            )

        except Exception as e:

            _logger.exception(

                f"[EXCEL CREATE] "

                f"INVALID AI JSON "

                f"| {str(e)}"
            )

            return


        if not ai_pages:

            _logger.warning(
                "[EXCEL CREATE] EMPTY AI"
            )

            return


        ai_page = ai_pages[0]

        products = ai_page.get(
            "products",
            []
        )


        _logger.warning(

            f"[EXCEL CREATE] "

            f"RAW PRODUCTS={len(products)}"
        )


        if not products:

            return


        # =====================================================
        # MODELS
        # =====================================================

        product_obj = self.env[
            'product.template'
        ]

        category_obj = self.env[
            'product.category'
        ]

        attribute_obj = self.env[
            'product.attribute'
        ]

        attribute_value_obj = self.env[
            'product.attribute.value'
        ]

        line_obj = self.env[
            'product.template.attribute.line'
        ]


        # =====================================================
        # ROOT CATEGORY
        # =====================================================

        parent_category = category_obj.search([

            ('name', '=', "All Products")

        ], limit=1)


        if not parent_category:

            parent_category = category_obj.create({

                'name': "All Products"

            })


        # =====================================================
        # CATEGORY MAP
        # =====================================================

        CATEGORY_MAPPING = {

            "t-shirt": "Apparel",
            "shirt": "Apparel",
            "polo": "Apparel",
            "bag": "Bags",
            "backpack": "Bags",
            "cap": "Headwear",
            "hat": "Headwear",
            "bottle": "Drinkware",
            "drinkware": "Drinkware",
            "pen": "Stationery",
            "notebook": "Stationery",
            "powerbank": "Electronics",
            "charger": "Electronics",
            "laptop": "Electronics",
        }


        # =====================================================
        # GROUP PRODUCTS
        # =====================================================

        grouped_products = {}


        for p in products:

            raw_name = (
                p.get("name") or ""
            ).strip()


            variant_group = (
                p.get("variant_group")
            )


            if variant_group:

                group_id = str(
                    variant_group
                ).strip().upper()

            else:

                match = re.search(

                    r'(?:Product\s*)?([A-Z]*\d+)',

                    raw_name,

                    re.I
                )


                if match:

                    group_id = (
                        match.group(1)
                        .upper()
                    )

                else:

                    group_id = (
                        raw_name.upper()
                    )


            grouped_products.setdefault(

                group_id,

                []

            ).append(p)


        grouped_keys = list(
            grouped_products.keys()
        )


        _logger.warning(

            f"[EXCEL GROUPS] "

            f"TOTAL={len(grouped_keys)}"
        )


        # =====================================================
        # BATCH GROUPS
        # =====================================================

        BATCH_SIZE = 10

        start = (
            self.excel_created_index or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(grouped_keys)
        )


        _logger.warning(

            f"[EXCEL BATCH] "

            f"{start} → {end}"
        )


        created_count = 0
        merged_count = 0


        # =====================================================
        # PROCESS GROUPS
        # =====================================================

        for group_idx in range(start, end):

            try:

                group_id = grouped_keys[
                    group_idx
                ]

                group_items = grouped_products[
                    group_id
                ]


                _logger.warning(

                    f"[EXCEL GROUP] "

                    f"{group_id} "

                    f"| items={len(group_items)}"
                )


                main_product = (
                    group_items[0]
                )

                fingerprint = self._build_vendor_fingerprint(
                    main_product
                )


                name = (

                    main_product.get(
                        "name"
                    ) or ""

                ).strip()


                description = (

                    main_product.get(
                        "description"
                    ) or ""
                )


                raw_category = (

                    main_product.get(
                        "category"
                    ) or ""

                ).lower()


                mapped_category = (
                    "General"
                )


                for key, val in CATEGORY_MAPPING.items():

                    if key in raw_category:

                        mapped_category = val

                        break


                category = category_obj.search([

                    (
                        'name',
                        '=',
                        mapped_category
                    ),

                    (
                        'parent_id',
                        '=',
                        parent_category.id
                    )

                ], limit=1)


                if not category:

                    category = category_obj.create({

                        'name':
                            mapped_category,

                        'parent_id':
                            parent_category.id
                    })


                # ================================================
                # FIND BY PRODUCT CODE FIRST
                # ================================================

                vendor_id = self.partner_id.id if self.partner_id else False

                product = False

                # =====================================================
                # 1. STRICT FINGERPRINT MATCH
                # =====================================================

                if (
                    'vendor_fingerprint' in product_obj._fields
                    and vendor_id
                ):

                    product = product_obj.search([

                        (
                            'vendor_fingerprint',
                            '=',
                            fingerprint
                        ),

                        (
                            'vendor_id',
                            '=',
                            vendor_id
                        )

                    ], limit=1)


                    if product:

                        _logger.warning(

                            f"[FINGERPRINT MATCH] "

                            f"{group_id} "

                            f"| vendor={vendor_id} "

                            f"| product_id={product.id}"
                        )

                 # =====================================================
                # 2. FALLBACK SKU MATCH
                # =====================================================

                if not product and vendor_id:

                    product = product_obj.search([

                        (
                            'default_code',
                            '=',
                            group_id
                        ),

                        (
                            'vendor_id',
                            '=',
                            vendor_id
                        )

                    ], limit=1)


                    if product:

                        _logger.warning(

                            f"[SKU MATCH] "

                            f"{group_id} "

                            f"| vendor={vendor_id} "

                            f"| product_id={product.id}"
                        )

                is_new_product = False

                if product:

                    _logger.warning(
                        f"[EXCEL DUPLICATE FOUND] "
                        f"{group_id} | vendor={vendor_id} | product_id={product.id}"
                    )

                else:
                    is_new_product = True


                # =================================================
                # CREATE PARENT
                # =================================================

                if is_new_product:

                    vals = {

                        'name': name,

                        'default_code':
                            group_id,

                        'description_sale':
                            description,

                       'type': 'consu',

                        'categ_id':
                            category.id,

                        'sale_ok': True,

                        'website_published':
                            False,

                        # =====================================
                        # SAVE VENDOR LINK
                        # =====================================

                        'vendor_id':
                            vendor_id,

                        'list_price':
                            self._safe_float(
                                main_product.get("price")
                            ),

                        'vendor_fingerprint': fingerprint,

                        'vendor_import_job_id': self.id,
                    }


                    image = main_product.get(
                        "image"
                    )


                    if image:

                        vals[
                            'image_1920'
                        ] = image


                    product = product_obj.create(
                        vals
                    )

                    # ✅ SAFE TRANSLATION CALL (PLUG-IN)
                    self._apply_product_translation(product)
                    created_count += 1


                    _logger.warning(

                        f"[EXCEL CREATED] "

                        f"{group_id} "

                        f"| vendor={vendor_id}"
                    )

                else:

                    merged_count += 1

                    # =====================================
                    # TRANSLATE EXISTING PRODUCT TOO
                    # =====================================

                    self._apply_product_translation(product)

                    _logger.warning(

                        f"[EXCEL EXISTING PRODUCT] "

                        f"{group_id} "

                        f"| vendor={vendor_id} "

                        f"| product_id={product.id}"
                    )
              

                # ==================================================
                # VARIANTS
                # ==================================================

                for idx, item in enumerate(group_items):

                    # =============================================
                    # DETECT ATTRIBUTE TYPE
                    # =============================================

                    variant_attribute_name = "Variant"

                    if item.get("color") or item.get("colour"):

                        variant_attribute_name = "Color"

                    elif item.get("material"):

                        variant_attribute_name = "Material"

                    elif item.get("size"):

                        variant_attribute_name = "Size"

                    elif item.get("capacity"):

                        variant_attribute_name = "Capacity"

                    elif item.get("style"):

                        variant_attribute_name = "Style"


                    # =============================================
                    # DETECT ATTRIBUTE VALUE
                    # =============================================

                    attr_value = str(

                        item.get("color")

                        or item.get("colour")

                        or item.get("material")

                        or item.get("size")

                        or item.get("variant")

                        or item.get("capacity")

                        or item.get("style")

                        or f"Variant {idx+1}"

                    ).strip()

                    if not attr_value:

                        detected_color = self._detect_basic_image_color(
                            item.get("image")
                        )

                        if detected_color:

                            variant_attribute_name = "Color"

                            attr_value = detected_color

                            _logger.warning(

                                f"[IMAGE COLOR FALLBACK] "

                                f"{detected_color}"
                            )

                        else:


                            attr_value = (

                                item.get("vendor_code")

                                or

                                item.get("primary_code")

                                or

                                f"Code {idx+1}"
                            )

                    _logger.warning(

                        f"[VARIANT DETECTED] "

                        f"{variant_attribute_name} "

                        f"= {attr_value}"
                    )


                    # =============================================
                    # ATTRIBUTE
                    # =============================================

                    attribute = attribute_obj.search([

                        (
                            'name',
                            '=',
                            variant_attribute_name
                        )

                    ], limit=1)


                    if not attribute:

                        attribute = attribute_obj.create({

                            'name': variant_attribute_name

                        })


                        _logger.warning(

                            f"[ATTRIBUTE CREATED] "

                            f"{variant_attribute_name}"
                        )


                    # =============================================
                    # ATTRIBUTE VALUE
                    # =============================================

                    value = attribute_value_obj.search([

                        (
                            'name',
                            '=',
                            attr_value
                        ),

                        (
                            'attribute_id',
                            '=',
                            attribute.id
                        )

                    ], limit=1)

                    if not value:

                        value = attribute_value_obj.create({

                            'name': attr_value,

                            'attribute_id': attribute.id
                        })


                        # =========================================
                        # TRANSLATE VARIANT VALUE
                        # =========================================

                        try:

                            for lang_code in ['ru_RU', 'az_AZ']:

                                translated_variant = self._force_translate(

                                    attr_value,

                                    lang_code
                                )


                                if translated_variant:

                                    value.with_context(
                                        lang=lang_code
                                    ).write({

                                        'name': translated_variant
                                    })


                                    _logger.warning(

                                        f"[VARIANT TRANSLATED] "

                                        f"{attr_value} "

                                        f"-> "

                                        f"{translated_variant} "

                                        f"({lang_code})"
                                    )

                        except Exception as e:

                            _logger.warning(

                                f"[VARIANT TRANSLATION ERROR] "

                                f"{str(e)}"
                            )


                        _logger.warning(

                            f"[ATTRIBUTE VALUE CREATED] "

                            f"{attr_value}"
                        )


                    # =============================================
                    # TEMPLATE ATTRIBUTE LINE
                    # =============================================

                    line = line_obj.search([

                        (
                            'product_tmpl_id',
                            '=',
                            product.id
                        ),

                        (
                            'attribute_id',
                            '=',
                            attribute.id
                        )

                    ], limit=1)


                    if not line:

                        line = line_obj.create({

                            'product_tmpl_id': product.id,

                            'attribute_id': attribute.id,

                            'value_ids': [

                                (
                                    6,
                                    0,
                                    [value.id]
                                )
                            ]
                        })


                        _logger.warning(

                            f"[VARIANT LINE CREATED] "

                            f"{group_id}"
                        )

                    else:

                        if value.id not in line.value_ids.ids:

                            line.value_ids = [

                                (
                                    4,
                                    value.id
                                )
                            ]


                            _logger.warning(

                                f"[VARIANT ADDED] "

                                f"{group_id} "

                                f"| {attr_value}"
                            )


                    # =============================================
                    # VARIANT IMAGE
                    # =============================================

                    variant_record = self.env[
                        'product.product'
                    ].search([

                        (
                            'product_tmpl_id',
                            '=',
                            product.id
                        ),

                        (
                            'product_template_attribute_value_ids.product_attribute_value_id',
                            '=',
                            value.id
                        )

                    ], limit=1)


                    if variant_record:

                        variant_image = item.get(
                            "image"
                        )


                        if variant_image:

                            variant_record.image_1920 = (
                                variant_image
                            )


                            _logger.warning(

                                f"[VARIANT IMAGE] "

                                f"{group_id} "

                                f"| {attr_value}"
                            )


                # =================================================
                # SAVE PROGRESS
                # =================================================

                self.excel_created_index = (
                    group_idx + 1
                )


                self._safe_commit_progress()
                

                _logger.warning(

                    f"[EXCEL SAVE] "

                    f"index="

                    f"{self.excel_created_index}"
                )


            except Exception as e:

                _logger.exception(

                    f"[EXCEL GROUP ERROR] "

                    f"group_idx={group_idx} "

                    f"| {str(e)}"
                )

                self.env.cr.rollback()


        # =====================================================
        # FINAL LOG
        # =====================================================

        _logger.warning(

            f"[EXCEL COMPLETE] "

            f"created={created_count} "

            f"| merged={merged_count}"
        )


        # ======================================================
        # NEXT STATE
        # ======================================================

        if self.excel_created_index >= len(grouped_keys):

            _logger.warning(

                "[EXCEL FLOW] "

                "GROUP BATCH COMPLETE"
            )

            # =========================================
            # FULL IMPORT COMPLETED
            # =========================================

            if self.is_excel_parsed:

                _logger.warning(
                    "[EXCEL IMPORT COMPLETE] ✅"
                )

                # =========================================
                # FINAL RESET
                # =========================================

                self.excel_created_index = 0

                self.excel_ai_index = 0

                self.ai_response = False

                self.state = 'done'

                # cleanup URL queue
                self.excel_url_processing = False

                self.excel_url_queue = False

                self.excel_url_index = 0

            # =========================================
            # MORE PARSE ROWS REMAIN
            # =========================================

            else:

                _logger.warning(

                    "[EXCEL FLOW] "

                    "RETURN TO excel_parsing"
                )

                # IMPORTANT:
                # KEEP CURRENT AI STATE
                # for next parse batch

                self.state = 'excel_parsing'

                _logger.warning(

                    "[EXCEL FLOW] "

                    f"NEXT PARSE INDEX="

                    f"{self.excel_parse_index}"
                )

        else:

            self.state = 'excel_creating'


        self._safe_commit_progress()


    #====Excel variant mapping==================================
    def _detect_basic_image_color(self, image_data):

        try:

            import base64
            from io import BytesIO

            from PIL import Image

            img = Image.open(
                BytesIO(
                    base64.b64decode(image_data)
                )
            ).convert("RGB")

            img = img.resize((50, 50))

            colors = img.getcolors(
                50 * 50
            )

            if not colors:
                return False

            dominant = max(
                colors,
                key=lambda x: x[0]
            )[1]

            r, g, b = dominant


            # =====================================
            # BASIC COLOR MAPPING
            # =====================================

            if r > 200 and g > 200 and b > 200:
                return "White"

            if r < 60 and g < 60 and b < 60:
                return "Black"

            if r > 180 and g < 120 and b < 80:
                return "Orange"

            if r > 180 and g < 80 and b < 80:
                return "Red"

            if b > 150 and r < 120:
                return "Blue"

            if g > 140 and r < 120:
                return "Green"

            if r > 150 and g > 150 and b < 120:
                return "Yellow"

            return "Standard"

        except Exception as e:

            _logger.warning(

                f"[COLOR DETECTION FAILED] "

                f"{str(e)}"
            )

            return False


    #-----URL API FLOW-------------------------------------------

    def scrape_with_playwright(self):

        from playwright.sync_api import sync_playwright
        import subprocess
        import os

        _logger.warning(f"PLAYWRIGHT SCRAPE → {self.data_url}")

        #✅ CHECK IF BROWSER EXISTS FIRST
        browser_path = os.path.expanduser("~/.cache/ms-playwright")

        if not os.path.exists(browser_path):
            _logger.warning("PLAYWRIGHT → INSTALLING BROWSER (FIRST RUN)")

            try:
                subprocess.run(
                    ["python", "-m", "playwright", "install", "chromium"],
                    check=True
                )
                _logger.warning("PLAYWRIGHT BROWSER INSTALLED")
            except Exception as e:
                _logger.error(f"PLAYWRIGHT INSTALL FAILED → {str(e)}")
                return

        products = []

        with sync_playwright() as p:

            browser = p.chromium.launch(headless=True)
            page = browser.new_page()

            page.goto(self.data_url, timeout=60000)
            page.wait_for_timeout(5000)

            # ✅ Cookie handling
            try:
                page.locator("button:has-text('Accept')").click(timeout=3000)
            except:
                pass

            items = page.query_selector_all("a, div")

            _logger.warning(f"PLAYWRIGHT ELEMENTS FOUND → {len(items)}")

            for item in items[:200]:

                try:
                    text = item.inner_text().strip()

                    if not text or len(text) < 5:
                        continue

                    img_el = item.query_selector("img")
                    img_url = img_el.get_attribute("src") if img_el else None

                    if img_url and img_url.startswith("//"):
                        img_url = "https:" + img_url

                    img_base64 = None

                    if img_url:
                        try:
                            res = requests.get(img_url, timeout=10)
                            if res.status_code == 200:
                                img_base64 = base64.b64encode(res.content).decode("utf-8")
                        except:
                            pass

                    products.append({
                        "name": text[:120],
                        "image": img_base64
                    })

                except:
                    continue

            browser.close()

        _logger.warning(f"PLAYWRIGHT PRODUCTS → {len(products)}")

        if not products:
            _logger.error("PLAYWRIGHT FAILED → NO PRODUCTS")
            return

        pages = [{
            "page": 1,
            "text": "\n".join([p["name"] for p in products]),
            "images": [p["image"] for p in products if p["image"]]
        }]

        self.extracted_text = json.dumps(pages)

        _logger.warning(f"PLAYWRIGHT DONE → {len(products)} PRODUCTS")

    # =====================================================
    # CRON PROCESSOR
    # =====================================================
    def run_pending_jobs(self):

        from odoo import fields

        _logger.warning(
            "🔥 CRON HEARTBEAT → RUNNING"
        )


        active_states = [

            'draft',

            'excel_parsing',
            'excel_ai',
            'excel_creating',

            'pdf_extracting',
            'pdf_ai',
            'pdf_creating',

            'url_scraping',
            'url_ai',
            'url_creating',

            # =====================================
            # AUTO-RECOVER INTERRUPTED JOBS
            # =====================================

            'failed',
        ]


        # =================================================
        # LOAD JOBS
        # =================================================

        jobs = self.search(

            [('state', 'in', active_states)],

            order="id desc"
        )


        _logger.warning(

            f"[CRON] ACTIVE JOBS "

            f"→ {len(jobs)}"
        )


        # =================================================
        # REMOVE DUPLICATES
        # =================================================

        seen = {}

        duplicates = self.env[
            'vendor.import.job'
        ]

        for j in jobs:

            # =====================================
            # AUTO RECOVER FAILED JOBS
            # =====================================

            if j.state == 'failed':

                _logger.warning(

                    f"[AUTO RECOVER] "

                    f"job={j.id}"
                )

                try:

                    if j.last_created_page:

                        j.state = 'pdf_creating'

                    elif j.last_ai_page:

                        j.state = 'pdf_ai'

                    elif j.current_page:

                        j.state = 'pdf_extracting'

                    else:

                        j.state = 'draft'

                    j.lock = False

                    self.env.cr.commit()

                    _logger.warning(

                        f"[AUTO RECOVER OK] "

                        f"job={j.id} "

                        f"| state={j.state}"
                    )

                except Exception as e:

                    _logger.exception(

                        f"[AUTO RECOVER FAILED] "

                        f"{str(e)}"
                    )

            sig = j.upload_signature
           
            if not sig:

                continue


            if sig not in seen:

                seen[sig] = j

            else:

                if j.id > seen[sig].id:

                    duplicates |= seen[sig]

                    seen[sig] = j

                else:

                    duplicates |= j


        if duplicates:

            _logger.warning(

                f"[CRON] REMOVING DUPLICATES "

                f"→ {len(duplicates)}"
            )

            try:

                duplicates.unlink()

                self.env.cr.commit()

                _logger.warning(
                    "[CRON] DUPLICATES REMOVED"
                )

            except Exception as e:

                _logger.exception(

                    f"[CRON ERROR] "

                    f"DUPLICATE DELETE FAILED "

                    f"→ {str(e)}"
                )


        # =================================================
        # RECOVER STALE LOCKS
        # =================================================

        stale_jobs = self.search([

            ('state', 'in', active_states),

            ('lock', '=', True)

        ])


        for stale in stale_jobs:

            try:

                delta = (

                    fields.Datetime.now()

                    - stale.write_date

                ).total_seconds()

            except Exception:

                delta = 0


            _logger.warning(

                f"[LOCK CHECK] "

                f"job={stale.id} "

                f"| seconds={delta}"
            )


            if delta > 60:

                _logger.warning(

                    f"[STALE LOCK RESET] "

                    f"job={stale.id}"
                )

                try:

                    stale.lock = False

                    self.env.cr.commit()

                except Exception as e:

                    _logger.exception(

                        f"[STALE LOCK ERROR] "

                        f"{str(e)}"
                    )


        # =================================================
        # GET NEXT JOB
        # =================================================

        job = self.search(

            [

                ('state', 'in', active_states),

                ('lock', '=', False)

            ],

            order="create_date asc, id asc",

            limit=1
        )


        if not job:

            _logger.warning(

                "[CRON] NO AVAILABLE JOBS "
                "(all locked or done)"
            )

            return


        # =================================================
        # PROCESS
        # =================================================

        try:

            # =============================================
            # LOCK
            # =============================================

            job.lock = True

            self.env.cr.commit()


            _logger.warning(

                f"[CRON] JOB LOCKED "

                f"| job={job.id}"
            )


            # =============================================
            # SAFER CHAIN
            # =============================================

            MAX_CHAIN = 1


            for step in range(MAX_CHAIN):

                # =========================================
                # REFRESH
                # =========================================

                try:

                    job.invalidate_cache()

                except Exception:

                    pass


                job = self.env[
                    'vendor.import.job'
                ].browse(job.id)


                _logger.warning(

                    f"[CHAIN] STEP "

                    f"{step + 1} "

                    f"| state={job.state}"
                )


                # =========================================
                # STOP STATES
                # =========================================

                if job.state == 'done':

                    _logger.warning(

                        f"[CHAIN STOP] "

                        f"terminal state "

                        f"→ {job.state}"
                    )

                    break


                # =========================================
                # TRACK BEFORE
                # =========================================

                previous_state = (
                    job.state
                )

                previous_page = (
                    job.current_page or 0
                )

                previous_ai_page = (
                    job.last_ai_page or 0
                )

                previous_created = (
                    job.last_created_page or 0
                )

                previous_excel_ai = (
                    job.excel_ai_index or 0
                )

                previous_excel_created = (
                    job.excel_created_index or 0
                )

            
                previous_url_batch = (
                    job.url_batch_index or 0
                )

                previous_url_created = (
                    job.last_processed_product_index or 0
                )
             

                _logger.warning(

                    f"[CHAIN BEFORE] "

                    f"state={previous_state} "

                    f"| extract={previous_page} "

                    f"| ai={previous_ai_page} "

                    f"| create={previous_created} "

                    f"| excel_ai={previous_excel_ai} "

                    f"| excel_create={previous_excel_created}"
                )


                # =========================================
                # PROCESS
                # =========================================

                try:

                    job._process_step()

                except Exception as e:

                    _logger.exception(

                        f"[PROCESS ERROR] "

                        f"job={job.id} "

                        f"| {str(e)}"
                    )

                    try:

                        job.state = 'failed'

                        self.env.cr.commit()

                    except Exception:

                        _logger.warning(

                            "[PROCESS ERROR] "

                            "FAILED SAVE FAILED"
                        )

                    break


                # =========================================
                # REFRESH AFTER
                # =========================================

                try:

                    job.invalidate_cache()

                except Exception:

                    pass


                job = self.env[
                    'vendor.import.job'
                ].browse(job.id)


                _logger.warning(

                    f"[CHAIN AFTER] "

                    f"state={job.state} "

                    f"| extract={job.current_page} "

                    f"| ai={job.last_ai_page} "

                    f"| create={job.last_created_page} "

                    f"| excel_ai={job.excel_ai_index} "

                    f"| excel_create={job.excel_created_index}"
                )


                # =========================================
                # PROGRESS DETECTION
                # =========================================

                progress_detected = False


                # PDF extract progress

                if (

                    (job.current_page or 0)

                    >

                    previous_page

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] PDF "

                        f"{previous_page}"

                        f" → "

                        f"{job.current_page}"
                    )


                # PDF AI progress

                if (

                    (job.last_ai_page or 0)

                    >

                    previous_ai_page

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] PDF AI "

                        f"{previous_ai_page}"

                        f" → "

                        f"{job.last_ai_page}"
                    )


                # PDF create progress

                if (

                    (job.last_created_page or 0)

                    !=

                    previous_created

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] PDF CREATE "

                        f"{previous_created}"

                        f" → "

                        f"{job.last_created_page}"
                    )


                # PDF create state continuation

                elif job.state == 'pdf_creating':

                    progress_detected = True

                    _logger.warning(

                        "[PROGRESS] PDF CREATE LOOP ACTIVE"
                    )
            

                # Excel AI progress

                if (

                    (job.excel_ai_index or 0)

                    >

                    previous_excel_ai

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] EXCEL AI "

                        f"{previous_excel_ai}"

                        f" → "

                        f"{job.excel_ai_index}"
                    )


                # Excel create progress

                if (

                    (job.excel_created_index or 0)

                    >

                    previous_excel_created

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] EXCEL CREATE "

                        f"{previous_excel_created}"

                        f" → "

                        f"{job.excel_created_index}"
                    )

             
                # =========================================
                # URL AI progress
                # =========================================

                if (

                    (job.url_batch_index or 0)

                    >

                    previous_url_batch

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] URL AI "

                        f"{previous_url_batch}"

                        f" → "

                        f"{job.url_batch_index}"
                    )


                # =========================================
                # URL create progress
                # =========================================

                if (

                    (job.last_processed_product_index or 0)

                    >

                    previous_url_created

                ):

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] URL CREATE "

                        f"{previous_url_created}"

                        f" → "

                        f"{job.last_processed_product_index}"
                    )


                # =========================================
                # APIFY WAIT STATE
                # =========================================

                if job.state == 'url_scraping':

                    progress_detected = True

                    _logger.warning(

                        "[PROGRESS] APIFY WAITING"
                    )
              


                # state transition

                if previous_state != job.state:

                    progress_detected = True

                    _logger.warning(

                        f"[PROGRESS] STATE "

                        f"{previous_state}"

                        f" → "

                        f"{job.state}"
                    )


                # =========================================
                # STOP IF NO PROGRESS
                # =========================================

                if not progress_detected:

                    _logger.warning(

                        "[CHAIN STOP] "

                        "NO PROGRESS DETECTED"
                    )

                    break


                _logger.warning(

                    "[CHAIN CONTINUE] "

                    "PROGRESS DETECTED"
                )


                # =========================================
                # COMMIT
                # =========================================

                try:

                    self.env.cr.commit()

                    _logger.warning(
                        "[CHAIN] COMMIT OK"
                    )

                except Exception as e:

                    _logger.exception(

                        f"[CHAIN COMMIT ERROR] "

                        f"{str(e)}"
                    )

                    break


            _logger.warning(
                "[CRON] PROCESS LOOP COMPLETE"
            )


        except Exception as e:

            _logger.exception(

                f"[CRON FATAL ERROR] "

                f"{str(e)}"
            )


            try:

                self.env.cr.rollback()

                _logger.warning(
                    "[CRON] ROLLBACK OK"
                )

            except Exception:

                _logger.warning(
                    "[CRON] ROLLBACK FAILED"
                )


        finally:

            # =============================================
            # UNLOCK
            # =============================================

            try:

                if (

                    job

                    and

                    job.exists()

                ):

                    job.lock = False

                    self.env.cr.commit()


                    _logger.warning(

                        f"[CRON] JOB UNLOCKED "

                        f"| job={job.id}"
                    )

            except Exception:

                _logger.warning(

                    "[CRON] UNLOCK FAILED"
                )


        _logger.warning(
            "[CRON] RUN COMPLETE"
        )

        return


   #=============flask setup/installation=================== 
    def ping_flask_server(self):
      
        try:
            requests.get("https://pdf-extractor-staging.onrender.com", timeout=10)
            _logger.info("FLASK PING SUCCESS")
        except Exception:
            _logger.warning("FLASK PING FAILED")


    # #---------------clean_scraped_blocks-------------------------------
    def _clean_scraped_blocks(self, raw_blocks):
        """
        Clean Apify output before sending to AI
        """

        cleaned = []
        seen = set()

        for item in raw_blocks:

            text = (item.get("text") or "").strip()
            image = item.get("image")

            # ❌ REMOVE NOISE
            if not text:
                continue

            if len(text) < 15:
                continue

            if any(x in text.lower() for x in [
                "privacy", "cookie", "terms", "login",
                "menu", "navigation", "home"
            ]):
                continue

            # ❌ REMOVE DUPLICATES
            key = text[:120]  # allow more variation

            if key in seen:
                continue

            seen.add(key)

            # cleaned.append({
            #     "text": text,
            #     "image": image
            # })

            cleaned.append({
                "text": text,
                "image": image,
                "price": item.get("price", ""),
                "stock": item.get("stock", "")
            })

        return cleaned

    #---------------normalizer-------------------------------

    def _normalize_url_data(self, items):

        blocks = []

        for item in items:

            # =====================================================
            # FORMAT 1 → ORIGINAL WORKING FORMAT
            # =====================================================
            # {
            #   "text": "...",
            #   "image": "..."
            # }
            # =====================================================

            if item.get("text"):

                text = (item.get("text") or "").strip()
                image = item.get("image")

                # 🔥 STRICT VALIDATION
                if not text or len(text) < 5:
                    continue

                if (
                    image and
                    isinstance(image, str) and
                    not image.startswith("http")
                ):
                    image = None

                blocks.append({
                    "text": text,
                    "image": image,
                    "price": item.get("price", ""),
                    "stock": item.get("stock", "")
                })

                continue

            # =====================================================
            # FORMAT 2 → STRUCTURED FORMAT
            # =====================================================

            if item.get("type") == "PRODUCTS":

                for sub in item.get("items", []):

                    text = (
                        sub.get("text") or ""
                    ).strip()

                    image = sub.get("image")

                    if not text or len(text) < 5:
                        continue

                    if (
                        image and
                        isinstance(image, str) and
                        not image.startswith("http")
                    ):
                        image = None

                    # blocks.append({
                    #     "text": text,
                    #     "image": image
                    # })

                    blocks.append({
                        "text": text,
                        "image": image,
                        "price": sub.get("price", ""),
                        "stock": sub.get("stock", "")
                    })

            # =====================================================
            # DEBUG TYPES
            # =====================================================

            elif item.get("type") in [
                "EMPTY",
                "BLOCKED"
            ]:

                _logger.error(
                    f"URL DEBUG → "
                    f"{item.get('reason')}"
                )

        _logger.warning(f"NORMALIZED BLOCKS → {len(blocks)}")

        # =====================================================
        # 🔥 SPLIT INTO MULTIPLE PAGES (CRITICAL FIX)
        # =====================================================

        PAGE_SIZE = 20  # 🔥 prevents AI overload

        pages = []

        for i in range(0, len(blocks), PAGE_SIZE):

            chunk = blocks[i:i + PAGE_SIZE]

            pages.append({
                "page": len(pages) + 1,
                "blocks": chunk
            })

        _logger.warning(f"NORMALIZED PAGES → {len(pages)}")

        return pages
    

    #======apify url fetch/scrapp products=====================
    
    def _run_apify_actor(self, url):

        token = self.env['ir.config_parameter'].sudo().get_param('apify.api_token')

        if not token:
            raise Exception("Apify API token not configured")

        #ACTOR_ID = "selectad~my-actor"
        ACTOR_ID = "princ_adex~my-actor"

        # =====================================================
        # 🔥 STEP 1: START ACTOR (ONLY IF NOT STARTED)
        # =====================================================

        if not getattr(self, "apify_run_id", False):

            run_url = f"https://api.apify.com/v2/acts/{ACTOR_ID}/runs?token={token}"

            payload = {
                "startUrls": [{"url": url}]
            }

            headers = {
                "Content-Type": "application/json"
            }

            response = requests.post(run_url, json=payload, headers=headers, timeout=30)

            if response.status_code != 201:
                raise Exception(f"Apify run failed: {response.text}")

            run_data = response.json()

            # ✅ SAVE FOR NEXT CRON
            self.apify_run_id = run_data["data"]["id"]
            self.apify_dataset_id = run_data["data"]["defaultDatasetId"]

            _logger.warning(f"APIFY STARTED → RUN ID {self.apify_run_id}")

            # 🔥 IMPORTANT: STOP HERE (NON-BLOCKING)
            return None

        # =====================================================
        # 🔥 STEP 2: CHECK STATUS
        # =====================================================

        status_url = f"https://api.apify.com/v2/actor-runs/{self.apify_run_id}?token={token}"

        status_res = requests.get(status_url, timeout=20).json()
        status = status_res["data"]["status"]

        _logger.warning(f"APIFY STATUS → {status}")

        if status in ["RUNNING", "READY"]:
            _logger.warning("APIFY STILL RUNNING → WAIT NEXT CRON")
            return None

        if status in ["FAILED", "ABORTED", "TIMED-OUT"]:
            raise Exception(f"Apify run failed with status: {status}")

        # =====================================================
        # 🔥 STEP 3: FETCH DATA (ONLY WHEN DONE)
        # =====================================================

        dataset_url = f"https://api.apify.com/v2/datasets/{self.apify_dataset_id}/items"

        params = {
            "token": token,
            "limit": 1000,
            "clean": "true"
        }

        dataset_res = requests.get(dataset_url, params=params, timeout=30)

        if dataset_res.status_code != 200:
            raise Exception(f"Failed to fetch dataset: {dataset_res.text}")

        data = dataset_res.json()

        _logger.warning(f"APIFY ITEMS FETCHED → {len(data)}")

        if not data:
            _logger.warning("APIFY RETURNED EMPTY → MARK JOB AS DONE")

            self.state = 'done'   # 🔥 STOP LOOP COMPLETELY
            self._safe_commit_progress()
            return

        # 🔥 CLEAN UP (IMPORTANT)
        self.apify_run_id = False
        self.apify_dataset_id = False

        return data

    #=======validation===================
    def validate_ai_output(products):
        for p in products:
            if "variants" in p:
                if not isinstance(p["variants"], list):
                    p["variants"] = []

                for v in p["variants"]:
                    if "attributes" not in v:
                        v["attributes"] = {"Variant": "Default"}

        return products
    
    #=======keep cron alive================
    def keep_alive(self):
        _logger.warning("KEEP ALIVE PING")

   
   #=========gloat numbers=============
    def _safe_float(self, value):

        try:

            if value is None:
                return 0.0

            value = str(value)

            value = value.replace('$', '')
            value = value.replace('€', '')
            value = value.replace('£', '')
            value = value.replace(',', '')

            return float(value.strip())

        except:
            return 0.0

    #======product translate ==========================
    
    def translate_global_views(self, target_lang):

        from openai import OpenAI

        api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')

        if not api_key:
            _logger.warning("❌ Missing OpenAI API key")
            return

        client = OpenAI(api_key=api_key)

        # 🔥 FETCH SOURCE STRINGS FROM VIEWS (SAFE WAY)
        views = self.env['ir.ui.view'].sudo().search([
            ('arch_db', '!=', False)
        ], limit=20)

        _logger.warning(f"🌍 GLOBAL VIEW TRANSLATION START → {target_lang}")
        _logger.warning(f"🔍 Views found → {len(views)}")

        for view in views:

            try:
                text = view.name or ''
                if not text:
                    continue

                prompt = f"""
                Translate to {target_lang}:

                {text}
                """

                response = client.responses.create(
                    model="gpt-4.1-mini",
                    input=prompt
                )

                translated = response.output_text.strip()

                if translated:
                    view.with_context(lang=target_lang).write({
                        'name': translated
                    })

                    _logger.warning(f"✅ VIEW {view.name} → {translated}")

            except Exception as e:
                _logger.warning(f"❌ Failed: {str(e)}")
            







#====================================================
# Most recent pdf variant color 95% stable
#====================================================
from odoo import models, fields, api
import base64
import logging
import io
import requests
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import json
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openai import OpenAI
import re
import fitz
import hashlib
import psycopg2

from PIL import (
    Image,
    ImageOps,
    ImageChops
)

import cv2
import numpy as np
from odoo.exceptions import AccessError

 

_logger = logging.getLogger(__name__)

class ProductTemplate(models.Model):

    _inherit = 'product.template'

    vendor_id = fields.Many2one(
        'res.partner',
        string="Vendor"
    )

    vendor_fingerprint = fields.Char(
        index=True,
        copy=False
    )

    vendor_import_job_id = fields.Many2one(
        'vendor.import.job',
        string='Vendor Import Job',
        index=True,
        ondelete='set null'
    )
    

    vendor_stock_qty = fields.Integer()

    is_vendor_purged = fields.Boolean(
        default=False
    )


    # =============================================
    # AUTO ASSIGN VENDOR DURING CREATE
    # =============================================

    @api.model
    def create(self, vals):

        user = self.env.user

        # Vendor user creating manually from UI
        if (
            user.has_group(
                'gift_product_configurator.group_product_vendor'
            )
            and not vals.get('vendor_id')
        ):

            vals['vendor_id'] = user.partner_id.id

        return super().create(vals)

    # ==========================================
    # PROTECT OTHER VENDOR PRODUCTS
    # ==========================================

    def write(self, vals):

        user = self.env.user

        if user.has_group(
            'gift_product_configurator.group_product_vendor'
        ):

            for product in self:

                if (
                    product.vendor_id
                    and
                    product.vendor_id != user.partner_id
                ):

                    raise AccessError(
                        "You can only edit your own products."
                    )

        return super().write(vals)

    # ==========================================
    # PROTECT DELETE
    # ==========================================

    def unlink(self):

        user = self.env.user

        if user.has_group(
            'gift_product_configurator.group_product_vendor'
        ):

            for product in self:

                if (
                    product.vendor_id
                    and
                    product.vendor_id != user.partner_id
                ):

                    raise AccessError(
                        "You can only delete your own products."
                    )

        return super().unlink()
    

#=========== ✅ Extend existing model=================
class ResPartner(models.Model):
    _inherit = 'res.partner'

    #Vendor user role
    is_vendor_user = fields.Boolean(
        string="Vendor User",
        default=False
    )

class VendorImportJob(models.Model):

    _name = "vendor.import.job"
    _description = "Vendor Import Job"

    partner_id = fields.Many2one("res.partner", string="Vendor")  # ✅ LINK instead

    name = fields.Char(default="Vendor Data Import")
    extra_info = fields.Text()

    pdf_file = fields.Binary()
    excel_file = fields.Binary()
    logo_file = fields.Binary()

    extracted_text = fields.Text()
    current_page = fields.Integer(
        string="Current PDF Page",
        default=0
    )
    total_pages = fields.Integer(string="Total Pages", default=0)
    last_ai_page = fields.Integer(string="Last AI Page", default=0)
    ai_response = fields.Text()
    priority = fields.Integer(default=10)
    excel_created_index = fields.Integer(
        string="Excel Created Index",
        default=0
    )

    apify_run_id = fields.Char()
    apify_dataset_id = fields.Char()
   
    last_processed_product_index = fields.Integer(default=0)
    last_created_page = fields.Integer(default=0)
    lock = fields.Boolean(default=False)
    completion_email_sent = fields.Boolean(
        default=False
    )
    is_excel_parsed = fields.Boolean(default=False)
    excel_ai_index = fields.Integer(default=0)
    upload_signature = fields.Char(string="Upload Signature")
    processed_group_ids = fields.Text(default="[]")

    url_total_batches = fields.Integer(default=0)
    url_batch_index = fields.Integer(default=0)
    data_url = fields.Char()
    url_parse_index = fields.Integer(
        string="URL Parse Index",
        default=0
    )
    url_blocks_json = fields.Text(
        string="URL Blocks JSON"
    )
        
    excel_parse_index = fields.Integer(
        default=0
    )

    source_type = fields.Selection([
        ("pdf", "PDF"),
        ("excel", "Excel"),
        ("url", "URL"),
    ])


    excel_url_index = fields.Integer(
        default=0
    )

    excel_url_processing = fields.Boolean(
        default=False
    )

    completion_email_sent = fields.Boolean(
        default=False
    )

    stage_retry_count = fields.Integer(
        default=0
    )

    last_error = fields.Text()

    failed_at = fields.Datetime()

    last_known_state = fields.Char(
        string="Last Known State",
        default=""
    )

    failure_reason = fields.Text()
    ai_retry_count = fields.Integer(default=0)
   
    state = fields.Selection([
        ('draft', 'Draft'),
        ('processing', 'Processing'),
        ('review', 'Vendor Review'),
        ('done', 'Completed'),
        ('error', 'Error'),
        ('failed', 'Failed'),

         #New
        ('url_scraping', 'URL Scraping'),
        ('url_ai', 'URL AI'),
        ('url_creating', 'URL Creating'),

        ('pdf_extracting', 'PDF Extracting'),
        ('pdf_ai', 'PDF AI'),
        ('pdf_creating', 'PDF Creating'),

        ('excel_parsing', 'Excel Parsing'),
        ('excel_ai', 'Excel AI'),
        ('excel_creating', 'Excel Creating'),

    ], default='draft')


     #============================= MAIN FLOW (process steps) =====================

    def process_import(self):

        _logger.warning(f"PROCESS START → Job {self.id}")

        try:

            self._process_step()

        except Exception as e:
            _logger.error(f"PROCESS FAILED → {str(e)}")
            self.state = "failed"


    #=============Safe commit======================================================
    def _safe_commit_progress(self):

        try:

            self.flush_recordset()

        except Exception as flush_error:

            _logger.warning(
                f"FLUSH FAILED → {flush_error}"
            )

        try:

            self.env.cr.commit()

        except Exception as commit_error:

            _logger.warning(
                f"COMMIT FAILED → {commit_error}"
            )


    #========vendor completion email notification==========
    def send_completion_email(self, failed=False, error_message=None):

        self.ensure_one()

        _logger.warning(
            f"[EMAIL] START → job={self.id}"
        )

        # =============================================
        # VALIDATE VENDOR EMAIL
        # =============================================

        if not self.partner_id:

            _logger.error(
                f"[EMAIL] FAILED → NO PARTNER | job={self.id}"
            )

            return False

        if not self.partner_id.email:

            _logger.error(
                f"[EMAIL] FAILED → NO VENDOR EMAIL "
                f"| vendor={self.partner_id.id}"
            )

            return False

        # ============================================
        # DETECT ENVIRONMENT
        # ============================================

        base_url = self.env[
            'ir.config_parameter'
        ].sudo().get_param('web.base.url', '')

        is_staging = (
            'staging' in (base_url or '').lower()
        )

        _logger.warning(
            f"[EMAIL] ENVIRONMENT → "
            f"{'STAGING' if is_staging else 'PRODUCTION'}"
        )

        # ============================================
        # STAGING SKIP
        # ============================================

        if is_staging:

            _logger.warning(
                f"[EMAIL] SKIPPED → STAGING ENVIRONMENT "
                f"| vendor={self.partner_id.email}"
            )

            return False

        # ============================================
        # FIND SMTP SERVER
        # ============================================

        mail_server = self.env[
            'ir.mail_server'
        ].sudo().search([
            ('active', '=', True)
        ], limit=1)

        if not mail_server:

            _logger.error(
                "[EMAIL] FAILED → "
                "NO ACTIVE OUTGOING MAIL SERVER"
            )

            return False

        _logger.warning(
            f"[EMAIL] SMTP SERVER → "
            f"{mail_server.name}"
        )

        # ============================================
        # SUBJECT
        # ============================================

        if failed:

            subject = (
                f"Import Failed - {self.name}"
            )

            status_text = "Failed"

            error_html = f"""
                <br/><br/>
                <b>Error:</b><br/>
                {error_message or 'Unknown processing error'}
            """

        else:

            subject = (
                f"Import Completed - {self.name}"
            )

            status_text = "Completed"

            error_html = ""

        # ============================================
        # EMAIL BODY
        # ============================================

        body = f"""
            <div>
                <p>
                    Hello {self.partner_id.name},
                </p>

                <p>
                    Your vendor import job has finished processing.
                </p>

                <table border="1" cellpadding="6" cellspacing="0">

                    <tr>
                        <td><b>Job</b></td>
                        <td>{self.name}</td>
                    </tr>

                    <tr>
                        <td><b>Source</b></td>
                        <td>{self.source_type}</td>
                    </tr>

                    <tr>
                        <td><b>Status</b></td>
                        <td>{status_text}</td>
                    </tr>

                    <tr>
                        <td><b>Date</b></td>
                        <td>{self.create_date}</td>
                    </tr>

                </table>

                {error_html}

                <br/><br/>

                Regards
            </div>
        """

        # ============================================
        # CREATE MAIL
        # ============================================

        try:

            mail_values = {

                'subject': subject,

                'body_html': body,

                'email_to': self.partner_id.email,

                'email_from': mail_server.smtp_user or self.env.user.email,

                'auto_delete': False,
            }

            _logger.warning(
                f"[EMAIL] CREATING MAIL → "
                f"{self.partner_id.email}"
            )

            mail = self.env[
                'mail.mail'
            ].sudo().create(mail_values)

            _logger.warning(
                f"[EMAIL] MAIL CREATED → "
                f"mail_id={mail.id}"
            )

        except Exception as create_error:

            _logger.exception(
                f"[EMAIL] CREATE FAILED → "
                f"{str(create_error)}"
            )

            return False

        # ============================================
        # SEND
        # ============================================

        try:

            _logger.warning(
                f"[EMAIL] SENDING → mail_id={mail.id}"
            )

            mail.sudo().send()

            self.env.cr.commit()

            _logger.warning(
                f"[EMAIL] SUCCESS → "
                f"{self.partner_id.email}"
            )

            self.completion_email_sent = True

            self.env.cr.commit()

            return True

        except Exception as send_error:

            _logger.exception(
                f"[EMAIL] SEND FAILED → "
                f"{str(send_error)}"
            )

            try:

                mail.write({
                    'state': 'exception'
                })

                self.env.cr.commit()

            except Exception:
                pass

            return False
    
    #========vendor failed notification==============
    def _send_failed_processing_email(

        self,

        error_message=None
    ):

        self.ensure_one()

        try:

            # =====================================
            # FAILURE CONTEXT
            # =====================================

            stage = (

                self.last_known_state

                or

                self.state

                or

                "unknown"
            )

            retry_count = (

                self.retry_count

                or

                0
            )

            full_error = f"""

    Stage:
    {stage}

    Retries:
    {retry_count}

    Error:
    {error_message or 'Unknown processing failure'}
    """

            _logger.warning(

                f"[FAILED EMAIL] "

                f"job={self.id} "

                f"stage={stage} "

                f"retry={retry_count}"
            )

            return self.send_completion_email(

                failed=True,

                error_message=full_error
            )

        except Exception as e:

            _logger.exception(

                f"[FAILED EMAIL ERROR] "

                f"{str(e)}"
            )

            return False

    #============Processing Jobs===================================================
    def _process_step(self):

        import json
        import re

        self.ensure_one()

        # =================================================
        # SAFETY
        # =================================================

        if self.pdf_file and self.excel_file:

            _logger.error(
                "[PROCESS STEP] BOTH PDF AND EXCEL PROVIDED"
            )

            self.state = "failed"

            self._safe_commit_progress()

            return


        _logger.warning(
            f"[PROCESS STEP] → state={self.state}"
        )


        # =================================================
        # DONE
        # =================================================

        if self.state == 'done':

            _logger.warning(
                f"JOB {self.id} ALREADY DONE ✅"
            )

            return


        # =================================================
        # REVIEW RECOVERY
        # =================================================

        if self.state == 'review':

            _logger.warning(

                f"[REVIEW RECOVERY] "

                f"retry={self.stage_retry_count}/5 "

                f"last_state={self.last_known_state}"
            )

            # =============================================
            # HARD FAILURE LIMIT
            # =============================================

            if self.stage_retry_count >= 5:

                _logger.error(

                    f"[JOB FAILED PERMANENTLY] "

                    f"job={self.id}"
                )

                self.state = 'failed'

                self.failed_at = fields.Datetime.now()

                self._safe_commit_progress()

                return

            # =============================================
            # RESTORE LAST KNOWN STATE
            # =============================================

            if self.last_known_state:

                self.state = self.last_known_state

            else:

                self.state = 'failed'

            self._safe_commit_progress()

            return


        # =================================================
        # URL FLOW
        # =================================================
        

        if self.data_url:

            _logger.warning(
                "FLOW → URL"
            )


            # =============================================
            # START
            # =============================================

            if self.state == 'draft':

                _logger.warning(
                    "[URL FLOW] START"
                )

                self.stage_retry_count = 0

                self.last_known_state = 'url_scraping'
                self.state = 'url_scraping'

                self._safe_commit_progress()

                return


            # =============================================
            # SCRAPE
            # =============================================

            if self.state == 'url_scraping':

                # =========================================
                # RECOVERY
                # =========================================

                if self.url_blocks_json:

                    _logger.warning(

                        "[URL RECOVERY] "

                        "USING SAVED BLOCKS"
                    )

                    self.stage_retry_count = 0

                    self.last_known_state = 'url_scraping'
                    self.state = 'url_scraping'

                    self._safe_commit_progress()

                    return


                _logger.warning(

                    f"[URL SCRAPE] "

                    f"SENDING TO APIFY "

                    f"| {self.data_url}"
                )


                previous_extract = bool(
                    self.extracted_text
                )


                result = self.parse_url()


                # =========================================
                # APIFY STILL PROCESSING
                # =========================================

                if result is True:

                    _logger.warning(

                        "[APIFY STATUS] "

                        "WAITING FOR RESPONSE"
                    )

                    return


                # =========================================
                # APIFY RESPONSE READY
                # =========================================

                _logger.warning(
                    "[APIFY STATUS] RESPONSE RECEIVED"
                )


                if (

                    self.extracted_text

                    and

                    not previous_extract

                ):

                    _logger.warning(
                        "URL EXTRACTION SUCCESS → url_ai"
                    )
                    self.stage_retry_count = 0

                    self.last_known_state = 'url_ai'
                    self.state = 'url_ai'

                    self._safe_commit_progress()

                    return


                # =========================================
                # FAILED
                # =========================================

                if self.state == 'failed':

                    _logger.warning(
                        "[URL SCRAPE FAILED]"
                    )

                    self._safe_commit_progress()

                    return


                _logger.warning(
                    "[URL SCRAPE] NO DATA EXTRACTED"
                )

                self.state = 'failed'

                self._safe_commit_progress()

                return


            # =============================================
            # AI
            # =============================================

            if self.state == 'url_ai':

                previous_batch = (
                    self.url_batch_index or 0
                )


                _logger.warning(

                    f"[URL AI START] "

                    f"batch={previous_batch}"
                )


                try:

                    self.send_to_openai_url()

                except Exception as e:

                    _logger.exception(

                        f"URL AI FAILED → {str(e)}"
                    )

                    self.stage_retry_count += 1

                    self.last_error = str(e)

                    self.last_known_state = 'url_ai'

                    self.state = 'review'

                    self._safe_commit_progress()

                    return


                new_batch = (
                    self.url_batch_index or 0
                )


                _logger.warning(

                    f"[URL AI CHECK] "

                    f"{previous_batch} -> {new_batch}"
                )


                # =========================================
                # PROGRESS DETECTED
                # =========================================

                if new_batch > previous_batch:

                    _logger.warning(
                        "[URL AI] PROGRESS SAVED"
                    )

                elif self.state != 'url_creating':

                    _logger.warning(
                        "[URL AI] NO PROGRESS DETECTED"
                    )

                    self.state = 'failed'


                self._safe_commit_progress()

                return


            # =============================================
            # CREATE
            # =============================================

            if self.state == 'url_creating':

                if not self.ai_response:

                    self.state = 'failed'

                    _logger.warning(
                        "URL CREATE FAILED → NO AI RESPONSE"
                    )


                    self._safe_commit_progress()

                    return


                previous_index = (
                    self.last_processed_product_index or 0
                )


                _logger.warning(

                    f"[URL CREATE START] "

                    f"{previous_index}"
                )


                try:

                    self.create_products_url()

                except Exception as e:

                    _logger.exception(

                        f"URL CREATE FAILED → {str(e)}"
                    )

                    self.stage_retry_count += 1

                    self.last_error = str(e)

                    self.last_known_state = 'url_creating'

                    self.state = 'review'

                    self._safe_commit_progress()

                    return


                new_index = (
                    self.last_processed_product_index or 0
                )


                _logger.warning(

                    f"[URL CREATE CHECK] "

                    f"{previous_index} -> {new_index}"
                )


                try:

                    total = len(
                        json.loads(
                            self.ai_response or "[]"
                        )
                    )

                except Exception:

                    total = 0


                _logger.warning(

                    f"[URL TOTAL PRODUCTS] "

                    f"{total}"
                )


                if new_index >= total:
                    self.stage_retry_count = 0
                   

                    if not self.completion_email_sent:

                        self.send_completion_email()
                   
                   
                    self.state = 'done'

                    _logger.warning(
                        "URL COMPLETE ✅"
                    )


                elif new_index > previous_index:

                    self.stage_retry_count = 0

                    self.last_known_state = 'url_creating'
                    self.state = 'url_creating'

                    _logger.warning(
                        "[URL CREATE] CONTINUE"
                    )

                else:

                    _logger.warning(
                        "[URL CREATE] NO PROGRESS"
                    )

                    self.state = 'failed'


                self._safe_commit_progress()

                return


        # =================================================
        # EXCEL FLOW
        # =================================================

        if self.excel_file and not self.pdf_file:

            _logger.warning(
                "FLOW → EXCEL"
            )


            # =============================================
            # START
            # =============================================

            if self.state == 'draft':

                self.stage_retry_count = 0

                self.last_known_state = 'excel_parsing'

                self.state = 'excel_parsing'

                self._safe_commit_progress()

                return


            # =============================================
            # PARSE
            # =============================================

            if self.state == 'excel_parsing':

                previous_index = (
                
                    self.excel_parse_index or 0
                )

                _logger.warning(

                    f"[EXCEL PARSE START] "

                    f"previous_index={previous_index}"
                )


                self.parse_excel()


                new_index = (
                   
                    self.excel_parse_index or 0
                )


                _logger.warning(

                    f"[EXCEL PARSE CHECK] "

                    f"{previous_index} -> {new_index}"
                )


                # ==========================================
                # NEW ROWS FOUND
                # ==========================================

                if new_index > previous_index:
                    
                    self.state = 'excel_ai'

                    _logger.warning(

                        "[EXCEL STATE CHANGE] "

                        f"{self.id} -> excel_ai"
                    )

                    _logger.warning(
                        "[EXCEL PARSE] NEW BATCH READY → excel_ai"
                    )

                    self._safe_commit_progress()

                    return


                else:

                    _logger.warning(
                        "[EXCEL PARSE] NO NEW ROWS"
                    )

                    if self.is_excel_parsed:

                        _logger.warning(
                            "[EXCEL PARSE COMPLETE] → excel_ai"
                        )

                        # successful stage transition
                      
                        self.stage_retry_count = 0

                        self.last_known_state = 'excel_ai'

                        self.state = 'excel_ai'

                        # cleanup optional URL enrichment state
                        self.excel_url_processing = False

                        self.excel_url_index = 0

                  
                    else:

                        _logger.warning(
                            "[EXCEL PARSE STALLED]"
                        )

                        self.stage_retry_count += 1

                        self.last_error = "Excel parse stalled"

                        self.last_known_state = 'excel_parsing'

                        self.state = 'review' 

                self._safe_commit_progress()

                return


            # =============================================
            # AI
            # =============================================

            if self.state == 'excel_ai':

                try:
                    previous_ai_index = self.excel_ai_index or 0

                    _logger.warning(

                        f"[EXCEL AI BEFORE] "

                        f"{previous_ai_index}"
                    )

                    self.send_to_openai_excel()

                    new_ai_index = self.excel_ai_index or 0

                    ai_progress_detected = (
                        new_ai_index > previous_ai_index
                    )
                                        
                    _logger.warning(

                        f"[EXCEL AI AFTER] "

                        f"{new_ai_index}"
                    )

                except Exception as e:

                    _logger.exception(

                        f"EXCEL AI FAILED → {str(e)}"
                    )

                    self.stage_retry_count += 1

                    self.last_error = str(e)

                    self.last_known_state = 'excel_ai'

                    self.state = 'review'

                    self._safe_commit_progress()

                    return


                try:

                    extracted_rows = json.loads(
                        self.extracted_text or "[]"
                    )

                    total_rows = len(extracted_rows)

                except Exception:

                    total_rows = 0


                _logger.warning(

                    f"[EXCEL AI TOTAL ROWS] "

                    f"{total_rows}"
                )


                _logger.warning(

                    f"[EXCEL AI STATE] "

                    f"{self.state}"
                )

                # ==========================================
                # AI COMPLETE → CREATE
                # ==========================================

                if (
                        self.state != 'failed'
                        and
                        (
                            ai_progress_detected
                            or
                            self.ai_response
                        )
                    ):

                    _logger.warning(
                        "[EXCEL AI COMPLETE] → excel_creating"
                    )

                    self.stage_retry_count = 0

                    self.last_known_state = 'excel_creating'

                    self.state = 'excel_creating'

                self.flush_recordset()
                self.env.cr.commit()

                return

            # =============================================
            # CREATE
            # =============================================


            #-----------EXCEL ROW PROCESSOR-------------
        
            if self.state == 'excel_creating':

                try:

                    self.create_products_excel()
                  
                except Exception as e:

                    _logger.exception(
                        f"EXCEL CREATE FAILED → {str(e)}"
                    )

                    self.stage_retry_count += 1

                    self.last_error = str(e)

                    self.last_known_state = 'excel_creating'

                    self.state = 'review'

                    self._safe_commit_progress()

                    return

                _logger.warning(

                    f"[EXCEL CREATE STATE] "

                    f"{self.state}"
                )

                self.flush_recordset()
                self.env.cr.commit()


            # =============================================
            # RETURN
            # =============================================

            if self.state == 'excel_creating' \
                    or self.excel_url_processing:

                return
        
        # =================================================
        # PDF FLOW
        # =================================================

        elif self.pdf_file:

            _logger.warning(
                "FLOW → PDF"
            )


            # =============================================
            # START
            # =============================================

            if self.state == 'draft':

                self.stage_retry_count = 0

                self.last_known_state = 'pdf_extracting'

                self.state = 'pdf_extracting'

                self._safe_commit_progress()

                return


            # =============================================
            # EXTRACT
            # =============================================

            if self.state == 'pdf_extracting':

                try:

                    self.extract_pdf()

                except Exception as e:

                    _logger.exception(

                        f"PDF EXTRACT FAILED → {str(e)}"
                    )

                    self.stage_retry_count += 1

                    self.last_error = str(e)

                    self.last_known_state = 'pdf_extracting'

                    self.state = 'review'

                    self._safe_commit_progress()

                    return


                if (

                    (self.current_page or 0)

                    <

                    (self.total_pages or 0)

                ):

                    _logger.warning(

                        f"PDF EXTRACTION CONTINUES "

                        f"→ PAGE "

                        f"{self.current_page}/"

                        f"{self.total_pages}"
                    )

                    self.state = 'pdf_extracting'

                else:

                    _logger.warning(
                        "PDF EXTRACTION COMPLETE → pdf_ai"
                    )

                    self.state = 'pdf_ai'


                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # PDF AI
            # =============================================

            if self.state == 'pdf_ai':

                try:

                    self.send_to_openai_pdf()

                except Exception as e:

                    _logger.exception(

                       f"PDF AI FAILED → {str(e)}"
                    )

                    self.stage_retry_count += 1

                    self.last_error = str(e)

                    self.last_known_state = 'pdf_ai'

                    self.state = 'review'

                    self._safe_commit_progress()

                    return


                page_total = self.env[
                    'vendor.import.page'
                ].search_count([

                    ('job_id', '=', self.id)

                ])


                _logger.warning(

                    f"[PDF AI CHECK] "

                    f"{self.last_ai_page}/"

                    f"{page_total}"
                )


                if (

                    (self.last_ai_page or 0)

                    <

                    page_total

                ):

                    _logger.warning(

                        f"PDF AI CONTINUES "

                        f"→ {self.last_ai_page}/"

                        f"{page_total}"
                    )

                    self.stage_retry_count = 0

                    self.last_known_state = 'pdf_ai'

                    self.state = 'pdf_ai'

                else:

                    _logger.warning(
                        "PDF AI COMPLETE → pdf_creating"
                    )

                    self.stage_retry_count = 0

                    self.last_known_state = 'pdf_creating'
                    
                    self.state = 'pdf_creating'


                self.flush_recordset()
                self.env.cr.commit()

                return


            # =============================================
            # PDF CREATE
            # =============================================

            if self.state == 'pdf_creating':

                try:

                    self.create_products_pdf()

                except Exception as e:

                    _logger.exception(
                        f"PDF CREATE FAILED → {str(e)}"
                    )

                    self.stage_retry_count += 1

                    self.last_error = str(e)

                    self.last_known_state = 'pdf_creating'

                    self.state = 'review'

                    self._safe_commit_progress()

                    return

                try:

                    total_ai_pages = len(
                        json.loads(
                            self.ai_response or "[]"
                        )
                    )

                except Exception:

                    total_ai_pages = 0


                if (

                    (self.last_created_page or 0)

                    <

                    total_ai_pages

                ):

                    _logger.warning(

                        f"PDF CREATE CONTINUES "

                        f"→ {self.last_created_page}/"

                        f"{total_ai_pages}"
                    )

                    self.stage_retry_count = 0

                    self.last_known_state = 'pdf_creating'
                    self.state = 'pdf_creating'

                else:

                    _logger.warning(
                        "PDF COMPLETE ✅"
                    )

                    self.stage_retry_count = 0
                    self.state = 'done'

                    if not self.completion_email_sent:

                        self.send_completion_email()


                self.flush_recordset()
                self.env.cr.commit()

                return

     #=========Variant swatch 1======================================
   
    #=========Variant color swatch logic===========================================

    COLOR_HEX_MAP = {

        # =====================================
        # BASIC COLORS
        # =====================================

        "black": "#000000",
        "white": "#FFFFFF",
        "red": "#FF0000",
        "green": "#008000",
        "blue": "#0066CC",
        "yellow": "#FFD700",
        "orange": "#FF6600",
        "purple": "#800080",
        "pink": "#FF69B4",
        "brown": "#8B4513",

        # =====================================
        # GREY / GRAY FAMILY
        # =====================================

        "grey": "#808080",
        "gray": "#808080",
        "light grey": "#D3D3D3",
        "light gray": "#D3D3D3",
        "dark grey": "#555555",
        "dark gray": "#555555",
        "charcoal": "#36454F",
        "charcoal grey": "#36454F",
        "charcoal gray": "#36454F",
        "ash": "#B2BEB5",
        "ash grey": "#B2BEB5",
        "ash gray": "#B2BEB5",
        "heather": "#9AA0A6",
        "heather grey": "#A9A9A9",
        "heather gray": "#A9A9A9",
        "grey marl": "#A9A9A9",
        "gray marl": "#A9A9A9",

        # =====================================
        # BLUE FAMILY
        # =====================================

        "navy": "#000080",
        "navy blue": "#000080",
        "royal blue": "#4169E1",
        "light blue": "#ADD8E6",
        "sky blue": "#87CEEB",
        "ice blue": "#BFEFFF",
        "powder blue": "#B0E0E6",

        # =====================================
        # GREEN FAMILY
        # =====================================

        "lime": "#32CD32",
        "lime green": "#32CD32",
        "neon green": "#39FF14",

        # =====================================
        # WHITE / NATURAL FAMILY
        # =====================================

        "off white": "#F8F8F8",
        "natural": "#F5F5DC",
        "cream": "#FFFDD0",
        "ivory": "#FFFFF0",
        "stone": "#D2C29D",
        "sand": "#C2B280",
        "beige": "#F5F5DC",

        # =====================================
        # METALLIC COLORS
        # =====================================

        "silver": "#C0C0C0",
        "gold": "#D4AF37",

        # =====================================
        # SPECIAL COLORS
        # =====================================

        "burgundy": "#800020",
        "wine": "#722F37",

        # =====================================
        # MULTI COLORS
        # =====================================

        "multi": "#CCCCCC",
        "multi-color": "#CCCCCC",
        "multicolor": "#CCCCCC",
    }


    # =========================================
    # REUSABLE ATTRIBUTE ENGINE
    # =========================================

    def _get_or_create_attribute_and_value(

        self,

        attr_name,

        attr_value
    ):

        attr_name = str(
            attr_name or ""
        ).strip()

        attr_value = str(
            attr_value or ""
        ).strip()

        # =====================================
        # ATTRIBUTE NORMALIZATION
        # =====================================
        normalized_attr = attr_name.lower().strip()

        is_color_attribute = normalized_attr in [

            'color',

            'colour',

            'colors',

            'colourway',

            'color name'
        ]

        attribute = self.env[
            'product.attribute'
        ].search([

            ('name', '=', attr_name)

        ], limit=1)

        # =====================================
        # CREATE ATTRIBUTE
        # =====================================

        if not attribute:

            attribute_vals = {

                'name': attr_name
            }

            # =====================================
            # COLOR SWATCH SUPPORT
            # =====================================

            if is_color_attribute:

                attribute_vals[
                    'display_type'
                ] = 'color'

            attribute = self.env[
                'product.attribute'
            ].create(attribute_vals)
            

            _logger.warning(

                f"[ATTRIBUTE CREATED] "

                f"{attr_name}"
            )


        # =====================================
        # FORCE EXISTING COLOR ATTRIBUTE
        # INTO SWATCH MODE
        # =====================================


        if (

            is_color_attribute

            and

            attribute.display_type != 'color'
        ):

            attribute.display_type = 'color'

            _logger.warning(

                "[COLOR ATTRIBUTE UPDATED] "

                f"{attribute.name}"
            )

        # =====================================
        # CREATE VALUE
        # =====================================

        value = self.env[
            'product.attribute.value'
        ].search([

            ('name', '=', attr_value),

            ('attribute_id', '=', attribute.id)

        ], limit=1)

        if not value:

            value_vals = {

                'name': attr_value,

                'attribute_id': attribute.id
            }

            # =====================================
            # HTML COLOR SUPPORT
            # =====================================

            if is_color_attribute:

                # =====================================
                # SMART COLOR NORMALIZATION
                # =====================================

                normalized_color = " ".join(

                    attr_value
                    .lower()
                    .replace("-", " ")
                    .replace("_", " ")
                    .split()
                )

                # =====================================
                # COLOR ALIAS NORMALIZATION
                # =====================================

                COLOR_ALIASES = {

                    'lt blue': 'light blue',
                    'dk blue': 'navy blue',
                    'dk navy': 'navy blue',
                    'royal': 'royal blue',
                    'lime': 'lime green',
                    'charcoal marl': 'charcoal',
                    'heather navy': 'navy blue',
                    'heather blue': 'blue',
                    'heather grey': 'grey',
                    'heather gray': 'gray',
                    'sky': 'sky blue',
                    'off white': 'white',
                    'natural': 'beige',
                }

                normalized_color = COLOR_ALIASES.get(

                    normalized_color,

                    normalized_color
                )

                color_hex = None

                _logger.warning(

                    f"[COLOR NORMALIZED] "

                    f"raw={attr_value} "

                    f"normalized={normalized_color}"
                )

                # =====================================
                # DIRECT MATCH
                # =====================================

                color_hex = self.COLOR_HEX_MAP.get(
                    normalized_color
                )

                # ======================================
                # PARTIAL MATCH FALLBACK
                # ======================================

                if not color_hex:

                    # =====================================
                    # REMOVE SECONDARY COLOR CONTEXT
                    # =====================================

                    primary_color_text = normalized_color

                    split_keywords = [

                        " with ",
                        " trim",
                        " piping",
                        " contrast",
                        "/",
                        "&",
                        ","
                    ]

                    for splitter in split_keywords:

                        if splitter in primary_color_text:

                            primary_color_text = (
                                primary_color_text
                                .split(splitter)[0]
                                .strip()
                            )

                    _logger.warning(

                        f"[PRIMARY COLOR PARSED] "

                        f"raw={normalized_color} "

                        f"primary={primary_color_text}"
                    )

                    # =====================================
                    # LONGEST COLOR MATCH PRIORITY
                    # =====================================

                    best_match = None

                    best_length = 0

                    for key, hex_value in self.COLOR_HEX_MAP.items():

                        if key in primary_color_text:

                            if len(key) > best_length:

                                best_match = (key, hex_value)

                                best_length = len(key)

                    if best_match:

                        matched_key, matched_value = best_match

                        color_hex = matched_value

                        _logger.warning(

                            f"[BEST COLOR MATCH] "

                            f"{attr_value} "

                            f"→ {matched_key} "

                            f"→ {matched_value}"
                        )

                # =====================================
                # SAFE FALLBACK COLORS
                # =====================================

                if not color_hex:

                    if "white" in normalized_color:

                        color_hex = "#F8F8F8"

                    elif "grey" in normalized_color \
                            or "gray" in normalized_color:

                        color_hex = "#808080"

                    elif "black" in normalized_color:

                        color_hex = "#000000"

                    elif "navy" in normalized_color:

                        color_hex = "#000080"

                    elif "blue" in normalized_color:

                        color_hex = "#0066CC"

                    elif "green" in normalized_color:

                        color_hex = "#008000"

                    elif "red" in normalized_color:

                        color_hex = "#FF0000"

                    elif "yellow" in normalized_color:

                        color_hex = "#FFD700"

                    elif "purple" in normalized_color:

                        color_hex = "#800080"

                    elif "orange" in normalized_color:

                        color_hex = "#FF6600"

                # =====================================
                # APPLY HTML COLOR
                # =====================================

                if color_hex:

                    value_vals[
                        'html_color'
                    ] = color_hex

                    _logger.warning(

                        f"[COLOR HEX ASSIGNED] "

                        f"{attr_value} "

                        f"→ {color_hex}"
                    )

                else:

                    _logger.warning(

                        f"[COLOR HEX MISSING] "

                        f"{attr_value}"
                    )


            value = self.env[
                'product.attribute.value'
            ].create(value_vals)

        # =====================================
        # PATCH EXISTING COLOR VALUES
        # =====================================

        elif is_color_attribute:

            existing_html = (
                value.html_color or ""
            ).strip()

            if not existing_html:

                normalized_color = " ".join(

                    attr_value
                    .lower()
                    .replace("-", " ")
                    .replace("_", " ")
                    .split()
                )

                COLOR_ALIASES = {

                    'lt blue': 'light blue',
                    'dk blue': 'navy blue',
                    'dk navy': 'navy blue',
                    'royal': 'royal blue',
                    'lime': 'lime green',
                    'charcoal marl': 'charcoal',
                    'heather navy': 'navy blue',
                    'heather blue': 'blue',
                    'heather grey': 'grey',
                    'heather gray': 'gray',
                    'sky': 'sky blue',
                    'off white': 'white',
                    'natural': 'beige',
                }

                normalized_color = COLOR_ALIASES.get(

                    normalized_color,

                    normalized_color
                )

                color_hex = self.COLOR_HEX_MAP.get(
                    normalized_color
                )

                # =====================================
                # FALLBACK PARTIAL MATCH
                # =====================================

                if not color_hex:

                    best_match = None
                    best_length = 0

                    for key, hex_value in self.COLOR_HEX_MAP.items():

                        if key in normalized_color:

                            if len(key) > best_length:

                                best_match = (
                                    key,
                                    hex_value
                                )

                                best_length = len(key)

                    if best_match:

                        matched_key, matched_value = best_match

                        color_hex = matched_value

                        _logger.warning(

                            f"[PATCH COLOR MATCH] "

                            f"{attr_value} "

                            f"→ {matched_key} "

                            f"→ {matched_value}"
                        )

                # =====================================
                # APPLY PATCHED HTML COLOR
                # =====================================

                if color_hex:

                    value.write({

                        'html_color': color_hex
                    })

                    _logger.warning(

                        f"[PATCH EXISTING COLOR] "

                        f"{attr_value} "

                        f"→ {color_hex}"
                    )

                else:

                    _logger.warning(

                        f"[PATCH FAILED NO HEX] "

                        f"{attr_value}"
                    )

        return attribute, value


    #------------parse url-----------------------------------

    def parse_url(self):

        import json

        _logger.warning(f"APIFY SCRAPE → {self.data_url}")

        raw_data = self._run_apify_actor(self.data_url)

        # =====================================================
        # APIFY STILL RUNNING
        # =====================================================

        if raw_data is None:

            _logger.warning(
                "APIFY NOT READY → WAIT NEXT CRON"
            )

            self.state = "url_scraping"

            return True

        # =====================================================
        # EMPTY RAW RESPONSE
        # =====================================================

        if not raw_data:

            _logger.error(
                "APIFY FAILED → EMPTY DATASET"
            )

            self.state = "failed"

            return

        # =====================================================
        # SAFE DEBUG LOG
        # =====================================================

        try:

            _logger.warning(
                f"RAW APIFY ITEMS → {len(raw_data)}"
            )

        except Exception:
            pass

        # =====================================================
        # HANDLE STRUCTURED RESPONSES
        # =====================================================

        first = raw_data[0] if raw_data else {}

        response_type = first.get("type")

        # =====================================================
        # BLOCKED
        # =====================================================

        if response_type == "BLOCKED":

            reason = first.get(
                "reason",
                "Unknown block detected"
            )

            status_code = first.get(
                "status_code"
            )

            _logger.error(
                f"URL BLOCKED → {reason}"
            )

            if status_code:
                _logger.error(
                    f"BLOCK STATUS CODE → {status_code}"
                )

            self.state = "failed"

            return

        # =====================================================
        # EMPTY
        # =====================================================

        if response_type == "EMPTY":

            reason = first.get(
                "reason",
                "No products extracted"
            )

            debug = first.get("debug", {})

            _logger.error(
                f"URL EXTRACTION EMPTY → {reason}"
            )

            if debug:

                _logger.error(
                    f"PAGE TITLE → {debug.get('title')}"
                )

                _logger.error(
                    f"IMAGES FOUND → {debug.get('images_found')}"
                )

                _logger.error(
                    f"LINKS FOUND → {debug.get('links_found')}"
                )

                _logger.error(
                    f"POSSIBLE PRODUCT BLOCKS → "
                    f"{debug.get('possible_product_blocks')}"
                )

                _logger.error(
                    f"COOKIE DETECTED → "
                    f"{debug.get('cookie_detected')}"
                )

                preview = debug.get(
                    'body_preview',
                    ''
                )

                _logger.error(
                    f"BODY PREVIEW → {preview[:300]}"
                )

            self.state = "failed"

            return

        # ===================================================
        # PRODUCTS
        # ===================================================

        structured_data = []

        for block in raw_data:

            # ==============================================
            # FORMAT 1 → ORIGINAL EB FORMAT
            # ==============================================

            if block.get("text"):

                structured_data.append({
                    "text": block.get("text"),
                    "image": block.get("image")
                })

                continue

            # ==============================================
            # FORMAT 2 → STRUCTURED FORMAT
            # ==============================================

            if block.get("type") == "PRODUCTS":

                items = block.get("items", [])

                if not items:
                    continue

                structured_data.extend(items)

        # =====================================================
        # NO PRODUCTS AFTER PARSE
        # =====================================================

        if not structured_data:

            _logger.error(
                "NO VALID PRODUCTS FOUND AFTER PARSING"
            )

            self.state = "failed"

            return


        # ============================================
        # URL BATCHING
        # ============================================

        BATCH_SIZE = 40

        start = (
            self.url_parse_index or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(structured_data)
        )

        total_structured = len(structured_data)

        structured_data = structured_data[
            start:end
        ]


        _logger.warning(

            f"[URL PARSE BATCH] "

            f"{start} -> {end} "

            f"| total={len(normalized if 'normalized' in locals() else structured_data)}"
        )


        # =====================================================
        # NORMALIZE
        # =====================================================

        normalized = self._normalize_url_data(
            structured_data
        )

        # ============================================
        # SAVE URL PARSE PROGRESS
        # ============================================

        self.url_parse_index = end


        _logger.warning(

            f"[URL PARSE SAVE] "

            f"{self.url_parse_index}"
        )


        if not normalized:

            _logger.error(
                "NORMALIZATION FAILED → EMPTY DATA"
            )

            self.state = "failed"

            return

        # =====================================================
        # STORE SAFELY
        # =====================================================

        #payload = json.dumps(normalized)

        # 🔥 LIMIT STORAGE SIZE
        #self.extracted_text = payload[:50000]

        payload = json.dumps(normalized)

        # ============================================
        # PERSIST URL BLOCKS
        # ============================================

        self.url_blocks_json = payload

        # compatibility
        self.extracted_text = payload[:50000]


        _logger.warning(

            f"[URL STORE] "

            f"saved_blocks={len(normalized)}"
        )


        _logger.warning(
            f"APIFY DONE → {len(normalized)} ITEMS"
        )

        # =====================================================
        # MOVE TO NEXT STEP
        # =====================================================

        if self.url_parse_index >= total_structured:
            _logger.warning(
                "[URL PARSE] FINAL BATCH READY"
            )

        else:

            _logger.warning(
                "[URL PARSE] MORE BATCHES REMAIN"
            )


        self.state = "url_ai"

    # ======================================================
    # LIGHTWEIGHT URL ENRICHMENT (excel url backup)
    # ======================================================

    def _extract_url_product_data(

        self,

        product_url
    ):

        try:

            _logger.warning(

                f"[URL ENRICHMENT START] "

                f"{product_url}"
            )

            raw_data = self._run_apify_actor(
                product_url
            )

            # =========================================
            # APIFY STILL RUNNING
            # =========================================

            if raw_data is None:

                _logger.warning(

                    "[EXCEL URL ENRICHMENT] "

                    "APIFY NOT READY"
                )

                return {}

            if not raw_data:

                _logger.warning(

                    "[EXCEL URL ENRICHMENT] "

                    "EMPTY RESPONSE"
                )

                return {}

            first = raw_data[0] if raw_data else {}

            # =========================================
            # BLOCKED / EMPTY
            # =========================================

            if first.get("type") in [

                "BLOCKED",

                "EMPTY"
            ]:

                _logger.warning(

                    f"[EXCEL URL ENRICHMENT] "

                    f"INVALID TYPE "

                    f"{first.get('type')}"
                )

                return {}

            # =========================================
            # NORMALIZE
            # =========================================

            structured_data = []

            for block in raw_data:

                if block.get("text"):

                    structured_data.append({

                        "text": block.get("text"),

                        "image": block.get("image")
                    })

                    continue

                if block.get("type") == "PRODUCTS":

                    structured_data.extend(

                        block.get("items", [])
                    )

            normalized = self._normalize_url_data(
                structured_data
            )

            if not normalized:

                return {}

            first_product = normalized[0]

            _logger.warning(

                f"[EXCEL URL ENRICHMENT SUCCESS] "

                f"{product_url}"
            )

            return {

                "name":
                    first_product.get("name"),

                "description":
                    first_product.get(
                        "description"
                    ),

                "category":
                    first_product.get(
                        "category"
                    ),

                "images":
                    first_product.get(
                        "images",
                        []
                    )
            }

        except Exception as e:

            _logger.warning(

                f"[EXCEL URL ENRICHMENT FAILED] "

                f"{str(e)}"
            )

            return {}


    #------excel parsing method---------------
    
    def parse_excel(self):

        _logger.warning(
            "EXCEL → START PARSING (BATCH MODE)"
        )

        excel_bytes = base64.b64decode(
            self.excel_file
        )

        wb = load_workbook(
            filename=BytesIO(excel_bytes)
        )

        headers = {
            "User-Agent": "Mozilla/5.0"
        }

        pages = []

        # =====================================
        # SAFE BATCH CONTROL
        # =====================================

        BATCH_SIZE = 8
        MAX_BUFFER = 150


        start_index = (
            self.excel_parse_index or 0
        )


        current_count = 0

        global_index = 0

        _logger.warning(
            f"EXCEL RESUME FROM INDEX "
            f"→ {start_index}"
        )


        # =====================================
        # TOTAL ROWS
        # =====================================

        total_rows = 0

        for sheet in wb.worksheets:

            for idx, row in enumerate(

                sheet.iter_rows()

            ):

                if idx == 0:
                    continue

                row_text_parts = [

                    str(cell.value or "").strip()

                    for cell in row

                    if str(
                        cell.value or ""
                    ).strip()
                ]

                if not row_text_parts:
                    continue

                total_rows += 1


        _logger.warning(
            f"[DEBUG] REAL TOTAL ROWS "
            f"→ {total_rows}"
        )


        # =====================================
        # MAIN LOOP
        # =====================================

        for sheet in wb.worksheets:

            _logger.warning(
                f"PROCESSING SHEET → "
                f"{sheet.title}"
            )

            image_loader = SheetImageLoader(
                sheet
            )


            for idx, row in enumerate(

                sheet.iter_rows()

            ):

                if idx == 0:
                    continue


                row_text_parts = [

                    str(cell.value or "").strip()

                    for cell in row

                    if str(
                        cell.value or ""
                    ).strip()
                ]


                if not row_text_parts:
                    continue

                # =================================
                # GLOBAL INDEX TRACKING
                # =================================

                global_index += 1


                # =================================
                # SKIP OLD ROWS
                # =================================

                if global_index <= start_index:
                    continue


                # =================================
                # BATCH LIMIT
                # =================================

                if current_count >= BATCH_SIZE:

                    _logger.warning(
                        "BATCH LIMIT REACHED "
                        "→ NEXT CRON"
                    )

                    break


                # =================================
                # PRICE/STOCK DETECTION
                # =================================

                price = ""
                stock = ""

                numeric_candidates = []


                for col_idx, cell in enumerate(row):

                    raw_val = str(
                        cell.value or ""
                    ).strip()

                    if not raw_val:
                        continue


                    # skip ranges
                    if (
                        "-" in raw_val
                        and not raw_val.startswith("-")
                    ):
                        continue


                    try:

                        clean = raw_val.replace(
                            ",",
                            "."
                        )

                        clean = re.sub(

                            r"[^\d.]",

                            "",

                            clean
                        )


                        if not clean:
                            continue


                        num = float(clean)

                        is_real_decimal = False


                        if "." in clean:

                            decimal_part = (
                                clean.split(".")[-1]
                            )

                            if decimal_part not in [

                                "0",

                                "00"
                            ]:

                                is_real_decimal = True


                        numeric_candidates.append({

                            "col": col_idx,

                            "num": num,

                            "raw": raw_val,

                            "is_decimal":
                                is_real_decimal
                        })

                    except:
                        continue


                # =================================
                # PRICE
                # =================================

                price_candidates = [

                    x for x in numeric_candidates

                    if (

                        x["is_decimal"]

                        and

                        0 < x["num"] < 1000
                    )
                ]


                best_price = None


                if price_candidates:

                    best_price = sorted(

                        price_candidates,

                        key=lambda x: x["col"]

                    )[-1]

                    price = str(
                        best_price["num"]
                    )


                # =================================
                # STOCK
                # =================================

                if best_price:

                    price_col = (
                        best_price["col"]
                    )

                    stock_candidates = []


                    for item in numeric_candidates:

                        if item["is_decimal"]:
                            continue


                        val = item["num"]


                        if val > 9999:
                            continue


                        if item["col"] >= price_col:
                            continue


                        stock_candidates.append(
                            item
                        )


                    if stock_candidates:

                        best_stock = max(

                            stock_candidates,

                            key=lambda x: x["num"]
                        )

                        stock = str(

                            int(
                                best_stock["num"]
                            )
                        )


                _logger.warning(
                    f'''
                    EXCEL RAW ROW →

                    TEXT=
                    {" | ".join(row_text_parts)}

                    PRICE={price}

                    STOCK={stock}
                    '''
                )


                # =================================
                # ROW TEXT
                # =================================

                row_text = f"""
                ROW_DATA:
                {" | ".join(row_text_parts)}

                RULE:
                - THIS IS EXACTLY ONE PRODUCT
                - DO NOT SPLIT THIS ROW
                - THIS ROW MAY BE A VARIANT
                - USE SIMILAR ID/SKU
                """

                row_images = []


                # ==================================
                # EMBEDDED IMAGE
                # ==================================

                for cell in row:

                    try:

                        if image_loader.image_in(
                            cell.coordinate
                        ):

                            pil_img = (
                                image_loader.get(
                                    cell.coordinate
                                )
                            )

                            buffer = BytesIO()

                            pil_img.save(
                                buffer,
                                format="JPEG"
                            )

                            img_base64 = (
                                base64.b64encode(

                                    buffer.getvalue()

                                ).decode("utf-8")
                            )

                            row_images.append(
                                img_base64
                            )

                            break

                    except:
                        continue


                # =================================
                # URL IMAGE
                # =================================

                if not row_images:

                    for cell in row:

                        val = str(
                            cell.value or ""
                        ).strip()

                        if val.startswith("http"):

                            try:

                                response = requests.get(

                                    val,

                                    headers=headers,

                                    timeout=5
                                )

                                if (

                                    response.status_code
                                    == 200

                                    and

                                    "image"

                                    in response.headers.get(
                                        "Content-Type",
                                        ""
                                    )
                                ):

                                    img_base64 = (
                                        base64.b64encode(

                                            response.content

                                        ).decode("utf-8")
                                    )

                                    row_images.append(
                                        img_base64
                                    )

                                    break

                            except:
                                continue


                # =================================
                # STORE
                # =================================

                pages.append({

                    "page": global_index,

                    "text": row_text,

                    "images": row_images,

                    "row_index": global_index,

                    "price": price,

                    "stock": stock,
                })


                current_count += 1


                # =================================
                # MEMORY SAFETY
                # =================================

                if len(pages) >= MAX_BUFFER:

                    _logger.warning(
                        f"EXCEL SAFETY BREAK "
                        f"→ {len(pages)} rows"
                    )

                    break


            if (
                current_count >= BATCH_SIZE
                or
                len(pages) >= MAX_BUFFER
            ):
                break


        # =====================================
        # STORE
        # =====================================

        existing = []

        if self.extracted_text:

            try:

                existing = json.loads(
                    self.extracted_text
                )

            except:
                existing = []


        # combined = existing + pages
        
        existing_map = {

            item.get("row_index"): item

            for item in existing
        }


        for item in pages:

            existing_map[
                item.get("row_index")
            ] = item


        combined = sorted(

            existing_map.values(),

            key=lambda x: x.get(
                "row_index",
                0
            )
        )
    

        self.extracted_text = json.dumps(
            combined
        )


        # =====================================
        # SAVE PROGRESS
        # =====================================

        new_index = (
            start_index
            +
            current_count
        )

        self.excel_parse_index = (
            new_index
        )

        _logger.warning(

            f"[EXCEL PARSE INDEX SAVE] "

            f"{self.excel_parse_index}"
        )


        # =====================================
        # DEBUG
        # =====================================

        remaining = max(
            total_rows - new_index,
            0
        )

        progress = round(

            (new_index / total_rows) * 100,

            2

        ) if total_rows else 0


        _logger.warning(
            f"[DEBUG] CURRENT INDEX "
            f"→ {new_index}"
        )

        _logger.warning(
            f"[DEBUG] REMAINING ROWS "
            f"→ {remaining}"
        )

        _logger.warning(
            f"[DEBUG] PROGRESS "
            f"→ {progress}%"
        )

        _logger.warning(
            f"EXCEL NEW INDEX "
            f"→ {new_index}"
        )

        _logger.warning(
            f"EXCEL BATCH STORED "
            f"→ {len(pages)} rows"
        )


        # =====================================
        # COMPLETION FLAG ONLY
        # =====================================

        if new_index >= total_rows:

            _logger.warning(
                "EXCEL → PARSING COMPLETED ✅"
            )

            self.is_excel_parsed = True

        else:

            _logger.warning(
                "EXCEL → MORE DATA REMAIN "
                "→ NEXT CRON"
            )

       
        wb.close()


    # =====================================================
    # REMOVE TEXT AREAS
    # =====================================================

    def _trim_catalog_whitespace(self, pil_image):
        
        original_image = pil_image

        try:

            # =====================================
            # SAFE BACKGROUND ESTIMATION
            # =====================================

            corners = [

                pil_image.getpixel((0, 0)),
                pil_image.getpixel((pil_image.width - 1, 0)),
                pil_image.getpixel((0, pil_image.height - 1)),
                pil_image.getpixel((
                    pil_image.width - 1,
                    pil_image.height - 1
                )),
            ]

            # average corner color
            avg_corner = tuple(

                int(sum(c[i] for c in corners) / 4)

                for i in range(len(corners[0]))
            )

            bg = Image.new(
                pil_image.mode,
                pil_image.size,
                avg_corner
            )

            diff = ImageChops.difference(
                pil_image,
                bg
            )

            # =====================================
            # REDUCE OVER-TRIMMING
            # =====================================

            # convert to grayscale for stable trim mask
            diff = diff.convert("L")

            diff = diff.point(

                lambda p: 255 if p > 18 else 0
            )

            bbox = diff.getbbox()

            if bbox:

                left, top, right, bottom = bbox

                padding = 12

                left = max(0, left - padding)
                top = max(0, top - padding)

                right = min(
                    pil_image.width,
                    right + padding
                )

                bottom = min(
                    pil_image.height,
                    bottom + padding
                )

                pil_image = pil_image.crop(

                    (left, top, right, bottom)
                )

            # =====================================
            # POST-TRIM SAFETY
            # =====================================

            if (
                pil_image.width < 60
                or
                pil_image.height < 60
            ):

                return None

            return pil_image

        except Exception as e:

            _logger.warning(

                f"[TRIM FAILED] {str(e)}"
            )

            return original_image


    # =====================================================
    # SEGMENT CATALOG PAGE INTO CLEAN PRODUCT ASSETS
    # =====================================================

    def _segment_catalog_images(self, images):

        segmented_images = []

        if not images:
            return segmented_images

        for img_b64 in images:

            try:

                img_data = base64.b64decode(img_b64)

                pil_image = Image.open(
                    BytesIO(img_data)
                ).convert("RGB")

                original_width, original_height = pil_image.size

                # =========================================
                # CONVERT TO OPENCV
                # =========================================

                cv_image = cv2.cvtColor(
                    np.array(pil_image),
                    cv2.COLOR_RGB2BGR
                )

                gray = cv2.cvtColor(
                    cv_image,
                    cv2.COLOR_BGR2GRAY
                )

                # =========================================
                # THRESHOLD
                # =========================================

                _, thresh = cv2.threshold(
                    gray,
                    245,
                    255,
                    cv2.THRESH_BINARY_INV
                )

                # =========================================
                # DILATION
                # =========================================

                kernel = cv2.getStructuringElement(
                    cv2.MORPH_RECT,
                    (9, 9)
                )

                dilated = cv2.dilate(
                    thresh,
                    kernel,
                    iterations=2
                )

                # =========================================
                # FIND CONTOURS
                # =========================================

                contours, _ = cv2.findContours(
                    dilated,
                    cv2.RETR_EXTERNAL,
                    cv2.CHAIN_APPROX_SIMPLE
                )

                filtered_contours = []

                for contour in contours:

                    area = cv2.contourArea(contour)

                    if area < 2500:
                        continue

                    x, y, w, h = cv2.boundingRect(contour)

                    # reject ultra-thin text columns
                    if w < 65 or h < 65:
                        continue

                    ratio = w / float(h)

                    # reject long text strips
                    if ratio > 4.5 or ratio < 0.22:
                        continue

                    filtered_contours.append(contour)

                contours = filtered_contours[:40]

                candidate_crops = []

                for contour in contours:

                    x, y, w, h = cv2.boundingRect(contour)

                    # =====================================
                    # SIZE FILTERS
                    # =====================================

                    if w < 120 or h < 120:
                        continue

                    # reject huge full page
                    if (
                        w > original_width * 0.95
                        and
                        h > original_height * 0.95
                    ):
                        continue

                    area = w * h

                    # reject tiny fragments
                    if area < 25000:
                        continue

                    # =====================================
                    # CROP
                    # =====================================

                    pad = 12

                    x1 = max(x - pad, 0)
                    y1 = max(y - pad, 0)
                    x2 = min(x + w + pad, original_width)
                    y2 = min(y + h + pad, original_height)

                    crop = pil_image.crop(
                        (x1, y1, x2, y2)
                    )

                    crop = self._trim_catalog_whitespace(
                        crop
                    )

                    # =====================================
                    # VALIDATE
                    # =====================================

                    if not self._is_valid_product_crop(crop):
                        continue

                    # =====================================
                    # OCR-LIKE TEXT REJECTION
                    # =====================================

                    crop_gray = crop.convert("L")

                    crop_arr = np.array(crop_gray)

                    dark_pixels = np.mean(
                        crop_arr < 90
                    )

                    if dark_pixels < 0.01:
                        continue

                    # =====================================
                    # IMAGE ANALYSIS
                    # =====================================

                    crop_width, crop_height = crop.size

                    crop_area = crop_width * crop_height

                    page_area = (
                        original_width * original_height
                    )

                    coverage_ratio = (
                        crop_area / float(page_area)
                    )

                    # =====================================
                    # COLLAGE DETECTION
                    # =====================================

                    is_collage = False

                    if len(filtered_contours) >= 6:

                        is_collage = True

                    # =====================================
                    # CENTER DETECTION
                    # =====================================

                    centered_object = False

                    crop_center_x = x + (w / 2.0)
                    crop_center_y = y + (h / 2.0)

                    page_center_x = (
                        original_width / 2.0
                    )

                    page_center_y = (
                        original_height / 2.0
                    )

                    distance_x = abs(
                        crop_center_x - page_center_x
                    )

                    distance_y = abs(
                        crop_center_y - page_center_y
                    )

                    if (

                        distance_x < original_width * 0.18

                        and

                        distance_y < original_height * 0.18
                    ):

                        centered_object = True
                   
                    # =====================================
                    # HERO SCORE
                    # =====================================

                    human_penalty = 0

                    hero_score = 0
                    gallery_score = 0

                    # =====================================
                    # HERO IMAGE PRIORITY
                    # =====================================

                    # big clean product bonus
                    hero_score += int(
                        coverage_ratio * 140
                    )

                    # centered ecommerce product
                    if centered_object:
                        hero_score += 55

                    # portrait product bonus
                    if crop_height > crop_width:
                        hero_score += 18

                    # =====================================
                    # EDGE DENSITY
                    # =====================================

                    edge_density = cv2.Canny(
                        crop_arr,
                        80,
                        160
                    ).mean()

                    # =====================================
                    # BACKGROUND ANALYSIS
                    # =====================================

                    background_ratio = np.mean(
                        crop_arr > 235
                    )

                    # =====================================
                    # CLEAN HERO DETECTION
                    # =====================================

                    # strong ecommerce isolated render
                    if (
                        centered_object
                        and
                        background_ratio > 0.45
                        and
                        not is_collage
                    ):
                        hero_score += 120

                    # medium clean product
                    elif (
                        background_ratio > 0.30
                        and
                        not is_collage
                    ):
                        hero_score += 60

                    # dark/lifestyle penalty
                    if background_ratio < 0.12:
                        hero_score -= 55

                    # excessive visual noise
                    if edge_density > 55:
                        hero_score -= 35

                    # =====================================
                    # HUMAN / LIFESTYLE DETECTION
                    # =====================================

                    rgb_arr = np.array(crop)

                    r = rgb_arr[:, :, 0]
                    g = rgb_arr[:, :, 1]
                    b = rgb_arr[:, :, 2]

                    skin_mask = (

                        (r > 95)

                        &

                        (g > 40)

                        &

                        (b > 20)

                        &

                        (r > g)

                        &

                        (r > b)

                        &

                        (np.abs(r - g) > 15)
                    )

                    skin_ratio = np.mean(skin_mask)

                    if skin_ratio > 0.28:

                        human_penalty = 40

                    hero_score -= human_penalty

                    # balanced edge density
                    if 8 < edge_density < 35:
                        hero_score += 25

                    # =====================================
                    # GALLERY SCORE
                    # =====================================

                    gallery_score = 0

                    # gallery accepts collages/grids
                    if is_collage:
                        gallery_score += 45

                    # preserve useful thumbnails
                    gallery_score += int(
                        coverage_ratio * 90
                    )

                    # clean product bonus
                    if background_ratio > 0.18:
                        gallery_score += 35

                    # acceptable noise
                    if edge_density < 75:
                        gallery_score += 20

                    # retain portrait products
                    if crop_height > crop_width:
                        gallery_score += 12

                    # moderate penalty only
                    if skin_ratio > 0.38:
                        gallery_score -= 20

                    # avoid total garbage
                    if background_ratio < 0.05:
                        gallery_score -= 40

                    # =====================================
                    # FINAL SCORE
                    # =====================================

                    score = hero_score

                    # =====================================
                    # SAVE
                    # =====================================

                    buffer = BytesIO()

                    crop.save(
                        buffer,
                        format="JPEG",
                        quality=92
                    )

                    encoded = base64.b64encode(
                        buffer.getvalue()
                    ).decode("utf-8")


                    candidate_crops.append({

                        "image": encoded,

                        "score": hero_score,

                        "hero_score": hero_score,

                        "gallery_score": gallery_score,

                        "is_collage": is_collage,

                        "centered_object": centered_object,

                        "background_ratio": background_ratio
                    })

                    _logger.warning(

                        f"[CROP DETECTED] "

                        f"{w}x{h} "

                        f"| hero={hero_score} "

                        f"| gallery={gallery_score} "

                        f"| collage={is_collage}"
                    )

                # =========================================
                # FALLBACK
                # =========================================

                if not candidate_crops:

                    buffer = BytesIO()

                    pil_image.save(
                        buffer,
                        format="JPEG"
                    )

                    encoded = base64.b64encode(
                        buffer.getvalue()
                    ).decode("utf-8")

                    candidate_crops.append({

                        "image": encoded,

                        "score": 10,

                        "is_collage": False
                    })

                segmented_images.extend(
                    candidate_crops
                )

            except Exception as e:

                _logger.warning(
                    f"[SEGMENTATION FAILED] {str(e)}"
                )

        # =============================================
        # DEDUPE
        # =============================================

        deduped = []
        hashes = {}

        for asset in segmented_images:

            try:

                img = asset.get("image")

                image_hash = hashlib.md5(
                    img.encode("utf-8")
                ).hexdigest()


                existing_score = hashes.get(
                    image_hash
                )

                if existing_score == asset.get(
                    "score"
                ):
                    continue

                hashes[image_hash] = asset.get(
                    "score"
                )

                deduped.append(asset)

            except Exception:
                continue

        return deduped


    # =====================================================
    # VARIANTS IMAGES CONTROLLER/DETECTOR
    # =====================================================

    def _split_grid_products(self, image):

        try:

            import cv2
            import numpy as np
            import base64

            gray = cv2.cvtColor(
                image,
                cv2.COLOR_BGR2GRAY
            )

            thresh = cv2.adaptiveThreshold(
                gray,
                255,
                cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY_INV,
                15,
                3
            )

            contours, _ = cv2.findContours(
                thresh,
                cv2.RETR_EXTERNAL,
                cv2.CHAIN_APPROX_SIMPLE
            )

            kernel = cv2.getStructuringElement(
                cv2.MORPH_RECT,
                (3, 3)
            )

            thresh = cv2.morphologyEx(
                thresh,
                cv2.MORPH_CLOSE,
                kernel,
                iterations=1
            )

            results = []

            for contour in contours:

                area = cv2.contourArea(contour)

                if area < 1600:

                    _logger.warning(

                        f"[GRID REJECT] "

                        f"small area={area}"
                    )

                    continue

                x, y, w, h = cv2.boundingRect(contour)

                if w < 55 or h < 55:

                    _logger.warning(

                        f"[GRID REJECT] "

                        f"tiny size={w}x{h}"
                    )

                    continue

                ratio = w / float(h)

                # reject text strips

                if ratio > 6.5 or ratio < 0.12:

                    _logger.warning(

                        f"[GRID REJECT] "

                        f"ratio={ratio}"
                    )

                    continue

                sub = image[
                    y:y+h,
                    x:x+w
                ]

                from PIL import Image

                pil_sub = Image.fromarray(
                    cv2.cvtColor(
                        sub,
                        cv2.COLOR_BGR2RGB
                    )
                )


                if not self._is_valid_product_crop(
                    pil_sub
                ):

                    _logger.warning(

                        f"[GRID REJECT] "

                        f"invalid crop "
                        f"size={w}x{h}"
                    )

                    continue

                success, buffer = cv2.imencode(
                    '.jpg',
                    sub
                )

              
                if not success:

                    _logger.warning(

                        "[GRID REJECT] "

                        "encode failed"
                    )

                    continue

                encoded = base64.b64encode(
                    buffer
                ).decode()

                score = self._score_segmented_image(
                    encoded
                )

                dominant = self._get_dominant_color_name(
                    encoded
                )

                _logger.warning(

                    f"[GRID ACCEPT] "

                    f"score={score} "

                    f"color={dominant} "

                    f"size={w}x{h}"
                )

                results.append({

                    "image": encoded,

                    "score": score,

                    "width": w,

                    "height": h,

                    "is_collage": False
                })


            # =====================================
            # DEBUG BEFORE SORT
            # =====================================

            _logger.warning(

                f"[GRID SPLIT RAW] "

                f"total={len(results)}"
            )

            for idx, item in enumerate(results):

                _logger.warning(

                    f"[GRID RAW ITEM] "

                    f"idx={idx} "

                    f"score={item.get('score')} "

                    f"size={item.get('width')}x{item.get('height')}"
                )

            # =====================================
            # SORT BEST FIRST
            # =====================================

            results = sorted(

                results,

                key=lambda x: x.get(
                    "score",
                    0
                ),

                reverse=True
            )

            # =====================================
            # DEBUG AFTER SORT
            # =====================================

            for idx, item in enumerate(results):

                _logger.warning(

                    f"[GRID SORTED ITEM] "

                    f"idx={idx} "

                    f"score={item.get('score')} "
                )

            # =====================================
            # NEVER HARD-TRIM VALID VARIANTS
            # =====================================

            _logger.warning(

                f"[GRID FINAL COUNT] "

                f"returned={len(results)}"
            )

            return results

        except Exception as e:

            _logger.warning(
                f"[GRID SPLIT FAILED] {str(e)}"
            )

            return []


    # =====================================================
    # VALIDATE CROPPED IMAGE
    # =====================================================

    def _is_valid_product_crop(self, pil_image):

        try:

            width, height = pil_image.size

            # too tiny
    
            if width < 75 or height < 75:
                return False

            # aspect safety
            ratio = width / float(height)

            if ratio > 7 or ratio < 0.10:
                return False

            # reject ultra-thin strips
            if width < 40 or height < 40:
                return False

            gray = pil_image.convert("L")

            arr = np.array(gray)

            # blank image rejection
            if arr.std() < 7:
                return False

            # excessive dark block rejection
    
            dark_pixels = np.mean(
                arr < 12
            )

            # only reject nearly solid dark blocks
            if dark_pixels > 0.985:
                return False

            # =====================================
            # TEXTURE VALIDATION
            # =====================================

            pixel_std = np.std(arr)

            if pixel_std < 5:
                return False

            return True

        except Exception:

            return False


    #=========VALIDATE AI IMAGE====================================
    def _is_valid_ai_image(self, image_data):
        
        try:
            import numpy as np

            if not image_data:
                return False

            import base64
            import io

            from PIL import Image

            # remove data url prefix
            if ',' in image_data:
                image_data = image_data.split(',')[1]

            decoded = base64.b64decode(image_data)

            # =====================================
            # SAFETY LIMIT
            # =====================================

            if len(decoded) > 15 * 1024 * 1024:

                _logger.warning(

                    "[INVALID AI IMAGE] image too large"
                )

                return False

            img = Image.open(
                io.BytesIO(decoded)
            )

            img.verify()
            
            # reopen after verify
            img = Image.open(
                io.BytesIO(decoded)
            ).convert("RGB")

            width, height = img.size

            # =====================================
            # REJECT VERY SMALL IMAGES
            # =====================================

            if width < 80 or height < 80:

                _logger.warning(

                    f"[INVALID AI IMAGE] "

                    f"tiny image {width}x{height}"
                )

                return False

            np_img = np.array(img)

            # =====================================
            # REJECT MOSTLY BLANK IMAGES
            # =====================================

            # white_ratio = np.mean(
            #     np_img > 245
            # )

            white_pixels = np.all(
                np_img > 245,
                axis=2
            )

            white_ratio = np.mean(
                white_pixels
            )

            if white_ratio > 0.985:

                _logger.warning(

                    "[INVALID AI IMAGE] blank image"
                )

                return False
            
            # =====================================
            # REJECT EXTREME DARK FRAMES
            # =====================================

            dark_pixels = np.all(
                np_img < 8,
                axis=2
            )

            dark_ratio = np.mean(
                dark_pixels
            )

            if dark_ratio > 0.985:

                _logger.warning(

                    "[INVALID AI IMAGE] dark frame"
                )

                return False

            return True

        except Exception as e:

            _logger.warning(

                f"[AI IMAGE VALIDATION FAILED] "

                f"{str(e)}"
            )

            return False
        
    # ---------------- Extract PDF ----------------
 
    def extract_pdf(self):

        import gc
        import json
        import io
        import re
        import fitz
        import base64
        import requests

        _logger.warning(
            f"[PDF EXTRACT] START "
            f"| job={self.id}"
        )

        MAX_RETRIES = 3

        # balanced batch size
        BATCH_SIZE = 3

        doc = None

        try:

            pdf_bytes = base64.b64decode(
                self.pdf_file
            )

        except Exception as e:

            _logger.exception(
                f"[PDF EXTRACT ERROR] "
                f"PDF DECODE FAILED "
                f"| {str(e)}"
            )

            self.state = "failed"

            return


        # =========================================
        # OPEN PDF
        # =========================================

        try:

            doc = fitz.open(
                stream=pdf_bytes,
                filetype="pdf"
            )

        except Exception as e:

            _logger.exception(
                f"[PDF EXTRACT ERROR] "
                f"PDF OPEN FAILED "
                f"| {str(e)}"
            )

            self.state = "failed"

            return


        try:

            total_pages = len(doc)

            self.total_pages = total_pages


            _logger.warning(
                f"[PDF EXTRACT] "
                f"TOTAL PAGES={total_pages}"
            )


            # =====================================
            # CRASH SAFE RECOVERY
            # =====================================

            existing_pages = self.env[
                'vendor.import.page'
            ].search([

                ('job_id', '=', self.id)

            ], order='page_number desc', limit=1)


            if existing_pages:

                # move to NEXT page
                start_page = (
                    existing_pages.page_number
                )

                _logger.warning(
                    f"[PDF RECOVERY] "
                    f"LAST SAVED PAGE="
                    f"{existing_pages.page_number}"
                )

            else:

                start_page = (
                    self.current_page or 0
                )

                _logger.warning(
                    f"[PDF RECOVERY] "
                    f"NO SAVED PAGES"
                )


            # =====================================
            # SAFETY CLAMP
            # =====================================

            if start_page >= total_pages:

                start_page = total_pages


            end_page = min(
                start_page + BATCH_SIZE,
                total_pages
            )


            _logger.warning(
                f"[PDF BATCH] "
                f"START={start_page + 1} "
                f"| END={end_page}"
            )


            processed_count = 0


            # =====================================
            # PROCESS PAGES
            # =====================================

            for i in range(start_page, end_page):

                _logger.warning(
                    f"[PDF PAGE] "
                    f"START PAGE={i + 1}"
                )

                page_success = False


                # =================================
                # SKIP IF ALREADY EXISTS
                # =================================

                existing = self.env[
                    'vendor.import.page'
                ].search([

                    ('job_id', '=', self.id),

                    ('page_number', '=', i + 1)

                ], limit=1)


                if existing:

                    _logger.warning(
                        f"[PDF PAGE] "
                        f"SKIP EXISTING "
                        f"| page={i + 1}"
                    )

                    self.current_page = i + 1

                    continue


                for attempt in range(MAX_RETRIES):

                    single_pdf = None
                    pdf_bytes_io = None

                    try:

                        _logger.warning(
                            f"[PDF API] "
                            f"PAGE={i + 1} "
                            f"| ATTEMPT={attempt + 1}"
                        )


                        # =========================
                        # SINGLE PAGE PDF
                        # =========================

                        single_pdf = fitz.open()

                        single_pdf.insert_pdf(

                            doc,

                            from_page=i,

                            to_page=i
                        )


                        pdf_bytes_io = io.BytesIO()

                        single_pdf.save(
                            pdf_bytes_io
                        )

                        pdf_bytes_io.seek(0)


                        # =========================
                        # API CALL
                        # =========================

                        response = requests.post(

                            "https://pdf-extractor-staging.onrender.com/extract",

                            files={

                                "file": (

                                    "page.pdf",

                                    pdf_bytes_io,

                                    "application/pdf"
                                )
                            },

                            timeout=45
                        )


                        _logger.warning(
                            f"[PDF API] "
                            f"STATUS="
                            f"{response.status_code} "
                            f"| page={i + 1}"
                        )


                        if response.status_code != 200:

                            continue


                        page_data = response.json()


                        # =========================
                        # RESPONSE FORMAT
                        # =========================

                        if isinstance(page_data, dict):

                            pages = page_data.get(
                                "pages",
                                []
                            )

                        elif isinstance(
                            page_data,
                            list
                        ):

                            pages = page_data

                        else:

                            pages = []


                        if not pages:

                            _logger.warning(
                                f"[PDF PAGE] "
                                f"EMPTY RESPONSE "
                                f"| page={i + 1}"
                            )

                            continue


                        normalized_blocks = []


                        # =========================
                        # NORMALIZE
                        # =========================

                        for p in pages:

                            text = p.get(
                                "text",
                                ""
                            )

                            images = p.get(
                                "images",
                                []
                            )

                            # ===========================
                            # CLEAN CATALOG SEGMENTATION
                            # ===========================

                            images = self._segment_catalog_images(
                                images
                            )


                            if (
                                not text
                                and
                                not images
                            ):
                                continue


                            price = ""

                            stock = ""


                            price_match = re.search(

                                r'(\$|€|£)\s?\d+[.,]?\d*',

                                text
                            )


                            if price_match:

                                price = (
                                    price_match.group(0)
                                )


                            stock_match = re.search(

                                r'(stock|available)'
                                r'\s*:?\s*'
                                r'(\d+)'
                                r'\s*(pcs|pieces)?',

                                text,

                                re.I
                            )


                            if stock_match:

                                stock = (
                                    stock_match.group(2)
                                )


                            normalized_blocks.append({

                                "page": i + 1,

                                "text": text,

                                "price": price,

                                "stock": stock,

                                "images": images
                            })


                        if not normalized_blocks:

                            _logger.warning(
                                f"[PDF PAGE] "
                                f"NO VALID BLOCKS "
                                f"| page={i + 1}"
                            )

                            continue


                        # ===========================
                        # SAVE PAGE
                        # ===========================


                        all_page_images = []

                        for block in normalized_blocks:

                            all_page_images.extend(
                                block.get("images", [])
                            )

                            self._safe_commit_progress()

                        self.env[
                            'vendor.import.page'
                        ].create({

                            'job_id': self.id,

                            'page_number': i + 1,

                            'extracted_json': json.dumps(
                                normalized_blocks
                            ),

                            'page_images_json': json.dumps(
                                all_page_images
                            )
                        })


                        _logger.warning(
                            f"[PDF PAGE] "
                            f"SAVED "
                            f"| page={i + 1}"
                        )


                        self.current_page = i + 1

                        processed_count += 1

                        page_success = True

                        break


                    except Exception as e:

                        _logger.exception(
                            f"[PDF PAGE ERROR] "
                            f"page={i + 1} "
                            f"| {str(e)}"
                        )


                    finally:

                        try:

                            if pdf_bytes_io:
                                pdf_bytes_io.close()

                        except Exception:
                            pass


                        try:

                            if single_pdf:
                                single_pdf.close()

                        except Exception:
                            pass


                if not page_success:

                    _logger.error(
                        f"[PDF PAGE FAILED] "
                        f"page={i + 1}"
                    )


            # =====================================
            # SAVE BATCH ONCE
            # =====================================

            _logger.warning(
                f"[PDF BATCH] "
                f"PROCESSED="
                f"{processed_count}"
            )


            if self.current_page < total_pages:

                self.state = "pdf_extracting"

            else:

                self.state = "pdf_ai"


            try:

                self._safe_commit_progress()

                _logger.warning(
                    f"[PDF SAVE] "
                    f"SUCCESS "
                    f"| state={self.state} "
                    f"| current={self.current_page}"
                )

            except Exception as e:

                _logger.exception(
                    f"[PDF SAVE ERROR] "
                    f"{str(e)}"
                )


        finally:

            try:

                if doc:
                    doc.close()

            except Exception:
                pass


            gc.collect()

            _logger.warning(
                "[PDF GC] COMPLETE"
            )


    # ---------------- Send to OPENAI URL ----------------
    def send_to_openai_url(self):

        import re
        import json
        import math

        api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')

        if not api_key:
            raise Exception("OpenAI API key not configured")

        client = OpenAI(api_key=api_key)

        # ================= LOAD PAGES =================
        try:
            # pages = json.loads(self.extracted_text or "[]")

            pages = json.loads(

                self.url_blocks_json

                or

                self.extracted_text

                or

                "[]"
            )

        except Exception:
            _logger.error("INVALID extracted_text JSON")
            return

        if not pages:
            _logger.error("NO PAGES TO PROCESS")
            return

        # ================= LOAD EXISTING =================
        existing_products = []
        if self.ai_response:
            try:
                data = json.loads(self.ai_response)
                if isinstance(data, list):
                    existing_products = data
            except Exception as e:
                _logger.warning(f"AI RESPONSE LOAD FAILED → {str(e)}")
                existing_products = []

        current_batch = self.url_batch_index or 0

        # ================= FLATTEN =================
        all_blocks = [b for p in pages for b in p.get("blocks", [])]

        _logger.warning(f"RAW BLOCKS → {len(all_blocks)}")

        # ================= CLEAN =================
        cleaned_blocks = self._clean_scraped_blocks(all_blocks)

        _logger.warning(f"CLEAN BLOCKS → {len(cleaned_blocks)}")
        _logger.warning(f"REMOVED BLOCKS → {len(all_blocks) - len(cleaned_blocks)}")

        cleaned_blocks = sorted(cleaned_blocks, key=lambda x: (x.get("text") or "")[:50])

        # ================= BATCH =================
        BLOCK_BATCH_SIZE = 8

        batched_blocks = [
            cleaned_blocks[i:i + BLOCK_BATCH_SIZE]
            for i in range(0, len(cleaned_blocks), BLOCK_BATCH_SIZE)
        ]

        total_batches = len(batched_blocks)
        self.url_total_batches = total_batches

        _logger.warning(f"TOTAL BLOCK BATCHES → {total_batches}")
        _logger.warning(f"CURRENT BATCH → {current_batch}")

        # ================= STOP IF DONE =================
        if current_batch >= total_batches:
            _logger.warning("ALL URL BATCHES PROCESSED ✅")
            self.state = "url_creating"
            return

        # ================= PROCESS ONE BATCH =================
        block_batch = batched_blocks[current_batch]

        _logger.warning(f"PROCESSING BLOCK COUNT → {len(block_batch)}")
        _logger.warning(f"AI → PROCESSING BLOCK BATCH {current_batch + 1}")

      
        combined_text = "\n\n---\n\n".join([
            f"""
            TEXT:
            {b.get('text','')}

            PRICE:
            {b.get('price','')}

            STOCK:
            {b.get('stock','')}

            IMAGE_URL:
            {b.get('image','')}
            """

            for b in block_batch
        ])


        if not combined_text.strip():
            _logger.warning("EMPTY COMBINED TEXT → SKIP")
            self.url_batch_index += 1
            return

        if len(combined_text) > 15000:
            combined_text = combined_text[:15000]
            _logger.warning("TEXT TRIMMED → PREVENT TOKEN OVERFLOW")

        # ================= PROMPT =================
        prompt = f"""
            You are a highly precise e-commerce product extraction engine.

            You are processing raw scraped website content.

            =====================================
            YOUR GOAL
            =====================================

            Extract ALL individual products from the text.

            IMPORTANT:
            - A single block may contain MULTIPLE products → you MUST split them
            - Each product MUST be returned separately
            - NEVER merge multiple products into one

            =====================================
            HOW TO IDENTIFY A PRODUCT
            =====================================

            A product usually contains:
            - product name
            - price (e.g. $, £, €, numbers)
            - specs or description
            - sometimes reviews

            Strong indicators:
            - currency symbols ($, £, €)
            - model names (Dell, Lenovo, etc.)
            - numbers like GB, inch, SSD, RAM, etc.

            =====================================
            STRICT RULES
            =====================================

            1. RETURN ONLY VALID JSON ARRAY
            2. NO explanation
            3. NO markdown
            4. NO text outside JSON

            5. EACH product must be unique
            6. REMOVE duplicates
            7. DO NOT skip products
            8. SPLIT combined text into multiple products

            =====================================
            REMOVE THIS NOISE
            =====================================

            Ignore anything related to:
            - navigation (Home, Login, Pricing)
            - cookies/privacy
            - footer text
            - categories only (no product info)
            - repeated headers

           =====================================
            VARIANT DETECTION
            =====================================

            IMPORTANT:

            Use PRODUCT IMAGES as the PRIMARY
            source for detecting variants.

            Also use:
            - product title
            - description
            - SKU
            - repeated patterns
            - packaging labels
            - text printed on product
            - visible size/capacity markings

            Detect REAL differences such as:
            - color
            - material
            - finish
            - texture
            - pattern
            - lid type
            - bottle type
            - packaging
            - shape
            - capacity
            - dimensions
            - style
            - design
            - print variation

            IMPORTANT RULES:

            1. NEVER generate:
            - Variant 1
            - Variant 2
            - Default
            - Standard
            - Option A
            - Option B

            2. ALWAYS return meaningful
            attribute names and values.

            GOOD EXAMPLES:

            {{
            "Color": "Black"
            }}

            {{
            "Material": "Bamboo"
            }}

            {{
            "Capacity": "750ml"
            }}

            {{
            "Design": "Football Print"
            }}

            {{
            "Finish": "Matte Silver"
            }}

            3. If a SINGLE IMAGE contains
            MULTIPLE product colors/designs:

            Create SEPARATE variants for EACH
            visible product variation.

            Example:
            - black bottle
            - blue bottle
            - red bottle

            MUST become:

            [
            {{
                "attributes": {{
                "Color": "Black"
                }}
            }},
            {{
                "attributes": {{
                "Color": "Blue"
                }}
            }},
            {{
                "attributes": {{
                "Color": "Red"
                }}
            }}
            ]

            4. If products differ by:
            - artwork
            - printed graphics
            - pattern
            - branding
            - sports design
            - texture

            Use:
            {{
            "Design": "..."
            }}

            5. If products differ mainly by:
            - size
            - dimensions
            - capacity

            Use:
            {{
            "Size": "..."
            }}

            OR

            {{
            "Capacity": "..."
            }}

            6. NEVER invent attributes that
            cannot be visually or textually
            supported.

            7. If uncertainty exists:
            Prefer:
            - Color
            - Design
            - Material
            - Capacity

            based on strongest visible evidence.

            8. If NO meaningful difference exists:
            Return ONE variant only.

            9. IMPORTANT:
            When multiple products appear in
            one image, treat each visible
            variation as a separate variant,
            even if no explicit text exists.

            10. Preserve consistency across
            all variants for the same product.

            BAD EXAMPLE:
            [
            {{
                "attributes": {{
                "Variant": "Variant 1"
                }}
            }}
            ]

            GOOD EXAMPLE:
            [
            {{
                "attributes": {{
                "Color": "White"
                }}
            }},
            {{
                "attributes": {{
                "Color": "Black"
                }}
            }}
            ]


            =====================================
            OUTPUT FORMAT
            =====================================

            [
                {{
                    "name": "Clean product name",
                    "description": "Short product description (max 30 words)",
                    "category": "Best guess category",
                    "price": "",
                    "stock": "",
                    "image": "image_url_or_null",
                    "variants": [
                                {{
                                    "attributes": {{
                                        "Variant": ""
                                    }},
                                    "image_index": 0,
                                    "stock": null
                                }}
                            ]
                }}
            ]

            =====================================
            EXTRA RULES
            =====================================

            - Keep names SHORT and CLEAN
            - Description must be concise
            - Infer category intelligently
            - If no image exists → return null
            - If price exists → extract it
            - If stock exists → extract it
            - NEVER invent stock or price
            - If unsure → still extract

            =====================================
            TEXT TO PROCESS
            =====================================

        {combined_text}
        """

        # ================= OPENAI =================
        try:
            response = client.responses.create(
                model="gpt-4.1-mini",
                input=prompt,
                temperature=0,
                timeout=60
            )

            result = response.output_text.strip()
            result = re.sub(r"^```(?:json)?|```$", "", result).strip()

            parsed = json.loads(result)

            if isinstance(parsed, list):

                cleaned = [p for p in parsed if p.get("name")]

                _logger.warning(f"AI RETURNED → {len(cleaned)} PRODUCTS")

                # 🔥 DEDUPE BY NAME
                existing_map = {p.get("name"): p for p in existing_products}

                for p in cleaned:
                    existing_map[p.get("name")] = p

                existing_products = list(existing_map.values())

                _logger.warning(f"TOTAL ACCUMULATED → {len(existing_products)}")

            else:
                _logger.warning("AI RESPONSE NOT LIST")

        except Exception as e:
            _logger.warning(f"AI ERROR → {str(e)}")
            return

        # ================= SAVE =================
        self.ai_response = json.dumps(existing_products)
        self.url_batch_index = current_batch + 1

        _logger.warning(f"URL AI PROGRESS → {self.url_batch_index}/{self.url_total_batches}")

        # ================= STATE =================
        if self.url_batch_index < self.url_total_batches:
            self.state = "url_ai"
        else:
            _logger.warning("URL AI FINISHED ALL BATCHES")
            self.state = "url_creating"

        # 🔥 IMPORTANT: COMMIT FOR CRON CONTINUITY

        try:

            self.env.cr.commit()

        except Exception as commit_error:

            _logger.warning(
                f"COMMIT SKIPPED → {commit_error}"
            )

        return


    # =========== PDF OPENAI ================================
    
    def send_to_openai_pdf(self):

        import json

        api_key = self.env[
            'ir.config_parameter'
        ].sudo().get_param(
            'openai.api.key'
        )

        if not api_key:

            raise Exception(
                "OpenAI API key not configured"
            )


        client = OpenAI(
            api_key=api_key
        )


        _logger.warning(
            "[PDF AI] START"
        )


        # =====================================================
        # LOAD PAGE RECORDS
        # =====================================================

        page_records = self.env[
            'vendor.import.page'
        ].search([

            ('job_id', '=', self.id)

        ], order='page_number asc')


        total_available_pages = len(
            page_records
        )


        _logger.warning(

            f"[PDF AI] "

            f"TOTAL PAGE RECORDS="

            f"{total_available_pages}"
        )


        if total_available_pages <= 0:

            _logger.warning(
                "[PDF AI] "
                "NO PAGE RECORDS FOUND"
            )

            return


        # =====================================================
        # LOAD EXISTING AI RESPONSE
        # =====================================================

        existing_pages = []


        if self.ai_response:

            try:

                loaded = json.loads(
                    self.ai_response
                )

                if isinstance(
                    loaded,
                    list
                ):

                    existing_pages = loaded


            except Exception as e:

                _logger.warning(

                    f"[PDF AI] "

                    f"LOAD EXISTING FAILED "

                    f"| {str(e)}"
                )

                existing_pages = []


        # =====================================================
        # FIND ALREADY PROCESSED PAGES
        # =====================================================

        processed_pages = set()


        for p in existing_pages:

            page_num = p.get("page")

            if page_num:

                processed_pages.add(
                    page_num
                )


        _logger.warning(

            f"[PDF AI] "

            f"PROCESSED PAGES="

            f"{sorted(list(processed_pages))}"
        )


        # =====================================================
        # FIND NEXT UNPROCESSED PAGE
        # =====================================================

        next_record = None


        for record in page_records:

            if (

                record.page_number

                not in processed_pages

            ):

                next_record = record

                break


        # =====================================================
        # ALL COMPLETE
        # =====================================================

        if not next_record:

            _logger.warning(
                "[PDF AI] COMPLETE ✅"
            )

            self.last_ai_page = (
                total_available_pages
            )

            self.state = "pdf_creating"

            self.flush_recordset()

            self.env.cr.commit()

            return


        _logger.warning(

            f"[PDF AI] "

            f"PROCESSING PAGE "

            f"{next_record.page_number}"
        )


        # =====================================================
        # LOAD PAGE DATA
        # =====================================================

        try:

            page_blocks = json.loads(

                next_record.extracted_json
                or
                "[]"
            )

        except Exception as e:

            _logger.warning(

                f"[PDF AI] "

                f"PAGE LOAD FAILED "

                f"| PAGE "

                f"{next_record.page_number} "

                f"| {str(e)}"
            )

            return


        if not page_blocks:

            _logger.warning(

                f"[PDF AI] "

                f"EMPTY PAGE BLOCKS "

                f"| PAGE "

                f"{next_record.page_number}"
            )

            return


        # =====================================================
        # BUILD PAGE DATA
        # =====================================================

        page_text = "\n".join([

            p.get("text", "")

            for p in page_blocks

        ])


        page_images = []

        for p in page_blocks:

            raw_images = p.get("images", [])

            # =====================================
            # NORMALIZE STRUCTURED ASSETS
            # =====================================

            for img in raw_images:

                if isinstance(img, dict):

                    if img.get("image"):

                        page_images.append(img)

                elif isinstance(img, str):

                    page_images.append({

                        "image": img,

                        "score": 0,

                        "is_collage": False
                    })

        # =====================================================
        # VALIDATE PAGE IMAGES
        # =====================================================

        valid_page_images = []

        for asset in page_images:

            try:

                # segmented assets are now dicts
                if isinstance(asset, dict):

                    image_data = asset.get(
                        "image"
                    )

                else:

                    image_data = asset

                if not image_data:
                    continue

                if not self._is_valid_ai_image(
                    image_data
                ):

                    _logger.warning(

                        f"[PDF AI] INVALID IMAGE "

                        f"| PAGE "

                        f"{next_record.page_number}"
                    )

                    continue

                valid_page_images.append(
                    asset
                )

            except Exception as e:

                _logger.warning(

                    f"[PDF AI IMAGE ERROR] "

                    f"{str(e)}"
                )

        page_images = valid_page_images
        # =========================================
        # REBUILD CLEAN IMAGE INDEX MAP
        # =========================================

        normalized_page_images = []

        for idx, asset in enumerate(page_images):

            if isinstance(asset, dict):

                asset["clean_index"] = idx

                normalized_page_images.append(asset)


        page_images = normalized_page_images

        _logger.warning(

            f"[PDF AI IMAGES] "

            f"PAGE={next_record.page_number} "

            f"| valid={len(page_images)}"
        )

        # =====================================
        # NO VALID IMAGE FAILSAFE
        # =====================================

        if not page_images:

            _logger.warning(

                f"[PDF NO VALID IMAGE] "

                f"PAGE={next_record.page_number}"
            )

            existing_map = {}

            for p in existing_pages:

                existing_map[
                    p.get("page")
                ] = p

            existing_map[
                next_record.page_number
            ] = {

                "page": next_record.page_number,

                "products": [],

                "images": [],

                "failed": True,

                "reason": "no_valid_images"
            }

            combined_pages = sorted(

                list(existing_map.values()),

                key=lambda x: x.get(
                    "page",
                    0
                )
            )

            self.ai_response = json.dumps(
                combined_pages
            )

            self.last_ai_page = len(
                combined_pages
            )

            self.state = "pdf_ai"

            self.flush_recordset()

            self.env.cr.commit()

            return

        page_price = ""

        page_stock = ""


        for p in page_blocks:

            if (

                not page_price

                and

                p.get("price")

            ):

                page_price = (
                    p.get("price")
                )


            if (

                not page_stock

                and

                p.get("stock")

            ):

                page_stock = (
                    p.get("stock")
                )


       # =====================================================
        # PROMPT
        # =====================================================

        prompt = f"""
        You are an AI ecommerce catalog extraction engine.

        Analyze:
        - catalog page text
        - detected catalog images

        Extract ALL visible products accurately.

        ==================================================
        OUTPUT RULES
        ==================================================

        Return ONLY valid JSON array.

        No markdown.
        No explanations.
        No extra text.

        ==================================================
        CORE EXTRACTION RULES
        ==================================================

        This input represents ONE catalog page.

        A page may contain:
        - one product
        - multiple products
        - one product with variants
        - isolated thumbnails
        - grouped color variations


        IMPORTANT:

        If products appear without clear title blocks,
        still extract them.

        If multiple standalone products appear on one page:
        - infer product grouping visually
        - detect visible variants
        - create products even if title is missing

        Pens, bottles, shirts, caps and accessories
        must NEVER be ignored simply because:
        - title is small
        - products are grouped
        - products are arranged in grid layout


        IMPORTANT:

        Aggressively detect ALL visible ecommerce products.

        If a catalog page shows:
        - 8 shirts
        - 10 caps
        - 6 bottles

        extract ALL visible variants.

        Prefer over-detection rather than missing products.

        ==================================================
        PRODUCT GROUPING RULES
        ==================================================

        Group products as variants ONLY when:
        - same product shape
        - same structure
        - same dimensions
        - same branding
        - same item or product fall on same page
        - only color/material/size changes

        Examples:
        - same cap in different colors
        - same polo shirt in different colors
        - same bottle in different colors

        Otherwise:
        create separate products.

        ==================================================
        TITLE RULES
        ==================================================

        Use the TRUE MAIN PRODUCT TITLE.

        Main title is usually:
        - largest heading
        - top heading
        - dominant catalog title
        - visible product headline

        DO NOT use:
        - material-only text
        - bullet features
        - specifications
        - dimensions
        - marketing phrases

        GOOD:
        - 5 PANEL CAP
        - SOL'S PERFECT MEN POLO SHIRT PIQUÉ 180
        - Wireless Charging Pad

        BAD:
        - Heavy Brushed 100% Cotton
        - Rib 1x1 collar and cuffs

        ==================================================
        DESCRIPTION RULES
        ==================================================

        Build rich ecommerce descriptions using:
        - subtitle
        - features
        - specifications
        - bullet points
        - dimensions
        - materials
        - capacities
        - branding info
        - packaging info

        Combine useful text naturally.

        ==================================================
        PRICE RULES
        ==================================================

        Aggressively search for price.

        Prices and stocks may appear:
        - near title
        - beside variants
        - inside tables
        - inside text blocks
        - in corners

        Detect:
        - $
        - €
        - £
        - ₦
        - USD
        - EUR
        - GBP

        Examples:
        - $2.99
        - USD 4.25
        - €8.50

        If no price exists:
        return empty string.

        Extract:
        - visible product price
        - visible stock quantity
        - visible product code
        ==================================================
        MOST CRITICAL
        ==================================================
        NO PRODUCT SHOULD MISS OUT OR BE IGNORED EXCEPT BLANK PAGE, 
        TOTAL NUMBER OF PAGES SHOULD GENERAGES SAME NUMBERS OF 
        PRODUCTS WITH EACH PRODUCTS HAS IT'S VARIANTS WHERE 
        MULTIPLE ITEMS APPEAR 
        ON SINGLE PAGE ACCURATELY. 
        ==================================================
        IMAGE RULES
        ==================================================

        Prefer professional ecommerce images:
        - isolated products
        - clean background
        - centered products
        - full visibility

        Avoid:
        - large text blocks
        - banners
        - lifestyle scenes
        - infographic layouts

        If isolated variants exist:
        prefer them over model/lifestyle photos.

        ==================================================
        STOCK EXTRACTION RULES:
        ==================================================
        PRICE EXTRACTION IS CRITICAL.

        Extract: 
        - stock quantity ONLY when
        actual available inventory is explicitly stated.

        - visible stock quantity

        Examples:
        - "Stock: 11 pcs"
        - "Available: 25"
        - "In stock: 8"

        DO NOT extract:
        - delivery times
        - MOQ

        If no real stock quantity exists:
        set:

        "stock_qty": 0

        ==================================================
        VARIANT RULES
        ==================================================

        Each variant should contain:

        {{
            "attributes": {{
                "Color": "",
                "Size": ""
            }},

            "image_index": null,
            "price": "",
            "stock_qty": 0
        }}

        ==================================================
        OUTPUT FORMAT
        ==================================================

        [
            {{
                "name": "",
                "subtitle": "",
                "description": "",
                "bullet_features": [],
                "material": "",
                "dimensions": "",
                "stock_qty": 0,
                "price": "",
                "currency": "",
                "product_code": "",
                "hero_image_index": null,
                "gallery_image_indexes": [],
                "variants": []
            }}
        ]

        ==================================================
        PAGE TEXT
        ==================================================

        {page_text}

        ==================================================
        DETECTED PRICE
        ==================================================

        {page_price}

        ==================================================
        DETECTED STOCK
        ==================================================

        {page_stock}
        """

        # =====================================================
        # AI CALL
        # =====================================================

        try:
            

            MAX_IMAGES = 24

            image_inputs = []


            def ai_visibility_score(asset):

                base = asset.get(
                    "score",
                    0
                )

                # =====================================
                # BOOST CLEAN ISOLATED PRODUCTS
                # =====================================

                if not asset.get("is_collage"):

                    base *= 1.35

                # =====================================
                # BOOST SMALL/MEDIUM PRODUCT SHOTS
                # =====================================

                width = int(

                    asset.get(
                        "width",
                        0
                    )

                    or 0
                )

                height = int(

                    asset.get(
                        "height",
                        0
                    )

                    or 0
                )

                area = width * height

                if area < 350000:

                    base *= 1.25

                # =====================================
                # PENALIZE HUMAN/LIFESTYLE IMAGES
                # =====================================

                dominant = str(

                    asset.get(
                        "dominant_color"
                    )

                    or ""

                ).lower()

                if dominant in [
                    "skin",
                    "beige"
                ]:

                    base *= 0.7

                return base


            sorted_page_images = sorted(

                page_images,

                key=ai_visibility_score,

                reverse=True
            )

            for idx, asset in enumerate(
                sorted_page_images[:MAX_IMAGES]
            ):

                _logger.warning(

                    f"[AI IMAGE INPUT] "

                    f"rank={idx} "

                    f"score={asset.get('score')} "

                    f"clean={asset.get('clean_index')} "

                    f"collage={asset.get('is_collage')}"
                )

            for asset in sorted_page_images[:MAX_IMAGES]:

                try:

                    # =====================================
                    # SUPPORT DICT ASSETS
                    # =====================================

                    if isinstance(asset, dict):

                        image_data = asset.get(
                            "image"
                        )

                    else:

                        image_data = asset

                    if not image_data:
                        continue

                    image_inputs.append({

                        "type": "input_image",

                        "image_url":

                            f"data:image/jpeg;base64,{image_data}"
                    })

                except Exception as e:

                    _logger.warning(

                        f"[IMAGE INPUT BUILD FAILED] "

                        f"{str(e)}"
                    )


            response = client.responses.create(

                model="gpt-4.1",

                input=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": prompt
                        }
                    ] + image_inputs
                }],

                timeout=60
            )


            result = (
                response.output_text or ""
            ).strip()


            result = result.replace(
                "```json",
                ""
            )

            result = result.replace(
                "```",
                ""
            ).strip()


            if not result:

                raise Exception(
                    "EMPTY AI RESPONSE"
                )

            try:

                parsed = json.loads(result)

            except Exception as e:

                _logger.warning(

                    f"[PDF AI JSON FAILED] "

                    f"PAGE={next_record.page_number} "

                    f"| {str(e)}"
                )

                _logger.warning(

                    f"[PDF AI RAW OUTPUT] "

                    f"{result[:1200]}"
                )

                _logger.warning(

                    f"[PDF AI PAGE SKIPPED] "

                    f"PAGE={next_record.page_number} "

                    f"| INVALID JSON"
                )

                # =====================================
                # MARK PAGE AS SKIPPED
                # =====================================

                existing_map = {}

                for p in existing_pages:

                    existing_map[
                        p.get("page")
                    ] = p

                existing_map[
                    next_record.page_number
                ] = {

                    "page": next_record.page_number,

                    "products": [],

                    "images": [],

                    "failed": True
                }

                combined_pages = sorted(

                    list(existing_map.values()),

                    key=lambda x: x.get(
                        "page",
                        0
                    )
                )

                self.ai_response = json.dumps(
                    combined_pages
                )

                self.last_ai_page = len(
                    combined_pages
                )

                self._safe_commit_progress()

                return

            if not parsed:

                _logger.warning(

                    f"[PDF AI EMPTY RESPONSE] "

                    f"PAGE={next_record.page_number}"
                )


                _logger.warning(

                    f"[PDF AI PAGE SKIPPED] "

                    f"PAGE={next_record.page_number} "

                    f"| EMPTY RESPONSE"
                )

                existing_map = {}

                for p in existing_pages:

                    existing_map[
                        p.get("page")
                    ] = p

                existing_map[
                    next_record.page_number
                ] = {

                    "page": next_record.page_number,

                    "products": [],

                    "images": [],

                    "failed": True
                }

                combined_pages = sorted(

                    list(existing_map.values()),

                    key=lambda x: x.get(
                        "page",
                        0
                    )
                )

                self.ai_response = json.dumps(
                    combined_pages
                )

                self.last_ai_page = len(
                    combined_pages
                )

                self._safe_commit_progress()

                return


            if not isinstance(
                parsed,
                list
            ):

                parsed = []


        except Exception as e:

            _logger.warning(

                f"[PDF AI] "

                f"FAILED "

                f"| PAGE "

                f"{next_record.page_number} "

                f"| {str(e)}"
            )

            return


        # =====================================================
        # SMART IMAGE MATCHING
        # =====================================================

        for prod in parsed:

            try:

                product_name = (
                    prod.get("name")
                    or ""
                )


                best_index = prod.get(
                    "hero_image_index"
                )

                # =====================================
                # VALIDATE CLEAN INDEX
                # =====================================

                valid_indexes = [

                    a.get("clean_index")

                    for a in page_images

                    if isinstance(a, dict)
                ]

                if (

                    best_index is None

                    or

                    not isinstance(best_index, int)

                    or

                    best_index not in valid_indexes
                ):

                    best_index = (
                        self.match_image_index_with_ai(
                            product_name,
                            page_images
                        )
                    )

                    if isinstance(best_index, int):

                        try:

                            matched_asset = page_images[
                                best_index
                            ]

                            if isinstance(matched_asset, dict):

                                best_index = matched_asset.get(
                                    "clean_index"
                                )

                        except Exception:
                            pass

                if best_index is not None:

                    prod["hero_image_index"] = (
                        best_index
                    )

                    _logger.warning(

                        f"[PDF HERO INDEX] "

                        f"{product_name} "

                        f"-> {best_index}"
                    )

            except Exception as e:

                _logger.warning(

                    f"[PDF IMAGE MATCH FAILED] "

                    f"{str(e)}"

                )

        # =====================================================
        # MERGE RESULTS
        # =====================================================

        existing_map = {}


        for p in existing_pages:

            existing_map[
                p.get("page")
            ] = p


        existing_map[
            next_record.page_number
        ] = {

            "page": next_record.page_number,

            "products": parsed,

            "images": page_images
        }


        combined_pages = sorted(

            list(existing_map.values()),

            key=lambda x: x.get(
                "page",
                0
            )
        )


        # =====================================================
        # SAVE
        # =====================================================

        self.ai_response = json.dumps(
            combined_pages
        )


        self.last_ai_page = len(
            combined_pages
        )


        _logger.warning(

            f"[PDF AI] "

            f"PAGE SAVED "

            f"| PAGE "

            f"{next_record.page_number}"
        )


        # =====================================================
        # NEXT STATE
        # =====================================================

        if (

            self.last_ai_page

            <

            total_available_pages

        ):

            self.state = "pdf_ai"

            _logger.warning(

                f"[PDF AI] CONTINUE "

                f"{self.last_ai_page}/"

                f"{total_available_pages}"
            )

        else:

            _logger.warning(
                "[PDF AI] COMPLETE ✅"
            )

            self.state = "pdf_creating"


        self.flush_recordset()
        self.env.cr.commit()

        return
    
   
    #===========Excel Open AI================================
    def send_to_openai_excel(self):

        import json

        api_key = self.env[
            'ir.config_parameter'
        ].sudo().get_param(
            'openai.api.key'
        )

        if not api_key:

            raise Exception(
                "OpenAI API key not configured"
            )


        client = OpenAI(
            api_key=api_key
        )


        _logger.warning(
            "[EXCEL AI] START"
        )


        # =====================================================
        # LOAD EXTRACTED DATA
        # =====================================================

        try:

            pages = json.loads(
                self.extracted_text or "[]"
            )

        except Exception as e:

            _logger.error(

                f"[EXCEL AI] "

                f"INVALID extracted_text JSON "

                f"| {str(e)}"
            )

            return


        if not pages:

            _logger.error(
                "[EXCEL AI] NO ROWS TO PROCESS"
            )

            return


        _logger.warning(

            f"[EXCEL AI] "

            f"TOTAL ROWS={len(pages)}"
        )


        # =====================================================
        # BATCH
        # =====================================================

        BATCH_SIZE = 5

        start = (
            self.excel_ai_index or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(pages)
        )


        batch = pages[start:end]


        _logger.warning(

            f"[EXCEL AI BATCH] "

            f"{start} → {end}"
        )


        # =====================================================
        # LOAD EXISTING PRODUCTS
        # =====================================================

        existing_products = []


        if self.ai_response:

            try:

                existing_ai = json.loads(
                    self.ai_response
                )

                if (
                    isinstance(existing_ai, list)
                    and existing_ai
                ):

                    existing_products = (
                        existing_ai[0].get(
                            "products",
                            []
                        )
                    )


                _logger.warning(

                    f"[EXCEL AI] "

                    f"EXISTING PRODUCTS="

                    f"{len(existing_products)}"
                )

            except Exception as e:

                _logger.warning(

                    f"[EXCEL AI] "

                    f"FAILED LOAD EXISTING "

                    f"| {str(e)}"
                )

                existing_products = []


        new_products = []


        # =====================================================
        # PROCESS ROWS
        # =====================================================

        for idx, row in enumerate(

            batch,

            start=start

        ):

            try:

                row_text = row.get(
                    "text",
                    ""
                )

                row_price = row.get(
                    "price",
                    ""
                )

                row_stock = row.get(
                    "stock",
                    ""
                )

                images = row.get(
                    "images",
                    []
                )


                _logger.warning(

                    f"[EXCEL AI ROW] "

                    f"idx={idx} "

                    f"| images={len(images)}"
                )

                prompt = f"""
                You are a structured Excel product parser.

                Each input represents EXACTLY ONE ROW = ONE PRODUCT.

                =====================================
                COLUMN UNDERSTANDING (CRITICAL)
                =====================================

                The row could contain mixed values like:

                - ID (e.g. 94601, 12345)
                - Range (e.g. 2-66, 11-00)
                - Stock numbers
                - Prices
                - Links (http...)
                - Image references

                YOU MUST:

                1. IDENTIFY PRODUCT ID
                - Usually numeric (e.g. 94601)
                - Column name may vary:
                    - KOD
                    - SKU
                    - ID
                    - CODE

                2. IDENTIFY PRODUCT NAME
                - MUST NOT be:
                    - pure numbers
                    - ranges
                    - links
                    - dates
                    - headers

                - Product names should describe the ACTUAL product type.

               GOOD:
                - Sports Bottle
                - Metal Pen
                - Travel Mug
                - Drawstring Bag

                If the Excel already contains a valid product name:
                - preserve and use it

                If the Excel does NOT contain a real product name:
                - intelligently generate one using:
                    - Product <ID>
                    - category clues
                    - image appearance
                    - surrounding row data

                Fallback naming is allowed when necessary.

                GOOD fallback examples:
                - Product 94601
                - Bottle 94646
                - Pen 92070

                However:

                If rows belong to the SAME variant_group,
                you MUST still detect and extract the REAL variant difference.

                Example:

                Parent:
                Product 94646

                Variants:
                - White
                - Orange
                - Black

                DO NOT return:
                - Variant 1
                - Variant 2

                when a real difference can be visually or textually identified.
                
                =====================================
                VARIANT GROUPING (VERY IMPORTANT)
                =====================================

                - SAME ID = SAME variant_group
                - DIFFERENT ID = DIFFERENT PRODUCT
                - NEVER leave variant_group empty

                Rows sharing the same:
                - ID
                - grouped code
                - SKU group

                should be treated as variants of ONE parent product.

                 =====================================
                VARIANT DETECTION
                =====================================

                If rows share same PRODUCT ID:

                → they belong to the SAME product family.

                IMPORTANT:

                Use PRODUCT IMAGES as the PRIMARY
                source for identifying variants.

                Look for visual differences such as:

                - color
                - material
                - finish
                - lid type
                - texture
                - shape
                - capacity
                - packaging

                THEN use nearby codes/numbers
                as supporting evidence.

                Example:

                Rows may contain:

                106
                103
                128

                These MAY represent:
                - color codes
                - material codes
                - size codes

                DO NOT assume globally.

                Infer meaning from:
                - image differences
                - repeated patterns
                - product appearance

                If uncertain:

                Use safe fallback:

                "attributes": {{
                    "Vendor Code": "106"
                }}

                NEVER return:
                - Variant 1
                - Variant 2
                - Variant 3

                ALWAYS return meaningful attributes.

                =====================================
                VISUAL DIFFERENCE DETECTION
                =====================================

                If product images exist:

                You MUST visually inspect the images
                to identify the distinguishing feature.

                Example:

                If grouped products show:
                - white bottle
                - orange bottle
                - black bottle

                Return:

                {{
                    "name": "Sports Bottle",
                    "color": "White"
                }}

                {{
                    "name": "Sports Bottle",
                    "color": "Orange"
                }}

                DO NOT return:
                - Variant 1
                - Variant 2
                - Product 94601

                =====================================
                PARENT PRODUCT CONSISTENCY
                =====================================

                When multiple rows belong to the same
                variant_group:

                - The parent product name MUST remain consistent.
                - ONLY the variant fields should change.

                GOOD:

                Sports Bottle
                → White
                → Orange
                → Black

                BAD:

                White Bottle
                Orange Bottle
                Black Bottle

                =====================================
                ATTRIBUTE EXTRACTION
                =====================================

                Put distinguishing values into:

                - color
                - material
                - size
                - capacity
                - style

                Only use generic "Variant"
                if absolutely no real difference can be detected.

                =====================================
                PRICE & STOCK
                =====================================

                - Extract numeric price carefully
                - Extract stock carefully
                - Ignore ranges like:
                    - 2-66
                    - 11-00

                =====================================
                LINKS
                =====================================

                If a row contains a product URL:
                - preserve it
                - never use URL as product name

                =====================================
                OUTPUT FORMAT
                =====================================

                [
                    {{
                        "name": "",
                        "description": "",
                        "category": "",
                        "price": "",
                        "stock": "",
                        "variant_group": "",
                        "color": "",
                        "material": "",
                        "size": "",
                        "capacity": "",
                        "style": "",
                        "url": "",
                        "variants": [
                            {{
                                "attributes": {{
                                    "Variant": ""
                                }},
                                "image_index": 0,
                                "stock": null
                            }}
                        ]
                    }}
                ]

                =====================================
                IMPORTANT RULES
                =====================================

                - Return ONLY valid JSON
                - No markdown
                - No explanations
                - No comments
                - No trailing commas

                ROW TEXT:
                {row_text}

                DETECTED PRICE:
                {row_price}

                DETECTED STOCK:
                {row_stock}
                """

                response = client.responses.create(

                    model="gpt-4.1-mini",

                    input=prompt,

                    timeout=60
                )


                result = (
                    response.output_text or ""
                ).strip()


                result = result.replace(
                    "```json",
                    ""
                )

                result = result.replace(
                    "```",
                    ""
                ).strip()


                if not result:

                    raise Exception(
                        "EMPTY AI RESPONSE"
                    )


                parsed = json.loads(
                    result
                )


                if (

                    isinstance(parsed, list)

                    and parsed

                ):

                    parsed = parsed[0]


                if not isinstance(
                    parsed,
                    dict
                ):

                    _logger.warning(

                        f"[EXCEL AI] "

                        f"INVALID STRUCTURE "

                        f"| idx={idx}"
                    )

                    continue


                # =================================================
                # IMAGE
                # =================================================

                if images:

                    parsed["image"] = (
                        images[0]
                    )


                # =================================================
                # DEBUG
                # =================================================

                _logger.warning(

                    f"[EXCEL AI PRODUCT] "

                    f"name={parsed.get('name')} "

                    f"| group={parsed.get('variant_group')}"
                )


                new_products.append(
                    parsed
                )


            except Exception as e:

                _logger.exception(

                    f"[EXCEL AI ERROR] "

                    f"idx={idx} "

                    f"| {str(e)}"
                )


        # =====================================================
        # MERGE PRODUCTS SAFELY
        # =====================================================

        combined_products = (
            existing_products
            +
            new_products
        )


        _logger.warning(

            f"[EXCEL AI MERGE] "

            f"existing={len(existing_products)} "

            f"| new={len(new_products)} "

            f"| total={len(combined_products)}"
        )


        # =====================================================
        # SAVE
        # =====================================================

        self.ai_response = json.dumps([{

            "page": 1,

            "products": combined_products

        }])


        self.excel_ai_index = end


        _logger.warning(

            f"[EXCEL AI SAVE] "

            f"{self.excel_ai_index}/"

            f"{len(pages)}"
        )


        # =====================================================
        # NEXT STATE
        # =====================================================

        if end < len(pages):

            self.state = "excel_ai"

        else:

            _logger.warning(
                "[EXCEL AI COMPLETE]"
            )

            self.state = (
                "excel_creating"
            )


        self.flush_recordset()

        self.env.cr.commit()

        return
    

    #-----------scoring image before picking best/quality image (inage logic)-------------
    def pick_best_image(self, images):

                best_img = None
                best_score = 0
                seen_hashes = set()

                for img in images:

                    try:
                        img_bytes = base64.b64decode(img)
                        size = len(img_bytes)

                        # ❌ Skip tiny images (logos/icons)
                        if size < 10000:
                            continue

                        # ❌ Skip extremely large (full lifestyle pages)
                        if size > 800000:
                            continue

                        # ================= IMAGE ANALYSIS =================
                        pil_img = Image.open(BytesIO(img_bytes))
                        width, height = pil_img.size

                        # ❌ Skip extremely small resolution
                        if width < 150 or height < 150:
                            continue

                        # ❌ Skip extreme aspect ratios (banners, strips)
                        aspect_ratio = width / height if height else 1

                        if aspect_ratio > 3 or aspect_ratio < 0.3:
                            continue

                        # ================= DUPLICATE CHECK =================
                        img_hash = hash(img_bytes[:100])  # fast partial hash

                        if img_hash in seen_hashes:
                            continue

                        seen_hashes.add(img_hash)

                        # ================= SCORING =================
                        score = 0

                        # ✅ Prefer medium-good sizes
                        if 20000 < size < 300000:
                            score += 200

                        # ✅ Prefer square-ish product images
                        if 0.7 < aspect_ratio < 1.5:
                            score += 150

                        # ✅ Prefer decent resolution
                        if width > 400 and height > 400:
                            score += 150

                        # ❌ Penalize too wide/tall
                        if aspect_ratio > 2 or aspect_ratio < 0.5:
                            score -= 100

                        # ❌ Penalize very large images (likely lifestyle)
                        if size > 500000:
                            score -= 150

                        # Base size contribution
                        score += size / 2000

                        # ================= SELECT BEST =================
                        if score > best_score:
                            best_score = score
                            best_img = img

                    except Exception:
                        continue

                return best_img

    #=================Centralized Rusable Image=======================
    def _prepare_asset_pool(self, images):

        prepared = []

        seen = {}

        for asset in (images or []):

            try:

                if not asset:
                    continue

                # =====================================
                # SUPPORT OLD + NEW FORMAT
                # =====================================

                if isinstance(asset, dict):

                    img = asset.get("image")

                    score = asset.get(
                        "score",
                        0
                    )

                    is_collage = asset.get(
                        "is_collage",
                        False
                    )

                else:

                    img = asset

                    score = 0

                    is_collage = False

                if not img:

                    _logger.warning(
                        "[ASSET SKIPPED] EMPTY IMAGE"
                    )

                    continue

                # =====================================
                # SKIP EXTREMELY LOW SCORES
                # =====================================

                if score <= -500:

                    continue

                image_hash = hashlib.md5(

                    img.encode('utf-8')

                ).hexdigest()

                # =====================================
                # SAFE COLOR DETECTION
                # =====================================

                dominant_color = ""

                # =====================================
                # SAFE DEFAULT SCORES
                # =====================================
                hero_score = score
                gallery_score = score

                try:

                    dominant_color = (

                        self._detect_dominant_color(
                            img
                        ) or ""
                    )

                    # =====================================
                    # CLEAN PRODUCT BOOST
                    # =====================================

                    if not is_collage:

                        gallery_score += 25

                    # =====================================
                    # COLLAGE SUPPRESSION
                    # =====================================

                    if is_collage:

                        hero_score -= 30

                except Exception as color_error:

                    _logger.warning(

                        f"[COLOR DETECT FAILED] "

                        f"{str(color_error)}"
                    )

                # =====================================
                # ONLY REMOVE TRUE DUPLICATES
                # =====================================

                existing_asset = seen.get(
                    image_hash
                )

                # SAME HASH + SAME SCORE + SAME COLLAGE
                # = real duplicate only

                if existing_asset:

                    if (

                        abs(

                            existing_asset.get(
                                "score",
                                0
                            ) - score

                        ) <= 5

                        and

                        existing_asset.get(
                            "is_collage"
                        ) == is_collage

                        and

                        existing_asset.get(
                            "dominant_color"
                        ) == dominant_color

                    ):

                        _logger.warning(

                            f"[ASSET SKIPPED] TRUE DUPLICATE"
                        )

                        continue

                _logger.warning(

                    f"[ASSET DEBUG] "

                    f"type={type(asset)} "

                    f"score={score} "

                    f"collage={is_collage} "

                    f"color={dominant_color}"
                )


                prepared.append({

                    "image": img,

                    "score": score,

                    # =====================================
                    # SEPARATED SCORING SYSTEM
                    # =====================================

                    "hero_score": hero_score,

                    "gallery_score": gallery_score,

                    "is_collage": is_collage,

                    "dominant_color":
                        dominant_color,

                    # =====================================
                    # DIMENSIONS
                    # =====================================

                    "width": (

                        asset.get("width", 0)

                        if isinstance(asset, dict)

                        else 0
                    ),

                    "height": (

                        asset.get("height", 0)

                        if isinstance(asset, dict)

                        else 0
                    ),
                })

                seen[image_hash] = {

                    "score": score,

                    "is_collage": is_collage,

                    "dominant_color": dominant_color
                }

                _logger.warning(

                    f"[ASSET ADDED] "

                    f"score={score} "

                    f"collage={is_collage} "

                    f"color={dominant_color}"
                )

            except Exception as e:

                _logger.warning(

                    f"[ASSET POOL ERROR] "

                    f"{str(e)}"
                )

        # =====================================
        # SORT BEST FIRST
        # =====================================

        prepared = sorted(

            prepared,

            key=lambda x: (

                x.get(
                    "gallery_score",
                    x.get(
                        "score",
                        0
                    )
                ),

                x.get(
                    "hero_score",
                    0
                ),

                not x.get(
                    "is_collage",
                    False
                )

            ),

            reverse=True
        )

        # =====================================
        # REBUILD INDEXES AFTER SORT
        # =====================================

        for idx, asset in enumerate(prepared):

           asset["index"] = idx

           asset["clean_index"] = idx

        _logger.warning(

            f"[ASSET POOL READY] "

            f"{len(prepared)} assets"
        )

        return prepared


    # =======================================
    # ADVANCED DOMINANT COLOR DETECTION
    # =======================================
    
    def _detect_dominant_color(

        self,

        image_base64
    ):

        try:

            import base64
            import colorsys
            import numpy as np

            from io import BytesIO
            from PIL import Image

            image_data = base64.b64decode(
                image_base64
            )

            image = Image.open(

                BytesIO(image_data)

            ).convert("RGB")

            image = image.resize((120, 120))

            pixels = np.array(image)

            # =====================================
            # REMOVE VERY BRIGHT BACKGROUND
            # =====================================

            pixels = pixels.reshape(-1, 3)

            filtered_pixels = []

            for r, g, b in pixels:

                # remove white bg
                if r > 235 and g > 235 and b > 235:
                    continue

                filtered_pixels.append([r, g, b])

            if not filtered_pixels:
                return "white"

            pixels = np.array(filtered_pixels)

            avg = pixels.mean(axis=0)

            r, g, b = avg

            # =====================================
            # DARK COLOR ANALYSIS
            # =====================================

            brightness = np.mean(

                pixels,

                axis=1
            )

            dark_pixels_ratio = np.mean(
                brightness < 75
            )

            very_dark_ratio = np.mean(
                brightness < 45
            )

            # dominant blue inside dark pixels
            dark_blue_ratio = np.mean(

                (
                    pixels[:, 2] > pixels[:, 0] * 1.15
                )

                &

                (
                    pixels[:, 2] > pixels[:, 1] * 1.10
                )

                &

                (
                    brightness < 90
                )
            )

            # =====================================
            # TRUE BLACK DETECTION
            # =====================================

            if (

                very_dark_ratio > 0.24

                or

                (
                    dark_pixels_ratio > 0.40

                    and

                    abs(r - g) < 24

                    and

                    abs(g - b) < 24
                )
            ):

                return "black"

            # =====================================
            # DARK NAVY DETECTION
            # =====================================

            if dark_blue_ratio > 0.18:

                return "navy"

            # =====================================
            # RGB → HSV
            # =====================================

            h, s, v = colorsys.rgb_to_hsv(

                r / 255.0,
                g / 255.0,
                b / 255.0
            )

            h = h * 360
            s = s * 100
            v = v * 100

            # =====================================
            # BLACK
            # =====================================

            if v < 18:
                return "black"

            # =====================================
            # WHITE
            # =====================================

            if v > 92 and s < 10:
                return "white"

         
            # =====================================
            # GREY DETECTION
            # =====================================

            if s < 15:

                return "grey"

            # =====================================
            # RED
            # =====================================

            if h < 15 or h >= 345:
                return "red"

            # =====================================
            # ORANGE
            # =====================================

            if 15 <= h < 40:
                return "orange"

            # =====================================
            # YELLOW
            # =====================================

            if 40 <= h < 70:
                return "yellow"

            # =====================================
            # GREEN
            # =====================================

            if 70 <= h < 170:
                return "green"

            # =====================================
            # BLUE
            # =====================================

            if 170 <= h < 260:

                if v < 45:
                    return "navy"

                if s < 35:
                    return "light blue"

                return "blue"

            # =====================================
            # PURPLE
            # =====================================

            if 260 <= h < 320:
                return "purple"

            # =====================================
            # PINK
            # =====================================

            if 320 <= h < 345:
                return "pink"

            return "unknown"

        except Exception as e:

            _logger.warning(

                f"[DOMINANT COLOR FAILED] "

                f"{str(e)}"
            )

            return "unknown"


    # =====================================
    # PROFESSIONAL VARIANT IMAGE MATCHER
    # =====================================
  
    def _match_variant_image(
         self,
        variant,
        asset_pool,
        used_asset_indexes=None
    ):

        try:
            if used_asset_indexes is None:

                used_asset_indexes = set()

            if not asset_pool:
                return False

            best_asset = None

            best_score = -999

            if used_asset_indexes is None:

                used_asset_indexes = set()
            variant_text = ""

            attributes = variant.get(
                "attributes",
                {}
            )

            if isinstance(attributes, dict):

                variant_text = " ".join([

                    str(v)

                    for v in attributes.values()

                ]).lower()

            # =====================================
            # SCORE ASSETS
            # =====================================

            for asset in asset_pool:

                if asset.get("clean_index") in used_asset_indexes:
                        continue

                asset_score = 0

                # =====================================
                # UNUSED ASSET BONUS
                # =====================================

                asset_score += 25

                # =====================================
                # START FROM GALLERY SCORE
                # =====================================

                asset_score += asset.get(
                    "gallery_score",
                    asset.get(
                        "score",
                        0
                    )
                )


                dominant_color = str(

                    asset.get(
                        "dominant_color",
                        ""
                    )

                    or ""

                ).lower()

                # =====================================
                # BOOST CLEAN ISOLATED PRODUCTS
                # =====================================

                if not asset.get("is_collage"):

                    asset_score += 35

                # =====================================
                # SMALL/MEDIUM PRODUCT BOOST
                # =====================================

                width = int(
                    asset.get("width", 0) or 0
                )

                height = int(
                    asset.get("height", 0) or 0
                )

                area = width * height

                if 10000 < area < 350000:

                    asset_score += 55

                if dominant_color == "unknown":

                    asset_score -= 18

                # ---------------------------------
                # COLLAGE PENALTY
                # ---------------------------------

                if asset.get("is_collage"):

                    asset_score -= 12

                # ---------------------------------
                # COLOR MATCHING
                # ---------------------------------

                color_map = [

                    "red",
                    "blue",
                    "navy",
                    "green",
                    "lime",
                    "yellow",
                    "orange",
                    "white",
                    "black",
                    "gray",
                    "grey",
                    "light_grey",
                    "charcoal",
                    "silver",
                    "purple",
                    "pink",
                    "brown"
                ]

                normalized_variant_text = variant_text.lower()

                for color in color_map:

                    if color not in normalized_variant_text:
                        continue

                    # exact match
                    if color == dominant_color:

                        asset_score += 180

                    # navy/blue distinction
                    elif (
                        color == "navy"
                        and
                        dominant_color == "blue"
                    ):
                        asset_score += 40

                    # gray/grey normalization

                    elif (

                        color in ["gray", "grey"]

                        and

                        dominant_color in [
                            "gray",
                            "grey",
                            "light_grey"
                        ]
                    ):

                        asset_score += 120

                    # white/silver/light handling
                    elif (

                        color == "white"

                        and

                        dominant_color in [
                            "white",
                            "light_grey"
                        ]
                    ):

                        asset_score += 90

                    # dark product approximation
                    elif (

                        color == "black"

                        and

                        dominant_color in [
                            "black",
                            "navy"
                        ]
                    ):

                        asset_score += 90

                # ------------------------------------
                # HERO BONUS
                # ------------------------------------

                if asset.get("score", 0) >= 70:

                    asset_score += 4

                # ---------------------------------
                # BEST MATCH
                # ---------------------------------

                if asset_score > best_score:

                    best_score = asset_score

                    best_asset = asset

            # =====================================
            # SAFE FALLBACK
            # =====================================

            if not best_asset:

                remaining_assets = [

                    a

                    for a in asset_pool

                    if (
                        a.get("clean_index")
                        not in used_asset_indexes
                    )
                ]

                fallback_assets = (
                    remaining_assets
                    if remaining_assets
                    else asset_pool
                )

                best_asset = sorted(

                    fallback_assets,

                    key=lambda x: (

                        x.get(
                            "gallery_score",
                            0
                        ),

                        x.get(
                            "hero_score",
                            x.get(
                                "score",
                                0
                            )
                        )

                    ),

                    reverse=True

                )[0]

            if best_asset:

                used_asset_indexes.add(

                    best_asset.get("clean_index")
                )

            return best_asset

        except Exception as e:

            _logger.warning(

                f"[VARIANT MATCH FAILED] "

                f"{str(e)}"
            )

            return False


    #======score_segmented_image ==========================
    def _score_segmented_image(

        self,

        image_base64
    ):

        try:

            import base64
            import io

            import numpy as np

            from PIL import Image

            image_bytes = base64.b64decode(
                image_base64
            )

            img = Image.open(

                io.BytesIO(image_bytes)

            ).convert("RGB")

            width, height = img.size

            # ==========================================
            # REJECT VERY SMALL CROPS
            # ==========================================


            if width < 110 or height < 110:

                return -999

            np_img = np.array(img)

            score = 0

            # ==========================================
            # LARGE IMAGE BONUS
            # ==========================================

            area = width * height

            # ==========================================
            # MODERATE SIZE PREFERENCE
            # ==========================================

            if 15000 <= area <= 450000:

                score += 75

            elif area > 450000:

                score += 35

            else:

                score += 15


            # ==========================================
            # DARK OBJECT ANALYSIS
            # ==========================================

            gray_img = np.mean(
                np_img,
                axis=2
            )

            very_dark_ratio = np.mean(
                gray_img < 22
            )

            # ==========================================
            # ONLY REJECT TRUE SOLID DARK BLOCKS
            # ==========================================

            pixel_std_gray = np.std(gray_img)

            if (

                very_dark_ratio > 0.82

                and

                pixel_std_gray < 18
            ):

                score -= 35

            # ==========================================
            # GOOD PRODUCT ASPECT BONUS
            # ==========================================

            aspect = width / float(height)

            if 0.45 <= aspect <= 2.2:

                score += 50

            # ==========================================
            # CLEAN BACKGROUND BONUS
            # ==========================================
            white_pixels = np.all(
                np_img > 230,
                axis=2
            )

            white_ratio = np.mean(
                white_pixels
            )

            # moderate ecommerce bonus only
            if white_ratio > 0.45:

                score += 14

            elif white_ratio > 0.25:

                score += 8

            # ==========================================
            # DARK PRODUCT PRESERVATION
            # ==========================================

            dark_presence = np.mean(
                gray_img < 55
            )

            # preserve legitimate dark apparel
            if (
                0.18 < dark_presence < 0.75
                and
                pixel_std_gray > 16
            ):

                score += 18

            score += white_ratio * 24
           
            # ==========================================
            # DETAIL / TEXTURE BONUS
            # ==========================================

            pixel_std = np.std(np_img)

            if pixel_std > 45:

                score += 28

            elif pixel_std > 30:

                score += 15

            return score

        except Exception as e:

            _logger.warning(

                f"[IMAGE SCORE ERROR] "

                f"{str(e)}"
            )

            return 0    


    #=============variant color enhancement 1=================
    def _get_dominant_color_name(

        self,

        image_base64
    ):

        try:

            import base64
            import io
            import numpy as np

            from PIL import Image

            image_bytes = base64.b64decode(
                image_base64
            )

            img = Image.open(

                io.BytesIO(image_bytes)

            ).convert("RGB")

            img = img.resize((80, 80))

            np_img = np.array(img)

            pixels = np_img.reshape(
                (-1, 3)
            )

            # =====================================
            # REMOVE VERY LIGHT BACKGROUND PIXELS
            # =====================================

            filtered_pixels = []

            for px in pixels:

                pr, pg, pb = px

                # skip white/light background
                if (
                    pr > 235
                    and
                    pg > 235
                    and
                    pb > 235
                ):

                    continue

                filtered_pixels.append(px)

            # fallback if filtering too aggressive
            if not filtered_pixels:

                filtered_pixels = pixels

            filtered_pixels = np.array(
                filtered_pixels
            )

            # =====================================
            # USE MEDIAN FOR STABILITY
            # =====================================

            median = np.median(
                filtered_pixels,
                axis=0
            )

            r, g, b = median

            # =====================================
            # COLOR CLASSIFICATION
            # =====================================

  
            if r > 200 and g > 200 and b > 200:
                return "white"

            if (
                abs(r - g) < 18
                and
                abs(g - b) < 18
                and
                210 <= r <= 242
            ):
                return "light_grey"
          
            if r > 160 and g < 120 and b < 120:
                return "red"

            if r > 180 and g > 180 and b < 120:
                return "yellow"

            if (
                b > r * 1.12
                and
                b > g * 1.08
                and
                b < 110
            ):
                return "navy"

            if (
                b > r * 1.08
                and
                b > g * 1.05
            ):
                return "blue"
            
            
            if (
                g > r * 1.05
                and
                g > b * 1.03
            ):
                return "green"
            
            if (
                r > 110
                and
                b > 110
                and
                abs(r - b) < 60
            ):
                return "purple"
            
            if r > 150 and g > 120 and b < 100:
                return "orange"
            
            if r < 85 and g < 85 and b < 85:
                return "black"

            if (
                abs(r - g) < 22
                and
                abs(g - b) < 22
                and
                70 < r < 210
            ):
                return "grey"


            return "unknown"

        except Exception as e:

            _logger.warning(

                f"[COLOR DETECTION ERROR] "

                f"{str(e)}"
            )

            return "unknown"


    #=================Centralized Rusable Image resolver==============

    def _resolve_asset_image(
        self,
        asset_pool,
        index
    ):

        try:

            if index is None:
                return False

            for asset in asset_pool:

                if isinstance(asset, dict):

                    if asset.get("index") == index:

                        return asset.get("image")

                elif isinstance(asset, str):

                    return asset

            return False

        except Exception as e:

            _logger.warning(

                f"[RESOLVE ASSET ERROR] "

                f"{str(e)}"
            )

            return False

    #============marchin AI===================================================
    # =====================================================
    # LEGACY IMAGE PAYLOAD MATCHER
    # Deprecated after migration to
    # index-based asset orchestration.
    # Keep temporarily for rollback safety.
    # =====================================================
    def match_image_with_ai(self, product_name, images):

        api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')
        client = OpenAI(api_key=api_key)

        if not images:
            return None

        # limit images for performance
        images = images[:5]
        image_inputs = [
        {
                "type": "input_image",
                "image_url": f"data:image/jpeg;base64,{img}"
            }
            for img in images
        ]

        prompt = f"""
        You are an expert product image matcher.

        Select the image that BEST represents this product:

        PRODUCT:
        {product_name}

        RULES:
        - Return ONLY the index (0-based integer)
        - No explanation
        - No text

            PRIORITY:
        1. Prefer isolated product on plain/white background
        2. Prefer centered single-product image
        3. Prefer image showing full product clearly
        4. Prefer clean studio product photos
        5. Prefer catalog hero product image
        6. Avoid lifestyle scenes if isolated image exists
        7. Avoid collages whenever possible
        8. Avoid infographic layouts
        9. Avoid text-heavy graphics
        10. Avoid multi-product overview images
        11. Avoid images containing large text blocks

        DO NOT PICK:
        - logos
        - icons
        - banners
        - cropped fragments
        - specification charts
        - text-heavy graphics
        - tiny thumbnails
        """

        try:
            response = client.responses.create(
                model="gpt-4.1",
                input=[{
                    "role": "user",
                    "content": [{"type": "input_text", "text": prompt}] + image_inputs
                }],
                timeout=30
            )

            result = response.output_text.strip()

            index = int(result)

            if 0 <= index < len(images):
                return images[index]

        except Exception as e:
            _logger.warning(f"AI IMAGE MATCH FAILED → {str(e)}")

        return None
    
    #======== returning images indexes========================================
    def match_image_index_with_ai( self, product_name, images):

        api_key = self.env[
            'ir.config_parameter'
        ].sudo().get_param(
            'openai.api.key'
        )

        client = OpenAI(api_key=api_key)

        if not images:
            return None


        filtered_images = []

        for img in images:

            try:

                if not img:
                    continue

                img_lower = img.lower()

                bad_keywords = [

                    "banner",

                    "lifestyle",

                    "infographic",

                    "specification",

                    "sizechart",

                    "dimensions"
                ]

                if any(
                    k in img_lower
                    for k in bad_keywords
                ):
                    continue

                filtered_images.append(img)

            except Exception:
                continue

        if filtered_images:

            images = filtered_images

        images = images[:8]

        image_inputs = []

        for idx, img in enumerate(images):

            image_inputs.append({
                "type": "input_text",
                "text": f"IMAGE INDEX: {idx}"
            })

            image_inputs.append({
                "type": "input_image",
                "image_url":
                    f"data:image/jpeg;base64,{img}"
            })

        prompt = f"""
        You are an ecommerce
        product image selector.

        PRODUCT:
        {product_name}

        Return ONLY the BEST
        image index.

        PRIORITY:
        - isolated product
        - plain background
        - centered object
        - clean catalog render

        AVOID:
        - people
        - lifestyle scenes
        - infographics
        - collages
        - banners
        - text-heavy graphics

        Return ONLY integer index.
        """

        try:

            response = client.responses.create(

                model="gpt-4.1",

                input=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": prompt
                        }
                    ] + image_inputs
                }],

                timeout=30
            )

            result = (
                response.output_text or ""
            ).strip()

            index = int(result)

            if 0 <= index < len(images):

                return index

        except Exception as e:

            _logger.warning(

                f"[IMAGE INDEX MATCH FAILED] "

                f"{str(e)}"
            )

        return None

    #============enforce translation=========================================
    def _force_translate(self, text, target_lang):

        from openai import OpenAI

        if not text:
            return text

        try:

            api_key = self.env[
                'ir.config_parameter'
            ].sudo().get_param(
                'openai.api.key'
            )

            if not api_key:

                _logger.warning(
                    "[OPENAI TRANSLATE] MISSING API KEY"
                )

                return text


            client = OpenAI(
                api_key=api_key
            )


            prompt = f"""
            Translate the following text into {target_lang}.

            Rules:
            - Return ONLY the translated text
            - Preserve formatting
            - Preserve product terminology
            - Do not explain anything

            TEXT:
            {text}
            """


            response = client.responses.create(

                model="gpt-4.1-mini",

                input=prompt
            )


            translated = (
                response.output_text or ''
            ).strip()


            if not translated:

                _logger.warning(
                    "[OPENAI TRANSLATE EMPTY]"
                )

                return text


            _logger.warning(

                f"[OPENAI TRANSLATION SUCCESS] "

                f"lang={target_lang}"
            )

            return translated


        except Exception as e:

            _logger.warning(

                f"[OPENAI TRANSLATE FAILED] "

                f"{str(e)}"
            )

            return text

    #=========Translation new logic==========================================

    def _apply_product_translation(self, product):

        if not product:
            return

        name = product.name or ''
        desc = product.description_sale or ''

        # ----------------------------
        # DEBUG
        # ----------------------------
        _logger.warning(
            f"[TRANSLATION INPUT] product={product.id} | name={name} | desc_len={len(desc)}"
        )

        # ----------------------------
        # ALWAYS TRANSLATE NAME (cheap)
        # ----------------------------
        ru_name = self._force_translate(name, "ru")
        az_name = self._force_translate(name, "az")

        # ----------------------------
        # ONLY TRANSLATE DESCRIPTION IF EXISTS
        # ----------------------------
        if desc and len(desc.strip()) > 10:

            ru_desc = self._smart_translate(desc, "ru")
            az_desc = self._smart_translate(desc, "az")

        else:
            ru_desc = ''
            az_desc = ''

            _logger.warning(
                f"[TRANSLATION SKIPPED DESC] product={product.id}"
            )

        # ----------------------------
        # SAVE
        # ----------------------------
        product.with_context(lang='ru_RU').write({
            'name': ru_name,
            'description_sale': ru_desc
        })

        product.with_context(lang='az_AZ').write({
            'name': az_name,
            'description_sale': az_desc
        })


        # =========================================
        # DETECT REAL TRANSLATION
        # =========================================

        translation_changed = False

        try:

            # =====================================
            # SAFE LANGUAGE
            # =====================================

            lang_code = 'ru_RU'

            translated_product = product.with_context(
                lang=lang_code
            )

            translated_name = translated_product.name or ''

            original_name = product.name or ''

            translated_desc = (
                translated_product.description_sale or ''
            )

            original_desc = (
                product.description_sale or ''
            )

            # =====================================
            # NAME CHANGED
            # =====================================

            if translated_name != original_name:

                translation_changed = True

            # =====================================
            # DESCRIPTION CHANGED
            # =====================================

            if translated_desc != original_desc:

                translation_changed = True

        except Exception as e:

            _logger.warning(

                f"[TRANSLATION CHECK FAILED] "

                f"{str(e)}"
            )


        if translation_changed:


            # =========================================
            # SHOW REAL TRANSLATED VALUES
            # =========================================

            try:

                ru_name = product.with_context(
                    lang='ru_RU'
                ).name or ''

                az_name = product.with_context(
                    lang='az_AZ'
                ).name or ''

            except Exception:

                ru_name = ''
                az_name = ''


            _logger.warning(

                f"[TRANSLATION SUCCESS] "

                f"product={product.id} | "

                f"RU={ru_name[:120]} | "

                f"AZ={az_name[:120]}"
            )

        else:

            _logger.warning(

                f"[TRANSLATION NO-CHANGE] "

                f"product={product.id}"
            )


    #============product translation extended================
    def _smart_translate(self, text, lang):

        # fallback if needed
        if len(text) < 5:
            return self._force_translate(text, lang)

        try:
            api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')
            client = OpenAI(api_key=api_key)

            prompt = f"Translate to {lang} and improve clarity:\n{text}"

            response = client.responses.create(
                model="gpt-4.1-mini",
                input=prompt
            )

            return response.output_text.strip()

        except:
            return self._force_translate(text, lang)


    # ================= PRODUCT CREATION URL ==================

    def create_products_url(self):

        import requests
        import base64
        import json

        if not self.ai_response:
            _logger.warning("NO AI RESPONSE")
            return

        product_obj = self.env['product.template']
        category_obj = self.env['product.category']

        try:
            products = json.loads(self.ai_response)

            if isinstance(products, dict):
                products = [products]

        except Exception as e:
            _logger.error(f"INVALID AI JSON → {str(e)}")
            return

        TOTAL_PRODUCTS = len(products)
        start_index = self.last_processed_product_index or 0

        _logger.warning(f"TOTAL AI PRODUCTS → {TOTAL_PRODUCTS}")
        _logger.warning(f"START INDEX → {start_index}")

        created_count = 0
        skipped_count = 0

        MAX_PRODUCTS_PER_RUN = 5

        CATEGORY_MAPPING = {
            "t-shirt": "Apparel",
            "shirt": "Apparel",
            "polo": "Apparel",
            "bag": "Bags",
            "backpack": "Bags",
            "cap": "Headwear",
            "hat": "Headwear",
            "bottle": "Drinkware",
            "cup": "Drinkware",
            "drinkware": "Drinkware",
            "pen": "Stationery",
            "notebook": "Stationery",
            "powerbank": "Electronics",
            "charger": "Electronics",
            "laptop": "Electronics",
            "football": "Football Fever",
            "Wristband": "Football Fever",
            "sports t-shirt": "Football Fever",
            "sports towel": "Football Fever",
            "Sports Bottles": "Football Fever"
        }

        parent_category = category_obj.search([('name', '=', "All Products")], limit=1)
        #vendor_id = self.partner_id.id if self.partner_id else False
        vendor_id =  self.partner_id.id if self.partner_id else self.env.user.partner_id.id

        if not parent_category:
            parent_category = category_obj.create({'name': "All Products"})

        end_index = min(start_index + MAX_PRODUCTS_PER_RUN, TOTAL_PRODUCTS)

        _logger.warning(f"PROCESSING RANGE → {start_index} to {end_index}")

        for idx in range(start_index, end_index):

            product_data = products[idx]

            name = product_data.get("name")
            if not name:
                skipped_count += 1
                continue

            description = product_data.get("description", "")
            raw_category = (product_data.get("category") or "").lower()

            # ================= CATEGORY =================
            mapped_category = "General"
            for key, val in CATEGORY_MAPPING.items():
                if key in raw_category:
                    mapped_category = val
                    break

            category = category_obj.search([
                ('name', '=ilike', mapped_category),
                ('parent_id', '=', parent_category.id)
            ], limit=1)

            if not category:
                category = category_obj.create({
                    'name': mapped_category,
                    'parent_id': parent_category.id
                })

            # ================= FINGERPRINT=================

            variant_group = (

                product_data.get("variant_group")

                or

                name
            )

            variant_group = str(
                variant_group
            ).strip().upper()


            vendor_fingerprint = (
                f"{vendor_id}_{variant_group}"
            )


            # ================= DUPLICATE CHECK =================

            existing = product_obj.search([

                (
                    'vendor_fingerprint',
                    '=',
                    vendor_fingerprint
                )

            ], limit=1)


            if existing:

                _logger.warning(

                    f"[URL DUPLICATE SKIP] "

                    f"{vendor_fingerprint}"
                )

                skipped_count += 1

                continue

            vals = {
                'name': name.strip(),
                'description_sale': description,
                'type': 'consu',
                'categ_id': category.id,
                'sale_ok': True,
                'website_published': False,
                'vendor_id': vendor_id,
                'vendor_fingerprint': vendor_fingerprint,
                'vendor_import_job_id': self.id,
            }

            # ================= IMAGE =================
            image_url = product_data.get("image")

            if image_url and isinstance(image_url, str) and image_url.startswith("http"):

                try:
                    _logger.warning(f"FETCHING IMAGE → {image_url}")

                    res = requests.get(image_url, timeout=5, stream=True)

                    if res.status_code != 200:
                        _logger.warning(f"IMAGE HTTP ERROR → {res.status_code}")
                        continue

                    content_type = res.headers.get("Content-Type", "")
                    if "image" not in content_type:
                        _logger.warning(f"NOT AN IMAGE → {content_type}")
                        continue

                    content = res.raw.read(500000, decode_content=True)

                    if not content:
                        _logger.warning("EMPTY IMAGE CONTENT")
                        continue

                    vals['image_1920'] = base64.b64encode(content).decode("utf-8")

                    _logger.warning("IMAGE STORED SUCCESSFULLY")

                except Exception as e:
                    _logger.warning(f"IMAGE FAILED → {str(e)}")

            else:
                _logger.warning(f"NO VALID IMAGE URL → {image_url}")

            # ================= CREATE =================
            try:
                _logger.warning(
                    f"[URL CREATE PRODUCT] → "
                    f"NAME={name} | "
                    f"VENDOR_ID={vendor_id}"
                )

                # product_obj.create(vals)
                product = product_obj.with_context(
                    mail_create_nolog=True,
                    mail_notify_force_send=False,
                    tracking_disable=True
                ).create(vals)

                self._apply_product_translation(product)

                # =========================================
                # URL VARIANTS
                # =========================================

                variants = product_data.get(
                    "variants",
                    []
                )


                # =========================================
                # FALLBACK VARIANT
                # =========================================

                if not variants:

                    variants = [{

                        "attributes": {

                            "Variant": name
                        }

                    }]


                # =========================================
                # PROCESS VARIANTS
                # =========================================

                for variant in variants:

                    attributes = variant.get(
                        "attributes",
                        {}
                    )


                    for attr_name, attr_value in attributes.items():


                        if not attr_value:
                            continue


                        attr_value = str(attr_value).strip()


                        # =====================================
                        # NORMALIZE BAD VARIANTS
                        # =====================================

                        bad_variants = [

                            'variant 1',

                            'variant 2',

                            'variant 3',

                            'default',

                            'option a',

                            'option b'
                        ]


                        if attr_value.lower() in bad_variants:

                            attr_name = "Design"

                            attr_value = name


                        # =====================================
                        # ATTRIBUTE
                        # =====================================

                        attribute = self.env[
                            'product.attribute'
                        ].search([

                            ('name', '=', attr_name)

                        ], limit=1)

                        # =====================================
                        # TRANSLATE ATTRIBUTE NAME
                        # =====================================

                        try:

                            for lang_code in [

                                'ru_RU',

                                'az_AZ'
                            ]:

                                translated_attr = self._force_translate(

                                    str(attr_name),

                                    lang_code
                                )


                                if translated_attr:

                                    attribute.with_context(
                                        lang=lang_code
                                    ).write({

                                        'name': translated_attr
                                    })


                                    _logger.warning(

                                        f"[URL ATTRIBUTE TRANSLATED] "

                                        f"{attr_name} "

                                        f"-> "

                                        f"{translated_attr} "

                                        f"({lang_code})"
                                    )

                        except Exception as e:

                            _logger.warning(

                                f"[URL ATTRIBUTE TRANSLATION ERROR] "

                                f"{str(e)}"
                            )


                        if not attribute:

                            attribute = self.env[
                                'product.attribute'
                            ].create({

                                'name': attr_name
                            })


                        # =====================================
                        # ATTRIBUTE VALUE
                        # =====================================

                        value = self.env[
                            'product.attribute.value'
                        ].search([

                            ('name', '=', attr_value),

                            (
                                'attribute_id',
                                '=',
                                attribute.id
                            )

                        ], limit=1)


                        if not value:

                            value = self.env[
                                'product.attribute.value'
                            ].create({

                                'name': attr_value,

                                'attribute_id':
                                    attribute.id
                            })


                            # =================================
                            # TRANSLATE VARIANT VALUE
                            # =================================

                            try:

                                for lang_code in [

                                    'ru_RU',

                                    'az_AZ'
                                ]:

                                    translated_variant = (

                                        self._force_translate(

                                            str(attr_value),

                                            lang_code
                                        )
                                    )


                                    if translated_variant:

                                        value.with_context(
                                            lang=lang_code
                                        ).write({

                                            'name':
                                                translated_variant
                                        })


                                        _logger.warning(

                                            f"[URL VARIANT TRANSLATED] "

                                            f"{attr_value} "

                                            f"-> "

                                            f"{translated_variant} "

                                            f"({lang_code})"
                                        )

                            except Exception as e:

                                _logger.warning(

                                    f"[URL VARIANT TRANSLATION ERROR] "

                                    f"{str(e)}"
                                )


                        # =====================================
                        # ATTRIBUTE LINE
                        # =====================================

                        line = self.env[
                            'product.template.attribute.line'
                        ].search([

                            (
                                'product_tmpl_id',
                                '=',
                                product.id
                            ),

                            (
                                'attribute_id',
                                '=',
                                attribute.id
                            )

                        ], limit=1)


                        if not line:

                            self.env[
                                'product.template.attribute.line'
                            ].create({

                                'product_tmpl_id':
                                    product.id,

                                'attribute_id':
                                    attribute.id,

                                'value_ids': [(6, 0, [

                                    value.id

                                ])]
                            })

                        else:

                            if (

                                value.id

                                not in

                                line.value_ids.ids

                            ):

                                line.value_ids = [

                                    (4, value.id)

                                ]
               
                created_count += 1

            except Exception as e:
                _logger.error(f"CREATE FAILED → {name} | {str(e)}")
                skipped_count += 1
                continue

            if created_count % 10 == 0:
                self._safe_commit_progress()

        # ================= SAVE PROGRESS =================
        self.last_processed_product_index = end_index

        _logger.warning(f"CREATED THIS RUN → {created_count}")
        _logger.warning(f"SKIPPED THIS RUN → {skipped_count}")
        _logger.warning(f"NEXT START INDEX → {self.last_processed_product_index}")

        if self.last_processed_product_index >= TOTAL_PRODUCTS:
            _logger.warning("ALL PRODUCTS CREATED ✅")
            self.state = "done"
        else:
            _logger.warning("MORE PRODUCTS REMAIN → CONTINUE CREATION")
            self.state = "url_creating"

        self._safe_commit_progress()


    #==========create pdf product====================================

    def create_products_pdf(self):

        import json

        _logger.warning(
            "[PDF CREATE] START"
        )

        if not self.ai_response:

            _logger.warning(
                "[PDF CREATE] NO AI RESPONSE"
            )

            return

        try:

            ai_pages = json.loads(
                self.ai_response
            )

        except Exception as e:

            _logger.error(

                f"[PDF CREATE] INVALID AI JSON "

                f"| {str(e)}"
            )

            return

        if not isinstance(ai_pages, list):

            _logger.warning(
                "[PDF CREATE] INVALID AI FORMAT"
            )

            return

        product_obj = self.env[
            'product.template'
        ]

        category_obj = self.env[
            'product.category'
        ]

        stock_quant_obj = self.env[
            'stock.quant'
        ]

        stock_location = self.env[
            'stock.location'
        ].search([

            ('usage', '=', 'internal')

        ], limit=1)

        CATEGORY_MAPPING = {

            "t-shirt": "Apparel",
            "shirt": "Apparel",
            "polo": "Apparel",

            "bag": "Bags",
            "backpack": "Bags",

            "cap": "Headwear",
            "hat": "Headwear",

            "bottle": "Drinkware",
            "drinkware": "Drinkware",

            "pen": "Stationery",
            "notebook": "Stationery",

            "powerbank": "Electronics",
            "charger": "Electronics",
            "laptop": "Electronics",
        }

        parent_category = category_obj.search([

            ('name', '=', "All Products")

        ], limit=1)

        if not parent_category:

            parent_category = category_obj.create({

                'name': "All Products"

            })

        vendor_id = (

            self.partner_id.id
            if self.partner_id
            else self.env.user.partner_id.id
        )

        BATCH_SIZE = 3

        start = (
            self.last_created_page or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(ai_pages)
        )

        created_count = 0
        skipped_count = 0

        for page_index in range(start, end):

            try:

                page_data = ai_pages[
                    page_index
                ]

            except Exception as e:

                _logger.warning(

                    f"[PDF PAGE LOAD ERROR] "

                    f"{str(e)}"
                )

                continue

            page_number = page_data.get(
                "page"
            )

            page_record = self.env[
                'vendor.import.page'
            ].search([

                ('job_id', '=', self.id),

                ('page_number', '=', page_number)

            ], limit=1)


            # =====================================
            # LOAD AI-PERSISTED IMAGES
            # =====================================

            page_images = page_data.get(
                "images",
                []
            )

            if not page_images:

                _logger.warning(

                    f"[PDF CREATE] "

                    f"NO IMAGES FOUND "

                    f"| PAGE {page_number}"
                )


            products = page_data.get(
                "products",
                []
            )

            for product_data in products:

                # =====================================
                # PRODUCT IMAGE PREP
                # =====================================

                product_images = product_data.get(
                    "images",
                    []
                )

                _logger.warning(

                    f"[PRODUCT IMAGE COUNT] "

                    f"{product_data.get('name')} "

                    f"| images={len(product_images)}"
                )

                # =====================================
                # FALLBACK TO PAGE IMAGES
                # =====================================

                if not product_images:

                    # -----------------------------------------
                    # SAFE PAGE FALLBACK
                    # -----------------------------------------

                    product_images = []

                    variant_count = max(
                        len(product_data.get("variants", [])),
                        1
                    )

                    # prefer highest gallery assets
                    sorted_page_assets = sorted(

                        page_images,

                        key=lambda x: (

                            x.get("gallery_score", 0),

                            x.get("hero_score", 0)

                        ),

                        reverse=True
                    )

                    # allocate enough assets
                    allocation_size = min(

                        max(variant_count * 2, 4),

                        len(sorted_page_assets)
                    )

                    product_images = sorted_page_assets[
                        :allocation_size
                    ]

                    _logger.warning(

                        f"[SMART PAGE FALLBACK] "

                        f"{product_data.get('name')} "

                        f"| allocated={len(product_images)}"

                    )

                segmented_assets = []

                for img in product_images:

                    # ---------------------------------
                    # ALREADY STRUCTURED
                    # ---------------------------------

                    if isinstance(img, dict):

                        if img.get("image"):

                            segmented_assets.append(img)

                    # ---------------------------------
                    # RAW BASE64 FALLBACK
                    # ---------------------------------

                    elif isinstance(img, str):

                        segmented_assets.append({

                            "image": img,

                            "score": 0,

                            "is_collage": False
                        })

                # =====================================
                # BUILD ASSET POOL
                # =====================================

                asset_pool = self._prepare_asset_pool(
                    segmented_assets
                )

                _logger.warning(

                    f"[PDF ASSET POOL] "

                    f"product={product_data.get('name')} "

                    f"| assets={len(asset_pool)}"
                )

                try:

                    name = (

                        product_data.get(
                            "name"
                        )

                        or ""

                    ).strip()

                    if not name:

                        continue

                    raw_category = (

                        product_data.get(
                            "category"
                        ) or ""

                    ).lower()

                    variants = product_data.get(
                        "variants",
                        []
                    )

                    variant_group = (

                        product_data.get(
                            "variant_group"
                        )

                        or

                        name
                    )

                    variant_group = str(
                        variant_group
                    ).strip().upper()

                    category = (
                        self._get_or_create_pdf_category(

                            raw_category,

                            category_obj,

                            parent_category,

                            CATEGORY_MAPPING
                        )
                    )

                    vendor_fingerprint = (
                        f"{vendor_id}_{variant_group}"
                    )

                    product, created = (

                        self._get_or_create_pdf_product(

                            product_data,

                            variant_group,

                            vendor_id,

                            vendor_fingerprint,

                            category,

                            asset_pool,

                            product_obj
                        )
                    )


                    if created:

                        self._apply_product_translation(
                            product
                        )

                        self._create_pdf_gallery(

                            product,

                            product_data,

                            asset_pool
                        )

                        # =====================================
                        # APPLY REAL INVENTORY STOCK
                        # ONLY FOR STORABLE PRODUCTS
                        # =====================================

                        try:

                            stock_qty = int(

                                product_data.get(
                                    "stock_qty",
                                    0
                                ) or 0
                            )

                            # =====================================
                            # CONSUMABLE PRODUCTS:
                            # SKIP STOCK QUANTS
                            # =====================================

                            if stock_qty > 0:

                                quant = stock_quant_obj.search([

                                    (
                                        'product_id',
                                        '=',
                                        product.product_variant_id.id
                                    ),

                                    (
                                        'location_id',
                                        '=',
                                        stock_location.id
                                    )

                                ], limit=1)

                                if quant:

                                    quant.inventory_quantity = (
                                        stock_qty
                                    )

                                    quant.action_apply_inventory()

                                else:

                                    quant = stock_quant_obj.create({

                                        'product_id':
                                            product.product_variant_id.id,

                                        'location_id':
                                            stock_location.id,

                                        'inventory_quantity':
                                            stock_qty
                                    })

                                    quant.action_apply_inventory()

                                _logger.warning(

                                    f"[STOCK APPLIED] "

                                    f"{product.name} "

                                    f"| qty={stock_qty}"
                                )

                        except Exception as e:

                            _logger.warning(

                                f"[STOCK APPLY FAILED] "

                                f"{str(e)}"
                            )

                        created_count += 1

                    else:

                        skipped_count += 1

                    if not variants:

                        variants = [{

                            "attributes": {

                                "Variant": name
                            }

                        }]

                    # =======================================
                    # PASS 1:
                    # BUILD ALL ATTRIBUTE LINES FIRST
                    # =======================================

                    for variant in variants:

                        attributes = variant.get(
                            "attributes",
                            {}
                        )

                        for attr_name, attr_value in attributes.items():

                            if not attr_value:
                                continue

                            attribute, value = (

                                self._get_or_create_attribute_and_value(

                                    attr_name,

                                    attr_value
                                )
                            )

                            line = self.env[
                                'product.template.attribute.line'
                            ].search([

                                (
                                    'product_tmpl_id',
                                    '=',
                                    product.id
                                ),

                                (
                                    'attribute_id',
                                    '=',
                                    attribute.id
                                )

                            ], limit=1)

                            if not line:

                                self.env[
                                    'product.template.attribute.line'
                                ].create({

                                    'product_tmpl_id':
                                        product.id,

                                    'attribute_id':
                                        attribute.id,

                                    'value_ids': [(6, 0, [

                                        value.id

                                    ])]
                                })

                            else:

                                if (

                                    value.id

                                    not in

                                    line.value_ids.ids

                                ):

                                    line.value_ids = [

                                        (4, value.id)

                                    ]

                    # =======================================
                    # PASS 2:
                    # GENERATE ALL VARIANTS ONCE
                    # =======================================

                    product._create_variant_ids()

                    used_asset_indexes = set()

                    # =======================================
                    # PASS 3:
                    # MATCH REAL VARIANTS TO IMAGES
                    # =======================================

                    for variant in variants:

                        # =====================================
                        # MATCH REAL GENERATED VARIANT
                        # =====================================

                        variant_record = False

                        product_variants = (
                            product.product_variant_ids
                        )

                        variant_name = ""

                        attributes = variant.get(
                            "attributes",
                            {}
                        )

                        if isinstance(attributes, dict):

                            variant_name = " ".join([

                                str(v)

                                for v in attributes.values()

                            ]).lower()

                        for pv in product_variants:

                            combo = " ".join([

                                v.name.lower()

                                for v in (
                                    pv.product_template_variant_value_ids
                                )

                            ])

                            if combo:

                                combo_words = combo.split()

                                variant_words = (
                                    variant_name.split()
                                )

                                match_count = 0

                                for word in variant_words:

                                    if word in combo_words:

                                        match_count += 1


                                required_matches = max(

                                    1,

                                    min(
                                        len(variant_words),
                                        2
                                    )
                                )

                                if (

                                    variant_words

                                    and

                                    match_count >= required_matches
                                ):

                                    variant_record = pv
                                    break

                        # ---------------------------------
                        # SAFE FALLBACK
                        # ---------------------------------
                      
                        if (
                            not variant_record
                            and
                            len(product_variants) == 1
                        ):

                            variant_record = (
                                product_variants[0]
                            )

                        # =====================================
                        # PROFESSIONAL VARIANT IMAGE MATCHING
                        # =====================================

                        if variant_record:

                            try:
                                
                                matched_asset = self._match_variant_image(

                                    variant,

                                    asset_pool,

                                    used_asset_indexes
                                )

                                # =====================================
                                # APPLY
                                # =====================================

                                if matched_asset:

                                    variant_record.image_1920 = (
                                        matched_asset.get(
                                            "image"
                                        )
                                    )

                                    # used_asset_indexes.add(
                                    #     matched_asset.get("index")
                                    # )

                                    asset_index = (
                                        matched_asset.get("clean_index")
                                        if matched_asset.get("clean_index") is not None
                                        else matched_asset.get("index")
                                    )

                                    if asset_index is not None:

                                        used_asset_indexes.add(asset_index)

                                    _logger.warning(

                                        f"[VARIANT IMAGE APPLIED] "

                                        f"{variant_name} "

                                        f"| asset={matched_asset.get('index')}"
                                    )

                            except Exception as e:

                                _logger.warning(

                                    f"[VARIANT IMAGE FAILED] "

                                    f"{str(e)}"
                                )

                except Exception as e:

                    _logger.exception(

                        f"[PDF PRODUCT ERROR] "

                        f"{str(e)}"
                    )

                    continue

            try:

                self.last_created_page = (
                    page_index + 1
                )

                self._safe_commit_progress()

            except Exception as e:

                _logger.exception(

                    f"[PAGE COMMIT FAILED] "

                    f"{str(e)}"
                )

        _logger.warning(

            f"[PDF CREATE COMPLETE] "

            f"created={created_count} "

            f"| skipped={skipped_count}"
        )

        if self.last_created_page >= len(ai_pages):

            self.state = 'done'

        else:

            self.state = 'pdf_creating'

        self._safe_commit_progress()


    #==========pdf product PRODUCT CREATE/GET====================================
    
    def _get_or_create_pdf_product(

        self,

        product_data,

        variant_group,

        vendor_id,

        vendor_fingerprint,

        category,

        asset_pool,

        product_obj
    ):

        product = product_obj.search([

            (
                'vendor_fingerprint',
                '=',
                vendor_fingerprint
            )

        ], limit=1)

        if product:

            return product, False

        # =====================================
        # BUILD PROFESSIONAL DESCRIPTION
        # =====================================

        subtitle = (
            product_data.get("subtitle")
            or ""
        ).strip()

        description = (
            product_data.get("description")
            or ""
        ).strip()

        material = (
            product_data.get("material")
            or ""
        ).strip()

        dimensions = (
            product_data.get("dimensions")
            or ""
        ).strip()

        bullet_features = (
            product_data.get("bullet_features")
            or []
        )

        # =====================================
        # CLEAN BULLETS
        # =====================================

        clean_bullets = []

        for bullet in bullet_features:

            if not bullet:
                continue

            bullet = str(bullet).strip()

            if len(bullet) < 2:
                continue

            clean_bullets.append(bullet)

        # =====================================
        # BUILD HTML DESCRIPTION
        # =====================================

        description_parts = []

        if subtitle:

            description_parts.append(
                f"<h4>{subtitle}</h4>"
            )

        if description:

            description_parts.append(
                f"<p>{description}</p>"
            )

        if material:

            description_parts.append(
                f"<p><strong>Material:</strong> "
                f"{material}</p>"
            )

        if dimensions:

            description_parts.append(
                f"<p><strong>Dimensions:</strong> "
                f"{dimensions}</p>"
            )

        if clean_bullets:

            bullet_html = "".join([

                f"<li>{b}</li>"

                for b in clean_bullets
            ])

            description_parts.append(

                f"<ul>{bullet_html}</ul>"
            )

        rich_description = "<br/>".join(
            description_parts
        )

        vals = {

            'name': (
                product_data.get("name")
                or ""
            ).strip(),

           'default_code': (
                product_data.get("product_code")
                or variant_group
            ),

           'description_sale': rich_description,

            'type': 'consu',

            'categ_id': category.id,

            'sale_ok': True,

            'website_published': False,

            'vendor_id': vendor_id,

            'vendor_fingerprint':
                vendor_fingerprint,

            'vendor_import_job_id':
                self.id,

            'vendor_stock_qty': int(

                product_data.get(
                    "stock_qty",
                    0
                ) or 0
            ),

            'list_price': self._safe_parse_price(
                product_data.get("price")
            ),
        }

        hero_index = product_data.get(
            "hero_image_index"
        )
        
        # =====================================
        # PROFESSIONAL HERO IMAGE SELECTION
        # =====================================

        hero_asset = None

        # =====================================
        # AI SELECTED HERO
        # =====================================

        if hero_index is not None:

            for asset in asset_pool:

                if asset.get("clean_index") == hero_index:

                    # reject collages as hero
                    if asset.get("is_collage"):

                        continue

                    hero_asset = asset

                    break

        # =====================================
        # FALLBACK TO BEST CLEAN IMAGE
        # =====================================

        if not hero_asset:

            sorted_assets = sorted(

                asset_pool,

                key=lambda x: x.get(
                    "score",
                    0
                ),

                reverse=True
            )

            for asset in sorted_assets:

                # reject collage sheets
                if asset.get("is_collage"):
                    continue

                # require strong quality
                if asset.get("score", 0) >= 45:

                    hero_asset = asset

                    break

        # =====================================
        # FINAL SAFE FALLBACK
        # =====================================

        if not hero_asset and asset_pool:

            for asset in asset_pool:

                if not asset.get("is_collage"):

                    hero_asset = asset

                    break

            if not hero_asset:

                hero_asset = asset_pool[0]

        # =====================================
        # APPLY HERO IMAGE
        # =====================================

        if hero_asset:

            vals['image_1920'] = hero_asset.get(
                "image"
            )

            _logger.warning(

                f"[PDF HERO APPLIED] "

                f"score={hero_asset.get('score')} "

                f"color={hero_asset.get('dominant_color')}"
            )

        product = product_obj.with_context(

            mail_create_nolog=True,

            mail_notify_force_send=False,

            tracking_disable=True

        ).create(vals)

        return product, True


    #==========create pdf CATEGORY RESOLVER====================================
    
    def _get_or_create_pdf_category(

        self,

        raw_category,

        category_obj,

        parent_category,

        category_mapping
    ):

        mapped_category = "General"

        raw_category = (
            raw_category or ""
        ).lower()

        for key, val in category_mapping.items():

            if key in raw_category:

                mapped_category = val

                break

        category = category_obj.search([

            ('name', '=ilike', mapped_category),

            (
                'parent_id',
                '=',
                parent_category.id
            )

        ], limit=1)

        if not category:

            category = category_obj.create({

                'name': mapped_category,

                'parent_id': parent_category.id
            })

        return category

    #=========pdf product GALLERY CREATOR=======================
    def _create_pdf_gallery(

        self,

        product,

        product_data,

        asset_pool
    ):

        gallery_indexes = product_data.get(
            "gallery_image_indexes",
            []
        )

        # =====================================
        # FALLBACK GALLERY EXPANSION
        # =====================================

        if (

            len(gallery_indexes) < 3

            and

            asset_pool
        ):

            sorted_gallery_assets = sorted(

                asset_pool,

                key=lambda x: (

                    x.get(
                        "gallery_score",
                        x.get(
                            "score",
                            0
                        )
                    ),

                    x.get(
                        "hero_score",
                        0
                    )

                ),

                reverse=True
            )

            extra_indexes = [

                a.get("clean_index")

                for a in sorted_gallery_assets

                if (

                    a.get(
                        "gallery_score",
                        a.get(
                            "score",
                            0
                        )
                    ) >= 28

                )
            ]

            gallery_indexes.extend(
                extra_indexes
            )

            gallery_indexes = list(

                dict.fromkeys(
                    gallery_indexes
                )
            )[:10]

        used_images = set()
        used_hashes = set()

        existing_gallery = self.env[
            'product.image'
        ].search([

            ('product_tmpl_id', '=', product.id)

        ])

        for g in existing_gallery:

            if g.image_1920:

                used_hashes.add(

                    hashlib.md5(

                        g.image_1920.encode('utf-8')

                    ).hexdigest()

                )

        if product.image_1920:

            used_images.add(
                product.image_1920
            )

        # =====================================
        # ENSURE MINIMUM GALLERY RICHNESS
        # =====================================

        if len(gallery_indexes) < 4:

            fallback_indexes = [

                a.get("clean_index")

                for a in sorted(

                    asset_pool,

                    key=lambda x: (

                        x.get(
                            "gallery_score",
                            0
                        ),

                        x.get(
                            "score",
                            0
                        )

                    ),

                    reverse=True

                )

            ]

            gallery_indexes.extend(
                fallback_indexes
            )

            gallery_indexes = list(

                dict.fromkeys(
                    gallery_indexes
                )

            )[:10]

        for index in gallery_indexes:

            try:

                gallery_image = (
                    self._resolve_asset_image(

                        asset_pool,

                        index
                    )
                )

                if not gallery_image:
                    continue

                image_hash = hashlib.md5(

                    gallery_image.encode('utf-8')

                ).hexdigest()

                if image_hash in used_hashes:
                    continue

                self.env[
                    'product.image'
                ].create({

                    'name':
                        f"{product.name} Gallery",

                    'product_tmpl_id':
                        product.id,

                    'image_1920':
                        gallery_image
                })

                used_hashes.add(
                    image_hash
                )

            except Exception as e:

                _logger.warning(

                    f"[GALLERY IMAGE FAILED] "

                    f"{product.name} "

                    f"| {str(e)}"
                )
    
    #=========pdf product STOCK APPLY=======================
    def _apply_pdf_stock(

        self,

        variant_record,

        stock_qty,

        stock_quant_obj,

        stock_location
    ):

        if (

            not stock_qty

            or

            not variant_record

            or

            not stock_location
        ):

            return

        try:

            quant = stock_quant_obj.search([

                (
                    'product_id',
                    '=',
                    variant_record.id
                ),

                (
                    'location_id',
                    '=',
                    stock_location.id
                )

            ], limit=1)

            if quant:

                quant.inventory_quantity = (
                    stock_qty
                )

                quant.action_apply_inventory()

            else:

                quant = stock_quant_obj.create({

                    'product_id':
                        variant_record.id,

                    'location_id':
                        stock_location.id,

                    'inventory_quantity':
                            stock_qty
                })

                quant.action_apply_inventory()

            _logger.warning(

                f"[PDF STOCK SET] "

                f"{variant_record.display_name} "

                f"-> {stock_qty}"
            )

        except Exception as e:

            _logger.warning(

                f"[PDF STOCK FAILED] "

                f"{str(e)}"
            )

    #===============fingerprint================================
    def _build_vendor_fingerprint(self, product_data):

        import re
        import hashlib

        name = (
            product_data.get("name") or ""
        ).strip().lower()

        sku = (
            product_data.get("sku")
            or product_data.get("code")
            or product_data.get("product_code")
            or ""
        ).strip().lower()

        url = (
            product_data.get("url")
            or product_data.get("link")
            or ""
        ).strip().lower()

        # normalize
        def clean(v):
            return re.sub(r'[^a-z0-9]', '', v or '')

        base = "|".join([
            clean(name),
            clean(sku),
            clean(url),
        ])

        return hashlib.md5(
            base.encode("utf-8")
        ).hexdigest()

    #==========Excel url detect workflo=======================
    def _extract_product_url(self, row):

        possible_keys = [

            "url",
            "link",
            "product_url",
            "product link",
            "website",
            "href",

        ]

        for key in possible_keys:

            value = row.get(key)

            if not value:
                continue

            value = str(value).strip()

            if value.startswith(
                ("http://", "https://")
            ):
                return value

        return False
    
    #======Excel url detection router=======================
    def _route_excel_rows(self, products):

        normal_products = []
        url_products = []

        for row in products:

            url = self._extract_product_url(row)

            if url:

                row["detected_url"] = url

                url_products.append(row)

            else:

                normal_products.append(row)

        return {
            "normal": normal_products,
            "url": url_products,
        }

    
    #==========Excel URl queue logic========================
    
    def _queue_excel_urls(self, url_products):

        import json

        if not url_products:
            return

        seen = set()
        cleaned = []

        for row in url_products:

            url = row.get("detected_url")

            if not url:
                continue

            if url in seen:
                continue

            seen.add(url)

            cleaned.append(row)

        url_products = cleaned

        self.excel_url_queue = json.dumps(
            url_products
        )

        self.excel_url_processing = True

        self.excel_url_index = 0

    
    #============Excel URL processor==========================
    def process_excel_url_queue(self):

        import json

        if not self.excel_url_queue:

            _logger.warning(
                "[URL QUEUE] EMPTY"
            )

            return


        rows = json.loads(
            self.excel_url_queue
        )


        start = self.excel_url_index or 0

        BATCH_SIZE = 5

        end = min(
            start + BATCH_SIZE,
            len(rows)
        )


        _logger.warning(

            f"[URL QUEUE START] "

            f"{start} -> {end} | "

            f"total={len(rows)}"
        )

        vendor_id = (
             self.partner_id.id
            if self.partner_id
            else self.env.user.partner_id.id
        )

        for idx in range(start, end):

            try:

                row = rows[idx]

                product_url = row.get(
                    "detected_url"
                )


                if not product_url:

                    _logger.warning(
                        f"[URL QUEUE SKIP] "
                        f"NO URL AT INDEX {idx}"
                    )

                    continue

                existing_job = self.env[
                    'vendor.import.job'
                ].search([

                    ('data_url', '=', product_url),

                    ('state', '!=', 'failed')

                ], limit=1)


                if existing_job:

                    _logger.warning(

                        f"[URL JOB EXISTS] "

                        f"{product_url}"
                    )

                    # ====================================
                    # ADVANCE QUEUE INDEX
                    # ====================================

                    self.excel_url_index = idx + 1

                    self._safe_commit_progress()

                    continue

                # ====================================
                # CREATE ISOLATED URL JOB
                # ====================================

                new_job = self.env[
                    'vendor.import.job'
                ].create({

                    'name':
                        f"URL Import - {idx}",

                    'partner_id':
                        vendor_id,

                    'source_type':
                        'url',

                    'data_url':
                        product_url,

                    'state':
                        'url_scraping',
                })


                _logger.warning(

                    f"[URL JOB CREATED] "

                    f"job={new_job.id} | "

                    f"url={product_url}"
                )


                # ====================================
                # SAVE PROGRESS
                # ====================================

                self.excel_url_index = idx + 1

                self._safe_commit_progress()


            except Exception as e:

                _logger.exception(
                    f"[EXCEL URL ERROR] {str(e)}"
                )

                self.env.cr.rollback()

 
        # =========================================
        # COMPLETE
        # =========================================

        if self.excel_url_index >= len(rows):

            _logger.warning(
                "[URL QUEUE COMPLETE]"
            )

            self.excel_url_queue = False

            self.excel_url_processing = False

            self.excel_url_index = 0

            self._safe_commit_progress()

            self.env.invalidate_all()


    #==========create excel product=================================

    def create_products_excel(self):

        import json
        import re

        _logger.warning(
            "[EXCEL CREATE] START"
        )


        # =====================================================
        # VALIDATION
        # =====================================================

        if not self.ai_response:

            _logger.warning(
                "[EXCEL CREATE] NO AI RESPONSE"
            )

            return


        try:

            ai_pages = json.loads(
                self.ai_response or "[]"
            )

        except Exception as e:

            _logger.exception(

                f"[EXCEL CREATE] "

                f"INVALID AI JSON "

                f"| {str(e)}"
            )

            return


        if not ai_pages:

            _logger.warning(
                "[EXCEL CREATE] EMPTY AI"
            )

            return


        ai_page = ai_pages[0]

        products = ai_page.get(
            "products",
            []
        )


        _logger.warning(

            f"[EXCEL CREATE] "

            f"RAW PRODUCTS={len(products)}"
        )


        if not products:

            return


        # =====================================================
        # MODELS
        # =====================================================

        product_obj = self.env[
            'product.template'
        ]

        category_obj = self.env[
            'product.category'
        ]

        attribute_obj = self.env[
            'product.attribute'
        ]

        attribute_value_obj = self.env[
            'product.attribute.value'
        ]

        line_obj = self.env[
            'product.template.attribute.line'
        ]


        # =====================================================
        # ROOT CATEGORY
        # =====================================================

        parent_category = category_obj.search([

            ('name', '=', "All Products")

        ], limit=1)


        if not parent_category:

            parent_category = category_obj.create({

                'name': "All Products"

            })


        # =====================================================
        # CATEGORY MAP
        # =====================================================

        CATEGORY_MAPPING = {

            "t-shirt": "Apparel",
            "shirt": "Apparel",
            "polo": "Apparel",
            "bag": "Bags",
            "backpack": "Bags",
            "cap": "Headwear",
            "hat": "Headwear",
            "bottle": "Drinkware",
            "drinkware": "Drinkware",
            "pen": "Stationery",
            "notebook": "Stationery",
            "powerbank": "Electronics",
            "charger": "Electronics",
            "laptop": "Electronics",
        }


        # =====================================================
        # GROUP PRODUCTS
        # =====================================================

        grouped_products = {}


        for p in products:

            raw_name = (
                p.get("name") or ""
            ).strip()


            variant_group = (
                p.get("variant_group")
            )


            if variant_group:

                group_id = str(
                    variant_group
                ).strip().upper()

            else:

                match = re.search(

                    r'(?:Product\s*)?([A-Z]*\d+)',

                    raw_name,

                    re.I
                )


                if match:

                    group_id = (
                        match.group(1)
                        .upper()
                    )

                else:

                    group_id = (
                        raw_name.upper()
                    )


            grouped_products.setdefault(

                group_id,

                []

            ).append(p)


        grouped_keys = list(
            grouped_products.keys()
        )


        _logger.warning(

            f"[EXCEL GROUPS] "

            f"TOTAL={len(grouped_keys)}"
        )


        # =====================================================
        # BATCH GROUPS
        # =====================================================

        BATCH_SIZE = 10

        start = (
            self.excel_created_index or 0
        )

        end = min(

            start + BATCH_SIZE,

            len(grouped_keys)
        )


        _logger.warning(

            f"[EXCEL BATCH] "

            f"{start} → {end}"
        )


        created_count = 0
        merged_count = 0


        # =====================================================
        # PROCESS GROUPS
        # =====================================================

        for group_idx in range(start, end):

            try:

                group_id = grouped_keys[
                    group_idx
                ]

                group_items = grouped_products[
                    group_id
                ]


                _logger.warning(

                    f"[EXCEL GROUP] "

                    f"{group_id} "

                    f"| items={len(group_items)}"
                )


                main_product = (
                    group_items[0]
                )

                fingerprint = self._build_vendor_fingerprint(
                    main_product
                )


                name = (

                    main_product.get(
                        "name"
                    ) or ""

                ).strip()


                description = (

                    main_product.get(
                        "description"
                    ) or ""
                )


                raw_category = (

                    main_product.get(
                        "category"
                    ) or ""

                ).lower()


                mapped_category = (
                    "General"
                )


                for key, val in CATEGORY_MAPPING.items():

                    if key in raw_category:

                        mapped_category = val

                        break


                category = category_obj.search([

                    (
                        'name',
                        '=',
                        mapped_category
                    ),

                    (
                        'parent_id',
                        '=',
                        parent_category.id
                    )

                ], limit=1)


                if not category:

                    category = category_obj.create({

                        'name':
                            mapped_category,

                        'parent_id':
                            parent_category.id
                    })


                # ================================================
                # FIND BY PRODUCT CODE FIRST
                # ================================================

                vendor_id = self.partner_id.id if self.partner_id else False

                product = False

                # =====================================================
                # 1. STRICT FINGERPRINT MATCH
                # =====================================================

                if (
                    'vendor_fingerprint' in product_obj._fields
                    and vendor_id
                ):

                    product = product_obj.search([

                        (
                            'vendor_fingerprint',
                            '=',
                            fingerprint
                        ),

                        (
                            'vendor_id',
                            '=',
                            vendor_id
                        )

                    ], limit=1)


                    if product:

                        _logger.warning(

                            f"[FINGERPRINT MATCH] "

                            f"{group_id} "

                            f"| vendor={vendor_id} "

                            f"| product_id={product.id}"
                        )

                 # =====================================================
                # 2. FALLBACK SKU MATCH
                # =====================================================

                if not product and vendor_id:

                    product = product_obj.search([

                        (
                            'default_code',
                            '=',
                            group_id
                        ),

                        (
                            'vendor_id',
                            '=',
                            vendor_id
                        )

                    ], limit=1)


                    if product:

                        _logger.warning(

                            f"[SKU MATCH] "

                            f"{group_id} "

                            f"| vendor={vendor_id} "

                            f"| product_id={product.id}"
                        )

                is_new_product = False

                if product:

                    _logger.warning(
                        f"[EXCEL DUPLICATE FOUND] "
                        f"{group_id} | vendor={vendor_id} | product_id={product.id}"
                    )

                else:
                    is_new_product = True


                # =================================================
                # CREATE PARENT
                # =================================================

                if is_new_product:

                    vals = {

                        'name': name,

                        'default_code':
                            group_id,

                        'description_sale':
                            description,

                       'type': 'consu',

                        'categ_id':
                            category.id,

                        'sale_ok': True,

                        'website_published':
                            False,

                        # =====================================
                        # SAVE VENDOR LINK
                        # =====================================

                        'vendor_id':
                            vendor_id,

                        'list_price':
                            self._safe_float(
                                main_product.get("price")
                            ),

                        'vendor_fingerprint': fingerprint,

                        'vendor_import_job_id': self.id,
                    }


                    image = main_product.get(
                        "image"
                    )


                    if image:

                        vals[
                            'image_1920'
                        ] = image


                    product = product_obj.create(
                        vals
                    )

                    # ✅ SAFE TRANSLATION CALL (PLUG-IN)
                    self._apply_product_translation(product)
                    created_count += 1


                    _logger.warning(

                        f"[EXCEL CREATED] "

                        f"{group_id} "

                        f"| vendor={vendor_id}"
                    )

                else:

                    merged_count += 1

                    # =====================================
                    # TRANSLATE EXISTING PRODUCT TOO
                    # =====================================

                    self._apply_product_translation(product)

                    _logger.warning(

                        f"[EXCEL EXISTING PRODUCT] "

                        f"{group_id} "

                        f"| vendor={vendor_id} "

                        f"| product_id={product.id}"
                    )
              

                # ==================================================
                # VARIANTS
                # ==================================================

                for idx, item in enumerate(group_items):

                    # =============================================
                    # DETECT ATTRIBUTE TYPE
                    # =============================================

                    variant_attribute_name = "Variant"

                    if item.get("color") or item.get("colour"):

                        variant_attribute_name = "Color"

                    elif item.get("material"):

                        variant_attribute_name = "Material"

                    elif item.get("size"):

                        variant_attribute_name = "Size"

                    elif item.get("capacity"):

                        variant_attribute_name = "Capacity"

                    elif item.get("style"):

                        variant_attribute_name = "Style"


                    # =============================================
                    # DETECT ATTRIBUTE VALUE
                    # =============================================

                    attr_value = str(

                        item.get("color")

                        or item.get("colour")

                        or item.get("material")

                        or item.get("size")

                        or item.get("variant")

                        or item.get("capacity")

                        or item.get("style")

                        or f"Variant {idx+1}"

                    ).strip()

                    if not attr_value:

                        detected_color = self._detect_basic_image_color(
                            item.get("image")
                        )

                        if detected_color:

                            variant_attribute_name = "Color"

                            attr_value = detected_color

                            _logger.warning(

                                f"[IMAGE COLOR FALLBACK] "

                                f"{detected_color}"
                            )

                        else:


                            attr_value = (

                                item.get("vendor_code")

                                or

                                item.get("primary_code")

                                or

                                f"Code {idx+1}"
                            )

                    _logger.warning(

                        f"[VARIANT DETECTED] "

                        f"{variant_attribute_name} "

                        f"= {attr_value}"
                    )


                    # =============================================
                    # ATTRIBUTE
                    # =============================================

                    attribute = attribute_obj.search([

                        (
                            'name',
                            '=',
                            variant_attribute_name
                        )

                    ], limit=1)


                    if not attribute:

                        attribute = attribute_obj.create({

                            'name': variant_attribute_name

                        })


                        _logger.warning(

                            f"[ATTRIBUTE CREATED] "

                            f"{variant_attribute_name}"
                        )


                    # =============================================
                    # ATTRIBUTE VALUE
                    # =============================================

                    value = attribute_value_obj.search([

                        (
                            'name',
                            '=',
                            attr_value
                        ),

                        (
                            'attribute_id',
                            '=',
                            attribute.id
                        )

                    ], limit=1)

                    if not value:

                        value = attribute_value_obj.create({

                            'name': attr_value,

                            'attribute_id': attribute.id
                        })


                        # =========================================
                        # TRANSLATE VARIANT VALUE
                        # =========================================

                        try:

                            for lang_code in ['ru_RU', 'az_AZ']:

                                translated_variant = self._force_translate(

                                    attr_value,

                                    lang_code
                                )


                                if translated_variant:

                                    value.with_context(
                                        lang=lang_code
                                    ).write({

                                        'name': translated_variant
                                    })


                                    _logger.warning(

                                        f"[VARIANT TRANSLATED] "

                                        f"{attr_value} "

                                        f"-> "

                                        f"{translated_variant} "

                                        f"({lang_code})"
                                    )

                        except Exception as e:

                            _logger.warning(

                                f"[VARIANT TRANSLATION ERROR] "

                                f"{str(e)}"
                            )


                        _logger.warning(

                            f"[ATTRIBUTE VALUE CREATED] "

                            f"{attr_value}"
                        )


                    # =============================================
                    # TEMPLATE ATTRIBUTE LINE
                    # =============================================

                    line = line_obj.search([

                        (
                            'product_tmpl_id',
                            '=',
                            product.id
                        ),

                        (
                            'attribute_id',
                            '=',
                            attribute.id
                        )

                    ], limit=1)


                    if not line:

                        line = line_obj.create({

                            'product_tmpl_id': product.id,

                            'attribute_id': attribute.id,

                            'value_ids': [

                                (
                                    6,
                                    0,
                                    [value.id]
                                )
                            ]
                        })


                        _logger.warning(

                            f"[VARIANT LINE CREATED] "

                            f"{group_id}"
                        )

                    else:

                        if value.id not in line.value_ids.ids:

                            line.value_ids = [

                                (
                                    4,
                                    value.id
                                )
                            ]


                            _logger.warning(

                                f"[VARIANT ADDED] "

                                f"{group_id} "

                                f"| {attr_value}"
                            )


                    # =============================================
                    # VARIANT IMAGE
                    # =============================================

                    variant_record = self.env[
                        'product.product'
                    ].search([

                        (
                            'product_tmpl_id',
                            '=',
                            product.id
                        ),

                        (
                            'product_template_attribute_value_ids.product_attribute_value_id',
                            '=',
                            value.id
                        )

                    ], limit=1)


                    if variant_record:

                        variant_image = item.get(
                            "image"
                        )


                        if variant_image:

                            variant_record.image_1920 = (
                                variant_image
                            )


                            _logger.warning(

                                f"[VARIANT IMAGE] "

                                f"{group_id} "

                                f"| {attr_value}"
                            )


                # =================================================
                # SAVE PROGRESS
                # =================================================

                self.excel_created_index = (
                    group_idx + 1
                )


                self._safe_commit_progress()
                

                _logger.warning(

                    f"[EXCEL SAVE] "

                    f"index="

                    f"{self.excel_created_index}"
                )


            except Exception as e:

                _logger.exception(

                    f"[EXCEL GROUP ERROR] "

                    f"group_idx={group_idx} "

                    f"| {str(e)}"
                )

                self.env.cr.rollback()


        # =====================================================
        # FINAL LOG
        # =====================================================

        _logger.warning(

            f"[EXCEL COMPLETE] "

            f"created={created_count} "

            f"| merged={merged_count}"
        )


        # ======================================================
        # NEXT STATE
        # ======================================================

        if self.excel_created_index >= len(grouped_keys):

            _logger.warning(

                "[EXCEL FLOW] "

                "GROUP BATCH COMPLETE"
            )

            # =========================================
            # FULL IMPORT COMPLETED
            # =========================================

            if self.is_excel_parsed:

                _logger.warning(
                    "[EXCEL IMPORT COMPLETE] ✅"
                )

                # =========================================
                # FINAL RESET
                # =========================================

                self.excel_created_index = 0

                self.excel_ai_index = 0

                self.ai_response = False

                self.state = 'done'

                # cleanup URL queue
                self.excel_url_processing = False

                self.excel_url_queue = False

                self.excel_url_index = 0

            # =========================================
            # MORE PARSE ROWS REMAIN
            # =========================================

            else:

                _logger.warning(

                    "[EXCEL FLOW] "

                    "RETURN TO excel_parsing"
                )

                # IMPORTANT:
                # KEEP CURRENT AI STATE
                # for next parse batch

                self.state = 'excel_parsing'

                _logger.warning(

                    "[EXCEL FLOW] "

                    f"NEXT PARSE INDEX="

                    f"{self.excel_parse_index}"
                )

        else:

            self.state = 'excel_creating'


        self._safe_commit_progress()


    #=====excel group url update====================================

    def _enrich_group_with_url_data(

        self,

        group_items,

        url_cache=None
    ):

        if url_cache is None:

            url_cache = {}

        group_url = ""

        # =====================================
        # FIND FIRST VALID GROUP URL
        # =====================================

        for item in group_items:

            possible_url = (
                item.get("url")
                or
                item.get("product_url")
                or
                ""
            ).strip()

            if possible_url:

                group_url = possible_url

                break

        # =====================================
        # NO URL FOUND
        # =====================================

        if not group_url:

            _logger.warning(

                "[URL ENRICHMENT SKIPPED] "

                "NO URL FOUND"
            )

            return {}

        _logger.warning(

            f"[URL GROUP FOUND] "

            f"{group_url}"
        )

        # =====================================
        # CACHE HIT
        # =====================================

        if group_url in url_cache:

            _logger.warning(

                f"[URL CACHE HIT] "

                f"{group_url}"
            )

            return url_cache[group_url]

        try:

            _logger.warning(

                f"[URL ENRICHMENT START] "

                f"{group_url}"
            )

            url_data = self._extract_url_product_data(
                group_url
            ) or {}

            # =====================================
            # EMPTY RESPONSE
            # =====================================

            if not url_data:

                _logger.warning(

                    f"[URL ENRICHMENT EMPTY] "

                    f"{group_url}"
                )

                url_cache[group_url] = {}

                return {}

            url_cache[group_url] = url_data

            _logger.warning(

                f"[URL ENRICHMENT SUCCESS] "

                f"{group_url} "

                f"| keys={list(url_data.keys())}"
            )

            return url_data

        except Exception as e:

            _logger.warning(

                f"[URL ENRICHMENT FAILED] "

                f"{group_url} "

                f"| {str(e)}"
            )

            return {}

    #====Excel variant mapping==================================
    def _detect_basic_image_color(self, image_data):

        try:

            import base64
            from io import BytesIO

            from PIL import Image

            img = Image.open(
                BytesIO(
                    base64.b64decode(image_data)
                )
            ).convert("RGB")

            img = img.resize((50, 50))

            colors = img.getcolors(
                50 * 50
            )

            if not colors:
                return False

            dominant = max(
                colors,
                key=lambda x: x[0]
            )[1]

            r, g, b = dominant


            # =====================================
            # BASIC COLOR MAPPING
            # =====================================

            if r > 200 and g > 200 and b > 200:
                return "White"

            if r < 60 and g < 60 and b < 60:
                return "Black"

            if r > 180 and g < 120 and b < 80:
                return "Orange"

            if r > 180 and g < 80 and b < 80:
                return "Red"

            if b > 150 and r < 120:
                return "Blue"

            if g > 140 and r < 120:
                return "Green"

            if r > 150 and g > 150 and b < 120:
                return "Yellow"

            return "Standard"

        except Exception as e:

            _logger.warning(

                f"[COLOR DETECTION FAILED] "

                f"{str(e)}"
            )

            return False


    #-----URL API FLOW-------------------------------------------

    def scrape_with_playwright(self):

        from playwright.sync_api import sync_playwright
        import subprocess
        import os

        _logger.warning(f"PLAYWRIGHT SCRAPE → {self.data_url}")

        #✅ CHECK IF BROWSER EXISTS FIRST
        browser_path = os.path.expanduser("~/.cache/ms-playwright")

        if not os.path.exists(browser_path):
            _logger.warning("PLAYWRIGHT → INSTALLING BROWSER (FIRST RUN)")

            try:
                subprocess.run(
                    ["python", "-m", "playwright", "install", "chromium"],
                    check=True
                )
                _logger.warning("PLAYWRIGHT BROWSER INSTALLED")
            except Exception as e:
                _logger.error(f"PLAYWRIGHT INSTALL FAILED → {str(e)}")
                return

        products = []

        with sync_playwright() as p:

            browser = p.chromium.launch(headless=True)
            page = browser.new_page()

            page.goto(self.data_url, timeout=60000)
            page.wait_for_timeout(5000)

            # ✅ Cookie handling
            try:
                page.locator("button:has-text('Accept')").click(timeout=3000)
            except:
                pass

            items = page.query_selector_all("a, div")

            _logger.warning(f"PLAYWRIGHT ELEMENTS FOUND → {len(items)}")

            for item in items[:200]:

                try:
                    text = item.inner_text().strip()

                    if not text or len(text) < 5:
                        continue

                    img_el = item.query_selector("img")
                    img_url = img_el.get_attribute("src") if img_el else None

                    if img_url and img_url.startswith("//"):
                        img_url = "https:" + img_url

                    img_base64 = None

                    if img_url:
                        try:
                            res = requests.get(img_url, timeout=10)
                            if res.status_code == 200:
                                img_base64 = base64.b64encode(res.content).decode("utf-8")
                        except:
                            pass

                    products.append({
                        "name": text[:120],
                        "image": img_base64
                    })

                except:
                    continue

            browser.close()

        _logger.warning(f"PLAYWRIGHT PRODUCTS → {len(products)}")

        if not products:
            _logger.error("PLAYWRIGHT FAILED → NO PRODUCTS")
            return

        pages = [{
            "page": 1,
            "text": "\n".join([p["name"] for p in products]),
            "images": [p["image"] for p in products if p["image"]]
        }]

        self.extracted_text = json.dumps(pages)

        _logger.warning(f"PLAYWRIGHT DONE → {len(products)} PRODUCTS")

    # =====================================================
    # CRON PROCESSOR
    # =====================================================
    def run_pending_jobs(self):

        from odoo import fields

        _logger.warning(
            "🔥 CRON HEARTBEAT → RUNNING"
        )


        active_states = [

            'draft',
            'review',

            'excel_parsing',
            'excel_ai',
            'excel_creating',

            'pdf_extracting',
            'pdf_ai',
            'pdf_creating',

            'url_scraping',
            'url_ai',
            'url_creating',

        ]


        # =================================================
        # LOAD JOBS
        # =================================================

        jobs = self.search(

            [('state', 'in', active_states)],

            order="id desc"
        )


        _logger.warning(

            f"[CRON] ACTIVE JOBS "

            f"→ {len(jobs)}"
        )


        # =================================================
        # REMOVE DUPLICATES
        # =================================================

        seen = {}

        duplicates = self.env[
            'vendor.import.job'
        ]

        for j in jobs:

            sig = (

                str(j.partner_id.id)

                + "_"

                + str(j.upload_signature)
            )
           
            if not sig:

                continue


            if sig not in seen:

                seen[sig] = j

            else:

                if j.id > seen[sig].id:

                    duplicates |= seen[sig]

                    seen[sig] = j

                else:

                    duplicates |= j


        if duplicates:

            _logger.warning(

                f"[CRON] REMOVING DUPLICATES "

                f"→ {len(duplicates)}"
            )

            try:

                duplicates.unlink()

                self.env.cr.commit()

                _logger.warning(
                    "[CRON] DUPLICATES REMOVED"
                )

            except Exception as e:

                _logger.exception(

                    f"[CRON ERROR] "

                    f"DUPLICATE DELETE FAILED "

                    f"→ {str(e)}"
                )


        # =================================================
        # RECOVER STALE LOCKS
        # =================================================

        stale_jobs = self.search([

            ('state', 'in', active_states),

            ('lock', '=', True)

        ])


        for stale in stale_jobs:

            try:

                delta = (

                    fields.Datetime.now()

                    - stale.write_date

                ).total_seconds()

            except Exception:

                delta = 0


            _logger.warning(

                f"[LOCK CHECK] "

                f"job={stale.id} "

                f"| seconds={delta}"
            )


            if delta > 60:

                _logger.warning(

                    f"[STALE LOCK RESET] "

                    f"job={stale.id}"
                )

                try:

                    stale.lock = False

                    self.env.cr.commit()

                except Exception as e:

                    _logger.exception(

                        f"[STALE LOCK ERROR] "

                        f"{str(e)}"
                    )

        # =================================================
        # CONTINUOUS QUEUE PROCESSOR
        # =================================================

        while True:

        # =================================================
        # GET NEXT JOB
        # =================================================

            job = self.search(

                [

                    ('state', 'in', active_states),

                    ('lock', '=', False)

                ],

                order="create_date asc, id asc",

                limit=1
            )


            if not job:

                _logger.warning(

                    "[CRON] NO AVAILABLE JOBS "
                    "(all locked or done)"
                )

                return


            # =================================================
            # PROCESS
            # =================================================

            try:

                # =============================================
                # LOCK
                # =============================================

                job.lock = True

                self.env.cr.commit()


                _logger.warning(

                    f"[CRON] JOB LOCKED "

                    f"| job={job.id}"
                )


                # =============================================
                # SAFER CHAIN
                # =============================================

                MAX_CHAIN = 1


                for step in range(MAX_CHAIN):

                    # =========================================
                    # REFRESH
                    # =========================================

                    try:

                        job.invalidate_cache()

                    except Exception:

                        pass


                    job = self.env[
                        'vendor.import.job'
                    ].browse(job.id)


                    _logger.warning(

                        f"[CHAIN] STEP "

                        f"{step + 1} "

                        f"| state={job.state}"
                    )


                    # =========================================
                    # STOP STATES
                    # =========================================

                    if job.state == 'done':

                        _logger.warning(

                            f"[CHAIN STOP] "

                            f"terminal state "

                            f"→ {job.state}"
                        )

                        break


                    # =========================================
                    # TRACK BEFORE
                    # =========================================

                    previous_state = (
                        job.state
                    )

                    previous_page = (
                        job.current_page or 0
                    )

                    previous_ai_page = (
                        job.last_ai_page or 0
                    )

                    previous_created = (
                        job.last_created_page or 0
                    )

                    previous_excel_ai = (
                        job.excel_ai_index or 0
                    )

                    previous_excel_created = (
                        job.excel_created_index or 0
                    )

                
                    previous_url_batch = (
                        job.url_batch_index or 0
                    )

                    previous_url_created = (
                        job.last_processed_product_index or 0
                    )
                

                    _logger.warning(

                        f"[CHAIN BEFORE] "

                        f"state={previous_state} "

                        f"| extract={previous_page} "

                        f"| ai={previous_ai_page} "

                        f"| create={previous_created} "

                        f"| excel_ai={previous_excel_ai} "

                        f"| excel_create={previous_excel_created}"
                    )


                    # =========================================
                    # PROCESS
                    # =========================================

                    try:

                        # =========================================
                        # REAL PROCESSING
                        # =========================================

                        job._process_step()

                        # =========================================
                        # SUCCESS RESET
                        # =========================================

                        self.env.cr.commit()

                    except Exception as e:

                        _logger.exception(

                            f"[PROCESS ERROR] "

                            f"job={job.id} "

                            f"| {str(e)}"
                        )

                        # =====================================
                        # INFRASTRUCTURE FAILURE DETECTION
                        # =====================================

                        dead_cursor = any(

                            x in str(e).lower()

                            for x in [

                                "cursor already closed",

                                "connection already closed",

                                "closed cursor",

                                "interfaceerror"
                            ]
                        )

                        # =====================================
                        # DO NOT CONSUME BUSINESS RETRIES
                        # =====================================

                        if dead_cursor:

                            _logger.warning(

                                f"[INFRA FAILURE] "

                                f"job={job.id} "

                                f"| retry preserved"
                            )

                            return

                        try:

                            # =====================================
                            # RETRY TRACKING
                            # =====================================

                            job.stage_retry_count += 1

                            _logger.warning(

                                f"[JOB RETRY] "

                                f"job={job.id} "

                                f"| retry={job.stage_retry_count}/8"
                            )

                            # =====================================
                            # TERMINAL FAILURE
                            # =====================================

                            if job.stage_retry_count >= 8:

                                _logger.error(

                                    f"[JOB FAILED PERMANENTLY] "

                                    f"job={job.id}"
                                )

                                job.state = 'failed'

                                job.failed_at = fields.Datetime.now()

                                job.failure_reason = str(e)

                                self.env.cr.commit()

                                # =================================
                                # FAILED EMAIL NOTIFICATION
                                # =================================

                                try:

                                    _logger.warning(

                                        f"[FAILED EMAIL] "

                                        f"START → job={job.id}"
                                    )

                                    job._send_failed_processing_email(

                                        error_message=str(e)
                                    )

                                    _logger.warning(

                                        f"[FAILED EMAIL] "

                                        f"COMPLETE → job={job.id}"
                                    )

                                except Exception as email_error:

                                    _logger.exception(

                                        f"[FAILED EMAIL ERROR] "

                                        f"{str(email_error)}"
                                    )

                            else:

                                # =================================
                                # RECOVERY STATE
                                # =================================

                                job.state = 'review'

                                self.env.cr.commit()

                        except Exception:

                            _logger.warning(

                                "[PROCESS ERROR] "

                                "FAILED SAVE FAILED"
                            )

                        break

                    # =========================================
                    # REFRESH AFTER
                    # =========================================

                    try:

                        job.invalidate_cache()

                    except Exception:

                        pass


                    job = self.env[
                        'vendor.import.job'
                    ].browse(job.id)


                    _logger.warning(

                        f"[CHAIN AFTER] "

                        f"state={job.state} "

                        f"| extract={job.current_page} "

                        f"| ai={job.last_ai_page} "

                        f"| create={job.last_created_page} "

                        f"| excel_ai={job.excel_ai_index} "

                        f"| excel_create={job.excel_created_index}"
                    )


                    # =========================================
                    # PROGRESS DETECTION
                    # =========================================

                    progress_detected = False


                    # PDF extract progress

                    if (

                        (job.current_page or 0)

                        >

                        previous_page

                    ):

                        progress_detected = True

                        _logger.warning(

                            f"[PROGRESS] PDF "

                            f"{previous_page}"

                            f" → "

                            f"{job.current_page}"
                        )


                    # PDF AI progress

                    if (

                        (job.last_ai_page or 0)

                        >

                        previous_ai_page

                    ):

                        progress_detected = True

                        _logger.warning(

                            f"[PROGRESS] PDF AI "

                            f"{previous_ai_page}"

                            f" → "

                            f"{job.last_ai_page}"
                        )


                    # PDF create progress

                    if (

                        (job.last_created_page or 0)

                        !=

                        previous_created

                    ):

                        progress_detected = True

                        _logger.warning(

                            f"[PROGRESS] PDF CREATE "

                            f"{previous_created}"

                            f" → "

                            f"{job.last_created_page}"
                        )


                    # PDF create state continuation

                    elif job.state == 'pdf_creating':

                        progress_detected = True

                        _logger.warning(

                            "[PROGRESS] PDF CREATE LOOP ACTIVE"
                        )
                

                    # Excel AI progress

                    if (

                        (job.excel_ai_index or 0)

                        >

                        previous_excel_ai

                    ):

                        progress_detected = True

                        _logger.warning(

                            f"[PROGRESS] EXCEL AI "

                            f"{previous_excel_ai}"

                            f" → "

                            f"{job.excel_ai_index}"
                        )


                    # Excel create progress

                    if (

                        (job.excel_created_index or 0)

                        >

                        previous_excel_created

                    ):

                        progress_detected = True

                        _logger.warning(

                            f"[PROGRESS] EXCEL CREATE "

                            f"{previous_excel_created}"

                            f" → "

                            f"{job.excel_created_index}"
                        )

                
                    # =========================================
                    # URL AI progress
                    # =========================================

                    if (

                        (job.url_batch_index or 0)

                        >

                        previous_url_batch

                    ):

                        progress_detected = True

                        _logger.warning(

                            f"[PROGRESS] URL AI "

                            f"{previous_url_batch}"

                            f" → "

                            f"{job.url_batch_index}"
                        )


                    # =========================================
                    # URL create progress
                    # =========================================

                    if (

                        (job.last_processed_product_index or 0)

                        >

                        previous_url_created

                    ):

                        progress_detected = True

                        _logger.warning(

                            f"[PROGRESS] URL CREATE "

                            f"{previous_url_created}"

                            f" → "

                            f"{job.last_processed_product_index}"
                        )


                    # =========================================
                    # APIFY WAIT STATE
                    # =========================================

                    if job.state == 'url_scraping':

                        progress_detected = True

                        _logger.warning(

                            "[PROGRESS] APIFY WAITING"
                        )
                


                    # state transition

                    if previous_state != job.state:

                        progress_detected = True

                        _logger.warning(

                            f"[PROGRESS] STATE "

                            f"{previous_state}"

                            f" → "

                            f"{job.state}"
                        )


                    # =========================================
                    # STOP IF NO PROGRESS
                    # =========================================

                    if not progress_detected:

                        _logger.warning(

                            "[CHAIN STOP] "

                            "NO PROGRESS DETECTED"
                        )

                        break


                    _logger.warning(

                        "[CHAIN CONTINUE] "

                        "PROGRESS DETECTED"
                    )


                    # =========================================
                    # COMMIT
                    # =========================================

                    try:

                        self.env.cr.commit()

                        _logger.warning(
                            "[CHAIN] COMMIT OK"
                        )

                    except Exception as e:

                        _logger.exception(

                            f"[CHAIN COMMIT ERROR] "

                            f"{str(e)}"
                        )

                        break


                _logger.warning(
                    "[CRON] PROCESS LOOP COMPLETE"
                )


            except Exception as e:

                _logger.exception(

                    f"[CRON FATAL ERROR] "

                    f"{str(e)}"
                )


                try:

                    self.env.cr.rollback()

                    _logger.warning(
                        "[CRON] ROLLBACK OK"
                    )

                except Exception:

                    _logger.warning(
                        "[CRON] ROLLBACK FAILED"
                    )


            finally:

                # =============================================
                # UNLOCK
                # =============================================

                try:

                    # =============================================
                    # DEAD CURSOR SAFETY
                    # =============================================

                    if self.env.cr.closed:

                        _logger.warning(

                            "[CRON] SKIP UNLOCK "

                            "CURSOR CLOSED"
                        )

                    else:

                        if (

                            job

                            and

                            job.exists()

                        ):

                            job.lock = False

                            self.env.cr.commit()


                            _logger.warning(

                                f"[CRON] JOB UNLOCKED "

                                f"| job={job.id}"
                            )

                except Exception:

                    _logger.warning(

                        "[CRON] UNLOCK FAILED"
                    )


            _logger.warning(
                "[CRON] RUN COMPLETE"
            )

            break


   #=============flask setup/installation=================== 
    def ping_flask_server(self):
      
        try:
            requests.get("https://pdf-extractor-staging.onrender.com", timeout=10)
            _logger.info("FLASK PING SUCCESS")
        except Exception:
            _logger.warning("FLASK PING FAILED")


    # #---------------clean_scraped_blocks-------------------------------
    def _clean_scraped_blocks(self, raw_blocks):
        """
        Clean Apify output before sending to AI
        """

        cleaned = []
        seen = set()

        for item in raw_blocks:

            text = (item.get("text") or "").strip()
            image = item.get("image")

            # ❌ REMOVE NOISE
            if not text:
                continue

            if len(text) < 15:
                continue

            if any(x in text.lower() for x in [
                "privacy", "cookie", "terms", "login",
                "menu", "navigation", "home"
            ]):
                continue

            # ❌ REMOVE DUPLICATES
            key = text[:120]  # allow more variation

            if key in seen:
                continue

            seen.add(key)

            # cleaned.append({
            #     "text": text,
            #     "image": image
            # })

            cleaned.append({
                "text": text,
                "image": image,
                "price": item.get("price", ""),
                "stock": item.get("stock", "")
            })

        return cleaned

    #---------------normalizer-------------------------------

    def _normalize_url_data(self, items):

        blocks = []

        for item in items:

            # =====================================================
            # FORMAT 1 → ORIGINAL WORKING FORMAT
            # =====================================================
            # {
            #   "text": "...",
            #   "image": "..."
            # }
            # =====================================================

            if item.get("text"):

                text = (item.get("text") or "").strip()
                image = item.get("image")

                # 🔥 STRICT VALIDATION
                if not text or len(text) < 5:
                    continue

                if (
                    image and
                    isinstance(image, str) and
                    not image.startswith("http")
                ):
                    image = None

                blocks.append({
                    "text": text,
                    "image": image,
                    "price": item.get("price", ""),
                    "stock": item.get("stock", "")
                })

                continue

            # =====================================================
            # FORMAT 2 → STRUCTURED FORMAT
            # =====================================================

            if item.get("type") == "PRODUCTS":

                for sub in item.get("items", []):

                    text = (
                        sub.get("text") or ""
                    ).strip()

                    image = sub.get("image")

                    if not text or len(text) < 5:
                        continue

                    if (
                        image and
                        isinstance(image, str) and
                        not image.startswith("http")
                    ):
                        image = None

                    # blocks.append({
                    #     "text": text,
                    #     "image": image
                    # })

                    blocks.append({
                        "text": text,
                        "image": image,
                        "price": sub.get("price", ""),
                        "stock": sub.get("stock", "")
                    })

            # =====================================================
            # DEBUG TYPES
            # =====================================================

            elif item.get("type") in [
                "EMPTY",
                "BLOCKED"
            ]:

                _logger.error(
                    f"URL DEBUG → "
                    f"{item.get('reason')}"
                )

        _logger.warning(f"NORMALIZED BLOCKS → {len(blocks)}")

        # =====================================================
        # 🔥 SPLIT INTO MULTIPLE PAGES (CRITICAL FIX)
        # =====================================================

        PAGE_SIZE = 20  # 🔥 prevents AI overload

        pages = []

        for i in range(0, len(blocks), PAGE_SIZE):

            chunk = blocks[i:i + PAGE_SIZE]

            pages.append({
                "page": len(pages) + 1,
                "blocks": chunk
            })

        _logger.warning(f"NORMALIZED PAGES → {len(pages)}")

        return pages
    

    #======apify url fetch/scrapp products=====================
    
    def _run_apify_actor(self, url):

        token = self.env['ir.config_parameter'].sudo().get_param('apify.api_token')

        if not token:
            raise Exception("Apify API token not configured")

        #ACTOR_ID = "selectad~my-actor"
        ACTOR_ID = "princ_adex~my-actor"

        # =====================================================
        # 🔥 STEP 1: START ACTOR (ONLY IF NOT STARTED)
        # =====================================================

        if not getattr(self, "apify_run_id", False):

            run_url = f"https://api.apify.com/v2/acts/{ACTOR_ID}/runs?token={token}"

            payload = {
                "startUrls": [{"url": url}]
            }

            headers = {
                "Content-Type": "application/json"
            }

            response = requests.post(run_url, json=payload, headers=headers, timeout=30)

            if response.status_code != 201:
                raise Exception(f"Apify run failed: {response.text}")

            run_data = response.json()

            # ✅ SAVE FOR NEXT CRON
            self.apify_run_id = run_data["data"]["id"]
            self.apify_dataset_id = run_data["data"]["defaultDatasetId"]

            _logger.warning(f"APIFY STARTED → RUN ID {self.apify_run_id}")

            # 🔥 IMPORTANT: STOP HERE (NON-BLOCKING)
            return None

        # =====================================================
        # 🔥 STEP 2: CHECK STATUS
        # =====================================================

        status_url = f"https://api.apify.com/v2/actor-runs/{self.apify_run_id}?token={token}"

        status_res = requests.get(status_url, timeout=20).json()
        status = status_res["data"]["status"]

        _logger.warning(f"APIFY STATUS → {status}")

        if status in ["RUNNING", "READY"]:
            _logger.warning("APIFY STILL RUNNING → WAIT NEXT CRON")
            return None

        if status in ["FAILED", "ABORTED", "TIMED-OUT"]:
            raise Exception(f"Apify run failed with status: {status}")

        # =====================================================
        # 🔥 STEP 3: FETCH DATA (ONLY WHEN DONE)
        # =====================================================

        dataset_url = f"https://api.apify.com/v2/datasets/{self.apify_dataset_id}/items"

        params = {
            "token": token,
            "limit": 1000,
            "clean": "true"
        }

        dataset_res = requests.get(dataset_url, params=params, timeout=30)

        if dataset_res.status_code != 200:
            raise Exception(f"Failed to fetch dataset: {dataset_res.text}")

        data = dataset_res.json()

        _logger.warning(f"APIFY ITEMS FETCHED → {len(data)}")

        if not data:
            _logger.warning("APIFY RETURNED EMPTY → MARK JOB AS DONE")

            self.state = 'done'   # 🔥 STOP LOOP COMPLETELY
            self._safe_commit_progress()
            return

        # 🔥 CLEAN UP (IMPORTANT)
        self.apify_run_id = False
        self.apify_dataset_id = False

        return data


    #=======validation===================
    def validate_ai_output(products):
        for p in products:
            if "variants" in p:
                if not isinstance(p["variants"], list):
                    p["variants"] = []

                for v in p["variants"]:
                    if "attributes" not in v:
                        v["attributes"] = {"Variant": "Default"}

        return products
    
    #=======keep cron alive================
    def keep_alive(self):
        _logger.warning("KEEP ALIVE PING")

   
   #=========gloat numbers=============
    def _safe_float(self, value):

        try:

            if value is None:
                return 0.0

            value = str(value)

            value = value.replace('$', '')
            value = value.replace('€', '')
            value = value.replace('£', '')
            value = value.replace(',', '')

            return float(value.strip())

        except:
            return 0.0

    #======product translate ==========================
    
    def translate_global_views(self, target_lang):

        from openai import OpenAI

        api_key = self.env['ir.config_parameter'].sudo().get_param('openai.api.key')

        if not api_key:
            _logger.warning("❌ Missing OpenAI API key")
            return

        client = OpenAI(api_key=api_key)

        # 🔥 FETCH SOURCE STRINGS FROM VIEWS (SAFE WAY)
        views = self.env['ir.ui.view'].sudo().search([
            ('arch_db', '!=', False)
        ], limit=20)

        _logger.warning(f"🌍 GLOBAL VIEW TRANSLATION START → {target_lang}")
        _logger.warning(f"🔍 Views found → {len(views)}")

        for view in views:

            try:
                text = view.name or ''
                if not text:
                    continue

                prompt = f"""
                Translate to {target_lang}:

                {text}
                """

                response = client.responses.create(
                    model="gpt-4.1-mini",
                    input=prompt
                )

                translated = response.output_text.strip()

                if translated:
                    view.with_context(lang=target_lang).write({
                        'name': translated
                    })

                    _logger.warning(f"✅ VIEW {view.name} → {translated}")

            except Exception as e:
                _logger.warning(f"❌ Failed: {str(e)}")

    #===============pdf price helper====================
    def _safe_parse_price(self, value):

        try:

            import re

            if not value:
                return 1.0

            cleaned = re.sub(
                r'[^0-9.,]',
                '',
                str(value)
            )

            cleaned = cleaned.replace(',', '.')

            return float(cleaned)

        except Exception:

            return 1.0
